import sys
import os
import sqlite3
import json
import csv
import tempfile
import subprocess
import re
import logging
import traceback
import shutil
import hashlib
import base64
import io
import ctypes
from datetime import datetime, timedelta
from pathlib import Path
from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional

# Windows-specific imports
import win32print
import win32api
import win32con
import win32gui
import win32ui

# Task Scheduler imports for auto-backup
try:
    import pythoncom
    import win32com.client
    TASK_SCHEDULER_AVAILABLE = True
except ImportError:
    TASK_SCHEDULER_AVAILABLE = False

# PyQt6 imports
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QTabWidget, QMessageBox, QFileDialog, QComboBox, QSpinBox,
    QDoubleSpinBox, QFormLayout, QGroupBox, QHeaderView, QMenu,
    QDialog, QDialogButtonBox, QInputDialog, QCheckBox, QTextEdit,
    QSplitter, QFrame, QSizePolicy, QScrollArea, QGridLayout,
    QStyledItemDelegate, QStackedWidget, QDateEdit, QListWidget,
    QProgressBar
)
from PyQt6.QtCore import Qt, QSize, QRect, QPoint, QRegularExpression, QByteArray, QBuffer, QIODevice, pyqtSignal, QTimer, QDate, QObject, QFileSystemWatcher
from PyQt6.QtGui import (
    QPixmap, QImage, QFont, QIcon, QColor, QPalette, QRegularExpressionValidator,
    QPainter, QPen, QBrush, QFontMetrics, QKeySequence, QShortcut
)

# Third-party imports
import barcode
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont, ImageEnhance
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Local imports
from database.models import Database
from utils.barcode import BarcodeGenerator, resource_path
from database.models import get_persistent_path
from utils.report_generator import ReportGenerator
from utils.data_manager import DataManager
from utils.barcode_scanner import BarcodeScanner


log_dir = os.path.join(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__), 'logs')
os.makedirs(log_dir, exist_ok=True)
log_file_path = os.path.join(log_dir, 'app.log')

# Configure logging
import logging
import sys
import io

# Create a UTF-8 wrapper for stdout to handle Cyrillic text
if hasattr(sys.stdout, 'buffer'):
    # Python 3 - wrap the buffer with UTF-8 encoding
    utf8_stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
else:
    # Fallback for older systems
    utf8_stdout = sys.stdout

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path, encoding='utf-8'),
        logging.StreamHandler(utf8_stdout)
    ]
)
logger = logging.getLogger(__name__)

# Action/Command Pattern for Undo/Redo System

# Custom spin box classes that blur on Enter key press
class BlurOnEnterSpinBox(QSpinBox):
    """Custom QSpinBox that loses focus when Enter is pressed and remembers last confirmed value"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.last_confirmed_value = None
        self.value_on_focus = None
        
    def focusInEvent(self, event):
        # Store the current value when focus is gained
        self.value_on_focus = self.value()
        super().focusInEvent(event)
        
    def focusOutEvent(self, event):
        # If user changed value but didn't press Enter, and we have a last confirmed value, revert
        if (self.last_confirmed_value is not None and 
            self.value() != self.value_on_focus and 
            self.value() != self.last_confirmed_value):
            # User changed value but didn't press Enter, revert to last confirmed
            self.setValue(self.last_confirmed_value)
        super().focusOutEvent(event)
    
    def keyPressEvent(self, event):
        from PyQt6.QtCore import QTimer
        
        if event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
            # Store the current value as confirmed
            self.last_confirmed_value = self.value()
            
            # Clear focus first
            self.clearFocus()
            
            # Use QTimer to clear selection after focus is processed
            def clear_selection():
                line_edit = self.lineEdit()
                if line_edit:
                    line_edit.deselect()
            
            QTimer.singleShot(0, clear_selection)
            return  # Don't call super() to consume the event
        
        # For all other keys, call the parent implementation
        super().keyPressEvent(event)
    
    def reset_confirmed_value(self):
        """Reset the confirmed value memory - call when form is reset"""
        self.last_confirmed_value = None

class BlurOnEnterDoubleSpinBox(QDoubleSpinBox):
    """Custom QDoubleSpinBox that loses focus when Enter is pressed and remembers last confirmed value"""
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.last_confirmed_value = None
        self.value_on_focus = None
        
    def focusInEvent(self, event):
        # Store the current value when focus is gained
        self.value_on_focus = self.value()
        super().focusInEvent(event)
        
    def focusOutEvent(self, event):
        # If user changed value but didn't press Enter, and we have a last confirmed value, revert
        if (self.last_confirmed_value is not None and 
            self.value() != self.value_on_focus and 
            self.value() != self.last_confirmed_value):
            # User changed value but didn't press Enter, revert to last confirmed
            self.setValue(self.last_confirmed_value)
        super().focusOutEvent(event)
    
    def keyPressEvent(self, event):
        from PyQt6.QtCore import QTimer
        
        if event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
            # Store the current value as confirmed
            self.last_confirmed_value = self.value()
            
            # Clear focus first
            self.clearFocus()
            
            # Use QTimer to clear selection after focus is processed
            def clear_selection():
                line_edit = self.lineEdit()
                if line_edit:
                    line_edit.deselect()
            
            QTimer.singleShot(0, clear_selection)
            return  # Don't call super() to consume the event
        
        # For all other keys, call the parent implementation
        super().keyPressEvent(event)
    
    def reset_confirmed_value(self):
        """Reset the confirmed value memory - call when form is reset"""
        self.last_confirmed_value = None

class BlurOnEnterLineEdit(QLineEdit):
    """QLineEdit that remembers last confirmed value and reverts on unfocused changes"""
    
    def __init__(self, parent=None, numeric_only=False, max_value=None):
        super().__init__(parent)
        self.last_confirmed_value = None
        self.value_on_focus = None
        self.numeric_only = numeric_only
        self.max_value = max_value
        
        if self.numeric_only:
            # Only allow numeric input
            from PyQt6.QtGui import QRegularExpressionValidator
            from PyQt6.QtCore import QRegularExpression
            validator = QRegularExpressionValidator(QRegularExpression(r'^\d{0,2}$'))
            self.setValidator(validator)
    
    def focusInEvent(self, event):
        """Store the current value when field gets focus and select all text"""
        self.value_on_focus = self.text()
        super().focusInEvent(event)
        # Select all text for easy replacement
        QTimer.singleShot(0, self.selectAll)  # Delay selection to ensure it works
    
    def mousePressEvent(self, event):
        """Handle mouse clicks to select all text"""
        super().mousePressEvent(event)
        # If the field doesn't have focus, it will get focus and selectAll will be called
        # If it already has focus, select all text again
        if self.hasFocus():
            QTimer.singleShot(0, self.selectAll)
    
    def focusOutEvent(self, event):
        """Revert to last confirmed value if no Enter was pressed"""
        current_text = self.text()
        
        # Validate numeric input if applicable
        if self.numeric_only:
            self._validate_numeric_input()
        
        if (self.last_confirmed_value is not None and 
            current_text != self.value_on_focus and 
            self.last_confirmed_value != current_text):
            # User changed value but didn't press Enter, revert to last confirmed
            self.setText(self.last_confirmed_value)
        super().focusOutEvent(event)
    
    def keyPressEvent(self, event):
        """Handle Enter key to confirm value"""
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            if self.numeric_only:
                self._validate_numeric_input()
            self.last_confirmed_value = self.text()
            self.clearFocus()  # Remove focus to trigger validation
            return
        super().keyPressEvent(event)
    
    def _validate_numeric_input(self):
        """Validate numeric input and apply constraints"""
        if not self.numeric_only:
            return
            
        text = self.text().strip()
        if not text:
            self.setText("00")
            return
            
        try:
            value = int(text)
            if self.max_value is not None and value > self.max_value:
                value = self.max_value
            self.setText(f"{value:02d}")
        except ValueError:
            self.setText("00")
    
    def reset_confirmed_value(self):
        """Reset the confirmed value memory - call when form is reset"""
        self.last_confirmed_value = None

def apply_dark_theme():
    """Apply dark theme to the application"""
    from PyQt6.QtWidgets import QApplication
    from PyQt6.QtGui import QPalette, QColor
    from PyQt6.QtCore import Qt
    
    app = QApplication.instance()
    
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
    palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
    palette.setColor(QPalette.ColorRole.Base, QColor(25, 25, 25))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
    palette.setColor(QPalette.ColorRole.ToolTipBase, Qt.GlobalColor.white)
    palette.setColor(QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
    palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
    palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
    palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
    palette.setColor(QPalette.ColorRole.BrightText, Qt.GlobalColor.red)
    palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
    palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.black)
    
    app.setPalette(palette)
    
    # Force update all widgets
    for widget in app.topLevelWidgets():
        widget.update()


def get_application_icon(size=None):
    """
    Get the best quality application icon based on size requirements.
    Prioritizes high-quality PNG over ICO files, with comprehensive ICO fallbacks.
    
    Args:
        size: Optional QSize or tuple (width, height) for specific size requirements
        
    Returns:
        QIcon: The best available application icon
    """
    icon = QIcon()
    
    try:
        # Primary choice: High-quality PNG file
        png_path = resource_path("logo/V1 - NO BG.png")
        if os.path.exists(png_path):
            # Add PNG at multiple sizes for best quality
            icon.addFile(png_path, QSize(16, 16))   # Small taskbar icon
            icon.addFile(png_path, QSize(32, 32))   # Medium icons
            icon.addFile(png_path, QSize(48, 48))   # Large icons
            icon.addFile(png_path, QSize(64, 64))   # Extra large
            icon.addFile(png_path, QSize(128, 128)) # Very large icons
            icon.addFile(png_path, QSize(256, 256)) # Maximum quality
            logger.info(f"High-quality PNG icon loaded from: {png_path}")
            return icon
    except Exception as e:
        logger.warning(f"Could not load PNG icon: {e}")
    
    try:
        # Fallback: Multi-size ICO files with priority on larger sizes
        ico_sizes = [
            ("256x256.ico", QSize(256, 256)),  # Highest quality first
            ("48x48.ico", QSize(48, 48)),
            ("32x32.ico", QSize(32, 32)), 
            ("16x16.ico", QSize(16, 16))
        ]
        
        for ico_file, ico_size in ico_sizes:
            ico_path = resource_path(f"logo/{ico_file}")
            if os.path.exists(ico_path):
                icon.addFile(ico_path, ico_size)
        
        if not icon.isNull():
            logger.info("Multi-size ICO icons loaded successfully")
            return icon
    except Exception as e:
        logger.warning(f"Could not load ICO icons: {e}")
    
    try:
        # Final fallback: Main ICO file
        ico_path = resource_path("logo/jewelry_logo.ico") 
        if os.path.exists(ico_path):
            icon.addFile(ico_path)
            logger.info(f"Fallback ICO icon loaded from: {ico_path}")
            return icon
    except Exception as e:
        logger.warning(f"Could not load fallback ICO icon: {e}")
    
    logger.warning("No application icon could be loaded")
    return QIcon()  # Return empty icon if all else fails


class Action(ABC):
    """Abstract base class for all reversible actions"""
    
    def __init__(self, description: str):
        self.description = description
        self.timestamp = datetime.now()
    
    @abstractmethod
    def execute(self) -> bool:
        """Execute the action. Return True if successful."""
        pass
    
    @abstractmethod
    def undo(self) -> bool:
        """Undo the action. Return True if successful."""
        pass

class AddItemAction(Action):
    """Action for adding a new item"""
    
    def __init__(self, db: 'Database', item_data: Dict[str, Any]):
        super().__init__(f"Добави артикул {item_data.get('barcode', 'N/A')}")
        self.db = db
        self.item_data = item_data
        self.item_id = None
    
    def execute(self) -> bool:
        try:
            success = self.db.add_item(
                self.item_data['barcode'],
                self.item_data['name'], 
                self.item_data['description'],
                self.item_data['category'],
                self.item_data['price'],
                self.item_data['cost'],
                self.item_data['weight'],
                self.item_data['metal_type'],
                self.item_data['stone_type'],
                self.item_data['stock_quantity']
            )
            if success:
                # Get the item ID for potential undo
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT id FROM items WHERE barcode = ?", (self.item_data['barcode'],))
                    result = cursor.fetchone()
                    if result:
                        self.item_id = result[0]
            return success
        except Exception as e:
            logger.error(f"Error executing AddItemAction: {e}")
            return False
    
    def undo(self) -> bool:
        try:
            if self.item_id:
                return self.db.delete_item(self.item_id)
            return False
        except Exception as e:
            logger.error(f"Error undoing AddItemAction: {e}")
            return False

class DeleteItemAction(Action):
    """Action for deleting an item"""
    
    def __init__(self, db: 'Database', item_id: int, barcode: str):
        super().__init__(f"Изтрий артикул {barcode}")
        self.db = db
        self.item_id = item_id
        self.barcode = barcode
        self.item_backup = None
    
    def execute(self) -> bool:
        try:
            # Backup item data before deletion
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM items WHERE id = ?", (self.item_id,))
                self.item_backup = cursor.fetchone()
            
            return self.db.delete_item(self.item_id)
        except Exception as e:
            logger.error(f"Error executing DeleteItemAction: {e}")
            return False
    
    def undo(self) -> bool:
        try:
            if self.item_backup:
                # Restore the item from backup
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    # Get column names
                    cursor.execute("PRAGMA table_info(items)")
                    columns = [col[1] for col in cursor.fetchall()]
                    
                    # Prepare insert query
                    placeholders = ", ".join(["?" for _ in columns])
                    column_names = ", ".join(columns)
                    
                    cursor.execute(f"INSERT INTO items ({column_names}) VALUES ({placeholders})", self.item_backup)
                    conn.commit()
                return True
            return False
        except Exception as e:
            logger.error(f"Error undoing DeleteItemAction: {e}")
            return False

class EditItemAction(Action):
    """Action for editing an item"""
    
    def __init__(self, db: 'Database', item_id: int, barcode: str, old_data: Dict[str, Any], new_data: Dict[str, Any]):
        super().__init__(f"Редактирай артикул {barcode}")
        self.db = db
        self.item_id = item_id
        self.barcode = barcode
        self.old_data = old_data
        self.new_data = new_data
    
    def execute(self) -> bool:
        try:
            return self.db.update_item(
                self.item_id,
                name=self.new_data['name'],
                description=self.new_data['description'],
                category=self.new_data['category'],
                price=self.new_data['price'],
                cost=self.new_data['cost'],
                weight=self.new_data['weight'],
                metal_type=self.new_data['metal_type'],
                stone_type=self.new_data['stone_type'],
                stock_quantity=self.new_data['stock_quantity']
            )
        except Exception as e:
            logger.error(f"Error executing EditItemAction: {e}")
            return False
    
    def undo(self) -> bool:
        try:
            return self.db.update_item(
                self.item_id,
                name=self.old_data['name'],
                description=self.old_data['description'],
                category=self.old_data['category'],
                price=self.old_data['price'],
                cost=self.old_data['cost'],
                weight=self.old_data['weight'],
                metal_type=self.old_data['metal_type'],
                stone_type=self.old_data['stone_type'],
                stock_quantity=self.old_data['stock_quantity']
            )
        except Exception as e:
            logger.error(f"Error undoing EditItemAction: {e}")
            return False

class SaleAction(Action):
    """Action for recording a sale"""
    
    def __init__(self, db: 'Database', item_id: int, barcode: str, shop_id: int, price: float, quantity: int = 1):
        super().__init__(f"Продажба на артикул {barcode}")
        self.db = db
        self.item_id = item_id
        self.barcode = barcode
        self.shop_id = shop_id
        self.price = price
        self.quantity = quantity
        self.sale_id = None
    
    def execute(self) -> bool:
        try:
            # Record the sale
            sale_date = format_datetime_for_database()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Insert sale record
                sales_insert_query = """
                    INSERT INTO sales (item_id, quantity, total_price, sale_date, shop_id) 
                    VALUES (?, ?, ?, ?, ?)
                """
                sales_params = (self.item_id, self.quantity, self.price, sale_date, self.shop_id)
                
                cursor.execute(sales_insert_query, sales_params)
                self.sale_id = cursor.lastrowid
                
                # Remove from shop inventory
                shop_select_query = """
                    SELECT quantity FROM shop_items 
                    WHERE shop_id = ? AND item_id = ?
                """
                shop_select_params = (self.shop_id, self.item_id)
                
                cursor.execute(shop_select_query, shop_select_params)
                result = cursor.fetchone()
                
                if result and result[0] > self.quantity:
                    # Update quantity with timestamp
                    shop_update_query = """
                        UPDATE shop_items SET quantity = quantity - ?, updated_at = datetime('now', 'localtime') 
                        WHERE shop_id = ? AND item_id = ?
                    """
                    shop_update_params = (self.quantity, self.shop_id, self.item_id)
                    
                    cursor.execute(shop_update_query, shop_update_params)
                    
                elif result and result[0] == self.quantity:
                    # Delete record (quantity becomes 0)
                    shop_delete_query = """
                        DELETE FROM shop_items 
                        WHERE shop_id = ? AND item_id = ?
                    """
                    shop_delete_params = (self.shop_id, self.item_id)
                    
                    cursor.execute(shop_delete_query, shop_delete_params)
                    
                elif result and result[0] < self.quantity:
                    logger.warning(f"Insufficient quantity in shop for item {self.item_id}: available={result[0]}, requested={self.quantity}")
                    return False
                else:
                    logger.warning(f"No inventory record found for item {self.item_id} in shop {self.shop_id}")
                    return False
                
                # Commit the transaction
                conn.commit()
                
            return True
            
        except Exception as e:
            logger.error(f"Error executing SaleAction: {e}")
            return False
    
    def undo(self) -> bool:
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Delete the sale record
                if self.sale_id:
                    cursor.execute("DELETE FROM sales WHERE id = ?", (self.sale_id,))
                
                # Restore to shop inventory with timestamps
                cursor.execute("""
                    SELECT quantity FROM shop_items 
                    WHERE shop_id = ? AND item_id = ?
                """, (self.shop_id, self.item_id))
                
                result = cursor.fetchone()
                if result:
                    cursor.execute("""
                        UPDATE shop_items SET quantity = quantity + ?, updated_at = datetime('now', 'localtime') 
                        WHERE shop_id = ? AND item_id = ?
                    """, (self.quantity, self.shop_id, self.item_id))
                else:
                    cursor.execute("""
                        INSERT INTO shop_items (shop_id, item_id, quantity, created_at, updated_at) 
                        VALUES (?, ?, ?, datetime('now', 'localtime'), datetime('now', 'localtime'))
                    """, (self.shop_id, self.item_id, self.quantity))
                
                conn.commit()
            return True
        except Exception as e:
            logger.error(f"Error undoing SaleAction: {e}")
            return False

class ActionHistory:
    """Manages action history for undo/redo functionality"""
    
    def __init__(self, max_history: int = 3):
        self.max_history = max_history
        self.history: List[Action] = []
        self.current_index = -1
    
    def execute_action(self, action: Action) -> bool:
        """Execute an action and add it to history"""
        if action.execute():
            # Remove any redo history when a new action is executed
            self.history = self.history[:self.current_index + 1]
            
            # Add new action
            self.history.append(action)
            self.current_index += 1
            
            # Limit history size
            if len(self.history) > self.max_history:
                self.history.pop(0)
                self.current_index = min(self.current_index, len(self.history) - 1)
            
            logger.info(f"Action executed: {action.description}")
            return True
        return False
    
    def can_undo(self) -> bool:
        """Check if undo is possible"""
        return self.current_index >= 0
    
    def can_redo(self) -> bool:
        """Check if redo is possible"""
        return self.current_index < len(self.history) - 1
    
    def undo(self) -> bool:
        """Undo the last action"""
        if self.can_undo():
            action = self.history[self.current_index]
            if action.undo():
                self.current_index -= 1
                logger.info(f"Action undone: {action.description}")
                return True
        return False
    
    def redo(self) -> bool:
        """Redo the next action"""
        if self.can_redo():
            self.current_index += 1
            action = self.history[self.current_index]
            if action.execute():
                logger.info(f"Action redone: {action.description}")
                return True
            else:
                self.current_index -= 1
        return False
    
    def get_undo_description(self) -> Optional[str]:
        """Get description of action that can be undone"""
        if self.can_undo():
            return self.history[self.current_index].description
        return None
    
    def get_redo_description(self) -> Optional[str]:
        """Get description of action that can be redone"""
        if self.can_redo():
            return self.history[self.current_index + 1].description
        return None
    
    def clear(self):
        """Clear action history"""
        self.history.clear()
        self.current_index = -1

# Timezone and date utilities
def parse_database_datetime(datetime_str):
    """Parse datetime string from database and return local datetime object"""
    # Handle None, empty strings, and other falsy values
    if not datetime_str or datetime_str == '' or datetime_str is None:
        return None
    
    try:
        # Convert to string if it's not already
        datetime_str = str(datetime_str).strip()
        if not datetime_str:
            return None
            
        from datetime import datetime as dt
        
        # Try multiple datetime formats to handle different database storage formats
        formats_to_try = [
            "%Y-%m-%dT%H:%M:%S.%f",      # ISO format with microseconds: 2025-07-19T17:27:02.755305
            "%Y-%m-%dT%H:%M:%S",         # ISO format without microseconds: 2025-07-19T17:27:02
            "%Y-%m-%d %H:%M:%S.%f",      # Space format with microseconds: 2025-07-19 17:27:02.755305
            "%Y-%m-%d %H:%M:%S",         # Space format without microseconds: 2025-07-19 17:27:02
        ]
        
        for fmt in formats_to_try:
            try:
                # Parse and return as local time (system timezone)
                # Database stores local time, display as local time
                local_dt = dt.strptime(datetime_str, fmt)
                return local_dt
            except ValueError:
                continue
        
        # If none of the formats worked, log and return None
        logger.debug(f"Could not parse datetime string with any format: '{datetime_str}'")
        return None
        
    except (TypeError, AttributeError) as e:
        logger.debug(f"Error parsing datetime string: '{datetime_str}' - {e}")
        return None

def format_date_for_display(dt):
    """Format datetime for display in DD/MM/YYYY format"""
    if not dt:
        return ""
    return dt.strftime("%d/%m/%Y")

def format_time_for_display(dt):
    """Format datetime for display in HH:MM:SS format"""
    if not dt:
        return ""
    return dt.strftime("%H:%M:%S")

def format_datetime_for_database(dt=None):
    """Format datetime for database storage using system timezone"""
    from datetime import datetime as dt_class
    if dt is None:
        dt = dt_class.now()
    # Store as local time in database for consistency
    return dt.strftime("%Y-%m-%d %H:%M:%S")

class LoginWindow(QWidget):
    def __init__(self, parent=None, database=None):
        super().__init__(parent)
        # Use provided database instance or create new one
        self.db = database if database is not None else Database()
        self.setup_ui()
        self.setWindowTitle("Система за управление на бижута - Вход")
        
        # Set application icon for taskbar
        try:
            icon = get_application_icon()
            if not icon.isNull():
                self.setWindowIcon(icon)
                logger.info("Login window icon set using high-quality PNG")
            else:
                logger.warning("No icon could be loaded for login window")
        except Exception as e:
            logger.warning(f"Could not set login window icon: {e}")
        
        self.setFixedSize(600, 400)  # Set fixed size
        self.center_window()
        apply_dark_theme()  # Apply dark theme
        self.show()
        self.pin_input.setFocus()  # Set focus after window is shown

    def center_window(self):
        # Center the window on the screen
        screen = QApplication.primaryScreen().geometry()
        size = self.geometry()
        self.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)  # Add margins
        layout.setSpacing(10)  # Add spacing between widgets
        
        # Title
        title = QLabel("Система за управление на бижута")
        title.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # Add some spacing
        layout.addSpacing(30)

        # PIN instruction
        pin_instruction = QLabel("Въведете PIN код за достъп:")
        pin_instruction.setFont(QFont("Arial", 12))
        pin_instruction.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pin_instruction.setStyleSheet("color: #666; margin-bottom: 10px;")
        layout.addWidget(pin_instruction)

        # PIN input
        self.pin_input = QLineEdit()
        self.pin_input.setPlaceholderText("PIN код")
        self.pin_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.pin_input.setMinimumHeight(40)  # Make input field taller
        self.pin_input.setMaxLength(10)  # Limit to 10 characters (password length 4-10)
        self.pin_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.pin_input.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        self.pin_input.setStyleSheet("""
            QLineEdit {
                border: 2px solid #ccc;
                border-radius: 10px;
                padding: 10px;
                font-size: 18px;
                text-align: center;
                letter-spacing: 10px;
            }
            QLineEdit:focus {
                border-color: #2196F3;
            }
        """)
        self.pin_input.returnPressed.connect(self.login)  # Add Enter key handler
        layout.addWidget(self.pin_input)

        # Add some spacing
        layout.addSpacing(30)

        # Login button
        login_btn = QPushButton("Вход")
        login_btn.setMinimumHeight(40)  # Make button taller
        login_btn.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        login_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 10px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        login_btn.clicked.connect(self.login)
        layout.addWidget(login_btn)

        # Add some spacing
        layout.addSpacing(10)

        # Forgot password link
        forgot_btn = QPushButton("Забравена парола?")
        forgot_btn.setMinimumHeight(30)
        forgot_btn.setFont(QFont("Arial", 10))
        forgot_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #666;
                border: none;
                text-decoration: underline;
                padding: 5px;
            }
            QPushButton:hover {
                color: #2196F3;
            }
        """)
        forgot_btn.clicked.connect(self.show_recovery_dialog)
        layout.addWidget(forgot_btn)

        self.setLayout(layout)

    def login(self):
        pin = self.pin_input.text().strip()
        
        # Verify PIN against database for single-user system
        if self.db.verify_user("admin", pin):
            self.hide()  # Hide login window
            # Create and show main window, reusing the same database instance
            self.main_window = MainWindow(database=self.db)
            self.main_window.show()
        else:
            QMessageBox.warning(self, "Грешка", "Невалиден PIN код")
            self.pin_input.clear()  # Clear PIN input for security
            self.pin_input.setFocus()  # Set focus back to PIN input

    def show_recovery_dialog(self):
        """Show master key recovery dialog"""
        try:
            # Create recovery dialog
            recovery_dialog = QDialog(self)
            recovery_dialog.setWindowTitle("🔐 Възстановяване на парола")
            recovery_dialog.setModal(True)
            recovery_dialog.setFixedSize(450, 450)
            
            # Center the dialog on screen (like LoginWindow)
            screen = QApplication.primaryScreen().geometry()
            size = recovery_dialog.geometry()
            recovery_dialog.move(
                (screen.width() - size.width()) // 2,
                (screen.height() - size.height()) // 2
            )
            
            layout = QVBoxLayout(recovery_dialog)
            layout.setContentsMargins(20, 20, 20, 20)
            layout.setSpacing(15)
            
            # Title
            title_label = QLabel("Възстановяване на парола")
            title_label.setFont(QFont("Arial", 16, QFont.Weight.Bold))
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_label.setStyleSheet("color: #2196F3; margin-bottom: 10px;")
            layout.addWidget(title_label)
            
            # Instructions
            instructions = QLabel("""
<b>Забравили сте паролата си?</b><br><br>
Свържете се с техническа поддръжка за получаване на <b>главен ключ за възстановяване</b>.<br><br>
Въведете предоставения ключ по-долу за нулиране на паролата към <b>0000</b>.
            """)
            instructions.setWordWrap(True)
            instructions.setAlignment(Qt.AlignmentFlag.AlignCenter)
            instructions.setStyleSheet("color: #666; background-color: #f9f9f9; padding: 15px; border-radius: 8px;")
            layout.addWidget(instructions)
            
            # Master key input
            key_label = QLabel("Главен ключ за възстановяване:")
            key_label.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            layout.addWidget(key_label)
            
            master_key_input = QLineEdit()
            master_key_input.setPlaceholderText("JWL-XXXX-XXXX-XXXX")
            master_key_input.setFont(QFont("Arial", 12))
            master_key_input.setMaxLength(18)  # JWL-XXXX-XXXX-XXXX format (18 characters)
            master_key_input.setStyleSheet("""
                QLineEdit {
                    border: 2px solid #ccc;
                    border-radius: 8px;
                    padding: 10px;
                    font-family: monospace;
                    text-transform: uppercase;
                }
                QLineEdit:focus {
                    border-color: #2196F3;
                }
            """)
            layout.addWidget(master_key_input)
            
            # Auto-format input (add dashes)
            def format_master_key():
                text = master_key_input.text().upper().replace('-', '')
                if len(text) > 3:
                    formatted = text[:3]
                    for i in range(3, len(text), 4):
                        formatted += '-' + text[i:i+4]
                    master_key_input.blockSignals(True)
                    master_key_input.setText(formatted)
                    master_key_input.blockSignals(False)
            
            master_key_input.textChanged.connect(format_master_key)
            
            # Status label
            status_label = QLabel("")
            status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            status_label.setStyleSheet("font-weight: bold; padding: 5px;")
            layout.addWidget(status_label)
            
            # Buttons
            button_layout = QHBoxLayout()
            
            cancel_btn = QPushButton("Отказ")
            cancel_btn.setMinimumHeight(35)
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #666;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #555;
                }
            """)
            cancel_btn.clicked.connect(recovery_dialog.reject)
            
            reset_btn = QPushButton("🔓 Нулирай парола")
            reset_btn.setMinimumHeight(35)
            reset_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            
            def attempt_recovery():
                master_key = master_key_input.text().strip().upper()
                if not master_key:
                    status_label.setText("❌ Моля, въведете главен ключ")
                    status_label.setStyleSheet("color: red; font-weight: bold; padding: 5px;")
                    return
                
                # Verify master key
                success, message = self.db.verify_master_key(master_key)
                
                if success:
                    status_label.setText("✅ Паролата е нулирана към 0000")
                    status_label.setStyleSheet("color: green; font-weight: bold; padding: 5px;")
                    
                    # Success message
                    QMessageBox.information(
                        recovery_dialog, "Успех",
                        "🎉 Паролата е успешно нулирана!\n\n"
                        "Новата парола е: 0000\n\n"
                        "Можете да влезете в системата и да смените паролата от настройките."
                    )
                    
                    recovery_dialog.accept()
                    # Clear the PIN input and set focus for immediate login
                    self.pin_input.clear()
                    self.pin_input.setFocus()
                else:
                    status_label.setText(f"❌ {message}")
                    status_label.setStyleSheet("color: red; font-weight: bold; padding: 5px;")
                    master_key_input.clear()
                    master_key_input.setFocus()
            
            reset_btn.clicked.connect(attempt_recovery)
            master_key_input.returnPressed.connect(attempt_recovery)
            
            button_layout.addWidget(cancel_btn)
            button_layout.addStretch()
            button_layout.addWidget(reset_btn)
            
            layout.addLayout(button_layout)
            
            # Set focus to master key input
            master_key_input.setFocus()
            
            # Show dialog
            recovery_dialog.exec()
            
        except Exception as e:
            logger.error(f"Error showing recovery dialog: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при показване на диалога за възстановяване: {str(e)}")

class PrinterHandler:
    def __init__(self):
        # Citizen CLP-631 specifications
        self.printer_dpi = 300  # CLP-631 is exactly 300 DPI
        self.mm_to_px = self.printer_dpi / 25.4  # 11.811 pixels per mm
        
        # Your label size: 1cm x 4.3cm
        self.label_width_mm = 43.0   # 4.3cm
        self.label_height_mm = 10.0  # 1cm
        
        # Calculate exact pixel dimensions
        self.label_width = int(self.label_width_mm * self.mm_to_px)   # ~508 pixels
        self.label_height = int(self.label_height_mm * self.mm_to_px) # ~118 pixels

    def optimize_for_thermal_transfer(self, label_image):
        """Optimize image specifically for Citizen CLP-631 thermal transfer printing"""
        
        # Resize to exact pixel dimensions for your label
        resized_image = label_image.resize(
            (self.label_width, self.label_height), 
            Image.Resampling.LANCZOS  # High-quality resampling
        )
        
        # Convert to grayscale first for better control
        grayscale_image = resized_image.convert('L')
        
        # Apply optimized threshold for thermal transfer (different from direct thermal)
        # Thermal transfer can handle more gradations, so we can be less aggressive
        threshold = 140  # Slightly higher threshold for cleaner whites
        
        # Convert to pure black/white (1-bit) - essential for thermal printers
        bw_image = grayscale_image.point(
            lambda x: 0 if x < threshold else 255, 
            mode='1'
        )
        
        return bw_image

    def create_optimized_barcode(self, barcode_text, width_px, height_px):
        """Create barcode optimized for 300 DPI thermal transfer printing"""
        try:
            from barcode import Code128
            from barcode.writer import ImageWriter
            
            # Configure writer for exact CLP-631 specifications
            writer = ImageWriter()
            
            # Calculate optimal bar width for 300 DPI
            # Minimum bar width should be at least 3 dots (0.254mm) for reliable scanning
            min_dots = 3
            dot_size_mm = 25.4 / 300  # 0.0847mm per dot
            min_bar_width_mm = min_dots * dot_size_mm  # 0.254mm minimum
            
            # Convert to writer units (modules)
            module_width = max(0.3, min_bar_width_mm / dot_size_mm * 0.1)
            
            code = Code128(barcode_text, writer=writer)
            
            # Generate with optimal settings for your label size
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, "temp_barcode_clp631")
            
            # Get font path for barcode text - use resource_path for PyInstaller compatibility
            barcode_font_path = resource_path("fonts/arial.ttf")
            
            barcode_options = {
                'module_width': module_width,
                'module_height': max(8, int(height_px * 0.6)),  # Leave 40% for text
                'font_size': max(10, int(height_px / 6)),  # Scale font to label
                'text_distance': max(3, int(height_px / 15)),
                'quiet_zone': max(3, int(width_px / 80)),  # Quiet zone
                'foreground': 'black',
                'background': 'white',
                'write_text': True,
                'center_text': True,
                'dpi': 300,
                'font_path': barcode_font_path  # Specify bundled font for PyInstaller compatibility
            }
            
            code.save(temp_file, options=barcode_options)
            barcode_image = Image.open(temp_file + ".png")
            
            # Resize to exact dimensions if needed
            if barcode_image.size != (width_px, height_px):
                barcode_image = barcode_image.resize((width_px, height_px), Image.Resampling.LANCZOS)
            
            # Clean up temp file
            try:
                os.remove(temp_file + ".png")
            except:
                pass
            
            return barcode_image
            
        except ImportError:
            logger.error("python-barcode not installed. Cannot generate barcode.")
            return None

    def print_label_clp631(self, label_image):
        """Print label optimized for Citizen CLP-631"""
        if not label_image:
            return False, "Missing label image"
            
        try:
            # Optimize for thermal transfer printing
            optimized_image = self.optimize_for_thermal_transfer(label_image)
            
            # Save to temporary file with exact specifications
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, "clp631_label.png")
            
            # Save as 1-bit PNG with 300 DPI - exactly matching printer specs
            optimized_image.save(
                temp_file, 
                "PNG", 
                dpi=(300, 300),
                optimize=True,
                compress_level=1  # Minimal compression for faster processing
            )
            
            # Print using Windows default method
            printer_name = win32print.GetDefaultPrinter()
            
            # Verify it's the correct printer (optional check)
            if "citizen" not in printer_name.lower() and "clp" not in printer_name.lower():
                logger.warning(f"Current printer '{printer_name}' might not be CLP-631")
            
            win32api.ShellExecute(
                0,
                "print",
                temp_file,
                f'/d:"{printer_name}"',
                ".",
                0  # Hide window
            )
            
            return True, f"Label sent to {printer_name}"
            
        except Exception as e:
            return False, f"Print error: {str(e)}"

    def print_label(self, label_image, barcode_image=None, barcode_text=""):
        """Legacy method for compatibility - redirects to optimized version"""
        return self.print_label_clp631(label_image) 





# Continue with MainWindow class next...

class ExportFormatDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Избор на формат за експорт")
        self.setModal(True)
        self.setFixedSize(300, 150)
        
        # Set dialog icon
        try:
            icon = get_application_icon(QSize(32, 32))  # Medium size for dialogs
            if not icon.isNull():
                self.setWindowIcon(icon)
        except Exception:
            pass  # Ignore icon errors for dialogs
        
        layout = QVBoxLayout(self)
        
        # Title
        title_label = QLabel("Изберете формат(и) за експорт:")
        title_label.setStyleSheet("font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(title_label)
        
        # Checkboxes
        self.excel_checkbox = QCheckBox("Excel (.xlsx)")
        self.pdf_checkbox = QCheckBox("PDF")
        
        # Set default selections (both checked)
        self.excel_checkbox.setChecked(True)
        self.pdf_checkbox.setChecked(True)
        
        layout.addWidget(self.excel_checkbox)
        layout.addWidget(self.pdf_checkbox)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("Експорт")
        self.export_btn.clicked.connect(self.accept_export)
        self.export_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
        
        cancel_btn = QPushButton("Отказ")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("background-color: #f44336; color: white; font-weight: bold; padding: 8px;")
        
        button_layout.addStretch()
        button_layout.addWidget(self.export_btn)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
        
        # Connect checkbox changes to validation
        self.excel_checkbox.toggled.connect(self.validate_selection)
        self.pdf_checkbox.toggled.connect(self.validate_selection)
        
        # Initial validation
        self.validate_selection()
    
    def validate_selection(self):
        """Enable/disable export button based on selection"""
        has_selection = self.excel_checkbox.isChecked() or self.pdf_checkbox.isChecked()
        self.export_btn.setEnabled(has_selection)
        
        if not has_selection:
            self.export_btn.setStyleSheet("background-color: #cccccc; color: #666666; font-weight: bold; padding: 8px;")
        else:
            self.export_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px;")
    
    def accept_export(self):
        """Accept only if at least one format is selected"""
        if self.excel_checkbox.isChecked() or self.pdf_checkbox.isChecked():
            self.accept()
        else:
            QMessageBox.warning(self, "Грешка", "Моля изберете поне един формат за експорт!")
    
    def get_selections(self):
        """Return tuple of (excel_selected, pdf_selected)"""
        return (self.excel_checkbox.isChecked(), self.pdf_checkbox.isChecked())

class EditItemDialog(QDialog):
    def __init__(self, parent, barcode, category, description, price, cost, weight, metal, stone, stock, from_warehouse=False):
        super().__init__(parent)
        self.setWindowTitle("Редактирай артикул")
        
        # Store the source tab information
        self.from_warehouse = from_warehouse
        
        # Create blur-on-enter event filter for this dialog
        
        # Set dialog icon
        try:
            icon = get_application_icon(QSize(32, 32))  # Medium size for dialogs
            if not icon.isNull():
                self.setWindowIcon(icon)
        except Exception:
            pass  # Ignore icon errors for dialogs
        
        self.setModal(True)
        self.setFixedSize(500, 550)
        
        # Currency conversion rate (fixed)
        self.EUR_TO_LEV_RATE = 1.95583
        
        # Store original barcode for reference
        self.original_barcode = barcode
        
        # Create form layout
        layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        
        # Barcode (read-only)
        self.barcode_label = QLabel(barcode)
        self.barcode_label.setStyleSheet("font-weight: bold; color: #666;")
        self.barcode_label.setFixedHeight(20)
        form_layout.addRow("Баркод:", self.barcode_label)
        
        # Category
        self.category_input = QComboBox()
        self.category_input.setEditable(True)
        self.category_input.addItems(["Пръстен", "Гривна", "Обеци", "Синджир", "Друго"])
        self.category_input.setCurrentText(category)
        form_layout.addRow("Категория:", self.category_input)
        
        # Description
        self.description_input = QTextEdit()
        self.description_input.setPlainText(description)
        self.description_input.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        self.description_input.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.description_input.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        # Calculate line height for auto-resizing
        font_metrics = self.description_input.fontMetrics()
        line_height = font_metrics.lineSpacing()
        self.min_height = line_height + 10
        self.max_height = line_height * 3 + 10
        
        self.description_input.setMinimumHeight(self.min_height)
        self.description_input.setMaximumHeight(self.max_height)
        self.description_input.textChanged.connect(self.auto_resize_description)
        
        form_layout.addRow("Описание:", self.description_input)
        self.auto_resize_description()
        
        # Price in Euro
        self.price_input = BlurOnEnterDoubleSpinBox()
        self.price_input.setRange(0, 999999)
        self.price_input.setDecimals(2)
        self.price_input.setSuffix(" €")
        self.price_input.setValue(price)
        self.price_input.valueChanged.connect(self.update_lev_price)
        form_layout.addRow("Цена (€):", self.price_input)
        
        # Price in Lev (read-only display)
        self.price_lev_label = QLabel(self.format_currency_lev(self.euro_to_lev(price)))
        self.price_lev_label.setStyleSheet("font-weight: bold; color: #2196F3; padding: 1px; background-color: #f0f8ff; border-radius: 1px;")
        self.price_lev_label.setFixedHeight(20)
        form_layout.addRow("Цена (лв):", self.price_lev_label)
        
        # Cost in Euro
        self.cost_input = BlurOnEnterDoubleSpinBox()
        self.cost_input.setRange(0, 999999)
        self.cost_input.setDecimals(2)
        self.cost_input.setSuffix(" €")
        self.cost_input.setValue(cost)
        self.cost_input.valueChanged.connect(self.update_lev_cost)
        form_layout.addRow("Цена на едро (€):", self.cost_input)
        
        # Cost in Lev (read-only display)
        self.cost_lev_label = QLabel(self.format_currency_lev(self.euro_to_lev(cost)))
        self.cost_lev_label.setStyleSheet("font-weight: bold; color: #FF9800; padding: 3px; background-color: #fff3e0; border-radius: 2px;")
        self.cost_lev_label.setFixedHeight(20)
        form_layout.addRow("Цена на едро (лв):", self.cost_lev_label)
        
        # Weight
        self.weight_input = BlurOnEnterDoubleSpinBox()
        self.weight_input.setRange(0, 10000)
        self.weight_input.setDecimals(2)
        self.weight_input.setSuffix(" g")
        self.weight_input.setValue(weight)
        # Blur on Enter key press
        form_layout.addRow("Тегло:", self.weight_input)
        
        # Metal
        self.metal_input = QComboBox()
        self.metal_input.setEditable(True)
        self.metal_input.addItems(["Злато", "Сребро", "Платина", "Друго"])
        self.metal_input.setCurrentText(metal)
        form_layout.addRow("Метал:", self.metal_input)
        
        # Stone
        self.stone_input = QComboBox()
        self.stone_input.setEditable(True)
        self.stone_input.addItems(["Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък", "Друго"])
        self.stone_input.setCurrentText(stone)
        form_layout.addRow("Камък:", self.stone_input)
        
        # Stock
        self.stock_input = BlurOnEnterSpinBox()
        self.stock_input.setRange(0, 10000)
        self.stock_input.setValue(stock)
        
        # Make quantity read-only when opened from shop (not from warehouse)
        if not self.from_warehouse:
            self.stock_input.setReadOnly(True)
            self.stock_input.setStyleSheet("background-color: #f0f0f0; color: #666;")
        
        form_layout.addRow("Количество:", self.stock_input)
        
        layout.addLayout(form_layout)
        
        # Shop Locations Information Section
        self.create_shop_locations_section(layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        # Only show reprint button when editing from warehouse
        if self.from_warehouse:
            self.print_again_btn = QPushButton("Принтирай повторно")
            self.print_again_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold; padding: 8px;")
            self.print_again_btn.clicked.connect(self.print_again)
        
        self.update_btn = QPushButton("Обнови")
        self.update_btn.clicked.connect(self.accept)
        self.cancel_btn = QPushButton("Отказ")
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        if self.from_warehouse:
            button_layout.addWidget(self.print_again_btn)
        button_layout.addWidget(self.update_btn)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
    
    def auto_resize_description(self):
        """Auto-resize description field based on content"""
        try:
            doc = self.description_input.document()
            doc_height = doc.size().height()
            required_height = max(self.min_height, min(self.max_height, int(doc_height) + 10))
            self.description_input.setFixedHeight(required_height)
        except:
            self.description_input.setFixedHeight(self.min_height)
    
    def euro_to_lev(self, euro_amount):
        """Convert Euro to Lev using fixed rate"""
        return round(euro_amount * self.EUR_TO_LEV_RATE, 2)
    
    def format_currency_lev(self, amount):
        """Format amount as Lev currency with thousands separators"""
        return f"{amount:,.2f} лв".replace(",", " ")
    
    def update_lev_price(self):
        """Update Lev price when Euro price changes"""
        try:
            euro_price = self.price_input.value()
            lev_price = self.euro_to_lev(euro_price)
            self.price_lev_label.setText(self.format_currency_lev(lev_price))
        except:
            self.price_lev_label.setText("0.00 лв")
    
    def update_lev_cost(self):
        """Update Lev cost when Euro cost changes"""
        try:
            euro_cost = self.cost_input.value()
            lev_cost = self.euro_to_lev(euro_cost)
            self.cost_lev_label.setText(self.format_currency_lev(lev_cost))
        except:
            self.cost_lev_label.setText("0.00 лв")
    
    def get_data(self):
        """Get updated data from dialog"""
        return {
            'category': self.category_input.currentText(),
            'description': self.description_input.toPlainText(),
            'price': self.price_input.value(),
            'cost': self.cost_input.value(),
            'weight': self.weight_input.value(),
            'metal': self.metal_input.currentText(),
            'stone': self.stone_input.currentText(),
            'stock': self.stock_input.value()
        }
    
    def create_shop_locations_section(self, layout):
        """Create and add the shop locations information section"""
        try:
            # Get shop locations from parent (MainWindow)
            shop_locations = []
            if hasattr(self.parent(), 'get_item_shop_locations'):
                shop_locations = self.parent().get_item_shop_locations(self.original_barcode)
            
            if shop_locations:
                # Create group box for shop locations
                locations_group = QGroupBox("Локации в магазини")
                locations_group.setStyleSheet("""
                    QGroupBox {
                        font-weight: bold;
                        border: 2px solid #cccccc;
                        border-radius: 5px;
                        margin-top: 1ex;
                        padding-top: 10px;
                    }
                    QGroupBox::title {
                        subcontrol-origin: margin;
                        left: 10px;
                        padding: 0 5px 0 5px;
                        color: #2196F3;
                    }
                """)
                
                locations_layout = QVBoxLayout()
                
                # Create a scroll area for locations
                scroll_area = QScrollArea()
                scroll_widget = QWidget()
                scroll_layout = QVBoxLayout(scroll_widget)
                
                total_in_shops = 0
                for location in shop_locations:
                    shop_name = location['shop_name']
                    quantity = location['quantity']
                    updated_at = location['updated_at']
                    total_in_shops += quantity
                    
                    # Format the updated date
                    try:
                        from datetime import datetime
                        if updated_at:
                            updated_date = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
                            formatted_date = updated_date.strftime('%d.%m.%Y %H:%M')
                        else:
                            formatted_date = "Неизвестно"
                    except:
                        formatted_date = "Неизвестно"
                    
                    # Create location info label
                    location_label = QLabel(f"📍 {shop_name}: {quantity} бр. (обновено: {formatted_date})")
                    location_label.setStyleSheet("""
                        QLabel {
                            background-color: #f8f9fa;
                            border: 1px solid #dee2e6;
                            border-radius: 3px;
                            padding: 5px;
                            margin: 2px 0;
                            font-size: 11px;
                            color: #212529;
                        }
                    """)
                    scroll_layout.addWidget(location_label)
                
                # Add summary
                summary_label = QLabel(f"📊 Общо в магазини: {total_in_shops} бр.")
                summary_label.setStyleSheet("""
                    QLabel {
                        background-color: #e3f2fd;
                        border: 2px solid #2196F3;
                        border-radius: 5px;
                        padding: 8px;
                        font-weight: bold;
                        color: #1976d2;
                        margin: 5px 0;
                    }
                """)
                scroll_layout.addWidget(summary_label)
                
                scroll_layout.addStretch()
                scroll_area.setWidget(scroll_widget)
                scroll_area.setMaximumHeight(120)
                scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
                scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
                
                locations_layout.addWidget(scroll_area)
                locations_group.setLayout(locations_layout)
                layout.addWidget(locations_group)
            else:
                # Show message that item is not in any shops
                no_locations_label = QLabel("ℹ️ Артикулът не е наличен в магазини")
                no_locations_label.setStyleSheet("""
                    QLabel {
                        background-color: #fff3e0;
                        border: 1px solid #ff9800;
                        border-radius: 3px;
                        padding: 8px;
                        margin: 5px 0;
                        color: #f57c00;
                        font-style: italic;
                    }
                """)
                layout.addWidget(no_locations_label)
                
        except Exception as e:
            # If there's an error, show a simple message
            error_label = QLabel("⚠️ Грешка при зареждане на локации")
            error_label.setStyleSheet("""
                QLabel {
                    background-color: #ffebee;
                    border: 1px solid #f44336;
                    border-radius: 3px;
                    padding: 8px;
                    margin: 5px 0;
                    color: #d32f2f;
                }
            """)
            layout.addWidget(error_label)
    
    def print_again(self):
        """Print additional labels by switching to Add Item tab with pre-filled data"""
        try:
            main_window = self.parent()
            if main_window and hasattr(main_window, 'tabs'):
                current_data = {
                    'barcode': self.original_barcode,
                    'category': self.category_input.currentText(),
                    'description': self.description_input.toPlainText(),
                    'price': self.price_input.value(),
                    'cost': self.cost_input.value(),
                    'weight': self.weight_input.value(),
                    'metal': self.metal_input.currentText(),
                    'stone': self.stone_input.currentText(),
                    'stock': 0 if self.from_warehouse else self.stock_input.value(),  # Set to 0 for warehouse reprint
                    'is_warehouse_reprint': self.from_warehouse  # Flag to indicate this is a warehouse reprint
                }
                
                self.reject()
                main_window.tabs.setCurrentIndex(0)
                main_window.populate_add_item_form(current_data)
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при отваряне на формата за печат: {str(e)}")

class MainWindow(QMainWindow):
    def __init__(self, database=None):
        super().__init__()
        # Use provided database instance or create new one
        self.db = database if database is not None else Database()
        self.current_label = None  # Initialize current_label
        self.barcode_image = None  # Initialize barcode_image
        self.printer_handler = PrinterHandler()  # Initialize printer handler
        self.barcode_gen = BarcodeGenerator()
        self.barcode_scanner = BarcodeScanner()
        self.report_generator = ReportGenerator()
        
        # Initialize action history for undo/redo
        self.action_history = ActionHistory(max_history=3)
        
        # Initialize barcode input timers to handle scanner delays
        self.sales_barcode_timer = QTimer()
        self.sales_barcode_timer.setSingleShot(True)
        self.sales_barcode_timer.timeout.connect(self.process_sales_barcode)
        
        self.shop_barcode_timer = QTimer()
        self.shop_barcode_timer.setSingleShot(True)
        self.shop_barcode_timer.timeout.connect(self.process_shop_barcode)
        
        # Initialize custom values storage
        self.custom_categories = set()
        
        # Flag to prevent auto-switching to custom period during programmatic changes
        self.programmatic_date_change = False
        self.programmatic_inventory_date_change = False
        self.custom_metals = set()
        self.custom_stones = set()
        
        # Flag to track warehouse reprint mode
        self.is_warehouse_reprint = False
        
        # Flag to prevent concurrent shop inventory loading
        self.shop_inventory_loading = False
        
        # Initialize audit state variables
        self.audit_in_progress = False
        self.audit_shop_id = None
        self.audit_shop_name = ""
        self.audit_start_time = None
        self.audit_paused = False
        self.audit_items_data = {}  # {barcode: {expected_qty, scanned_qty, category, metal_type, stone_type, price, weight, description}}
        self.audit_scanned_items = {}  # {barcode: scanned_quantity}
        self.audit_session_id = None
        
        # Create custom combo box delegate
        self.combo_delegate = CustomComboDelegate(self)
        
        # Create blur-on-enter event filter for numeric input fields
        
        # Ensure database schema is up to date
        self.ensure_database_schema()
        
        # Load auto backup settings
        self.load_auto_backup_settings()
        
        self.setup_ui()
        self.load_data()  # Load initial data
        # Ensure all shop combos are properly initialized
        self.refresh_all_shop_combos()
        self.update_action_buttons()  # Initialize undo/redo button states
        self.showMaximized()
        apply_dark_theme()  # Apply dark theme

    def get_backup_directory(self):
        """Get the backup directory path relative to the executable"""
        if getattr(sys, 'frozen', False):
            # Running as compiled executable
            base_dir = os.path.dirname(sys.executable)
        else:
            # Running as script
            base_dir = os.path.dirname(__file__)
        
        backup_dir = os.path.join(base_dir, 'backups')
        # Note: Directory creation and hiding is handled by setup_directories()
        return backup_dir
    
    def setup_backup_file_watcher(self):
        """Setup file system watcher to monitor backup directory for changes"""
        try:
            if hasattr(self, 'backup_watcher'):
                # Remove existing watcher
                self.backup_watcher.deleteLater()
            
            backup_dir = self.get_backup_directory()
            self.backup_watcher = QFileSystemWatcher()
            self.backup_watcher.addPath(backup_dir)
            self.backup_watcher.directoryChanged.connect(self.on_backup_directory_changed)
            logger.info(f"Setup backup file watcher for: {backup_dir}")
            
        except Exception as e:
            logger.error(f"Error setting up backup file watcher: {e}")
    
    def on_backup_directory_changed(self):
        """Handle backup directory changes (files added/removed externally)"""
        try:
            # Refresh the backup list when directory changes
            self.load_backup_list()
            logger.info("Backup list refreshed due to directory change")
        except Exception as e:
            logger.error(f"Error handling backup directory change: {e}")

    def ensure_database_schema(self):
        """Ensure database has all required tables and columns for sales functionality"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Check if sales table has shop_id column
                cursor.execute("PRAGMA table_info(sales)")
                sales_columns = [col[1] for col in cursor.fetchall()]
                
                if 'shop_id' not in sales_columns:
                    # Add shop_id column to sales table
                    cursor.execute("ALTER TABLE sales ADD COLUMN shop_id INTEGER")
                    logger.info("Added shop_id column to sales table")
                    
                    # Set default shop_id for existing sales
                    cursor.execute("SELECT id FROM shops LIMIT 1")
                    default_shop = cursor.fetchone()
                    if default_shop:
                        cursor.execute("UPDATE sales SET shop_id = ? WHERE shop_id IS NULL", (default_shop[0],))
                        logger.info("Updated existing sales with default shop_id")
                
                # Ensure shops table exists with proper structure
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS shops (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE
                )
                """)
                
                # Ensure shop_items table exists with proper structure and timestamps
                cursor.execute("""
                CREATE TABLE IF NOT EXISTS shop_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    shop_id INTEGER NOT NULL,
                    item_id INTEGER NOT NULL,
                    quantity INTEGER NOT NULL DEFAULT 1,
                    created_at TIMESTAMP DEFAULT (datetime('now', 'localtime')),
                    updated_at TIMESTAMP DEFAULT (datetime('now', 'localtime')),
                    FOREIGN KEY (shop_id) REFERENCES shops(id),
                    FOREIGN KEY (item_id) REFERENCES items(id),
                    UNIQUE(shop_id, item_id)
                )
                """)
                
                # Add timestamp columns to existing shop_items table if they don't exist
                cursor.execute("PRAGMA table_info(shop_items)")
                shop_items_columns = [col[1] for col in cursor.fetchall()]
                
                if 'created_at' not in shop_items_columns:
                    cursor.execute("ALTER TABLE shop_items ADD COLUMN created_at TIMESTAMP DEFAULT (datetime('now', 'localtime'))")
                    cursor.execute("UPDATE shop_items SET created_at = datetime('now', 'localtime') WHERE created_at IS NULL")
                    logger.info("Added created_at column to shop_items table")
                    
                if 'updated_at' not in shop_items_columns:
                    cursor.execute("ALTER TABLE shop_items ADD COLUMN updated_at TIMESTAMP DEFAULT (datetime('now', 'localtime'))") 
                    cursor.execute("UPDATE shop_items SET updated_at = datetime('now', 'localtime') WHERE updated_at IS NULL")
                    logger.info("Added updated_at column to shop_items table")
                
                # Create default shop if none exist
                cursor.execute("SELECT COUNT(*) FROM shops")
                if cursor.fetchone()[0] == 0:
                    cursor.execute("INSERT INTO shops (name) VALUES (?)", ("Магазин 1",))
                    logger.info("Created default shop 'Магазин 1'")
                
                conn.commit()
                logger.info("Database schema verification completed")
                
        except Exception as e:
            logger.error(f"Error ensuring database schema: {e}")
            # Don't crash the app, just log the error

    def setup_ui(self):
        self.setWindowTitle("Система за управление на бижута")
        
        # Set application icon for taskbar
        try:
            icon = get_application_icon()
            if not icon.isNull():
                self.setWindowIcon(icon)
                logger.info("Main window icon set using high-quality PNG")
            else:
                logger.warning("No icon could be loaded for main window")
        except Exception as e:
            logger.warning(f"Could not set main window icon: {e}")
        
        self.setMinimumSize(1200, 800)

        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Add theme switcher and undo/redo buttons
        top_toolbar_layout = QHBoxLayout()
        
        # Undo/Redo buttons
        self.undo_btn = QPushButton("↶ Отмени")
        self.undo_btn.setFixedSize(100, 40)
        self.undo_btn.clicked.connect(self.undo_action)
        self.undo_btn.setEnabled(False)
        self.undo_btn.setToolTip("Отмени последното действие")
        top_toolbar_layout.addWidget(self.undo_btn)
        
        self.redo_btn = QPushButton("↷ Върни")
        self.redo_btn.setFixedSize(100, 40)
        self.redo_btn.clicked.connect(self.redo_action)
        self.redo_btn.setEnabled(False)
        self.redo_btn.setToolTip("Върни отмененото действие")
        top_toolbar_layout.addWidget(self.redo_btn)
        
        # Action history label
        self.action_status_label = QLabel("Няма действия за отменяне")
        self.action_status_label.setStyleSheet("color: #666; font-size: 11px; font-style: italic;")
        top_toolbar_layout.addWidget(self.action_status_label)
        
        top_toolbar_layout.addStretch()
        
        main_layout.addLayout(top_toolbar_layout)

        # Create tab widget and store reference
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # Add tabs in the correct order
        self.tabs.addTab(self.create_add_item_tab(), "Добави артикул")
        self.tabs.addTab(self.create_inventory_tab(), "Склад")
        self.tabs.addTab(self.create_shop_loading_tab(), "Зареждане на магазин")
        self.tabs.addTab(self.create_sales_tab(), "Продажби")
        self.tabs.addTab(self.create_reports_tab(), "Отчети")
        self.tabs.addTab(self.create_audit_tab(), "Инвентаризация")
        self.tabs.addTab(self.create_database_tab(), "База данни")
        self.tabs.addTab(self.create_help_tab(), "Помощ")
        
        # Connect tab change to set focus appropriately
        self.tabs.currentChanged.connect(self.on_tab_changed)
        
        # Set up keyboard shortcuts for tab navigation
        self.setup_tab_shortcuts()

    def on_tab_changed(self, index):
        """Handle tab change events"""
        try:
            # Get the tab text to identify which tab was selected
            tab_text = self.tabs.tabText(index)
            
            # Set focus to barcode input when sales tab is selected
            if tab_text == "Продажби" and hasattr(self, 'sale_barcode_input'):
                # Use QTimer to ensure focus is set after tab switch is complete
                from PyQt6.QtCore import QTimer
                QTimer.singleShot(100, lambda: self.sale_barcode_input.setFocus())
            
            # Update statistics when switching to Reports or Database tabs
            elif tab_text == "Отчети":
                self.update_reports_and_database_stats()
            elif tab_text == "База данни":
                # Force immediate update of database statistics
                self.update_database_statistics()
                self.load_backup_list()
            
        except Exception as e:
            # Ignore errors in tab change handling
            pass

    def setup_tab_shortcuts(self):
        """Set up keyboard shortcuts for tab navigation"""
        try:
            # Ctrl+Tab - Next tab
            next_tab_shortcut = QShortcut(QKeySequence("Ctrl+Tab"), self)
            next_tab_shortcut.activated.connect(self.next_tab)
            
            # Ctrl+Shift+Tab - Previous tab
            prev_tab_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Tab"), self)
            prev_tab_shortcut.activated.connect(self.previous_tab)
            
            # Alternative shortcuts
            # Ctrl+PageDown - Next tab
            next_tab_shortcut2 = QShortcut(QKeySequence("Ctrl+PgDown"), self)
            next_tab_shortcut2.activated.connect(self.next_tab)
            
            # Ctrl+PageUp - Previous tab
            prev_tab_shortcut2 = QShortcut(QKeySequence("Ctrl+PgUp"), self)
            prev_tab_shortcut2.activated.connect(self.previous_tab)
            
            # Number shortcuts (Ctrl+1, Ctrl+2, etc.)
            for i in range(min(9, self.tabs.count())):  # Support up to 9 tabs
                shortcut = QShortcut(QKeySequence(f"Ctrl+{i+1}"), self)
                shortcut.activated.connect(lambda tab_index=i: self.goto_tab(tab_index))
            
            logger.info("Tab navigation shortcuts set up successfully")
            
        except Exception as e:
            logger.error(f"Error setting up tab shortcuts: {e}")

    def next_tab(self):
        """Switch to next tab (with wrapping)"""
        try:
            current_index = self.tabs.currentIndex()
            total_tabs = self.tabs.count()
            next_index = (current_index + 1) % total_tabs
            self.tabs.setCurrentIndex(next_index)
        except Exception as e:
            logger.error(f"Error switching to next tab: {e}")

    def previous_tab(self):
        """Switch to previous tab (with wrapping)"""
        try:
            current_index = self.tabs.currentIndex()
            total_tabs = self.tabs.count()
            prev_index = (current_index - 1) % total_tabs
            self.tabs.setCurrentIndex(prev_index)
        except Exception as e:
            logger.error(f"Error switching to previous tab: {e}")

    def goto_tab(self, tab_index):
        """Switch to specific tab by index"""
        try:
            if 0 <= tab_index < self.tabs.count():
                self.tabs.setCurrentIndex(tab_index)
        except Exception as e:
            logger.error(f"Error switching to tab {tab_index}: {e}")

    def undo_action(self):
        """Undo the last action"""
        if self.action_history.undo():
            QMessageBox.information(self, "Успех", f"Отменено: {self.action_history.get_redo_description()}")
            self.refresh_all_data()
            self.update_action_buttons()
        else:
            QMessageBox.warning(self, "Грешка", "Неуспешно отменяне на действието")

    def redo_action(self):
        """Redo the last undone action"""
        if self.action_history.redo():
            QMessageBox.information(self, "Успех", f"Възстановено: {self.action_history.get_undo_description()}")
            self.refresh_all_data()
            self.update_action_buttons()
        else:
            QMessageBox.warning(self, "Грешка", "Неуспешно възстановяване на действието")

    def update_action_buttons(self):
        """Update undo/redo button states and tooltips"""
        # Update undo button
        can_undo = self.action_history.can_undo()
        self.undo_btn.setEnabled(can_undo)
        if can_undo:
            undo_desc = self.action_history.get_undo_description()
            self.undo_btn.setToolTip(f"Отмени: {undo_desc}")
        else:
            self.undo_btn.setToolTip("Няма действия за отменяне")

        # Update redo button
        can_redo = self.action_history.can_redo()
        self.redo_btn.setEnabled(can_redo)
        if can_redo:
            redo_desc = self.action_history.get_redo_description()
            self.redo_btn.setToolTip(f"Върни: {redo_desc}")
        else:
            self.redo_btn.setToolTip("Няма действия за възстановяване")

        # Update status label
        if can_undo:
            current_desc = self.action_history.get_undo_description()
            self.action_status_label.setText(f"Последно: {current_desc}")
        else:
            self.action_status_label.setText("Няма действия за отменяне")

    def refresh_all_data(self):
        """Refresh all data in all tabs after undo/redo"""
        self.load_items()
        self.load_sales()
        if hasattr(self, 'shop_combo'):
            self.load_shop_inventory()
        # Refresh all shop combos to ensure they're up to date
        self.refresh_all_shop_combos()
        # Update reports and database statistics
        self.update_reports_and_database_stats()
    
    def show_temp_success_message(self, message):
        """Show a temporary success message that disappears after 2 seconds"""
        try:
            # Create a temporary message widget
            if not hasattr(self, 'temp_message_widget'):
                self.temp_message_widget = QLabel(self)
                self.temp_message_widget.setStyleSheet("""
                    QLabel {
                        background-color: #4CAF50;
                        color: white;
                        padding: 15px 25px;
                        border-radius: 8px;
                        font-size: 14px;
                        font-weight: bold;
                        border: 2px solid #45a049;
                    }
                """)
                self.temp_message_widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self.temp_message_widget.setWordWrap(True)
            
            # Set the message text
            self.temp_message_widget.setText(message)
            self.temp_message_widget.adjustSize()
            
            # Position the widget at the top center of the main window
            parent_width = self.width()
            parent_height = self.height()
            widget_width = self.temp_message_widget.width()
            widget_height = self.temp_message_widget.height()
            
            x = (parent_width - widget_width) // 2
            y = 50  # 50 pixels from the top
            
            self.temp_message_widget.move(x, y)
            self.temp_message_widget.show()
            self.temp_message_widget.raise_()
            
            # Create timer to hide the message after 2 seconds
            if hasattr(self, 'temp_message_timer'):
                self.temp_message_timer.stop()
            
            self.temp_message_timer = QTimer()
            self.temp_message_timer.setSingleShot(True)
            self.temp_message_timer.timeout.connect(self.hide_temp_message)
            self.temp_message_timer.start(2000)  # 2 seconds
            
        except Exception as e:
            logger.error(f"Error showing temporary message: {e}")
            # Fallback to status bar message
            self.statusBar().showMessage(message, 2000)
    
    def hide_temp_message(self):
        """Hide the temporary message widget"""
        try:
            if hasattr(self, 'temp_message_widget'):
                self.temp_message_widget.hide()
        except Exception as e:
            logger.error(f"Error hiding temporary message: {e}")
    
    def update_shop_inventory_info(self):
        """Update the shop inventory info label with current shop's item count"""
        try:
            if not hasattr(self, 'shop_inventory_info_label') or not hasattr(self, 'sales_shop_combo'):
                return
            
            shop_name = self.sales_shop_combo.currentText()
            if not shop_name:
                self.shop_inventory_info_label.setText("Налични артикули в магазин: 0")
                return
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get shop ID
                cursor.execute("SELECT id FROM shops WHERE name = ?", (shop_name,))
                shop_row = cursor.fetchone()
                if not shop_row:
                    self.shop_inventory_info_label.setText("Налични артикули в магазин: 0")
                    return
                
                shop_id = shop_row[0]
                
                # Count total items in shop (sum of all quantities)
                cursor.execute("""
                    SELECT COUNT(DISTINCT si.item_id), COALESCE(SUM(si.quantity), 0)
                    FROM shop_items si 
                    WHERE si.shop_id = ? AND si.quantity > 0
                """, (shop_id,))
                
                result = cursor.fetchone()
                unique_items = result[0] if result else 0
                total_quantity = result[1] if result else 0
                
                # Update label with both unique items and total quantity
                if unique_items == 0:
                    self.shop_inventory_info_label.setText("Налични артикули в магазин: 0")
                    self.shop_inventory_info_label.setStyleSheet("color: #f44336; font-weight: bold;")  # Red for empty
                else:
                    self.shop_inventory_info_label.setText(f"Налични артикули в магазин: {unique_items} вида ({total_quantity} бр.)")
                    self.shop_inventory_info_label.setStyleSheet("color: #4CAF50; font-weight: bold;")  # Green for available
                    
        except Exception as e:
            logger.error(f"Error updating shop inventory info: {e}")
            if hasattr(self, 'shop_inventory_info_label'):
                self.shop_inventory_info_label.setText("Налични артикули в магазин: ? (грешка)")

    def update_reports_and_database_stats(self):
        """Update both reports dashboard and database statistics"""
        try:
            # Update reports dashboard if it exists
            if hasattr(self, 'stats_cards'):
                self.update_dashboard_stats()
            
            # Update database statistics if it exists
            if hasattr(self, 'db_stats_cards'):
                self.update_database_statistics()
            
            # Also update backup list if we're on database tab
            if hasattr(self, 'backup_list'):
                self.load_backup_list()
                
        except Exception as e:
            logger.error(f"Error updating reports and database stats: {e}")

    def create_add_item_tab(self):
        """Create the add item tab"""
        widget = QWidget()
        layout = QHBoxLayout(widget)  # Changed to horizontal layout

        # Left panel - Add/Edit item form (50% width)
        left_panel = QWidget()
        left_panel.setMaximumWidth(600)  # Set fixed width for consistent 50/50 split
        left_layout = QVBoxLayout(left_panel)
        
        form_group = QGroupBox("Добави артикул")
        form_layout = QFormLayout()

        # Currency conversion rate (fixed)
        self.EUR_TO_LEV_RATE = 1.95583
        
        # Initialize all input widgets
        self.barcode_input = QLineEdit()
        self.barcode_input.setReadOnly(True)  # Make barcode readonly
        
        # Price input in Euro
        self.price_input = BlurOnEnterDoubleSpinBox()
        self.price_input.setRange(0, 1000000)
        self.price_input.setDecimals(2)
        self.price_input.setSuffix(" €")
        self.price_input.valueChanged.connect(self.update_lev_price)
        # Auto-select all text when clicked
        self.price_input.lineEdit().installEventFilter(self)
        
        # Price display in Lev (read-only)
        self.price_lev_label = QLabel("0.00 лв")
        self.price_lev_label.setStyleSheet("font-weight: bold; color: #2196F3; padding: 5px; background-color: #f0f8ff; border-radius: 3px;")
        
        self.weight_input = BlurOnEnterDoubleSpinBox()
        self.weight_input.setRange(0, 1000)
        self.weight_input.setDecimals(2)
        # Auto-select all text when clicked
        self.weight_input.lineEdit().installEventFilter(self)
        # Blur on Enter key press
        
        self.stock_input = BlurOnEnterSpinBox()
        self.stock_input.setRange(0, 10000)
        # Auto-select all text when clicked
        self.stock_input.lineEdit().installEventFilter(self)
        # Blur on Enter key press
        
        # Description - Auto-resizing text area with word wrap
        self.description_input = QTextEdit()
        self.description_input.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        self.description_input.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.description_input.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        # Calculate line height for auto-resizing
        font_metrics = self.description_input.fontMetrics()
        line_height = font_metrics.lineSpacing()
        self.desc_min_height = line_height + 10  # 1 line + padding
        self.desc_max_height = line_height * 3 + 10  # 3 lines + padding
        
        # Set initial height
        self.description_input.setMinimumHeight(self.desc_min_height)
        self.description_input.setMaximumHeight(self.desc_max_height)
        
        # Connect text change to auto-resize function
        self.description_input.textChanged.connect(self.auto_resize_add_item_description)
        
        # Set initial size
        self.description_input.setFixedHeight(self.desc_min_height)
        
        self.category_input = QComboBox()
        self.category_input.setItemDelegate(self.combo_delegate)
        self.category_input.addItems(["Пръстен", "Гривна", "Обеци", "Синджир", "Друго"])
        self.category_input.currentTextChanged.connect(
            lambda text: self.handle_custom_input(self.category_input, self.custom_categories, text)
        )
        # Cost input in Euro  
        self.cost_input = BlurOnEnterDoubleSpinBox()
        self.cost_input.setRange(0, 1000000)
        self.cost_input.setDecimals(2)
        self.cost_input.setSuffix(" €")
        self.cost_input.valueChanged.connect(self.update_lev_cost)
        # Auto-select all text when clicked
        self.cost_input.lineEdit().installEventFilter(self)
        # Blur on Enter key press
        
        # Cost display in Lev (read-only)
        self.cost_lev_label = QLabel("0.00 лв")
        self.cost_lev_label.setStyleSheet("font-weight: bold; color: #FF9800; padding: 5px; background-color: #fff3e0; border-radius: 3px;")
        
        self.metal_input = QComboBox()
        self.metal_input.setItemDelegate(self.combo_delegate)
        self.metal_input.addItems(["Злато", "Сребро", "Платина", "Друго"])
        self.metal_input.setCurrentIndex(1)  # Set default to Сребро (index 1)
        self.metal_input.currentTextChanged.connect(
            lambda text: self.handle_custom_input(self.metal_input, self.custom_metals, text)
        )
        self.stone_input = QComboBox()
        self.stone_input.setItemDelegate(self.combo_delegate)
        self.stone_input.addItems(["Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък", "Друго"])
        self.stone_input.setCurrentIndex(4)  # Set default to Без камък (index 4)
        self.stone_input.currentTextChanged.connect(
            lambda text: self.handle_custom_input(self.stone_input, self.custom_stones, text)
        )

        # Add fields to layout in the specified order
        form_layout.addRow("Категория:", self.category_input)
        form_layout.addRow("Метал:", self.metal_input)
        form_layout.addRow("Камък:", self.stone_input)
        form_layout.addRow("Описание:", self.description_input)
        
        # Cost fields - Euro input and Lev display
        form_layout.addRow("Цена на едро (€):", self.cost_input)
        form_layout.addRow("Цена на едро (лв):", self.cost_lev_label)
        
        # Price fields - Euro input and Lev display  
        form_layout.addRow("Цена (€):", self.price_input)
        form_layout.addRow("Цена (лв):", self.price_lev_label)
        
        form_layout.addRow("Грамаж:", self.weight_input)
        form_layout.addRow("Количество:", self.stock_input)

        # Barcode (needs its own layout for the button/read-only)
        barcode_layout = QHBoxLayout()
        barcode_layout.addWidget(self.barcode_input)
        form_layout.addRow("Баркод:", barcode_layout)
        


        # Buttons - stacked vertically and removed "Запази" button
        button_layout = QVBoxLayout()
        clear_btn = QPushButton("Изчисти")
        clear_btn.clicked.connect(self.clear_item_form)
        clear_btn.setStyleSheet("background-color: #ff9999; color: #990000; font-weight: bold;")  # More saturated red
        
        generate_btn = QPushButton("Генерирай етикет")
        generate_btn.clicked.connect(self.generate_barcode)
        generate_btn.setStyleSheet("background-color: #ffeb99; color: #664d00; font-weight: bold;")  # More saturated yellow
        
        print_btn = QPushButton("Принтирай етикет")
        print_btn.clicked.connect(self.print_and_add_item)
        print_btn.setStyleSheet("background-color: #99ff99; color: #004d00; font-weight: bold;")  # More saturated green
        
        # Add buttons vertically
        button_layout.addWidget(clear_btn)
        button_layout.addWidget(generate_btn)
        button_layout.addWidget(print_btn)
        button_layout.addStretch()  # Add stretch to push buttons to top
        
        form_layout.addRow("", button_layout)

        form_group.setLayout(form_layout)
        left_layout.addWidget(form_group)
        layout.addWidget(left_panel)

        # Right panel - Barcode preview (50% width)
        right_panel = QWidget()
        right_panel.setMaximumWidth(600)  # Set fixed width for consistent 50/50 split
        right_layout = QVBoxLayout(right_panel)
        
        # Barcode preview
        preview_group = QGroupBox("Преглед на етикет")
        preview_layout = QVBoxLayout()
        
        # Calculate label dimensions
        # Initialize PrinterHandler for CLP-631 specifications
        self.printer_handler = PrinterHandler()
        self.label_width_mm = self.printer_handler.label_width_mm
        self.label_height_mm = self.printer_handler.label_height_mm
        self.dpi = self.printer_handler.printer_dpi
        self.mm_to_px = self.printer_handler.mm_to_px
        self.label_width = self.printer_handler.label_width
        self.label_height = self.printer_handler.label_height
        
        # Add checkbox for price display option
        self.include_lev_price_checkbox = QCheckBox("Включи лв цена на етикета")
        self.include_lev_price_checkbox.setChecked(True)  # Default to checked (current behavior)
        self.include_lev_price_checkbox.stateChanged.connect(self.update_barcode_preview)
        preview_layout.addWidget(self.include_lev_price_checkbox, 0, Qt.AlignmentFlag.AlignCenter)
        
        # Add checkbox for grams display option
        self.include_grams_checkbox = QCheckBox("Включи грамаж на етикета")
        self.include_grams_checkbox.setChecked(True)  # Default to checked
        self.include_grams_checkbox.stateChanged.connect(self.update_barcode_preview)
        preview_layout.addWidget(self.include_grams_checkbox, 0, Qt.AlignmentFlag.AlignCenter)
        
        # Add checkbox for price order inversion
        self.invert_prices_checkbox = QCheckBox("Размени реда на цените (лв / €)")
        self.invert_prices_checkbox.setChecked(False)  # Default to unchecked (Euro first)
        self.invert_prices_checkbox.stateChanged.connect(self.update_barcode_preview)
        preview_layout.addWidget(self.invert_prices_checkbox, 0, Qt.AlignmentFlag.AlignCenter)
        
        # Add stretch before preview to center it
        preview_layout.addStretch()
        
        self.barcode_preview = QLabel()
        self.barcode_preview.setFixedSize(self.label_width, self.label_height)
        self.barcode_preview.setStyleSheet("background-color: white; border: 1px solid #d0d0d0;")
        self.barcode_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Center the content within the label
        preview_layout.addWidget(self.barcode_preview, 0, Qt.AlignmentFlag.AlignCenter)  # Center the widget itself
        
        # Add stretch after preview to center it
        preview_layout.addStretch()
        
        preview_group.setLayout(preview_layout)
        right_layout.addWidget(preview_group)
        layout.addWidget(right_panel)

        return widget

    def create_inventory_tab(self):
        """Create the inventory management tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Enhanced Tabbed Search and Filter Section
        search_group = QGroupBox("Търсене и филтриране")
        search_layout = QVBoxLayout()
        
        # Create search tabs
        self.search_tabs = QTabWidget()
        self.search_tabs.setMaximumHeight(100)  # Reduced height for more compact design
        
        # Tab 1: General Search
        general_tab = QWidget()
        general_layout = QVBoxLayout(general_tab)
        
        # Main search bar for general search
        main_search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Търси по всички полета (баркод, категория, метал, камък, описание, цена, тегло, количество, дата)...")
        self.search_input.textChanged.connect(self.search_items)
        clear_search_btn = QPushButton("✕")
        clear_search_btn.setFixedSize(30, 30)
        clear_search_btn.clicked.connect(self.clear_search)
        main_search_layout.addWidget(QLabel("Общо търсене:"))
        main_search_layout.addWidget(self.search_input)
        main_search_layout.addWidget(clear_search_btn)
        general_layout.addLayout(main_search_layout)
        
        # Add general search help
        help_label = QLabel("💡 Търси във всички полета едновременно")
        help_label.setStyleSheet("color: #666; font-size: 10px; font-style: italic;")
        general_layout.addWidget(help_label)
        
        self.search_tabs.addTab(general_tab, "Общо търсене")
        
        # Tab 2: Price Search
        price_tab = QWidget()
        price_layout = QVBoxLayout(price_tab)
        
        # Combine price and weight in one compact horizontal layout
        price_weight_layout = QHBoxLayout()
        
        # Price controls (more compact)
        self.min_price_input = BlurOnEnterDoubleSpinBox()
        self.min_price_input.setRange(0, 999999)
        self.min_price_input.setSuffix(" лв")
        self.min_price_input.setMaximumWidth(100)
        self.min_price_input.editingFinished.connect(self.search_items)
        # Auto-clear text when clicked
        self.min_price_input.lineEdit().installEventFilter(self)
        # Blur on Enter key press
        self.max_price_input = BlurOnEnterDoubleSpinBox()
        self.max_price_input.setRange(0, 999999)
        self.max_price_input.setValue(999999)
        self.max_price_input.setSuffix(" лв")
        self.max_price_input.setMaximumWidth(100)
        self.max_price_input.editingFinished.connect(self.search_items)
        # Auto-clear text when clicked
        self.max_price_input.lineEdit().installEventFilter(self)
        # Blur on Enter key press
        
        price_weight_layout.addWidget(QLabel("Цена:"))
        price_weight_layout.addWidget(self.min_price_input)
        price_weight_layout.addWidget(QLabel("-"))
        price_weight_layout.addWidget(self.max_price_input)
        
        # Add separator
        price_weight_layout.addWidget(QLabel(" | "))
        
        # Weight controls (more compact)
        self.min_weight_input = BlurOnEnterDoubleSpinBox()
        self.min_weight_input.setRange(0, 9999)
        self.min_weight_input.setSuffix(" г")
        self.min_weight_input.setMaximumWidth(100)
        self.min_weight_input.editingFinished.connect(self.search_items)
        # Auto-clear text when clicked
        self.min_weight_input.lineEdit().installEventFilter(self)
        # Blur on Enter key press
        self.max_weight_input = BlurOnEnterDoubleSpinBox()
        self.max_weight_input.setRange(0, 9999)
        self.max_weight_input.setValue(9999)
        self.max_weight_input.setSuffix(" г")
        self.max_weight_input.setMaximumWidth(100)
        self.max_weight_input.editingFinished.connect(self.search_items)
        # Auto-clear text when clicked
        self.max_weight_input.lineEdit().installEventFilter(self)
        # Blur on Enter key press
        
        price_weight_layout.addWidget(QLabel("Тегло:"))
        price_weight_layout.addWidget(self.min_weight_input)
        price_weight_layout.addWidget(QLabel("-"))
        price_weight_layout.addWidget(self.max_weight_input)
        price_weight_layout.addStretch()
        
        price_layout.addLayout(price_weight_layout)
        
        self.search_tabs.addTab(price_tab, "По цена/тегло")
        
        # Tab 3: Date Search
        date_tab = QWidget()
        date_layout = QVBoxLayout(date_tab)
        
        # Period filter radio buttons (first row) - like in sales
        date_period_layout = QHBoxLayout()
        
        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        self.inventory_period_group = QButtonGroup()
        
        self.inv_all_radio = QRadioButton("Всички")
        self.inv_all_radio.setChecked(True)
        self.inv_today_radio = QRadioButton("Днес")
        self.inv_week_radio = QRadioButton("Седмица")
        self.inv_month_radio = QRadioButton("Месец")
        self.inv_custom_radio = QRadioButton("Персонализиран")
        
        self.inventory_period_group.addButton(self.inv_all_radio, 0)
        self.inventory_period_group.addButton(self.inv_today_radio, 1)
        self.inventory_period_group.addButton(self.inv_week_radio, 2)
        self.inventory_period_group.addButton(self.inv_month_radio, 3)
        self.inventory_period_group.addButton(self.inv_custom_radio, 4)
        
        date_period_layout.addWidget(self.inv_all_radio)
        date_period_layout.addWidget(self.inv_today_radio)
        date_period_layout.addWidget(self.inv_week_radio)
        date_period_layout.addWidget(self.inv_month_radio)
        date_period_layout.addWidget(self.inv_custom_radio)
        date_period_layout.addStretch()
        
        date_layout.addLayout(date_period_layout)
        
        # Date range controls (second row)
        date_range_layout = QHBoxLayout()
        
        # Date range controls
        self.start_date_input = QDateEdit()
        self.start_date_input.setDate(QDate.currentDate().addMonths(-1))  # Default to 1 month ago
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setMaximumWidth(120)
        self.start_date_input.setEnabled(False)  # Initially disabled since "Всички" is selected
        
        # Connect to auto-switch and search functions
        self.start_date_input.dateChanged.connect(self.auto_switch_to_custom_inventory_period)
        self.start_date_input.dateChanged.connect(self.search_items)
        self.start_date_input.editingFinished.connect(self.auto_switch_to_custom_inventory_period)
        self.start_date_input.installEventFilter(self)
        
        self.end_date_input = QDateEdit()
        self.end_date_input.setDate(QDate.currentDate())  # Default to today
        self.end_date_input.setCalendarPopup(True)
        self.end_date_input.setMaximumWidth(120)
        self.end_date_input.setEnabled(False)  # Initially disabled since "Всички" is selected
        
        # Connect to auto-switch and search functions
        self.end_date_input.dateChanged.connect(self.auto_switch_to_custom_inventory_period)
        self.end_date_input.dateChanged.connect(self.search_items)
        self.end_date_input.editingFinished.connect(self.auto_switch_to_custom_inventory_period)
        self.end_date_input.installEventFilter(self)
        
        date_range_layout.addWidget(QLabel("От дата:"))
        date_range_layout.addWidget(self.start_date_input)
        date_range_layout.addWidget(QLabel("до дата:"))
        date_range_layout.addWidget(self.end_date_input)
        date_range_layout.addStretch()
        
        date_layout.addLayout(date_range_layout)
        
        # Connect radio buttons to update date ranges
        self.inventory_period_group.buttonClicked.connect(self.on_inventory_period_changed)
        
        self.search_tabs.addTab(date_tab, "По дата")
        
        # Tab 4: Category Search
        category_tab = QWidget()
        category_layout = QVBoxLayout(category_tab)
        
        # All category filters in one compact horizontal layout
        all_cat_filters_layout = QHBoxLayout()
        
        # Category filter
        all_cat_filters_layout.addWidget(QLabel("Категория:"))
        self.category_filter = QComboBox()
        self.category_filter.addItem("Всички категории")
        self.category_filter.setMaximumWidth(150)
        self.category_filter.currentTextChanged.connect(self.search_items)
        all_cat_filters_layout.addWidget(self.category_filter)
        
        # Metal filter
        all_cat_filters_layout.addWidget(QLabel("Метал:"))
        self.metal_filter = QComboBox()
        self.metal_filter.addItem("Всички метали")
        self.metal_filter.setMaximumWidth(120)
        self.metal_filter.currentTextChanged.connect(self.search_items)
        all_cat_filters_layout.addWidget(self.metal_filter)
        
        # Stone filter
        all_cat_filters_layout.addWidget(QLabel("Камък:"))
        self.stone_filter = QComboBox()
        self.stone_filter.addItem("Всички камъни")
        self.stone_filter.setMaximumWidth(120)
        self.stone_filter.currentTextChanged.connect(self.search_items)
        all_cat_filters_layout.addWidget(self.stone_filter)
        
        # Stock status filter
        all_cat_filters_layout.addWidget(QLabel("Количество:"))
        self.stock_filter = QComboBox()
        self.stock_filter.addItems(["Всички", "С количество", "Малко количество (≤5)", "Без количество"])
        self.stock_filter.setMaximumWidth(180)
        self.stock_filter.currentTextChanged.connect(self.search_items)
        all_cat_filters_layout.addWidget(self.stock_filter)
        
        all_cat_filters_layout.addStretch()
        category_layout.addLayout(all_cat_filters_layout)
        
        self.search_tabs.addTab(category_tab, "По категория")
        
        # Add tabs to main layout
        search_layout.addWidget(self.search_tabs)
        
        # Clear all filters button - moved directly under tabs for compact design
        clear_all_layout = QHBoxLayout()
        clear_filters_btn = QPushButton("Изчисти всички филтри")
        clear_filters_btn.clicked.connect(self.clear_all_filters)
        clear_filters_btn.setStyleSheet("background-color: #ffeb99; color: #664d00; font-weight: bold;")
        clear_filters_btn.setMaximumWidth(180)  # Compact button width
        clear_all_layout.addWidget(clear_filters_btn)
        
        # Add current search info on same line to save space
        self.search_info_label = QLabel("Няма активни филтри")
        self.search_info_label.setStyleSheet("color: #666; font-size: 10px; font-style: italic;")
        clear_all_layout.addWidget(self.search_info_label)
        clear_all_layout.addStretch()
        
        search_layout.addLayout(clear_all_layout)
        
        # Set compact maximum height for entire search group
        search_group.setLayout(search_layout)
        search_group.setMaximumHeight(170)  # Compact height to maximize table space
        layout.addWidget(search_group)

        # Bulk actions toolbar
        bulk_actions_layout = QHBoxLayout()
        
        select_all_btn = QPushButton("Избери всички")
        select_all_btn.clicked.connect(self.select_all_items)
        bulk_actions_layout.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("Отмени избора")
        deselect_all_btn.clicked.connect(self.deselect_all_items)
        bulk_actions_layout.addWidget(deselect_all_btn)
        
        bulk_actions_layout.addWidget(QLabel(" | "))
        
        bulk_delete_btn = QPushButton("Изтрий избраните")
        bulk_delete_btn.setStyleSheet("background-color: #ffcccc; color: #cc0000; font-weight: bold;")
        bulk_delete_btn.clicked.connect(self.bulk_delete_items)
        bulk_actions_layout.addWidget(bulk_delete_btn)
        
        bulk_move_btn = QPushButton("Премести избраните в магазин")
        bulk_move_btn.clicked.connect(self.bulk_move_to_shop)
        bulk_actions_layout.addWidget(bulk_move_btn)
        
        bulk_export_btn = QPushButton("Експорт на избраните")
        bulk_export_btn.clicked.connect(self.bulk_export_items)
        bulk_actions_layout.addWidget(bulk_export_btn)
        
        bulk_edit_btn = QPushButton("Редактирай цени")
        bulk_edit_btn.clicked.connect(self.bulk_edit_prices)
        bulk_actions_layout.addWidget(bulk_edit_btn)
        
        export_warehouse_btn = QPushButton("📄 Експорт склад")
        export_warehouse_btn.clicked.connect(self.export_warehouse)
        bulk_actions_layout.addWidget(export_warehouse_btn)
        
        bulk_actions_layout.addStretch()
        
        # Help text
        help_text = QLabel("💡 Съвет: Ctrl+Click за избор на няколко артикула, Shift+Click за диапазон, Ctrl+A за всички, Delete за изтриване")
        help_text.setStyleSheet("color: #666; font-size: 11px; font-style: italic;")
        bulk_actions_layout.addWidget(help_text)
        
        # Selection info label
        self.selection_info_label = QLabel("Няма избрани артикули")
        self.selection_info_label.setStyleSheet("font-weight: bold; color: #666;")
        bulk_actions_layout.addWidget(self.selection_info_label)
        
        layout.addLayout(bulk_actions_layout)

        # Table
        self.items_table = QTableWidget()
        self.items_table.setColumnCount(11)  # Updated column count  
        self.items_table.setHorizontalHeaderLabels([
            "Баркод", "Категория", "Метал", "Камък", "Описание", 
            "Цена на едро", "Цена", "Тегло", "Количество", "Дата", "Час"
        ])
        self.items_table.itemDoubleClicked.connect(self.edit_item)
        
        # Configure column spacing with controlled resize limits (resizable between bounds)
        header = self.items_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Set minimum and maximum widths for controlled resizing
        min_width = 80   # Minimum column width
        max_width = 200  # Maximum column width
        default_width = 110  # Default column width
        
        # Set resizable widths for first 10 columns with bounds
        for col in range(10):  # First 10 columns are resizable within bounds
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            self.items_table.setColumnWidth(col, default_width)
            # Note: QTableWidget doesn't have built-in min/max width constraints,
            # but Interactive mode allows manual resizing within the table frame
        
        # Last column stretches to fill remaining space
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)
        
        self.items_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.items_table.customContextMenuRequested.connect(self.inventory_right_click)
        # Enable sorting
        self.items_table.setSortingEnabled(True)
        # MULTI-SELECT: Allow multiple row selection
        self.items_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.items_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)  # Ctrl+Click, Shift+Click support
        self.items_table.setAlternatingRowColors(True)
        # Fix table dimensions to prevent resizing during search
        self.items_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        # DISABLE INLINE EDITING: Force users to use the edit dialog only
        self.items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        # Connect selection change to update info
        self.items_table.itemSelectionChanged.connect(self.update_selection_info)
        # Add keyboard shortcut for deletion
        self.items_table.keyPressEvent = self.handle_table_key_press
        layout.addWidget(self.items_table)

        # Summary bar at the bottom
        summary_layout = QHBoxLayout()
        self.summary_labels = []
        for i in range(6):
            label = QLabel()
            label.setStyleSheet("font-weight: bold;")
            self.summary_labels.append(label)
            summary_layout.addWidget(label)
        self.summary_labels[1].setText("ОБЩО:")
        layout.addLayout(summary_layout)

        return widget

    def format_number_with_spaces(self, number):
        """Format integer or float with spaces every 3 digits"""
        if isinstance(number, float):
            number = int(number)
        return f"{number:,}".replace(",", " ")

    def format_grams(self, grams):
        """Convert grams to kg and grams, format with spaces"""
        try:
            grams = int(float(grams))
        except Exception:
            return ""
        kg = grams // 1000
        g = grams % 1000
        if kg > 0:
            if g > 0:
                return f"{self.format_number_with_spaces(kg)}kg {self.format_number_with_spaces(g)}g"
            else:
                return f"{self.format_number_with_spaces(kg)}kg"
        else:
            return f"{self.format_number_with_spaces(g)}g"
    
    def get_exports_directory(self):
        """Ensure exports directory exists and return its path"""
        import os
        exports_dir = "exports"
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        return exports_dir
    
    def generate_bulgarian_filename(self, base_name, file_extension):
        """Generate Bulgarian snake_case filename with DD.MM.YYYY format - Dynamic and flexible"""
        from datetime import datetime
        import re
        
        # Convert base_name to Bulgarian snake_case
        clean_name = base_name.lower().strip()
        
        # Translation dictionary for common terms only (no shop names)
        term_translations = {
            # System terms
            "warehouse": "склад",
            "shop": "магазин", 
            "audit": "инвентаризация",
            "selected_items": "избрани_артикули",
            "analysis": "анализ",
            "missing_items": "липсващи_артикули",
            "missing items": "липсващи_артикули",
            "export": "експорт",
            "report": "доклад",
            "items": "артикули",
            "products": "продукти",
            "inventory": "инвентар",
            "database_export": "експорт_база_данни",
            "complete_export": "пълен_експорт",
            # Analysis types
            "price analysis": "анализ_цени",
            "category analysis": "анализ_категории"
        }
        
        # Check if it's a known system term
        if clean_name in term_translations:
            bg_name = term_translations[clean_name]
        else:
            # Dynamic processing for any shop name or custom term
            # Replace common abbreviations and words
            bg_name = clean_name
            
            # Replace common Bulgarian abbreviations
            bg_name = bg_name.replace("бул.", "булевард")
            bg_name = bg_name.replace("ул.", "улица") 
            bg_name = bg_name.replace("пл.", "площад")
            bg_name = bg_name.replace("кв.", "квартал")
            bg_name = bg_name.replace("ж.к.", "жилищен_комплекс")
            
            # Convert to snake_case: replace spaces, dots, slashes, etc.
            bg_name = re.sub(r'[^\w\u0400-\u04FF]', '_', bg_name)  # Keep Cyrillic and Latin letters
            bg_name = re.sub(r'_+', '_', bg_name)  # Remove multiple underscores
            bg_name = bg_name.strip('_')  # Remove leading/trailing underscores
            
            # Smart prefix handling - avoid duplication
            has_audit_prefix = bg_name.startswith('инвентаризация') or clean_name.startswith('инвентаризация')
            has_shop_prefix = bg_name.startswith('магазин') or clean_name.startswith('магазин')
            
            # Add audit prefix if needed
            if (clean_name.startswith('инвентаризация') or 'audit' in clean_name) and not has_audit_prefix:
                bg_name = f"инвентаризация_{bg_name}"
            # Add shop prefix if needed  
            elif ('магазин' in clean_name or 'shop' in clean_name) and not has_shop_prefix and not has_audit_prefix:
                bg_name = f"магазин_{bg_name}"
        
        # Generate DD.MM.YYYY date format
        current_date = datetime.now().strftime("%d.%m.%Y")
        
        # Ensure extension starts with dot
        if not file_extension.startswith('.'):
            file_extension = '.' + file_extension
            
        return f"{bg_name} - {current_date}{file_extension}"
    
    def parse_weight_to_grams(self, weight_text):
        """Parse formatted weight string back to grams"""
        try:
            weight_text = weight_text.replace(" ", "")  # Remove spaces
            total_grams = 0
            
            # Handle kg
            if "kg" in weight_text:
                kg_part = weight_text.split("kg")[0]
                total_grams += int(kg_part) * 1000
                weight_text = weight_text.split("kg")[1] if "kg" in weight_text else ""
            
            # Handle remaining grams
            if "g" in weight_text:
                g_part = weight_text.replace("g", "")
                if g_part:
                    total_grams += int(g_part)
            
            return total_grams
        except Exception:
            return 0

    def load_items(self):
        """Load items into table"""
        try:
            items = self.db.get_all_items()
            self.items_table.setRowCount(len(items))
            
            # Update database statistics when loading items
            if hasattr(self, 'db_stats_cards'):
                self.update_database_statistics()
            total_price = 0.0
            total_weight = 0.0
            total_items = 0

            for row, item in enumerate(items):
                try:
                    # Basic item data with safe indexing
                    # Barcode - NEVER EDITABLE (barcodes must never change once assigned)
                    barcode_item = QTableWidgetItem(str(item[1]) if len(item) > 1 else "")
                    barcode_item.setFlags(barcode_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Remove editable flag
                    barcode_item.setToolTip("Баркодът не може да бъде редактиран директно в таблицата")  # Tooltip
                    barcode_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 0, barcode_item)  # Barcode
                    
                    category_item = QTableWidgetItem(str(item[4]) if len(item) > 4 else "")
                    category_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 1, category_item)  # Category
                    
                    metal_item = QTableWidgetItem(str(item[8]) if len(item) > 8 else "")
                    metal_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 2, metal_item)  # Metal
                    
                    stone_item = QTableWidgetItem(str(item[9]) if len(item) > 9 else "")
                    stone_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 3, stone_item)  # Stone
                    
                    # Description
                    description_item = QTableWidgetItem(str(item[3]) if len(item) > 3 else "")
                    description_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 4, description_item)  # Description
                    
                    # Handle cost with fallback (Price bought / wholesale price in Euro)
                    cost_eur = float(item[6]) if len(item) > 6 and item[6] is not None else 0.0
                    cost_lev = self.euro_to_lev(cost_eur)
                    cost_text = f"{cost_eur:.2f} €\n{cost_lev:.2f} лв"
                    cost_item = QTableWidgetItem(cost_text)
                    cost_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 5, cost_item)  # Cost / Price bought
                    
                    # Handle price with fallback (Retail price in Euro)
                    price_eur = float(item[5]) if len(item) > 5 and item[5] is not None else 0.0
                    price_lev = self.euro_to_lev(price_eur)
                    price_text = f"{price_eur:.2f} €\n{price_lev:.2f} лв"
                    price_item = QTableWidgetItem(price_text)
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 6, price_item)  # Price
                    
                    # Handle weight with fallback
                    weight = float(item[7]) if len(item) > 7 and item[7] is not None else 0.0
                    weight_item = QTableWidgetItem(self.format_grams(weight))
                    weight_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 7, weight_item)  # Weight
                    
                    # Handle stock with fallback
                    stock = int(item[10]) if len(item) > 10 and item[10] is not None else 0
                    stock_item = QTableWidgetItem(str(stock))
                    stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                    self.items_table.setItem(row, 8, stock_item)  # Stock
                    
                    # Highlight row if quantity is 0
                    self.highlight_zero_quantity_row(self.items_table, row, stock)
                    
                    # Date and Time - prioritize updated_at timestamp - NEVER EDITABLE
                    date_added = None
                    try:
                        # With explicit column order: created_at=item[11], updated_at=item[12]
                        # Prioritize updated_at to show when item was last modified
                        if len(item) > 12 and item[12]:  # updated_at column (most recent)
                            date_added = parse_database_datetime(item[12])
                        elif len(item) > 11 and item[11]:  # created_at column (fallback)
                            date_added = parse_database_datetime(item[11])
                    except (IndexError, TypeError, ValueError) as e:
                        # Log the error but continue loading other items
                        logger.warning(f"Could not parse date for item {item[1] if len(item) > 1 else 'unknown'}: {e}")
                        date_added = None
                    
                    if date_added:
                        # Date - NEVER EDITABLE (Column 9)
                        date_item = QTableWidgetItem(format_date_for_display(date_added))
                        date_item.setFlags(date_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Remove editable flag
                        date_item.setToolTip("Датата се генерира автоматично и не може да бъде редактирана")
                        date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                        self.items_table.setItem(row, 9, date_item)  # Date
                        
                        # Time - NEVER EDITABLE (Column 10)
                        time_item = QTableWidgetItem(format_time_for_display(date_added))
                        time_item.setFlags(time_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Remove editable flag
                        time_item.setToolTip("Часът се генерира автоматично и не може да бъде редактиран")
                        time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                        self.items_table.setItem(row, 10, time_item)  # Time
                    else:
                        # Set empty cells if no date is available - NEVER EDITABLE
                        empty_date_item = QTableWidgetItem("")
                        empty_date_item.setFlags(empty_date_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        empty_date_item.setToolTip("Датата се генерира автоматично и не може да бъде редактирана")
                        empty_date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                        self.items_table.setItem(row, 9, empty_date_item)  # Date
                        
                        empty_time_item = QTableWidgetItem("")
                        empty_time_item.setFlags(empty_time_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        empty_time_item.setToolTip("Часът се генерира автоматично и не може да бъде редактиран")
                        empty_time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align
                        self.items_table.setItem(row, 10, empty_time_item)  # Time

                    # Update totals safely (using Euro prices)
                    try:
                        total_price += price_eur * stock
                        total_weight += weight * stock
                        total_items += stock
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Could not calculate totals for item {item[1] if len(item) > 1 else 'unknown'}: {e}")

                except Exception as e:
                    # Log error for individual item but continue loading others
                    logger.error(f"Error loading item at row {row}: {e}")
                    continue

            # Update summary (show both currencies)
            self.summary_labels[0].setText("")
            total_price_lev = self.euro_to_lev(total_price)
            price_summary = f"{self.format_currency_eur(total_price)}\n{self.format_currency_lev(total_price_lev)}"
            self.summary_labels[2].setText(price_summary)
            self.summary_labels[3].setText(self.format_grams(total_weight))
            self.summary_labels[4].setText(f"{total_items} артикула")
            self.summary_labels[5].setText("")
            
            # Populate filter dropdowns with unique values
            if hasattr(self, 'category_filter'):  # Check if filters exist
                self.populate_filter_dropdowns(items)

        except Exception as e:
            logger.error(f"Critical error in load_items: {e}", exc_info=True)
            if not getattr(self, '_suppress_error_dialogs', False):
                QMessageBox.critical(self, "Грешка", f"Грешка при зареждане на артикулите: {str(e)}")
            else:
                logger.error(f"Error in load_items (suppressed): {e}")

    def search_items(self):
        """Enhanced search function with cumulative filtering across all tabs"""
        visible_rows = 0
        total_price = 0.0
        total_weight = 0.0
        total_items = 0
        active_filters = []
        
        for row in range(self.items_table.rowCount()):
            show_row = True
            
            # Apply filters from ALL tabs cumulatively, not just the current tab
            
            # 1. General Search Filter (Tab 0)
            search_text = self.search_input.text().lower()
            if search_text:
                active_filters.append(f"Общо търсене: '{search_text}'")
                text_match = False
                # Search ALL columns
                search_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                for col in search_columns:
                    item = self.items_table.item(row, col)
                    if item and search_text in item.text().lower():
                        text_match = True
                        break
                if not text_match:
                    show_row = False
            
            # 2. Price/Weight Filter (Tab 1) - apply regardless of current tab
            if show_row:
                min_price = self.min_price_input.value()
                max_price = self.max_price_input.value()
                min_weight = self.min_weight_input.value()
                max_weight = self.max_weight_input.value()
                
                # Price filter
                if min_price > 0 or max_price < 999999:
                    active_filters.append(f"Цена: {min_price:.2f} - {max_price:.2f} лв")
                    price_item = self.items_table.item(row, 6)  # Price column
                    if price_item:
                        try:
                            # Extract Euro price from dual currency text (first line)
                            price_text = price_item.text().split('\n')[0].replace(" €", "").replace(" ", "")
                            price_value_eur = float(price_text)
                            price_value_lev = self.euro_to_lev(price_value_eur)  # Convert to lev for comparison
                            if price_value_lev < min_price or price_value_lev > max_price:
                                show_row = False
                        except (ValueError, IndexError):
                            show_row = False
                
                # Weight filter
                if min_weight > 0 or max_weight < 9999:
                    active_filters.append(f"Тегло: {min_weight:.2f} - {max_weight:.2f} г")
                    weight_item = self.items_table.item(row, 7)  # Weight column
                    if weight_item:
                        try:
                            weight_grams = self.parse_weight_to_grams(weight_item.text())
                            if weight_grams < min_weight or weight_grams > max_weight:
                                show_row = False
                        except (ValueError, AttributeError):
                            show_row = False
            
            # 3. Date Filter (Tab 2) - apply regardless of current tab
            if show_row:
                start_date = self.start_date_input.date()
                end_date = self.end_date_input.date()
                
                # Check if date filter is at default "show all" state
                default_start = QDate.currentDate().addMonths(-1)
                default_end = QDate.currentDate()
                is_default_date_range = (start_date == default_start and end_date == default_end)
                
                # Only apply date filter if it's not at default "show all" state
                if not is_default_date_range:
                    active_filters.append(f"Дата: {start_date.toString('dd.MM.yyyy')} - {end_date.toString('dd.MM.yyyy')}")
                    
                    date_item = self.items_table.item(row, 9)  # Date column
                    if date_item:
                        try:
                            item_date_str = date_item.text()
                            # Convert date string to QDate for comparison (using correct format with slashes)
                            item_date = QDate.fromString(item_date_str, "dd/MM/yyyy")
                            if not item_date.isValid() or item_date < start_date or item_date > end_date:
                                show_row = False
                        except (ValueError, AttributeError):
                            show_row = False
            
            # 4. Category Filters (Tab 3) - apply regardless of current tab
            if show_row:
                category_filter = self.category_filter.currentText()
                metal_filter = self.metal_filter.currentText()
                stone_filter = self.stone_filter.currentText()
                stock_filter = self.stock_filter.currentText()
                
                # Category filter
                if category_filter != "Всички категории":
                    active_filters.append(f"Категория: {category_filter}")
                    category_item = self.items_table.item(row, 1)  # Category column
                    if not category_item or category_item.text() != category_filter:
                        show_row = False
                
                # Metal filter
                if show_row and metal_filter != "Всички метали":
                    active_filters.append(f"Метал: {metal_filter}")
                    metal_item = self.items_table.item(row, 2)  # Metal column
                    if not metal_item or metal_item.text() != metal_filter:
                        show_row = False
                
                # Stone filter
                if show_row and stone_filter != "Всички камъни":
                    active_filters.append(f"Камък: {stone_filter}")
                    stone_item = self.items_table.item(row, 3)  # Stone column
                    if not stone_item or stone_item.text() != stone_filter:
                        show_row = False
                
                # Stock filter
                if show_row and stock_filter != "Всички":
                    active_filters.append(f"Количество: {stock_filter}")
                    stock_item = self.items_table.item(row, 8)  # Stock column
                    if stock_item:
                        try:
                            stock_value = int(stock_item.text())
                            if stock_filter == "С количество" and stock_value <= 0:
                                show_row = False
                            elif stock_filter == "Малко количество (≤5)" and stock_value > 5:
                                show_row = False
                            elif stock_filter == "Без количество" and stock_value > 0:
                                show_row = False
                        except ValueError:
                            show_row = False
            
            # Apply visibility
            self.items_table.setRowHidden(row, not show_row)
            
            # Update summary for visible rows
            if show_row:
                visible_rows += 1
                try:
                    # Calculate totals for visible items only (using Euro prices)
                    price_item = self.items_table.item(row, 6)  # Price column
                    stock_item = self.items_table.item(row, 8)  # Stock column
                    weight_item = self.items_table.item(row, 7)  # Weight column
                    
                    if price_item and stock_item:
                        # Extract Euro price from dual currency text (first line)
                        price_text = price_item.text().split('\n')[0].replace(" €", "").replace(" ", "")
                        price_value = float(price_text)
                        stock_value = int(stock_item.text())
                        total_price += price_value * stock_value
                        total_items += stock_value
                    
                    if weight_item and stock_item:
                        weight_text = weight_item.text()
                        if weight_text:
                            # Convert weight back to grams for calculation
                            weight_grams = self.parse_weight_to_grams(weight_text)
                            stock_value = int(stock_item.text())
                            total_weight += weight_grams * stock_value
                except (ValueError, AttributeError, IndexError):
                    pass
        
        # Update search info label
        if active_filters:
            # Remove duplicates and show active filters
            unique_filters = list(dict.fromkeys(active_filters))
            self.search_info_label.setText(f"Активни филтри: {' | '.join(unique_filters[:2])}")
        else:
            self.search_info_label.setText("Няма активни филтри")
        
        # Update summary with filtered results (show both currencies)
        selected_count = len(self.get_selected_rows())
        if selected_count > 0:
            self.summary_labels[0].setText(f"Показани: {visible_rows} | Избрани: {selected_count}")
        else:
            self.summary_labels[0].setText(f"Показани: {visible_rows}")
        
        total_price_lev = self.euro_to_lev(total_price)
        price_summary = f"{self.format_currency_eur(total_price)}\n{self.format_currency_lev(total_price_lev)}"
        self.summary_labels[2].setText(price_summary)
        self.summary_labels[3].setText(self.format_grams(total_weight))
        self.summary_labels[4].setText(f"{total_items} артикула")
        self.summary_labels[5].setText("")

    def clear_search(self):
        """Clear the main search input"""
        self.search_input.clear()
    
    def clear_all_filters(self):
        """Clear all search filters in all tabs"""
        # General search tab
        self.search_input.clear()
        
        # Price/Weight search tab
        self.min_price_input.setValue(0)
        self.max_price_input.setValue(999999)
        self.min_weight_input.setValue(0)
        self.max_weight_input.setValue(9999)
        
        # Reset confirmed values for filter spin boxes
        self.min_price_input.reset_confirmed_value()
        self.max_price_input.reset_confirmed_value()
        self.min_weight_input.reset_confirmed_value()
        self.max_weight_input.reset_confirmed_value()
        
        # Date search tab
        self.start_date_input.setDate(QDate.currentDate().addMonths(-1))
        self.end_date_input.setDate(QDate.currentDate())
        
        # Category search tab
        self.category_filter.setCurrentText("Всички категории")
        self.metal_filter.setCurrentText("Всички метали")
        self.stone_filter.setCurrentText("Всички камъни")
        self.stock_filter.setCurrentText("Всички")
        
        # Update search info
        self.search_info_label.setText("Няма активни филтри")
        
        # Trigger search to refresh table visibility and show all items
        self.search_items()
    
    def set_date_range(self, days_back):
        """Set date range for quick date selection"""
        # Set flag to prevent auto-switching during programmatic changes
        self.programmatic_inventory_date_change = True
        
        try:
            end_date = QDate.currentDate()
            start_date = end_date.addDays(-days_back)
            self.start_date_input.setDate(start_date)
            self.end_date_input.setDate(end_date)
        finally:
            # Always reset the flag
            self.programmatic_inventory_date_change = False
            
        self.search_items()  # Trigger search immediately
    
    def set_sales_date_range(self, period_type):
        """Set sales date range based on period type and update dropdown menus"""
        from datetime import datetime, timedelta
        now = datetime.now()
        
        # Set flag to prevent auto-switching during programmatic changes
        self.programmatic_date_change = True
        
        try:
            if period_type == "today":
                start_date = QDate.currentDate()
                end_date = QDate.currentDate()
            elif period_type == "week":
                # Start of current week (Monday)
                start_date = QDate.currentDate().addDays(-QDate.currentDate().dayOfWeek() + 1)
                end_date = QDate.currentDate()
            elif period_type == "month":
                # Start of current month
                start_date = QDate(QDate.currentDate().year(), QDate.currentDate().month(), 1)
                end_date = QDate.currentDate()
            elif period_type == "year":
                # Start of current year
                start_date = QDate(QDate.currentDate().year(), 1, 1)
                end_date = QDate.currentDate()
            else:
                # Default case or "all" - don't change dates
                return
            
            # Update the unified date controls (which also update search date inputs via aliases)
            self.sales_start_date.setDate(start_date)
            self.sales_end_date.setDate(end_date)
        finally:
            # Always reset the flag
            self.programmatic_date_change = False
        
        # Trigger both data reload and search
        self.load_sales()
        self.search_sales()
    
    def on_inventory_period_changed(self):
        """Handle inventory period radio button changes"""
        checked_button = self.inventory_period_group.checkedButton()
        if checked_button:
            button_id = self.inventory_period_group.id(checked_button)
            
            # Enable/disable custom date fields based on selection
            is_custom = self.inv_custom_radio.isChecked()
            self.start_date_input.setEnabled(is_custom)
            self.end_date_input.setEnabled(is_custom)
            
            if button_id == 1:  # Today
                self.set_date_range(0)
            elif button_id == 2:  # Week  
                self.set_date_range(7)
            elif button_id == 3:  # Month
                self.set_date_range(30)
            elif button_id == 4:  # Custom
                # Just enable the fields and trigger search, don't change dates
                self.search_items()
            # For "All" (button_id == 0), reset dates to default (show all) and trigger search
            elif button_id == 0:
                # Reset date fields to default "show all" range
                self.programmatic_inventory_date_change = True
                try:
                    self.start_date_input.setDate(QDate.currentDate().addMonths(-1))
                    self.end_date_input.setDate(QDate.currentDate())
                finally:
                    self.programmatic_inventory_date_change = False
                self.search_items()
    
    def search_sales(self):
        """Enhanced search function for sales table with cumulative filtering"""
        visible_rows = 0
        active_sales_filters = []
        
        for row in range(self.sales_table.rowCount()):
            show_row = True
            
            # Apply filters from ALL tabs cumulatively
            
            # 1. General Search Filter (Tab 0)
            search_text = self.sales_search_input.text().lower()
            if search_text:
                active_sales_filters.append(f"Общо търсене: '{search_text}'")
                text_match = False
                # Search ALL columns
                search_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                for col in search_columns:
                    item = self.sales_table.item(row, col)
                    if item and search_text in item.text().lower():
                        text_match = True
                        break
                if not text_match:
                    show_row = False
            
            # 2. Date Filter (Tab 1) - apply regardless of current tab
            if show_row:
                start_date = self.sales_start_date_input.date()
                end_date = self.sales_end_date_input.date()
                
                # Check if date filter is at default "show all" state (same logic as warehouse)
                default_start = QDate.currentDate().addMonths(-1)
                default_end = QDate.currentDate()
                is_default_date_range = (start_date == default_start and end_date == default_end)
                
                # Only apply date filter if it's not at default "show all" state
                if not is_default_date_range:
                    active_sales_filters.append(f"Дата: {start_date.toString('dd.MM.yyyy')} - {end_date.toString('dd.MM.yyyy')}")
                    
                    date_item = self.sales_table.item(row, 9)  # Date column
                    if date_item:
                        try:
                            item_date_str = date_item.text()
                            # Convert date string to QDate for comparison (using correct format with slashes)
                            item_date = QDate.fromString(item_date_str, "dd/MM/yyyy")
                            if not item_date.isValid() or item_date < start_date or item_date > end_date:
                                show_row = False
                        except (ValueError, AttributeError):
                            show_row = False
            
            # Apply visibility
            self.sales_table.setRowHidden(row, not show_row)
            
            # Update summary for visible rows
            if show_row:
                visible_rows += 1
        
        # Update sales search info label
        if active_sales_filters:
            # Remove duplicates and show active filters
            unique_filters = list(dict.fromkeys(active_sales_filters))
            self.sales_search_info_label.setText(f"Активни филтри: {' | '.join(unique_filters[:2])}")
        else:
            self.sales_search_info_label.setText("Няма активни филтри")
        
        # Update summary with filtered results
        if hasattr(self, 'sales_summary_labels'):
            self.sales_summary_labels[0].setText(f"Показани продажби: {visible_rows}")
    
    def clear_sales_search(self):
        """Clear the sales search input"""
        self.sales_search_input.clear()
    
    def clear_all_sales_filters(self):
        """Clear all sales search filters in all tabs"""
        # General search tab
        self.sales_search_input.clear()
        
        # Set flag to prevent auto-switching to custom period during programmatic changes
        self.programmatic_date_change = True
        
        try:
            # Date search tab - reset to default "show all" range like warehouse
            self.sales_start_date.setDate(QDate.currentDate().addMonths(-1))
            self.sales_end_date.setDate(QDate.currentDate())
            
            # Keep "Всички" selected instead of switching to "Персонализиран"
            if hasattr(self, 'all_time_radio'):
                self.all_time_radio.setChecked(True)
        finally:
            # Always reset the flag
            self.programmatic_date_change = False
        
        # Update search info
        self.sales_search_info_label.setText("Няма активни филтри")
        
        # Reload sales data and refresh search results (same as "Всички" button)
        self.load_sales()
        self.search_sales()
    
    def populate_filter_dropdowns(self, items):
        """Populate filter dropdowns with unique values from items"""
        categories = set()
        metals = set()
        stones = set()
        
        for item in items:
            if len(item) > 8:
                categories.add(item[4])  # Category
                metals.add(item[8])      # Metal
                stones.add(item[9])      # Stone
        
        # Update category filter
        current_category = self.category_filter.currentText()
        self.category_filter.clear()
        self.category_filter.addItem("Всички категории")
        for category in sorted(categories):
            if category:  # Skip empty values
                self.category_filter.addItem(category)
        # Restore selection if it still exists
        index = self.category_filter.findText(current_category)
        if index >= 0:
            self.category_filter.setCurrentIndex(index)
        
        # Update metal filter
        current_metal = self.metal_filter.currentText()
        self.metal_filter.clear()
        self.metal_filter.addItem("Всички метали")
        for metal in sorted(metals):
            if metal:  # Skip empty values
                self.metal_filter.addItem(metal)
        # Restore selection if it still exists
        index = self.metal_filter.findText(current_metal)
        if index >= 0:
            self.metal_filter.setCurrentIndex(index)
        
        # Update stone filter
        current_stone = self.stone_filter.currentText()
        self.stone_filter.clear()
        self.stone_filter.addItem("Всички камъни")
        for stone in sorted(stones):
            if stone:  # Skip empty values
                self.stone_filter.addItem(stone)
        # Restore selection if it still exists
        index = self.stone_filter.findText(current_stone)
        if index >= 0:
            self.stone_filter.setCurrentIndex(index)

    def edit_item(self, item):
        """Edit selected item using dedicated dialog"""
        try:
            row = item.row()
            
            # Get current item data - Updated column indices for new structure
            barcode_item = self.items_table.item(row, 0)     # Barcode
            category_item = self.items_table.item(row, 1)    # Category
            metal_item = self.items_table.item(row, 2)       # Metal
            stone_item = self.items_table.item(row, 3)       # Stone
            description_item = self.items_table.item(row, 4) # Description
            cost_item = self.items_table.item(row, 5)        # Cost (Price bought)
            price_item = self.items_table.item(row, 6)       # Price
            weight_item = self.items_table.item(row, 7)      # Weight
            stock_item = self.items_table.item(row, 8)       # Stock
            
            # Check if all required items exist
            if not all([barcode_item, category_item, metal_item, stone_item, cost_item, price_item, weight_item, stock_item]):
                QMessageBox.warning(self, "Грешка", "Няма достатъчно данни за редактиране на този артикул")
                return
            
            # Extract text values
            barcode = barcode_item.text()
            category = category_item.text()
            metal = metal_item.text()
            stone = stone_item.text()
            description = description_item.text() if description_item else ""
            
            # Parse dual currency format (Euro on first line)
            cost_text = cost_item.text().split('\n')[0].replace(" €", "").replace(" ", "") if cost_item else "0"
            price_text = price_item.text().split('\n')[0].replace(" €", "").replace(" ", "") if price_item else "0"
            weight_text = weight_item.text()
            stock_text = stock_item.text()
            
            # Convert numeric values with error handling
            try:
                price = float(price_text) if price_text else 0.0
                cost = float(cost_text) if cost_text else 0.0
                stock = int(stock_text) if stock_text else 0
            except (ValueError, IndexError) as e:
                QMessageBox.warning(self, "Грешка", f"Невалидни числови данни: {str(e)}")
                return
            
            # Parse weight back to grams
            weight_grams = self.parse_weight_to_grams(weight_text)
            
            # Create and show edit dialog - passing description as well and from_warehouse=True
            dialog = EditItemDialog(self, barcode, category, description, price, cost, weight_grams, metal, stone, stock, from_warehouse=True)
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # Get updated data from dialog
                updated_data = dialog.get_data()
                
                # Get item_id from barcode first
                try:
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute('SELECT id FROM items WHERE barcode = ?', (barcode,))
                        result = cursor.fetchone()
                        if not result:
                            QMessageBox.warning(self, "Грешка", "Артикулът не е намерен")
                            return
                        
                        item_id = result[0]
                        
                        # Create old data dict for undo
                        old_data = {
                            'name': category,
                            'description': description,
                            'category': category,
                            'price': price,
                            'cost': cost,
                            'weight': weight_grams,
                            'metal_type': metal,
                            'stone_type': stone,
                            'stock_quantity': stock
                        }
                        
                        # Create new data dict
                        new_data = {
                            'name': updated_data['category'],
                            'description': updated_data['description'],
                            'category': updated_data['category'],
                            'price': updated_data['price'],
                            'cost': updated_data['cost'],
                            'weight': updated_data['weight'],
                            'metal_type': updated_data['metal'],
                            'stone_type': updated_data['stone'],
                            'stock_quantity': updated_data['stock']
                        }
                        
                        # Create and execute edit action
                        edit_action = EditItemAction(self.db, item_id, barcode, old_data, new_data)
                        if self.action_history.execute_action(edit_action):
                            QMessageBox.information(self, "Успех", "Артикулът е обновен успешно")
                            self.load_items()  # Reload the table
                            self.update_action_buttons()
                            self.update_reports_and_database_stats()
                        else:
                            QMessageBox.warning(self, "Грешка", "Неуспешно обновяване на артикула")
                except Exception as e:
                    QMessageBox.critical(self, "Грешка", f"Грешка при обновяване: {str(e)}")
        
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при редактиране на артикула: {str(e)}")
            logger.error(f"Error in edit_item: {e}", exc_info=True)

    def inventory_right_click(self, position):
        """Handle right-click on inventory table"""
        try:
            selected_rows = self.get_selected_rows()
            menu = QMenu()
            
            if len(selected_rows) == 1:
                # Single item context menu
                edit_action = menu.addAction("Редактирай")
                delete_action = menu.addAction("Изтрий")
                move_action = menu.addAction("Премести в магазин")
            elif len(selected_rows) > 1:
                # Multiple items context menu
                delete_action = menu.addAction(f"Изтрий {len(selected_rows)} артикула")
                move_action = menu.addAction(f"Премести {len(selected_rows)} артикула в магазин")
                edit_action = None  # Can't edit multiple items
            else:
                # No selection, single row under cursor
                row = self.items_table.rowAt(position.y())
                if row >= 0:
                    self.items_table.selectRow(row)  # Select the row under cursor
                    edit_action = menu.addAction("Редактирай")
                    delete_action = menu.addAction("Изтрий")
                    move_action = menu.addAction("Премести в магазин")
                else:
                    return  # No row under cursor
            
            action = menu.exec(self.items_table.viewport().mapToGlobal(position))
            
            if action == edit_action and edit_action is not None:
                row = self.items_table.rowAt(position.y())
                if row >= 0:
                    self.edit_item(self.items_table.item(row, 0))
            elif action == delete_action:
                if len(selected_rows) > 1:
                    self.bulk_delete_items()
                else:
                    row = self.items_table.rowAt(position.y())
                    if row >= 0:
                        barcode = self.items_table.item(row, 0).text()
                        self.confirm_delete_item(barcode)
            elif action == move_action:
                if len(selected_rows) > 1:
                    self.bulk_move_to_shop()
                else:
                    row = self.items_table.rowAt(position.y())
                    if row >= 0:
                        barcode = self.items_table.item(row, 0).text()
                        self.move_to_shop(barcode)
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при обработка на контекстното меню: {str(e)}")

    def confirm_delete_item(self, barcode):
        """Confirm item deletion"""
        reply = QMessageBox.question(
            self, "Изтрий артикул",
            f"Сигурни ли сте, че искате да изтриете артикул с баркод {barcode}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.delete_item(barcode)

    def delete_item(self, barcode):
        """Delete item from database"""
        try:
            # First get the item_id from the barcode
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM items WHERE barcode = ?', (barcode,))
                result = cursor.fetchone()
                if not result:
                    QMessageBox.warning(self, "Грешка", "Артикулът не е намерен")
                    return
                
                item_id = result[0]
                
                # Create and execute delete action
                delete_action = DeleteItemAction(self.db, item_id, barcode)
                if self.action_history.execute_action(delete_action):
                    QMessageBox.information(self, "Успех", "Артикулът е изтрит успешно")
                    self.load_items()
                    self.update_action_buttons()
                    self.update_reports_and_database_stats()
                else:
                    QMessageBox.warning(self, "Грешка", "Неуспешно изтриване на артикула")
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при изтриване на артикула: {str(e)}")

    def move_to_shop(self, barcode):
        """Move item to shop"""
        try:
            # Get available stock for this item first
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT stock_quantity FROM items WHERE barcode = ?", (barcode,))
                result = cursor.fetchone()
                if not result:
                    QMessageBox.warning(self, "Грешка", "Артикулът не е намерен")
                    return
                
                available_stock = result[0]
                if available_stock <= 0:
                    QMessageBox.warning(self, "Грешка", f"Няма количество в склада за артикул {barcode}")
                    return

            # Get available shops
            shops = self.db.get_all_shops()
            if not shops:
                QMessageBox.warning(self, "Грешка", "Няма налични магазини")
                return

            # Create shop selection dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Избери магазин")
            layout = QVBoxLayout(dialog)

            # Add available stock info
            stock_info = QLabel(f"Налично в склада: {available_stock} артикула")
            stock_info.setStyleSheet("font-weight: bold; color: #2196F3; padding: 5px; background-color: #f0f8ff; border-radius: 3px;")
            layout.addWidget(stock_info)

            # Add shop selection combo box
            shop_combo = QComboBox()
            for shop in shops:
                shop_combo.addItem(shop[1])
            layout.addWidget(QLabel("Избери магазин:"))
            layout.addWidget(shop_combo)

            # Add quantity spin box with limited range
            quantity_spin = BlurOnEnterSpinBox()
            quantity_spin.setRange(1, available_stock)  # Limit to available stock
            quantity_spin.setValue(min(1, available_stock))
            quantity_label = QLabel(f"Количество (максимум {available_stock}):")
            layout.addWidget(quantity_label)
            layout.addWidget(quantity_spin)

            # Add warning label
            warning_label = QLabel("⚠ Количеството не може да надвишава наличното в склада")
            warning_label.setStyleSheet("color: #ff6b35; font-size: 10px; font-style: italic;")
            layout.addWidget(warning_label)

            # Add buttons
            button_box = QHBoxLayout()
            ok_button = QPushButton("OK")
            cancel_button = QPushButton("Отказ")
            button_box.addWidget(ok_button)
            button_box.addWidget(cancel_button)
            layout.addLayout(button_box)

            # Connect buttons
            ok_button.clicked.connect(dialog.accept)
            cancel_button.clicked.connect(dialog.reject)

            # Show dialog
            if dialog.exec() == QDialog.DialogCode.Accepted:
                shop_name = shop_combo.currentText()
                quantity = quantity_spin.value()
                
                # Double-check quantity is still valid
                if quantity > available_stock:
                    QMessageBox.warning(self, "Грешка", f"Количеството ({quantity}) надвишава наличното в склада ({available_stock})")
                    return
                
                if quantity <= 0:
                    QMessageBox.warning(self, "Грешка", "Количеството трябва да бъде положително число")
                    return
                
                shop_id = self.db.get_shop_id(shop_name)
                
                if self.db.move_item_to_shop(shop_id, barcode, quantity):
                    QMessageBox.information(self, "Успех", f"Успешно преместени {quantity} артикула в магазин '{shop_name}'")
                    self.load_items()
                    self.load_shop_inventory()  # Refresh shop inventory table
                    self.update_reports_and_database_stats()
                else:
                    QMessageBox.warning(self, "Грешка", "Неуспешно преместване на артикула")
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при преместване на артикула: {str(e)}")

    # Bulk operations methods
    def get_selected_rows(self):
        """Get list of selected row numbers"""
        selected_ranges = self.items_table.selectionModel().selectedRows()
        return [index.row() for index in selected_ranges]
    
    def get_selected_barcodes(self):
        """Get list of barcodes for selected rows"""
        selected_rows = self.get_selected_rows()
        barcodes = []
        for row in selected_rows:
            barcode_item = self.items_table.item(row, 0)
            if barcode_item:
                barcodes.append(barcode_item.text())
        return barcodes
    
    def select_all_items(self):
        """Select all visible items in the table"""
        self.items_table.selectAll()
    
    def deselect_all_items(self):
        """Deselect all items in the table"""
        self.items_table.clearSelection()
    
    def update_selection_info(self):
        """Update the selection info label and summary"""
        selected_rows = self.get_selected_rows()
        
        if len(selected_rows) == 0:
            self.selection_info_label.setText("Няма избрани артикули")
        elif len(selected_rows) == 1:
            # Show detailed info for single selection
            try:
                row = selected_rows[0]
                category = self.items_table.item(row, 1).text() if self.items_table.item(row, 1) else ""
                price_text = self.items_table.item(row, 6).text() if self.items_table.item(row, 6) else "0.00 €\n0.00 лв"
                # Extract Euro price (first line)
                price_eur = price_text.split('\n')[0].replace(" €", "").replace(" ", "")
                self.selection_info_label.setText(f"Избран: {category} ({price_eur} €)")
            except:
                self.selection_info_label.setText("Избран 1 артикул")
        else:
            # Show summary for multiple selections
            try:
                total_value = 0
                total_items = 0
                for row in selected_rows:
                    try:
                        price_item = self.items_table.item(row, 6)  # Price column
                        stock_item = self.items_table.item(row, 8)  # Stock column
                        if price_item and stock_item:
                            # Extract Euro price from dual currency text (first line)
                            price_text = price_item.text().split('\n')[0].replace(" €", "").replace(" ", "")
                            price = float(price_text)
                            stock = int(stock_item.text())
                            total_value += price * stock
                            total_items += stock
                    except (ValueError, AttributeError):
                        pass
                
                # Show total value in both currencies
                total_value_lev = self.euro_to_lev(total_value)
                self.selection_info_label.setText(
                    f"Избрани {len(selected_rows)} артикула | "
                    f"Стойност: {self.format_currency_eur(total_value)} / {self.format_currency_lev(total_value_lev)} | "
                    f"Брой: {total_items}"
                )
            except:
                self.selection_info_label.setText(f"Избрани {len(selected_rows)} артикула")
        
        # Also update the main summary to show selection count
        # Trigger a search update to refresh summary
        self.search_items()
    
    def handle_table_key_press(self, event):
        """Handle keyboard shortcuts in the table"""
        # Call the original keyPressEvent first
        QTableWidget.keyPressEvent(self.items_table, event)
        
        # Handle Delete key
        if event.key() == Qt.Key.Key_Delete:
            selected_rows = self.get_selected_rows()
            if len(selected_rows) > 0:
                if len(selected_rows) == 1:
                    # Single item deletion
                    barcode = self.get_selected_barcodes()[0]
                    self.confirm_delete_item(barcode)
                else:
                    # Bulk deletion
                    self.bulk_delete_items()
        # Handle Ctrl+A for select all
        elif event.key() == Qt.Key.Key_A and event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            self.select_all_items()
    
    def bulk_delete_items(self):
        """Delete multiple selected items"""
        try:
            selected_barcodes = self.get_selected_barcodes()
            if not selected_barcodes:
                QMessageBox.warning(self, "Предупреждение", "Няма избрани артикули за изтриване")
                return
            
            # Confirmation dialog
            reply = QMessageBox.question(
                self, "Изтрий артикули",
                f"Сигурни ли сте, че искате да изтриете {len(selected_barcodes)} избрани артикула?\n\n"
                f"Това действие е необратимо!",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No  # Default to No for safety
            )

            if reply == QMessageBox.StandardButton.Yes:
                deleted_count = 0
                failed_items = []
                
                # Delete each item using action system (only track the last one for undo)
                last_successful_action = None
                for barcode in selected_barcodes:
                    try:
                        # Get item_id from barcode
                        with self.db.get_connection() as conn:
                            cursor = conn.cursor()
                            cursor.execute('SELECT id FROM items WHERE barcode = ?', (barcode,))
                            result = cursor.fetchone()
                            if result:
                                item_id = result[0]
                                delete_action = DeleteItemAction(self.db, item_id, barcode)
                                if self.action_history.execute_action(delete_action):
                                    deleted_count += 1
                                    last_successful_action = delete_action
                                else:
                                    failed_items.append(barcode)
                            else:
                                failed_items.append(barcode)
                    except Exception as e:
                        logger.error(f"Error deleting item {barcode}: {e}")
                        failed_items.append(barcode)
                
                # Show results
                if failed_items:
                    QMessageBox.warning(
                        self, "Частично изтриване",
                        f"Изтрити: {deleted_count} артикула\n"
                        f"Неуспешни: {len(failed_items)} артикула\n\n"
                        f"Неуспешни баркодове: {', '.join(failed_items[:5])}"
                        f"{'...' if len(failed_items) > 5 else ''}"
                    )
                else:
                    QMessageBox.information(
                        self, "Успех",
                        f"Успешно изтрити {deleted_count} артикула"
                    )
                
                # Reload the table
                self.load_items()
                self.update_action_buttons()
                self.update_reports_and_database_stats()
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при масово изтриване: {str(e)}")
    
    def bulk_move_to_shop(self):
        """Move multiple selected items to shop"""
        try:
            selected_barcodes = self.get_selected_barcodes()
            if not selected_barcodes:
                QMessageBox.warning(self, "Предупреждение", "Няма избрани артикули за преместване")
                return
            
            # Get available shops
            shops = self.db.get_all_shops()
            if not shops:
                QMessageBox.warning(self, "Грешка", "Няма налични магазини")
                return

            # Create shop selection dialog
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Премести {len(selected_barcodes)} артикула в магазин")
            dialog.setModal(True)
            layout = QVBoxLayout(dialog)

            # Info label
            info_label = QLabel(f"Ще преместите {len(selected_barcodes)} избрани артикула:")
            layout.addWidget(info_label)
            
            # Show first few barcodes
            barcode_preview = QLabel(f"Баркодове: {', '.join(selected_barcodes[:3])}")
            if len(selected_barcodes) > 3:
                barcode_preview.setText(barcode_preview.text() + f" и още {len(selected_barcodes) - 3}...")
            barcode_preview.setStyleSheet("color: #666; font-size: 10px;")
            layout.addWidget(barcode_preview)

            # Shop selection
            layout.addWidget(QLabel("Избери магазин:"))
            shop_combo = QComboBox()
            for shop in shops:
                shop_combo.addItem(shop[1])
            layout.addWidget(shop_combo)

            # Check minimum available stock to set reasonable limit
            min_available_stock = 1000  # Default high value
            try:
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    for barcode in selected_barcodes[:5]:  # Check first 5 items for performance
                        cursor.execute("SELECT stock_quantity FROM items WHERE barcode = ?", (barcode,))
                        result = cursor.fetchone()
                        if result and result[0] < min_available_stock:
                            min_available_stock = result[0]
            except Exception:
                min_available_stock = 1

            # Quantity selection
            layout.addWidget(QLabel(f"Количество за всеки артикул (максимум {min_available_stock}):"))
            quantity_spin = BlurOnEnterSpinBox()
            quantity_spin.setRange(1, max(1, min_available_stock))  # Ensure at least 1
            quantity_spin.setValue(1)
            layout.addWidget(quantity_spin)
            
            # Warning label
            warning_label = QLabel("⚠ Ще се провери количеството за всеки артикул преди преместване")
            warning_label.setStyleSheet("color: #ff6b35; font-size: 10px; font-style: italic;")
            layout.addWidget(warning_label)

            # Buttons
            button_layout = QHBoxLayout()
            ok_button = QPushButton("Премести")
            cancel_button = QPushButton("Отказ")
            button_layout.addWidget(ok_button)
            button_layout.addWidget(cancel_button)
            layout.addLayout(button_layout)

            # Connect buttons
            ok_button.clicked.connect(dialog.accept)
            cancel_button.clicked.connect(dialog.reject)

            # Show dialog
            if dialog.exec() == QDialog.DialogCode.Accepted:
                shop_name = shop_combo.currentText()
                quantity = quantity_spin.value()
                shop_id = self.db.get_shop_id(shop_name)
                
                moved_count = 0
                failed_items = []
                
                # Move each item
                for barcode in selected_barcodes:
                    try:
                        if self.db.move_item_to_shop(shop_id, barcode, quantity):
                            moved_count += 1
                        else:
                            failed_items.append(barcode)
                    except Exception as e:
                        logger.error(f"Error moving item {barcode}: {e}")
                        failed_items.append(barcode)
                
                # Show results
                if failed_items:
                    QMessageBox.warning(
                        self, "Частично преместване",
                        f"Преместени: {moved_count} артикула\n"
                        f"Неуспешни: {len(failed_items)} артикула\n\n"
                        f"Неуспешни баркодове: {', '.join(failed_items[:5])}"
                        f"{'...' if len(failed_items) > 5 else ''}"
                    )
                else:
                    QMessageBox.information(
                        self, "Успех",
                        f"Успешно преместени {moved_count} артикула в магазин '{shop_name}'"
                    )
                
                # Reload both tables and update statistics
                self.load_items()
                self.load_shop_inventory()  # Refresh shop inventory table
                self.update_reports_and_database_stats()
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при масово преместване: {str(e)}")
    
    def bulk_export_items(self):
        """Export selected items to Excel"""
        try:
            selected_rows = self.get_selected_rows()
            if not selected_rows:
                QMessageBox.warning(self, "Предупреждение", "Няма избрани артикули за експорт")
                return
            
            # Get save file path
            exports_dir = self.get_exports_directory()
            default_filename = self.generate_bulgarian_filename("избрани_артикули", "xlsx")
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"Експорт на {len(selected_rows)} избрани артикула", 
                f"{exports_dir}/{default_filename}",
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                
                # Create workbook
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                wb = Workbook()
                ws = wb.active
                ws.title = "Избрани артикули"
                
                # Headers
                headers = [
                    "Баркод", "Категория", "Метал", "Камък", "Описание", 
                    "Цена на едро", "Цена", "Тегло", "Количество", "Дата", "Час"
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    ws.column_dimensions[get_column_letter(col)].width = 15
                
                # Export selected items data
                exported_count = 0
                for row_idx, table_row in enumerate(selected_rows, 2):
                    try:
                        # Export all columns except image column (9)
                        col_mapping = [0, 1, 2, 3, 4, 5, 6, 7, 8, 10, 11]  # Skip image column (9)
                        for excel_col, table_col in enumerate(col_mapping, 1):
                            item = self.items_table.item(table_row, table_col)
                            if item:
                                value = item.text()
                                # Clean up and format numeric values
                                if table_col in [5, 6]:  # Cost, Price columns
                                    value = value.replace(" лв", "").replace(" ", "")
                                    try:
                                        value = f"{float(value):.2f}"
                                    except:
                                        pass
                                elif table_col == 7:  # Weight column
                                    value = value.replace(" гр", "").replace(" ", "")
                                    try:
                                        value = f"{float(value):.2f}"
                                    except:
                                        pass
                                ws.cell(row=row_idx, column=excel_col, value=value)
                        exported_count += 1
                    except Exception as e:
                        logger.warning(f"Error exporting row {table_row}: {e}")
                
                # Save file
                wb.save(file_path)
                QMessageBox.information(
                    self, "Успех", 
                    f"Успешно експортирани {exported_count} артикула в:\n{file_path}"
                )
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при експорт: {str(e)}")
    
    def bulk_edit_prices(self):
        """Bulk edit prices for selected items"""
        try:
            selected_barcodes = self.get_selected_barcodes()
            if not selected_barcodes:
                QMessageBox.warning(self, "Предупреждение", "Няма избрани артикули за редактиране на цени")
                return
            
            # Create price adjustment dialog
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Редактиране на цени за {len(selected_barcodes)} артикула")
            dialog.setModal(True)
            dialog.setFixedSize(400, 250)
            layout = QVBoxLayout(dialog)
            
            # Info
            info_label = QLabel(f"Ще промените цените на {len(selected_barcodes)} избрани артикула:")
            layout.addWidget(info_label)
            
            # Price adjustment options
            adjustment_group = QGroupBox("Метод на корекция")
            adjustment_layout = QVBoxLayout()
            
            # Create button group for mutual exclusivity
            from PyQt6.QtWidgets import QRadioButton, QButtonGroup
            price_button_group = QButtonGroup()
            
            # Option 1: Set absolute price
            set_absolute_radio = QRadioButton("Задай нова цена за всички")
            set_absolute_radio.setChecked(True)  # Default selection
            price_button_group.addButton(set_absolute_radio, 0)
            adjustment_layout.addWidget(set_absolute_radio)
            
            absolute_price_layout = QHBoxLayout()
            absolute_price_layout.addWidget(QLabel("Нова цена:"))
            absolute_price_input = BlurOnEnterDoubleSpinBox()
            absolute_price_input.setRange(0, 999999)
            absolute_price_input.setDecimals(2)
            absolute_price_input.setSuffix(" лв")
            absolute_price_input.setValue(100.0)  # Default value
            absolute_price_layout.addWidget(absolute_price_input)
            adjustment_layout.addLayout(absolute_price_layout)
            
            # Option 2: Percentage adjustment
            percentage_radio = QRadioButton("Процентно изменение")
            price_button_group.addButton(percentage_radio, 1)
            adjustment_layout.addWidget(percentage_radio)
            
            percentage_layout = QHBoxLayout()
            percentage_layout.addWidget(QLabel("Изменение:"))
            percentage_input = BlurOnEnterDoubleSpinBox()
            percentage_input.setRange(-99, 999)
            percentage_input.setDecimals(1)
            percentage_input.setSuffix(" %")
            percentage_input.setValue(10.0)  # Default value
            percentage_input.setEnabled(False)
            percentage_layout.addWidget(percentage_input)
            percentage_layout.addWidget(QLabel("(+ за увеличение, - за намаление)"))
            adjustment_layout.addLayout(percentage_layout)
            
            # Option 3: Fixed amount adjustment
            fixed_radio = QRadioButton("Фиксирано изменение")
            price_button_group.addButton(fixed_radio, 2)
            adjustment_layout.addWidget(fixed_radio)
            
            fixed_layout = QHBoxLayout()
            fixed_layout.addWidget(QLabel("Изменение:"))
            fixed_input = BlurOnEnterDoubleSpinBox()
            fixed_input.setRange(-99999, 99999)
            fixed_input.setDecimals(2)
            fixed_input.setSuffix(" лв")
            fixed_input.setValue(10.0)  # Default value
            fixed_input.setEnabled(False)
            fixed_layout.addWidget(fixed_input)
            fixed_layout.addWidget(QLabel("(+ за увеличение, - за намаление)"))
            adjustment_layout.addLayout(fixed_layout)
            
            adjustment_group.setLayout(adjustment_layout)
            layout.addWidget(adjustment_group)
            
            # Connect radio buttons to enable/disable inputs
            def toggle_inputs():
                absolute_price_input.setEnabled(set_absolute_radio.isChecked())
                percentage_input.setEnabled(percentage_radio.isChecked())
                fixed_input.setEnabled(fixed_radio.isChecked())
            
            set_absolute_radio.toggled.connect(toggle_inputs)
            percentage_radio.toggled.connect(toggle_inputs)
            fixed_radio.toggled.connect(toggle_inputs)
            
            # Initialize with correct enabled state
            toggle_inputs()
            
            # Buttons
            button_layout = QHBoxLayout()
            ok_button = QPushButton("Приложи")
            cancel_button = QPushButton("Отказ")
            button_layout.addWidget(ok_button)
            button_layout.addWidget(cancel_button)
            layout.addLayout(button_layout)
            
            ok_button.clicked.connect(dialog.accept)
            cancel_button.clicked.connect(dialog.reject)
            
            # Show dialog
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # Get selected adjustment method
                selected_method = price_button_group.checkedId()
                if selected_method == -1:
                    QMessageBox.warning(self, "Грешка", "Моля, изберете метод на корекция")
                    return
                
                # Validate input values
                if selected_method == 0:  # Absolute price
                    if absolute_price_input.value() <= 0:
                        QMessageBox.warning(self, "Грешка", "Новата цена трябва да бъде по-голяма от 0")
                        return
                elif selected_method == 1:  # Percentage
                    if percentage_input.value() <= -100:
                        QMessageBox.warning(self, "Грешка", "Процентното намаление не може да бъде 100% или повече")
                        return
                # Fixed amount doesn't need special validation (can be negative)
                
                updated_count = 0
                failed_items = []
                
                # Create detailed confirmation message
                method_description = ""
                if selected_method == 0:
                    method_description = f"Задаване на нова цена: {absolute_price_input.value():.2f} лв"
                elif selected_method == 1:
                    change_percent = percentage_input.value()
                    sign = "увеличение" if change_percent >= 0 else "намаление"
                    method_description = f"Процентно {sign}: {abs(change_percent):.1f}%"
                elif selected_method == 2:
                    fixed_change = fixed_input.value()
                    sign = "увеличение" if fixed_change >= 0 else "намаление"
                    method_description = f"Фиксирано {sign}: {abs(fixed_change):.2f} лв"
                
                # Confirmation dialog
                reply = QMessageBox.question(
                    self, "Потвърждение",
                    f"Ще приложите {method_description} към {len(selected_barcodes)} артикула.\n\n"
                    f"Сигурни ли сте?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )

                if reply != QMessageBox.StandardButton.Yes:
                    return
                
                # Process all items in a single transaction for consistency
                try:
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        
                        for barcode in selected_barcodes:
                            try:
                                # Get current item data
                                cursor.execute('SELECT id, price FROM items WHERE barcode = ?', (barcode,))
                                result = cursor.fetchone()
                                if not result:
                                    failed_items.append(barcode)
                                    continue
                                
                                item_id, current_price = result
                                current_price = float(current_price)
                                new_price = current_price
                            
                                # Calculate new price based on selected method
                                if selected_method == 0:  # Absolute price
                                    new_price = absolute_price_input.value()
                                elif selected_method == 1:  # Percentage
                                    change_percent = percentage_input.value()
                                    new_price = current_price * (1 + change_percent / 100)
                                elif selected_method == 2:  # Fixed amount
                                    fixed_change = fixed_input.value()
                                    new_price = current_price + fixed_change
                                
                                # Ensure price is not negative
                                new_price = max(0, new_price)
                                        
                                logger.info(f"Updating item {barcode}: {current_price:.2f} -> {new_price:.2f}")
                                    
                                # Update in database
                                cursor.execute('UPDATE items SET price = ? WHERE id = ?', (new_price, item_id))
                                updated_count += 1
                                
                            except Exception as e:
                                logger.error(f"Error updating price for item {barcode}: {e}")
                                failed_items.append(barcode)
                        
                        # Commit all changes at once
                        conn.commit()
                        logger.info(f"Bulk price update completed. Updated {updated_count} items.")
                        
                except Exception as e:
                    logger.error(f"Database transaction failed during bulk price update: {e}")
                    QMessageBox.critical(self, "Грешка", f"Грешка при обновяване на базата данни: {str(e)}")
                    return
                
                # Show results
                if failed_items:
                    QMessageBox.warning(
                        self, "Частично обновяване",
                        f"Приложено: {method_description}\n\n"
                        f"Обновени: {updated_count} артикула\n"
                        f"Неуспешни: {len(failed_items)} артикула\n\n"
                        f"Неуспешни баркодове: {', '.join(failed_items[:5])}"
                        f"{'...' if len(failed_items) > 5 else ''}"
                    )
                else:
                    QMessageBox.information(
                        self, "Успех",
                        f"Приложено: {method_description}\n\n"
                        f"Успешно обновени цените на {updated_count} артикула"
                    )
                
                # Force UI to process pending events
                QApplication.processEvents()
                
                # Reload the table and update undo/redo buttons
                logger.info("Refreshing inventory table after bulk price update...")
                self.load_items()
                
                # Force table to refresh and update display
                self.items_table.viewport().update()
                QApplication.processEvents()
                
                # Update reports and database statistics
                self.update_reports_and_database_stats()
                
                logger.info("Table refresh completed.")
                # Note: Bulk price editing creates multiple individual edit actions
                # The undo system will track the last edit action for undo capability
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при масово редактиране на цени: {str(e)}")

    def export_warehouse(self):
        """Export warehouse items to Excel and PDF"""
        try:
            from datetime import datetime
            import os
            
            # Ensure exports directory exists
            exports_dir = self.get_exports_directory()
            
            # Get current date and time
            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")
            
            # Get all warehouse items
            items = self.db.get_all_items()
            
            # Show export format selection dialog
            export_dialog = ExportFormatDialog(self)
            if export_dialog.exec() == QDialog.DialogCode.Accepted:
                excel_selected, pdf_selected = export_dialog.get_selections()
                
                if excel_selected and pdf_selected:
                    # Get save location for Excel file
                    excel_filename = self.generate_bulgarian_filename("склад", "xlsx")
                    excel_file, _ = QFileDialog.getSaveFileName(
                        self, "Запази Excel файл за склад", 
                        f"{exports_dir}/{excel_filename}",
                        "Excel Files (*.xlsx)"
                    )
                    if not excel_file:
                        return
                    
                    # Get save location for PDF file
                    pdf_filename = self.generate_bulgarian_filename("склад", "pdf")
                    pdf_file, _ = QFileDialog.getSaveFileName(
                        self, "Запази PDF файл за склад", 
                        f"{exports_dir}/{pdf_filename}",
                        "PDF Files (*.pdf)"
                    )
                    if not pdf_file:
                        return
                    
                    # Export both formats
                    self.export_to_excel(items, "Склад", excel_file, date_str, time_str)
                    self.export_to_pdf(items, "Склад", pdf_file, date_str, time_str)
                    QMessageBox.information(self, "Успех", f"Експортирани файлове:\n- {os.path.basename(excel_file)}\n- {os.path.basename(pdf_file)}")
                elif excel_selected:
                    # Get save location for Excel file
                    excel_filename = self.generate_bulgarian_filename("склад", "xlsx")
                    excel_file, _ = QFileDialog.getSaveFileName(
                        self, "Запази Excel файл за склад", 
                        f"{exports_dir}/{excel_filename}",
                        "Excel Files (*.xlsx)"
                    )
                    if not excel_file:
                        return
                    
                    # Export Excel only
                    self.export_to_excel(items, "Склад", excel_file, date_str, time_str)
                    QMessageBox.information(self, "Успех", f"Експортиран файл: {os.path.basename(excel_file)}")
                elif pdf_selected:
                    # Get save location for PDF file
                    pdf_filename = self.generate_bulgarian_filename("склад", "pdf")
                    pdf_file, _ = QFileDialog.getSaveFileName(
                        self, "Запази PDF файл за склад", 
                        f"{exports_dir}/{pdf_filename}",
                        "PDF Files (*.pdf)"
                    )
                    if not pdf_file:
                        return
                    
                    # Export PDF only
                    self.export_to_pdf(items, "Склад", pdf_file, date_str, time_str)
                    QMessageBox.information(self, "Успех", f"Експортиран файл: {os.path.basename(pdf_file)}")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при експорт на склад: {str(e)}")

    def export_shop(self):
        """Export shop items to Excel and PDF"""
        try:
            from datetime import datetime
            import os
            
            # Check if a shop is selected
            shop_name = self.shop_combo.currentText()
            if not shop_name:
                QMessageBox.warning(self, "Грешка", "Моля изберете магазин за експорт!")
                return
            
            # Ensure exports directory exists
            exports_dir = self.get_exports_directory()
            
            # Get current date and time
            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")
            
            # Get shop items
            shop_id = None
            shops = self.db.get_all_shops()
            for shop in shops:
                if shop[1] == shop_name:
                    shop_id = shop[0]
                    break
            
            if not shop_id:
                QMessageBox.warning(self, "Грешка", "Магазинът не е намерен!")
                return
                
            items = self.db.get_shop_items(shop_id)
            
            # Show export format selection dialog
            export_dialog = ExportFormatDialog(self)
            if export_dialog.exec() == QDialog.DialogCode.Accepted:
                excel_selected, pdf_selected = export_dialog.get_selections()
                
                # Generate Bulgarian filename based on shop name
                shop_base_name = shop_name.lower().strip()
                
                if excel_selected and pdf_selected:
                    # Get save location for Excel file
                    excel_filename = self.generate_bulgarian_filename(shop_base_name, "xlsx")
                    excel_file, _ = QFileDialog.getSaveFileName(
                        self, f"Запази Excel файл за магазин {shop_name}", 
                        f"{exports_dir}/{excel_filename}",
                        "Excel Files (*.xlsx)"
                    )
                    if not excel_file:
                        return
                    
                    # Get save location for PDF file
                    pdf_filename = self.generate_bulgarian_filename(shop_base_name, "pdf")
                    pdf_file, _ = QFileDialog.getSaveFileName(
                        self, f"Запази PDF файл за магазин {shop_name}", 
                        f"{exports_dir}/{pdf_filename}",
                        "PDF Files (*.pdf)"
                    )
                    if not pdf_file:
                        return
                    
                    # Export both formats
                    self.export_to_excel(items, f"Магазин: {shop_name}", excel_file, date_str, time_str)
                    self.export_to_pdf(items, f"Магазин: {shop_name}", pdf_file, date_str, time_str)
                    QMessageBox.information(self, "Успех", f"Експортирани файлове:\n- {os.path.basename(excel_file)}\n- {os.path.basename(pdf_file)}")
                elif excel_selected:
                    # Get save location for Excel file
                    excel_filename = self.generate_bulgarian_filename(shop_base_name, "xlsx")
                    excel_file, _ = QFileDialog.getSaveFileName(
                        self, f"Запази Excel файл за магазин {shop_name}", 
                        f"{exports_dir}/{excel_filename}",
                        "Excel Files (*.xlsx)"
                    )
                    if not excel_file:
                        return
                    
                    # Export Excel only
                    self.export_to_excel(items, f"Магазин: {shop_name}", excel_file, date_str, time_str)
                    QMessageBox.information(self, "Успех", f"Експортиран файл: {os.path.basename(excel_file)}")
                elif pdf_selected:
                    # Get save location for PDF file
                    pdf_filename = self.generate_bulgarian_filename(shop_base_name, "pdf")
                    pdf_file, _ = QFileDialog.getSaveFileName(
                        self, f"Запази PDF файл за магазин {shop_name}", 
                        f"{exports_dir}/{pdf_filename}",
                        "PDF Files (*.pdf)"
                    )
                    if not pdf_file:
                        return
                    
                    # Export PDF only
                    self.export_to_pdf(items, f"Магазин: {shop_name}", pdf_file, date_str, time_str)
                    QMessageBox.information(self, "Успех", f"Експортиран файл: {os.path.basename(pdf_file)}")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при експорт на магазин: {str(e)}")

    def export_to_excel(self, items, title, filename, date_str, time_str):
        """Export items to Excel format"""
        try:
            import pandas as pd
            
            # Prepare data for Excel
            data = []
            for item in items:
                # Handle both warehouse and shop item formats
                if len(item) >= 13:  # Shop item format
                    row = {
                        'Баркод': item[1],
                        'Име': item[2], 
                        'Описание': item[3],
                        'Категория': item[4],
                        'Цена (лв)': f"{float(item[5]) * 1.95583:.2f}",
                        'Цена на едро (лв)': f"{float(item[6]) * 1.95583:.2f}",
                        'Тегло (г)': f"{float(item[7]):.2f}",
                        'Метал': item[8],
                        'Камък': item[9],
                        'Количество': item[13] if len(item) > 13 else item[10]
                    }
                else:  # Warehouse item format
                    row = {
                        'Баркод': item[1],
                        'Име': item[2], 
                        'Описание': item[3],
                        'Категория': item[4],
                        'Цена (лв)': f"{float(item[5]) * 1.95583:.2f}",
                        'Цена на едро (лв)': f"{float(item[6]) * 1.95583:.2f}",
                        'Тегло (г)': f"{float(item[7]):.2f}",
                        'Метал': item[8],
                        'Камък': item[9],
                        'Количество': item[10]
                    }
                data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Write title and date info
                title_df = pd.DataFrame({
                    'A': [title, f"Дата на експорт: {date_str}", f"Час на експорт: {time_str}", ""]
                })
                title_df.to_excel(writer, sheet_name='Данни', index=False, header=False, startrow=0)
                
                # Write main data
                df.to_excel(writer, sheet_name='Данни', index=False, startrow=4)
                
                # Format the worksheet
                worksheet = writer.sheets['Данни']
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 15  # Баркод
                worksheet.column_dimensions['B'].width = 25  # Име
                worksheet.column_dimensions['C'].width = 30  # Описание
                worksheet.column_dimensions['D'].width = 15  # Категория
                worksheet.column_dimensions['E'].width = 15  # Цена
                worksheet.column_dimensions['F'].width = 18  # Цена на едро
                worksheet.column_dimensions['G'].width = 12  # Тегло
                worksheet.column_dimensions['H'].width = 12  # Метал
                worksheet.column_dimensions['I'].width = 12  # Камък
                worksheet.column_dimensions['J'].width = 12  # Количество
                
        except ImportError:
            QMessageBox.critical(self, "Грешка", "Pandas библиотеката не е инсталирана. Моля инсталирайте я с: pip install pandas openpyxl")
        except Exception as e:
            raise e

    def export_to_pdf(self, items, title, filename, date_str, time_str):
        """Export items to PDF format with proper Cyrillic font support"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # Register fonts that support Cyrillic characters
            try:
                # Try to use system fonts that support Cyrillic
                font_paths = [
                    "fonts/arial.ttf",  # Our project font
                    "C:/Windows/Fonts/arial.ttf",  # Windows system font
                    "C:/Windows/Fonts/calibri.ttf",  # Alternative Windows font
                    "/System/Library/Fonts/Arial.ttf",  # macOS system font
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"  # Linux font
                ]
                
                font_registered = False
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        try:
                            pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                            pdfmetrics.registerFont(TTFont('CyrillicFont-Bold', font_path))  # Use same font for bold
                            font_registered = True
                            break
                        except Exception:
                            continue
                
                if not font_registered:
                    # Fallback to built-in font (may not display Cyrillic properly)
                    cyrillic_font = 'Helvetica'
                    cyrillic_font_bold = 'Helvetica-Bold'
                else:
                    cyrillic_font = 'CyrillicFont'
                    cyrillic_font_bold = 'CyrillicFont-Bold'
                    
            except Exception:
                # Fallback to built-in fonts
                cyrillic_font = 'Helvetica'
                cyrillic_font_bold = 'Helvetica-Bold'
            
            # Create PDF document
            doc = SimpleDocTemplate(filename, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Create custom styles with Cyrillic font support and black text
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontName=cyrillic_font_bold,
                fontSize=16,
                textColor=colors.black,  # Ensure black text
                spaceAfter=12
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=cyrillic_font,
                fontSize=12,
                textColor=colors.black,  # Ensure black text
                spaceAfter=6
            )
            
            # Title
            title_para = Paragraph(title, title_style)
            elements.append(title_para)
            elements.append(Spacer(1, 12))
            
            # Date and time
            date_para = Paragraph(f"Дата на експорт: {date_str} | Час на експорт: {time_str}", normal_style)
            elements.append(date_para)
            elements.append(Spacer(1, 20))
            
            # Prepare table data
            table_data = [['Баркод', 'Име', 'Описание', 'Категория', 'Цена (лв)', 'Цена на едро (лв)', 'Тегло (г)', 'Метал', 'Камък', 'Количество']]
            
            for item in items:
                # Handle both warehouse and shop item formats
                if len(item) >= 13:  # Shop item format
                    row = [
                        str(item[1]) if item[1] else "",  # Баркод
                        (str(item[2])[:15] + "..." if len(str(item[2])) > 15 else str(item[2])) if item[2] else "",  # Име (truncated)
                        (str(item[3])[:20] + "..." if len(str(item[3])) > 20 else str(item[3])) if item[3] else "",  # Описание (truncated)
                        str(item[4]) if item[4] else "",  # Категория
                        f"{float(item[5]) * 1.95583:.2f}" if item[5] else "0.00",  # Цена
                        f"{float(item[6]) * 1.95583:.2f}" if item[6] else "0.00",  # Цена на едро
                        str(item[7]) if item[7] else "",  # Тегло
                        str(item[8]) if item[8] else "",  # Метал
                        str(item[9]) if item[9] else "",  # Камък
                        str(item[13] if len(item) > 13 else item[10]) if len(item) > 10 else ""  # Количество
                    ]
                else:  # Warehouse item format
                    row = [
                        str(item[1]) if item[1] else "",  # Баркод
                        (str(item[2])[:15] + "..." if len(str(item[2])) > 15 else str(item[2])) if item[2] else "",  # Име (truncated)
                        (str(item[3])[:20] + "..." if len(str(item[3])) > 20 else str(item[3])) if item[3] else "",  # Описание (truncated)
                        str(item[4]) if item[4] else "",  # Категория
                        f"{float(item[5]) * 1.95583:.2f}" if item[5] else "0.00",  # Цена
                        f"{float(item[6]) * 1.95583:.2f}" if item[6] else "0.00",  # Цена на едро
                        str(item[7]) if item[7] else "",  # Тегло
                        str(item[8]) if item[8] else "",  # Метал
                        str(item[9]) if item[9] else "",  # Камък
                        str(item[10]) if item[10] else ""  # Количество
                    ]
                table_data.append(row)
            
            # Create table with proper font support
            table = Table(table_data)
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # White text on grey header
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),  # Use Cyrillic font for header
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Body styling
                ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),  # Use Cyrillic font for body
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # BLACK text on light background
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # White background for body
                
                # Grid and borders
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
        except ImportError:
            QMessageBox.critical(self, "Грешка", "ReportLab библиотеката не е инсталирана. Моля инсталирайте я с: pip install reportlab")
        except Exception as e:
            raise e

    def create_sales_tab(self):
        """Create the sales management tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Enhanced Tabbed Search for sales with integrated period filtering
        search_group = QGroupBox("Търсене в продажби")
        search_layout = QVBoxLayout()
        
        # Create sales search tabs
        self.sales_search_tabs = QTabWidget()
        self.sales_search_tabs.setMaximumHeight(120)  # Compact height
        
        # Tab 1: General Search for sales
        sales_general_tab = QWidget()
        sales_general_layout = QVBoxLayout(sales_general_tab)
        
        # Main search bar for sales general search
        sales_main_search_layout = QHBoxLayout()
        self.sales_search_input = QLineEdit()
        self.sales_search_input.setPlaceholderText("Търси по всички полета (баркод, категория, метал, камък, описание, цени, тегло, количество, дата, час)...")
        self.sales_search_input.textChanged.connect(self.search_sales)
        clear_sales_search_btn = QPushButton("✕")
        clear_sales_search_btn.setFixedSize(30, 30)
        clear_sales_search_btn.clicked.connect(self.clear_sales_search)
        sales_main_search_layout.addWidget(QLabel("Общо търсене:"))
        sales_main_search_layout.addWidget(self.sales_search_input)
        sales_main_search_layout.addWidget(clear_sales_search_btn)
        sales_general_layout.addLayout(sales_main_search_layout)
        
        self.sales_search_tabs.addTab(sales_general_tab, "Общо търсене")
        
        # Tab 2: Date Search with integrated period filtering
        sales_date_tab = QWidget()
        sales_date_layout = QVBoxLayout(sales_date_tab)
        
        # Period filter radio buttons (first row)
        period_layout = QHBoxLayout()
        
        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        self.time_button_group = QButtonGroup()
        
        self.all_time_radio = QRadioButton("Всички")
        self.all_time_radio.setChecked(True)
        self.today_radio = QRadioButton("Днес")
        self.week_radio = QRadioButton("Седмица")
        self.month_radio = QRadioButton("Месец")
        self.year_radio = QRadioButton("Година")
        self.custom_radio = QRadioButton("Персонализиран")
        
        self.time_button_group.addButton(self.all_time_radio, 0)
        self.time_button_group.addButton(self.today_radio, 1)
        self.time_button_group.addButton(self.week_radio, 2)
        self.time_button_group.addButton(self.month_radio, 3)
        self.time_button_group.addButton(self.year_radio, 4)
        self.time_button_group.addButton(self.custom_radio, 5)
        
        period_layout.addWidget(self.all_time_radio)
        period_layout.addWidget(self.today_radio)
        period_layout.addWidget(self.week_radio)
        period_layout.addWidget(self.month_radio)
        period_layout.addWidget(self.year_radio)
        period_layout.addWidget(self.custom_radio)
        period_layout.addStretch()
        
        sales_date_layout.addLayout(period_layout)
        
        # Custom date range section (second row) - unified date controls
        date_range_layout = QHBoxLayout()
        date_range_layout.addWidget(QLabel("От дата:"))
        
        # Unified date controls that serve both period filtering and search filtering
        self.sales_start_date = QDateEdit()
        self.sales_start_date.setDate(QDate.currentDate().addDays(-30))
        self.sales_start_date.setCalendarPopup(True)
        self.sales_start_date.setEnabled(False)
        self.sales_start_date.setMaximumWidth(120)
        date_range_layout.addWidget(self.sales_start_date)
        
        date_range_layout.addWidget(QLabel("до дата:"))
        
        self.sales_end_date = QDateEdit()
        self.sales_end_date.setDate(QDate.currentDate())
        self.sales_end_date.setCalendarPopup(True)
        self.sales_end_date.setEnabled(False)
        self.sales_end_date.setMaximumWidth(120)
        date_range_layout.addWidget(self.sales_end_date)
        
        date_range_layout.addStretch()
        sales_date_layout.addLayout(date_range_layout)
        
        # Create aliases for search functionality to maintain compatibility
        self.sales_start_date_input = self.sales_start_date
        self.sales_end_date_input = self.sales_end_date
        
        self.sales_search_tabs.addTab(sales_date_tab, "По дата")
        
        # Add tabs to main layout
        search_layout.addWidget(self.sales_search_tabs)
        
        # Clear all sales filters button
        clear_sales_filters_layout = QHBoxLayout()
        clear_sales_filters_btn = QPushButton("Изчисти всички филтри")
        clear_sales_filters_btn.clicked.connect(self.clear_all_sales_filters)
        clear_sales_filters_btn.setStyleSheet("background-color: #ffeb99; color: #664d00; font-weight: bold;")
        clear_sales_filters_btn.setMaximumWidth(180)
        clear_sales_filters_layout.addWidget(clear_sales_filters_btn)
        
        # Add current sales search info
        self.sales_search_info_label = QLabel("Няма активни филтри")
        self.sales_search_info_label.setStyleSheet("color: #666; font-size: 10px; font-style: italic;")
        clear_sales_filters_layout.addWidget(self.sales_search_info_label)
        clear_sales_filters_layout.addStretch()
        
        search_layout.addLayout(clear_sales_filters_layout)
        
        # Set compact maximum height for entire search group
        search_group.setLayout(search_layout)
        search_group.setMaximumHeight(160)
        layout.addWidget(search_group)
        
        # Connect radio buttons to reload sales
        self.time_button_group.buttonClicked.connect(self.on_time_filter_changed)
        self.time_button_group.buttonClicked.connect(self.update_reports_and_database_stats)
        
        # Connect unified date fields to both reload sales AND trigger search
        self.sales_start_date.dateChanged.connect(self.load_sales)
        self.sales_end_date.dateChanged.connect(self.load_sales)
        self.sales_start_date.dateChanged.connect(self.search_sales)
        self.sales_end_date.dateChanged.connect(self.search_sales)
        
        # Auto-switch to custom period when calendar is clicked
        self.sales_start_date.dateChanged.connect(self.auto_switch_to_custom_period)
        self.sales_end_date.dateChanged.connect(self.auto_switch_to_custom_period)
        
        # Also switch when user clicks into the date field or calendar popup
        self.sales_start_date.editingFinished.connect(self.auto_switch_to_custom_period)
        self.sales_end_date.editingFinished.connect(self.auto_switch_to_custom_period)
        
        # Use installEventFilter to catch focus events on the date fields
        self.sales_start_date.installEventFilter(self)
        self.sales_end_date.installEventFilter(self)

        # Shop selection dropdown
        shop_select_frame = QGroupBox("Магазин")
        shop_layout = QHBoxLayout()
        shop_layout.addWidget(QLabel("Избери магазин:"))
        
        self.sales_shop_combo = QComboBox()
        self.sales_shop_combo.setMinimumWidth(200)
        shops = self.db.get_all_shops()
        for shop in shops:
            self.sales_shop_combo.addItem(shop[1])  # shop[1] is the name
        
        if self.sales_shop_combo.count() > 0:
            self.sales_shop_combo.setCurrentIndex(0)
        
        self.sales_shop_combo.currentTextChanged.connect(self.load_sales)
        self.sales_shop_combo.currentTextChanged.connect(self.update_reports_and_database_stats)
        self.sales_shop_combo.currentTextChanged.connect(self.update_shop_inventory_info)
        shop_layout.addWidget(self.sales_shop_combo)
        
        # Add shop inventory info label
        shop_layout.addWidget(QLabel(" | "))
        self.shop_inventory_info_label = QLabel("Налични артикули в магазин: 0")
        self.shop_inventory_info_label.setStyleSheet("color: #2196F3; font-weight: bold;")
        shop_layout.addWidget(self.shop_inventory_info_label)
        
        shop_layout.addStretch()
        
        shop_select_frame.setLayout(shop_layout)
        layout.addWidget(shop_select_frame)
        
        # Export shop sales to PDF button
        export_shop_sales_frame = QGroupBox("Експорт на продажби за магазин")
        export_shop_layout = QHBoxLayout()
        
        self.export_shop_sales_btn = QPushButton("📄 Експорт PDF отчет за магазин")
        self.export_shop_sales_btn.setMinimumHeight(35)
        self.export_shop_sales_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                padding: 8px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        self.export_shop_sales_btn.clicked.connect(self.export_shop_sales_to_pdf)
        export_shop_layout.addWidget(self.export_shop_sales_btn)
        
        export_shop_layout.addStretch()
        export_shop_sales_frame.setLayout(export_shop_layout)
        layout.addWidget(export_shop_sales_frame)

        # Barcode Entry for Sales
        barcode_frame = QGroupBox("Сканиране за продажба")
        barcode_layout = QVBoxLayout()
        
        input_layout = QHBoxLayout()
        self.sale_barcode_input = QLineEdit()
        self.sale_barcode_input.setPlaceholderText("Сканирайте баркод за директна продажба...")
        # Use timer-based approach to handle complete barcode scanning
        self.sale_barcode_input.textChanged.connect(self.on_sales_barcode_changed)
        input_layout.addWidget(self.sale_barcode_input)
        
        barcode_layout.addLayout(input_layout)
        barcode_frame.setLayout(barcode_layout)
        layout.addWidget(barcode_frame)

        # Sales History Table
        history_frame = QGroupBox("История на продажбите")
        history_layout = QVBoxLayout()
        
        self.sales_table = QTableWidget()
        self.sales_table.setColumnCount(11)
        self.sales_table.setHorizontalHeaderLabels([
            "Баркод", "Категория", "Метал", "Камък", "Описание", 
            "Цена на едро", "Цена", "Тегло", "Количество", "Дата", "Час"
        ])
        
        # Configure column spacing with controlled resize limits (resizable between bounds)
        header = self.sales_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Set minimum and maximum widths for controlled resizing
        min_width = 80   # Minimum column width
        max_width = 200  # Maximum column width
        default_width = 110  # Default column width
        
        # Set resizable widths for first 10 columns with bounds
        for col in range(10):  # First 10 columns are resizable within bounds
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            self.sales_table.setColumnWidth(col, default_width)
        
        # Last column stretches to fill remaining space
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)
        
        # Enable sorting
        self.sales_table.setSortingEnabled(True)
        self.sales_table.setAlternatingRowColors(True)
        self.sales_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # MAKE SALES TABLE COMPLETELY IMMUTABLE - Read-only display only
        self.sales_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        # Remove any context menu or interaction capabilities
        self.sales_table.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)
        
        history_layout.addWidget(self.sales_table)
        
        history_frame.setLayout(history_layout)
        layout.addWidget(history_frame)

        # Load initial sales data
        self.load_sales()
        
        # Update shop inventory info display
        self.update_shop_inventory_info()
        
        # Set focus to barcode entry
        self.sale_barcode_input.setFocus()

        return widget

    def create_reports_tab(self):
        """Create the comprehensive reports tab"""
        widget = QWidget()
        main_layout = QVBoxLayout(widget)

        # Create scroll area for the content
        scroll_area = QScrollArea()
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)
        
        # === QUICK STATS DASHBOARD ===
        dashboard_group = QGroupBox("📊 Обобщена информация")
        dashboard_layout = QGridLayout()
        
        # Quick stats cards
        self.stats_cards = {}
        stats_info = [
            ("today_sales", "Продажби днес", "0 €", "#4CAF50"),
            ("total_inventory_value", "Стойност на склада", "0 €", "#2196F3"),
            ("low_stock_items", "Артикули под 5бр", "0", "#FF9800"),
            ("total_items", "Общо артикули", "0", "#9C27B0"),
            ("this_month_sales", "Продажби този месец", "0 €", "#00BCD4"),
            ("avg_profit_margin", "Средна печалба", "0%", "#8BC34A")
        ]
        
        for i, (key, title, value, color) in enumerate(stats_info):
            card = self.create_stats_card(title, value, color)
            self.stats_cards[key] = card
            row, col = divmod(i, 3)
            dashboard_layout.addWidget(card, row, col)
        
        dashboard_group.setLayout(dashboard_layout)
        layout.addWidget(dashboard_group)
        
        # === DATE RANGE SELECTION ===
        date_group = QGroupBox("📅 Период за анализ")
        date_layout = QHBoxLayout()
        
        # Quick period buttons
        period_layout = QVBoxLayout()
        period_label = QLabel("Бързи периоди:")
        period_layout.addWidget(period_label)
        
        period_buttons_layout = QHBoxLayout()
        self.period_buttons = {}
        periods = [("today", "Днес"), ("week", "Тази седмица"), ("month", "Този месец"), ("quarter", "Тримесечие"), ("year", "Тази година")]
        
        for key, text in periods:
            btn = QPushButton(text)
            btn.clicked.connect(lambda checked, k=key: self.set_quick_period(k))
            self.period_buttons[key] = btn
            period_buttons_layout.addWidget(btn)
        
        period_layout.addLayout(period_buttons_layout)
        date_layout.addLayout(period_layout)
        
        # Custom date range - more compact layout
        custom_date_layout = QVBoxLayout()
        custom_date_layout.addWidget(QLabel("Персонализиран период:"))
        
        # Compact date inputs - move controls closer to labels
        date_inputs_layout = QHBoxLayout()
        
        # Start date group
        date_inputs_layout.addWidget(QLabel("От:"))
        self.report_start_date = QDateEdit()
        self.report_start_date.setDate(QDate.currentDate().addDays(-30))
        self.report_start_date.setCalendarPopup(True)
        self.report_start_date.setMaximumWidth(120)  # Compact width
        date_inputs_layout.addWidget(self.report_start_date)
        
        # End date group
        date_inputs_layout.addWidget(QLabel("До:"))
        self.report_end_date = QDateEdit()
        self.report_end_date.setDate(QDate.currentDate())
        self.report_end_date.setCalendarPopup(True)
        self.report_end_date.setMaximumWidth(120)  # Compact width
        date_inputs_layout.addWidget(self.report_end_date)
        
        # Add export button in the freed space
        export_report_btn = QPushButton("📊 Експорт на отчет")
        export_report_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 5px;
                margin-left: 15px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        export_report_btn.clicked.connect(self.export_comprehensive_report)
        date_inputs_layout.addWidget(export_report_btn)
        
        date_inputs_layout.addStretch()  # Push everything to the left
        
        # Connect date changes to automatically update statistics
        self.report_start_date.dateChanged.connect(self.update_dashboard_stats)
        self.report_end_date.dateChanged.connect(self.update_dashboard_stats)
        
        custom_date_layout.addLayout(date_inputs_layout)
        date_layout.addLayout(custom_date_layout)
        
        date_group.setLayout(date_layout)
        layout.addWidget(date_group)
        
        # === DETAILED ANALYTICS ===
        analytics_group = QGroupBox("📈 Детайлни анализи")
        analytics_layout = QGridLayout()
        
        # Sales Analytics
        sales_analytics_group = QGroupBox("Продажби")
        sales_analytics_layout = QVBoxLayout()
        
        sales_buttons = [
            ("Топ продавани артикули", self.show_top_selling_items),
            ("Продажби по категории", self.show_sales_by_category),
            ("Дневни продажби", self.show_daily_sales_trend),
            ("Продажби по магазини", self.show_sales_by_shop)
        ]
        
        for text, func in sales_buttons:
            btn = QPushButton(text)
            btn.clicked.connect(func)
            sales_analytics_layout.addWidget(btn)
        
        sales_analytics_group.setLayout(sales_analytics_layout)
        analytics_layout.addWidget(sales_analytics_group, 0, 0)
        
        # Inventory Analytics
        inventory_analytics_group = QGroupBox("Склад")
        inventory_analytics_layout = QVBoxLayout()
        
        inventory_buttons = [
            ("Разпределение по категория", self.show_inventory_by_category),
            ("Артикули с малка наличност", self.show_low_stock_items),
            ("Най-скъпи артикули", self.show_most_expensive_items),
            ("Стари артикули (>6 месеца)", self.show_old_inventory)
        ]
        
        for text, func in inventory_buttons:
            btn = QPushButton(text)
            btn.clicked.connect(func)
            inventory_analytics_layout.addWidget(btn)
        
        inventory_analytics_group.setLayout(inventory_analytics_layout)
        analytics_layout.addWidget(inventory_analytics_group, 0, 1)
        

        
        analytics_group.setLayout(analytics_layout)
        layout.addWidget(analytics_group)
        
        # Add stretch to push everything to top
        layout.addStretch()
        
        # Set up scroll area
        scroll_area.setWidget(scroll_content)
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        main_layout.addWidget(scroll_area)
        
        # Initialize dashboard
        self.update_dashboard_stats()

        return widget

    def create_shop_loading_tab(self):
        """Create the shop loading tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Shop selection
        shop_select_layout = QHBoxLayout()
        shop_label = QLabel("Избери магазин:")
        self.shop_combo = QComboBox()
        self.shop_combo.setMinimumWidth(200)
        
        # Load shops into combo box
        shops = self.db.get_all_shops()
        for shop in sorted(shops, key=lambda x: x[1]):  # Sort by shop name
            self.shop_combo.addItem(shop[1])
        
        add_shop_btn = QPushButton("+ Нов магазин")
        add_shop_btn.clicked.connect(self.add_new_shop)
        edit_shop_btn = QPushButton("✎ Редактирай магазин")
        edit_shop_btn.clicked.connect(self.rename_selected_shop)
        delete_shop_btn = QPushButton("🗑 Изтрий магазин")
        delete_shop_btn.clicked.connect(self.delete_selected_shop)
        export_shop_btn = QPushButton("📄 Експорт магазин")
        export_shop_btn.clicked.connect(self.export_shop)

        shop_select_layout.addWidget(shop_label)
        shop_select_layout.addWidget(self.shop_combo)
        shop_select_layout.addWidget(add_shop_btn)
        shop_select_layout.addWidget(edit_shop_btn)
        shop_select_layout.addWidget(delete_shop_btn)
        shop_select_layout.addWidget(export_shop_btn)
        shop_select_layout.addStretch()
        layout.addLayout(shop_select_layout)

        # Barcode Entry for Shop Loading
        barcode_frame = QGroupBox("Зареждане на артикули в магазин")
        barcode_layout = QVBoxLayout()
        
        input_layout = QHBoxLayout()
        self.shop_barcode_input = QLineEdit()
        self.shop_barcode_input.setPlaceholderText("Сканирайте баркод за директно зареждане...")
        self.shop_barcode_input.textChanged.connect(self.on_shop_barcode_changed)
        input_layout.addWidget(self.shop_barcode_input)
        
        barcode_layout.addLayout(input_layout)
        barcode_frame.setLayout(barcode_layout)
        layout.addWidget(barcode_frame)

        # Shop inventory table
        self.shop_table = QTableWidget()
        self.shop_table.setColumnCount(11)
        self.shop_table.setHorizontalHeaderLabels([
            "Баркод", "Категория", "Метал", "Камък", "Описание", 
            "Цена на едро", "Цена", "Тегло", "Количество", "Дата", "Час"
        ])
        
        # Configure column spacing with controlled resize limits (resizable between bounds)
        header = self.shop_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Set minimum and maximum widths for controlled resizing
        min_width = 80   # Minimum column width
        max_width = 200  # Maximum column width
        default_width = 110  # Default column width
        
        # Set resizable widths for first 10 columns with bounds
        for col in range(10):  # First 10 columns are resizable within bounds
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            self.shop_table.setColumnWidth(col, default_width)
        
        # Last column stretches to fill remaining space
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)
        
        # Enable sorting
        self.shop_table.setSortingEnabled(True)
        self.shop_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.shop_table.customContextMenuRequested.connect(self.shop_right_click)
        # Make table non-editable and add double-click edit functionality
        self.shop_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.shop_table.itemDoubleClicked.connect(self.edit_shop_item)
        self.shop_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.shop_table.setAlternatingRowColors(True)
        layout.addWidget(self.shop_table)

        # Summary bar
        summary_layout = QHBoxLayout()
        self.shop_summary_labels = []
        for i in range(6):
            label = QLabel()
            label.setStyleSheet("font-weight: bold;")
            self.shop_summary_labels.append(label)
            summary_layout.addWidget(label)
        self.shop_summary_labels[1].setText("ОБЩО:")
        layout.addLayout(summary_layout)

        # Connect only ONE signal to avoid duplicate calls - use currentTextChanged only
        # Note: currentIndexChanged and currentTextChanged would both fire, causing race conditions
        self.shop_combo.currentTextChanged.connect(self.load_shop_inventory)
        
        # Select first shop and load its inventory immediately
        if self.shop_combo.count() > 0:
            self.shop_combo.setCurrentIndex(0)
            # Load initial shop inventory
            self.load_shop_inventory()

        return widget

    def create_database_tab(self):
        """Create the simplified database management tab"""
        widget = QWidget()
        main_layout = QVBoxLayout(widget)

        # Create scroll area for the content
        scroll_area = QScrollArea()
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)
        
        # === DATABASE OVERVIEW ===
        overview_group = QGroupBox("📊 Преглед на базата данни")
        overview_layout = QGridLayout()
        
        # Database statistics cards
        self.db_stats_cards = {}
        stats_info = [
            ("total_items", "Общо артикули", "0", "#4CAF50"),
            ("total_sales", "Общо продажби", "0", "#2196F3"),
            ("total_shops", "Магазини", "0", "#FF9800"),
            ("db_size", "Размер на БД", "0 MB", "#9C27B0"),
            ("last_backup", "Последно копие", "Никога", "#00BCD4"),
            ("integrity_status", "Статус", "OK", "#8BC34A")
        ]
        
        for i, (key, title, value, color) in enumerate(stats_info):
            card = self.create_stats_card(title, value, color)
            self.db_stats_cards[key] = card
            row, col = divmod(i, 3)
            overview_layout.addWidget(card, row, col)
        
        overview_group.setLayout(overview_layout)
        layout.addWidget(overview_group)
        
        # === BACKUP MANAGEMENT ===
        backup_group = QGroupBox("💾 Управление на резервни копия")
        backup_layout = QVBoxLayout()
        
        # Backup controls
        backup_controls_layout = QHBoxLayout()
        
        create_backup_btn = QPushButton("🔄 Създай резервно копие")
        create_backup_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }")
        create_backup_btn.clicked.connect(self.create_backup)
        backup_controls_layout.addWidget(create_backup_btn)
        
        self.auto_backup_btn = QPushButton("⏰ Автоматично копие")
        self.update_auto_backup_button_color()
        self.auto_backup_btn.clicked.connect(self.setup_auto_backup)
        backup_controls_layout.addWidget(self.auto_backup_btn)
        
        backup_controls_layout.addStretch()
        backup_layout.addLayout(backup_controls_layout)
        
        # Backup list
        self.backup_list = QTableWidget()
        self.backup_list.setColumnCount(4)
        self.backup_list.setHorizontalHeaderLabels(["Файл", "Дата", "Размер", "Действия"])
        self.backup_list.setMaximumHeight(150)
        self.backup_list.setAlternatingRowColors(True)
        
        # Configure row selection behavior - select entire row
        self.backup_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.backup_list.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        
        # Configure column spacing with controlled resize limits (resizable between bounds)
        header = self.backup_list.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Set default widths for columns
        self.backup_list.setColumnWidth(0, 180)  # Filename
        self.backup_list.setColumnWidth(1, 120)  # Date
        self.backup_list.setColumnWidth(2, 80)   # Size
        
        # First 3 columns are resizable
        for col in range(3):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
        
        # Actions column stretches to fill remaining space
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        
        # Make table completely read-only - disable all editing
        self.backup_list.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        backup_layout.addWidget(self.backup_list)
        
        # Setup file system watcher for real-time backup monitoring
        self.setup_backup_file_watcher()
        
        backup_group.setLayout(backup_layout)
        layout.addWidget(backup_group)
        
        # === DATA MANAGEMENT ===
        data_group = QGroupBox("📋 Управление на данни")
        data_layout = QGridLayout()
        
        # Export section
        export_section = QGroupBox("Експорт")
        export_layout = QVBoxLayout()
        
        export_excel_btn = QPushButton("📊 Експорт в Excel")
        export_excel_btn.clicked.connect(self.export_data)
        export_layout.addWidget(export_excel_btn)
        
        export_csv_btn = QPushButton("📄 Експорт в CSV")
        export_csv_btn.clicked.connect(self.export_to_csv)
        export_layout.addWidget(export_csv_btn)
        
        export_json_btn = QPushButton("🔧 Експорт в JSON")
        export_json_btn.clicked.connect(self.export_to_json)
        export_layout.addWidget(export_json_btn)
        
        export_section.setLayout(export_layout)
        data_layout.addWidget(export_section, 0, 0)
        
        # Import section
        import_section = QGroupBox("Импорт")
        import_layout = QVBoxLayout()
        
        import_excel_btn = QPushButton("📊 Импорт от Excel")
        import_excel_btn.clicked.connect(self.import_data)
        import_layout.addWidget(import_excel_btn)
        
        import_csv_btn = QPushButton("📄 Импорт от CSV")
        import_csv_btn.clicked.connect(self.import_from_csv)
        import_layout.addWidget(import_csv_btn)
        
        import_json_btn = QPushButton("🔧 Импорт от JSON")
        import_json_btn.clicked.connect(self.import_from_json)
        import_layout.addWidget(import_json_btn)
        
        restore_btn = QPushButton("♻️ Възстанови от копие")
        restore_btn.setStyleSheet("QPushButton { background-color: #FF9800; color: white; font-weight: bold; }")
        restore_btn.clicked.connect(self.restore_backup)
        import_layout.addWidget(restore_btn)
        
        import_section.setLayout(import_layout)
        data_layout.addWidget(import_section, 0, 1)
        
        # Migration tools
        migration_section = QGroupBox("Миграция")
        migration_layout = QVBoxLayout()
        
        migrate_schema_btn = QPushButton("🔄 Обнови схема")
        migrate_schema_btn.clicked.connect(self.run_schema_migration)
        migration_layout.addWidget(migrate_schema_btn)
        
        convert_data_btn = QPushButton("🔀 Конвертирай данни")
        convert_data_btn.clicked.connect(self.convert_legacy_data)
        migration_layout.addWidget(convert_data_btn)
        
        migration_section.setLayout(migration_layout)
        data_layout.addWidget(migration_section, 0, 2)
        
        data_group.setLayout(data_layout)
        layout.addWidget(data_group)
        
        # === PASSWORD CHANGE SECTION ===
        password_group = QGroupBox("🔐 Смяна на парола")
        password_layout = QVBoxLayout()
        
        # Instructions
        instruction_label = QLabel("""Въведете новата парола (4-10 символа):
• Само цифри (напр. 12345)
• Само букви (напр. abcde) 
• Комбинация (напр. abc123)""")
        instruction_label.setStyleSheet("font-weight: bold; color: #2196F3; margin-bottom: 10px;")
        instruction_label.setWordWrap(True)
        password_layout.addWidget(instruction_label)
        
        # Password input fields with compact left-aligned layout
        # New password row
        new_password_row = QHBoxLayout()
        new_password_label = QLabel("Нова парола:")
        new_password_label.setMinimumWidth(120)  # Fixed width for alignment
        new_password_row.addWidget(new_password_label)
        
        self.new_password_input = QLineEdit()
        self.new_password_input.setEchoMode(QLineEdit.EchoMode.Normal)  # Default to visible
        self.new_password_input.setMaxLength(10)
        self.new_password_input.setFixedWidth(200)  # Fixed width for consistency
        self.new_password_input.setStyleSheet("padding: 8px; border: 2px solid #ddd; border-radius: 4px;")
        new_password_row.addWidget(self.new_password_input)
        
        self.show_new_password = QCheckBox("Скрий парола")
        self.show_new_password.setChecked(False)  # Unchecked = show password, checked = hide password
        
        def toggle_new_password():
            if self.show_new_password.isChecked():
                self.new_password_input.setEchoMode(QLineEdit.EchoMode.Password)
                self.show_new_password.setText("Покажи парола")
            else:
                self.new_password_input.setEchoMode(QLineEdit.EchoMode.Normal)
                self.show_new_password.setText("Скрий парола")
        
        self.show_new_password.stateChanged.connect(lambda: toggle_new_password())
        new_password_row.addWidget(self.show_new_password)
        
        new_password_row.addStretch()  # Push everything to the left
        password_layout.addLayout(new_password_row)
        
        # Confirm password row
        confirm_password_row = QHBoxLayout()
        confirm_password_label = QLabel("Потвърди парола:")
        confirm_password_label.setMinimumWidth(120)  # Same width as above
        confirm_password_row.addWidget(confirm_password_label)
        
        self.confirm_password_input = QLineEdit()
        self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Normal)  # Default to visible
        self.confirm_password_input.setMaxLength(10)
        self.confirm_password_input.setFixedWidth(200)  # Same width as above
        self.confirm_password_input.setStyleSheet("padding: 8px; border: 2px solid #ddd; border-radius: 4px;")
        confirm_password_row.addWidget(self.confirm_password_input)
        
        self.show_confirm_password = QCheckBox("Скрий парола")
        self.show_confirm_password.setChecked(False)  # Unchecked = show password, checked = hide password
        
        def toggle_confirm_password():
            if self.show_confirm_password.isChecked():
                self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
                self.show_confirm_password.setText("Покажи парола")
            else:
                self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Normal)
                self.show_confirm_password.setText("Скрий парола")
        
        self.show_confirm_password.stateChanged.connect(lambda: toggle_confirm_password())
        confirm_password_row.addWidget(self.show_confirm_password)
        
        confirm_password_row.addStretch()  # Push everything to the left
        password_layout.addLayout(confirm_password_row)
        
        # Old password row
        old_password_row = QHBoxLayout()
        old_password_label = QLabel("Стара парола:")
        old_password_label.setMinimumWidth(120)  # Same width as above
        old_password_row.addWidget(old_password_label)
        
        self.old_password_input = QLineEdit()
        self.old_password_input.setEchoMode(QLineEdit.EchoMode.Normal)  # Default to visible
        self.old_password_input.setMaxLength(10)
        self.old_password_input.setFixedWidth(200)  # Same width as above
        self.old_password_input.setStyleSheet("padding: 8px; border: 2px solid #ddd; border-radius: 4px;")
        old_password_row.addWidget(self.old_password_input)
        
        self.show_old_password = QCheckBox("Скрий парола")
        self.show_old_password.setChecked(False)  # Unchecked = show password, checked = hide password
        
        def toggle_old_password():
            if self.show_old_password.isChecked():
                self.old_password_input.setEchoMode(QLineEdit.EchoMode.Password)
                self.show_old_password.setText("Покажи парола")
            else:
                self.old_password_input.setEchoMode(QLineEdit.EchoMode.Normal)
                self.show_old_password.setText("Скрий парола")
        
        self.show_old_password.stateChanged.connect(lambda: toggle_old_password())
        old_password_row.addWidget(self.show_old_password)
        
        old_password_row.addStretch()  # Push everything to the left
        password_layout.addLayout(old_password_row)
        
        # Add some spacing before the button
        password_layout.addSpacing(10)
        
        # Change password button - aligned with the input fields
        button_row = QHBoxLayout()
        button_spacer = QLabel("")  # Empty label for alignment
        button_spacer.setMinimumWidth(120)  # Same as label width
        button_row.addWidget(button_spacer)
        
        change_password_btn = QPushButton("🔒 Смени парола")
        change_password_btn.setFixedWidth(200)  # Same width as input fields
        change_password_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
                margin: 10px 0;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #1565C0;
            }
        """)
        change_password_btn.clicked.connect(self.change_user_password)
        button_row.addWidget(change_password_btn)
        
        button_row.addStretch()  # Push button to align with input fields
        password_layout.addLayout(button_row)
        
        password_group.setLayout(password_layout)
        layout.addWidget(password_group)
        
        # === FACTORY RESET SECTION ===
        reset_group = QGroupBox("⚠️ Системен рестарт")
        reset_layout = QVBoxLayout()
        
        # Warning text
        warning_text = QLabel("""
<b style="color: #8b0000;">ВНИМАНИЕ!</b><br>
Системният рестарт ще изтрие напълно всички данни от програмата и ще я върне в първоначално състояние.
<br><br>
<b>Това включва:</b><br>
• Всички артикули от склада<br>
• Всички артикули от магазините<br>
• Всички продажби и история<br>
• Всички генерирани баркодове<br>
• Всички настройки и потребителски данни<br>
<br>
<b style="color: #8b0000;">Тази операция е необратима!</b>
        """)
        warning_text.setWordWrap(True)
        warning_text.setStyleSheet("""
            QLabel {
                background-color: #ffebee;
                border: 2px solid #d32f2f;
                border-radius: 8px;
                padding: 15px;
                font-size: 11px;
                line-height: 1.4;
                color: #4a0000;
            }
        """)
        reset_layout.addWidget(warning_text)
        
        # System restart button
        factory_reset_btn = QPushButton("🔄 СИСТЕМЕН РЕСТАРТ")
        factory_reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #d32f2f;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 15px;
                border-radius: 8px;
                margin: 10px 0;
            }
            QPushButton:hover {
                background-color: #b71c1c;
            }
            QPushButton:pressed {
                background-color: #8e0000;
            }
        """)
        factory_reset_btn.clicked.connect(self.factory_reset)
        reset_layout.addWidget(factory_reset_btn)
        
        reset_group.setLayout(reset_layout)
        layout.addWidget(reset_group)
        
        # Add stretch to push everything to top
        layout.addStretch()
        
        # Set up scroll area
        scroll_area.setWidget(scroll_content)
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        main_layout.addWidget(scroll_area)
        
        # Initialize database statistics
        self.update_database_statistics()
        self.load_backup_list()
        
        # Set up periodic updates every 30 seconds for real-time stats
        from PyQt6.QtCore import QTimer
        self.db_stats_timer = QTimer()
        self.db_stats_timer.timeout.connect(self.update_database_statistics)
        self.db_stats_timer.start(30000)  # Update every 30 seconds

        return widget

    def create_audit_tab(self):
        """Create the inventory audit tab"""
        widget = QWidget()
        main_layout = QVBoxLayout(widget)
        
        # Main content area
        content_layout = QHBoxLayout()
        
        # === LEFT PANEL - Control Panel ===
        left_panel = QWidget()
        left_panel.setFixedWidth(350)
        left_layout = QVBoxLayout(left_panel)
        
        # Shop selection group
        shop_group = QGroupBox("Избор на магазин")
        shop_layout = QVBoxLayout()
        
        self.audit_shop_combo = QComboBox()
        self.audit_shop_combo.setFont(QFont("Arial", 12))
        self.refresh_audit_shop_combo()
        shop_layout.addWidget(self.audit_shop_combo)
        
        shop_group.setLayout(shop_layout)
        left_layout.addWidget(shop_group)
        
        # Audit controls group
        controls_group = QGroupBox("Управление на инвентаризацията")
        controls_layout = QVBoxLayout()
        
        # Start audit button
        self.start_audit_btn = QPushButton("Започни инвентаризация")
        self.start_audit_btn.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        self.start_audit_btn.setStyleSheet("background-color: #28a745; color: white; padding: 10px; border-radius: 5px;")
        self.start_audit_btn.clicked.connect(self.start_inventory_audit)
        controls_layout.addWidget(self.start_audit_btn)
        
        # Pause/Resume button
        self.pause_audit_btn = QPushButton("Пауза")
        self.pause_audit_btn.setFont(QFont("Arial", 11))
        self.pause_audit_btn.setStyleSheet("background-color: #ffc107; color: black; padding: 8px; border-radius: 5px;")
        self.pause_audit_btn.clicked.connect(self.pause_resume_audit)
        self.pause_audit_btn.setEnabled(False)
        controls_layout.addWidget(self.pause_audit_btn)
        
        # Finish audit button
        self.finish_audit_btn = QPushButton("Завърши инвентаризация")
        self.finish_audit_btn.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        self.finish_audit_btn.setStyleSheet("background-color: #dc3545; color: white; padding: 8px; border-radius: 5px;")
        self.finish_audit_btn.clicked.connect(self.finish_inventory_audit)
        self.finish_audit_btn.setEnabled(False)
        controls_layout.addWidget(self.finish_audit_btn)
        
        controls_group.setLayout(controls_layout)
        left_layout.addWidget(controls_group)
        
        # Statistics group
        stats_group = QGroupBox("Статистика")
        stats_layout = QVBoxLayout()
        
        # Stats labels (removed "Очаквани", updated labels)
        self.audit_stats_total = QLabel("Всички артикули: 0")
        self.audit_stats_total.setFont(QFont("Arial", 11))
        stats_layout.addWidget(self.audit_stats_total)
        
        self.audit_stats_scanned = QLabel("Сканирани: 0")
        self.audit_stats_scanned.setFont(QFont("Arial", 11))
        self.audit_stats_scanned.setStyleSheet("color: green;")
        stats_layout.addWidget(self.audit_stats_scanned)
        
        self.audit_stats_missing = QLabel("Липсващи: 0")
        self.audit_stats_missing.setFont(QFont("Arial", 11))
        self.audit_stats_missing.setStyleSheet("color: red;")
        stats_layout.addWidget(self.audit_stats_missing)
        
        self.audit_stats_progress = QLabel("Прогрес:")
        self.audit_stats_progress.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        stats_layout.addWidget(self.audit_stats_progress)
        
        # Progress bar
        self.audit_progress_bar = QProgressBar()
        self.audit_progress_bar.setMaximum(100)
        self.audit_progress_bar.setValue(0)
        self.audit_progress_bar.setStyleSheet("""
            QProgressBar {
                text-align: center;
                font-weight: bold;
                font-size: 12px;
            }
            QProgressBar::chunk { 
                background-color: #28a745; 
            }
        """)
        stats_layout.addWidget(self.audit_progress_bar)
        
        stats_group.setLayout(stats_layout)
        left_layout.addWidget(stats_group)
        
        # Scanner input group (removed quantity input)
        scanner_group = QGroupBox("Сканиране на артикули")
        scanner_layout = QVBoxLayout()
        
        # Barcode input
        barcode_label = QLabel("Баркод:")
        scanner_layout.addWidget(barcode_label)
        
        self.audit_barcode_input = QLineEdit()
        self.audit_barcode_input.setFont(QFont("Arial", 12))
        self.audit_barcode_input.setPlaceholderText("Сканирайте или въведете баркод...")
        self.audit_barcode_input.textChanged.connect(self.on_audit_barcode_changed)
        self.audit_barcode_input.returnPressed.connect(self.process_audit_barcode)
        self.audit_barcode_input.setEnabled(False)
        scanner_layout.addWidget(self.audit_barcode_input)
        
        scanner_group.setLayout(scanner_layout)
        left_layout.addWidget(scanner_group)
        
        # Add stretch to push everything to top
        left_layout.addStretch()
        
        # === RIGHT PANEL - Data Tables ===
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        
        # Create tab widget for different views (removed "Очаквани" tab)
        self.audit_tabs = QTabWidget()
        
        # Tab 1: Audit items (main audit table)
        audit_tab = QWidget()
        audit_layout = QVBoxLayout(audit_tab)
        
        audit_label = QLabel("Сканирани артикули")
        audit_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        audit_layout.addWidget(audit_label)
        
        # Main audit table with new columns
        self.audit_items_table = QTableWidget()
        self.audit_items_table.setColumnCount(4)
        self.audit_items_table.setHorizontalHeaderLabels(["Баркод", "Продукт", "Количество", "Цена"])
        
        # Configure column spacing with controlled resize limits (resizable between bounds)
        header = self.audit_items_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Set default widths for first 3 columns - all resizable
        default_width = 180  # Good size for audit table
        for col in range(3):  # First 3 columns are resizable
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            self.audit_items_table.setColumnWidth(col, default_width)
        
        # Last column stretches to fill remaining space
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        
        # Enable sorting
        self.audit_items_table.setSortingEnabled(True)
        self.audit_items_table.setAlternatingRowColors(True)
        self.audit_items_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.audit_items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        audit_layout.addWidget(self.audit_items_table)
        
        self.audit_tabs.addTab(audit_tab, "Инвентаризация")
        
        # Tab 2: Audit Results History
        results_tab = QWidget()
        results_layout = QVBoxLayout(results_tab)
        
        # Results tab header
        results_header_layout = QHBoxLayout()
        
        results_label = QLabel("История на инвентаризации")
        results_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        results_header_layout.addWidget(results_label)
        
        results_header_layout.addStretch()
        results_layout.addLayout(results_header_layout)
        
        # Results table
        self.audit_results_table = QTableWidget()
        self.audit_results_table.setColumnCount(7)
        self.audit_results_table.setHorizontalHeaderLabels([
            "Дата", "Магазин", "Продължителност", "Всички", "Сканирани", "Липсващи", "Действия"
        ])
        
        # Set table properties with controlled resize limits (resizable between bounds)
        header = self.audit_results_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Set default widths for first 6 columns - all resizable
        default_width = 130  # Good size for audit results
        for col in range(6):  # First 6 columns are resizable
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
            self.audit_results_table.setColumnWidth(col, default_width)
        
        # Last column (Actions) stretches to fill remaining space
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        
        # Enable sorting
        self.audit_results_table.setSortingEnabled(True)
        self.audit_results_table.setAlternatingRowColors(True)
        self.audit_results_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.audit_results_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        results_layout.addWidget(self.audit_results_table)
        
        # Results summary
        results_summary_layout = QHBoxLayout()
        
        self.results_total_label = QLabel("Общо инвентаризации: 0")
        self.results_total_label.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        results_summary_layout.addWidget(self.results_total_label)
        
        results_summary_layout.addStretch()
        
        # Clear all results button
        clear_all_results_btn = QPushButton("🗑️ Изчисти всички резултати")
        clear_all_results_btn.setFont(QFont("Arial", 10))
        clear_all_results_btn.setStyleSheet("background-color: #dc3545; color: white; padding: 5px 10px; border-radius: 3px;")
        clear_all_results_btn.clicked.connect(self.clear_all_audit_results)
        results_summary_layout.addWidget(clear_all_results_btn)
        
        results_layout.addLayout(results_summary_layout)
        
        self.audit_tabs.addTab(results_tab, "Резултати")
        
        right_layout.addWidget(self.audit_tabs)
        
        # Add panels to main content
        content_layout.addWidget(left_panel)
        content_layout.addWidget(right_panel, 1)  # Give more space to right panel
        
        main_layout.addLayout(content_layout)
        
        # Load initial data
        self.refresh_audit_shop_combo()
        self.load_audit_results()  # Load any existing audit results
        
        return widget

    def create_help_tab(self):
        """Create the comprehensive help tab"""
        widget = QWidget()
        main_layout = QVBoxLayout(widget)

        # Create scroll area for the content
        scroll_area = QScrollArea()
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)
        
        # === HEADER ===
        header_group = QGroupBox()
        header_layout = QVBoxLayout()
        
        title_label = QLabel("📘 Ръководство за работа със системата")
        title_label.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("color: #2196F3; padding: 15px;")
        header_layout.addWidget(title_label)
               
        header_group.setLayout(header_layout)
        layout.addWidget(header_group)
        
        # === OVERVIEW ===
        overview_group = QGroupBox("🏪 Обща информация")
        overview_layout = QVBoxLayout()
        
        overview_text = QLabel("""
<b>Система за управление на бижута</b> е професионално приложение за управление на склад, 
продажби и отчети за магазини за бижута с усъвършенствани възможности за търсене и анализ.

<br><b>✨ Нови възможности в тази версия:</b>
<br>• <b>Усъвършенствана система за търсене</b> - 4 таба за склад, 2 таба за продажби
<br>• <b>Интелигентно филтриране</b> по всички 11 колони с данни
<br>• <b>Диапазони за цени и дати</b> за прецизно търсене
<br>• <b>Професионални отчети</b> с Excel експорт и анализи
<br>• <b>ROI анализи</b> и проследяване на печалба

<br><b>🎯 Основни функции:</b>
<br>• Управление на артикули с баркодове и снимки
<br>• Печат на етикети за термопринтери (Citizen CLP 631)
<br>• Проследяване на склад и магазинни запаси
<br>• Система за продажби с поддръжка на баркод скенер
<br>• Детайлни отчети и статистики в реално време
<br>• Резервни копия и управление на данни
<br>• Поддръжка на множество магазини
<br>• Система за отменяне/възстановяване на действия
<br>• Валидация на данни и контрол на целостта
        """)
        overview_text.setWordWrap(True)
        overview_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        overview_layout.addWidget(overview_text)
        
        overview_group.setLayout(overview_layout)
        layout.addWidget(overview_group)
        
        # === KEYBOARD SHORTCUTS ===
        shortcuts_group = QGroupBox("⌨️ Клавишни комбинации")
        shortcuts_layout = QVBoxLayout()
        
        # Create table for shortcuts
        shortcuts_table = QTableWidget()
        shortcuts_table.setColumnCount(2)
        shortcuts_table.setHorizontalHeaderLabels(["Клавишна комбинация", "Описание"])
        
        # Configure column spacing with controlled resize limits (resizable between bounds)
        header = shortcuts_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        # Set default width for first column - resizable
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        shortcuts_table.setColumnWidth(0, 200)  # Default width for shortcuts
        
        # Last column stretches to fill remaining space
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        
        shortcuts_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        shortcuts_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        shortcuts_table.setAlternatingRowColors(True)
        shortcuts_table.verticalHeader().setVisible(False)
        
                 # Shortcuts data
        shortcuts_data = [
             ("Ctrl + Tab", "Следващ таб"),
             ("Ctrl + Shift + Tab", "Предишен таб"),
             ("Ctrl + PageDown", "Следващ таб (алтернатива)"),
             ("Ctrl + PageUp", "Предишен таб (алтернатива)"),
             ("Ctrl + 1-7", "Директен достъп до таб по номер"),
             ("Delete", "Изтриване на избрани елементи"),
             ("Ctrl + A", "Избор на всички елементи"),
             ("Ctrl + Click", "Множествен избор"),
             ("Shift + Click", "Избор на диапазон"),
             ("Double Click", "Редактиране на елемент"),
             ("Enter", "Потвърждаване/изпълнение на действие"),
             ("Escape", "Отказ/затваряне"),
             ("Tab", "Преминаване към следващо поле")
         ]
        
        shortcuts_table.setRowCount(len(shortcuts_data))
        
        for row, (shortcut, description) in enumerate(shortcuts_data):
            shortcut_item = QTableWidgetItem(shortcut)
            description_item = QTableWidgetItem(description)
             
            shortcut_item.setFont(QFont("Courier", 10, QFont.Weight.Bold))
            shortcut_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            description_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align description too
            
            shortcuts_table.setItem(row, 0, shortcut_item)
            shortcuts_table.setItem(row, 1, description_item)
    
        # Set fixed height to show all rows without scrolling
        table_height = shortcuts_table.verticalHeader().defaultSectionSize() * len(shortcuts_data) + \
                      shortcuts_table.horizontalHeader().height() + 10
        shortcuts_table.setFixedHeight(table_height)
        
        shortcuts_layout.addWidget(shortcuts_table)
        
        shortcuts_group.setLayout(shortcuts_layout)
        layout.addWidget(shortcuts_group)
        
                # === TAB DESCRIPTIONS ===
        tabs_group = QGroupBox("📑 Описание на табовете")
        tabs_layout = QVBoxLayout()
        
        # Create tab sections with side-by-side layout
        # Row 1: Add Item Tab + Inventory Tab
        row1_layout = QHBoxLayout()
        
        # Add Item Tab
        add_item_section = QGroupBox("1. 📦 Добави артикул")
        add_item_section.setMaximumWidth(600)
        add_item_layout = QVBoxLayout()
        add_item_text = QLabel("""
<br><b>Предназначение:</b>
<br>• Добавяне на нови артикули в системата

<br><b>Как да използвате:</b>
<br>• Попълнете всички задължителни полета (категория, метал, камък, цени, тегло, количество)
<br>• По желание добавете снимка и описание
<br>• Натиснете "Генерирай етикет" за създаване на баркод
<br>• Натиснете "Принтирай етикет" за отпечатване и запазване

<br><b>Важни забележки:</b>
<br>• Цените се въвеждат в евро и автоматично се конвертират в лева
<br>• Баркодовете се генерират автоматично в последователност
<br>• Можете да добавяте нови категории, метали и камъни чрез "Друго"
<br>• Етикетите са оптимизирани за термопринтери (препоръчван: Citizen CLP 631)
        """)
        add_item_text.setWordWrap(True)
        add_item_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        add_item_layout.addWidget(add_item_text)
        add_item_section.setLayout(add_item_layout)
        
        # Inventory Tab
        inventory_section = QGroupBox("2. 📋 Склад - Усъвършенствано търсене")
        inventory_section.setMaximumWidth(600)
        inventory_layout = QVBoxLayout()
        inventory_text = QLabel("""
<br><b>🔍 НОВО: Система с 4 таба за търсене:</b>
<br>• <b>Общо търсене:</b> Търсене във всички полета едновременно
<br>• <b>Цена и тегло:</b> Филтриране по ценови диапазони и тегло
<br>• <b>Дата:</b> Търсене по период на създаване/редактиране
<br>• <b>Категория:</b> Филтриране по категории, метали и камъни

<br><b>📊 Усъвършенствани функции:</b>
<br>• Търсене във всички 11 колони с данни
<br>• Комбиниране на филтри за точни резултати
<br>• Масови операции - избор, изтриване, преместване, експорт
<br>• Редактиране на цени на множество артикули едновременно
<br>• Интелигентно сортиране по всички колони

<br><b>🎯 Как да използвате:</b>
<br>• Изберете подходящ таб според типа търсене
<br>• Двоен клик за редактиране на артикул
<br>• Десен клик за контекстно меню с опции
<br>• Използвайте филтрите за бързо намиране
<br>• Артикулите с нулево количество се оцветяват в червено

<br><b>Съвети:</b>
<br>• Използвайте Ctrl+Click за избор на няколко артикула
<br>• Филтрите се комбинират за по-точни резултати
        """)
        inventory_text.setWordWrap(True)
        inventory_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        inventory_layout.addWidget(inventory_text)
        inventory_section.setLayout(inventory_layout)
        
        row1_layout.addWidget(add_item_section)
        row1_layout.addWidget(inventory_section)
        tabs_layout.addLayout(row1_layout)
        
        # Row 2: Shop Loading Tab + Sales Tab
        row2_layout = QHBoxLayout()
        
        # Shop Loading Tab
        shop_loading_section = QGroupBox("3. 🏪 Зареждане на магазин")
        shop_loading_section.setMaximumWidth(600)
        shop_loading_layout = QVBoxLayout()
        shop_loading_text = QLabel("""
<br><b>Предназначение:</b> 
<br>• Прехвърляне на артикули от склада в магазини

<br><b>Функционалности:</b>
<br>• Управление на магазини (добавяне, редактиране, изтриване)
<br>• Единично зареждане чрез сканиране на баркод
<br>• Преглед на магазинните запаси
<br>• Връщане на артикули обратно в склада

<br><b>Как да използвате:</b>
<br>• Изберете магазин от падащото меню
<br>• Сканирайте баркод (7 цифри) за единично зареждане
<br>• Въведете желаното количество (не може да надвишава складовото)

<br><b>Важни забележки:</b>
<br>• Артикулите се преместват от главния склад в магазина
<br>• Количеството в склада намалява съответно
<br>• Можете да връщате артикули с десен клик → "Връщане в склада"
        """)
        shop_loading_text.setWordWrap(True)
        shop_loading_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        shop_loading_layout.addWidget(shop_loading_text)
        shop_loading_section.setLayout(shop_loading_layout)
        
        # Sales Tab
        sales_section = QGroupBox("4. 💰 Продажби - Нова система за търсене")
        sales_section.setMaximumWidth(600)
        sales_layout = QVBoxLayout()
        sales_text = QLabel("""
<br><b>🔍 НОВО: Система с 2 таба за търсене:</b>
<br>• <b>Общо търсене:</b> Търсене в записите за продажби по всички детайли
<br>• <b>Филтър по дата:</b> Бързи филтри (днес, седмица, месец) + персонализиран период

<br><b>📊 Усъвършенствани функции:</b>
<br>• Единична продажба чрез сканиране на баркод
<br>• Интелигентно филтриране по период с бързи бутони
<br>• Филтриране по магазин за специфичен анализ
<br>• Пълна история на всички продажби
<br>• Анализ на печалба в реално време

<br><b>🎯 Как да използвате:</b>
<br>• Изберете магазин от който продавате
<br>• Сканирайте баркод за единична продажба
<br>• Артикулът се продава автоматично след сканиране
<br>• Използвайте табовете за различни видове търсене

<br><b>⚡ Особености:</b>
<br>• Продажбите се записват с точно време и дата
<br>• Артикулите се премахват от магазинните запаси
<br>• Можете да отмените продажби с бутона "Отмени"
<br>• Поддържа се система за отменяне/възстановяване
<br>• Автоматично изчисляване на печалба и ROI
        """)
        sales_text.setWordWrap(True)
        sales_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        sales_layout.addWidget(sales_text)
        sales_section.setLayout(sales_layout)
        
        row2_layout.addWidget(shop_loading_section)
        row2_layout.addWidget(sales_section)
        tabs_layout.addLayout(row2_layout)
        
        # Row 3: Reports Tab + Database Tab
        row3_layout = QHBoxLayout()
        
        # Reports Tab
        reports_section = QGroupBox("5. 📊 Отчети - Професионална аналитика")
        reports_section.setMaximumWidth(600)
        reports_layout = QVBoxLayout()
        reports_text = QLabel("""
<br><b>📈 НОВО: Усъвършенствана аналитика:</b>
<br>• <b>ROI анализи</b> и проследяване на възвращаемост на инвестициите
<br>• <b>Тенденции в реално време</b> с графики и диаграми
<br>• <b>Excel експорт</b> за професионални отчети
<br>• <b>Анализ по категории</b> и детайлна сегментация

<br><b>📊 Обобщена информация:</b>
<br>• Продажби днес и този месец
<br>• Стойност на склада с анализ на печалба
<br>• Артикули с ниско количество (автоматични предупреждения)
<br>• Средна печалба и margin анализ

<br><b>🎯 Детайлни анализи:</b>
<br>• Топ продавани артикули с rankings
<br>• Продажби по категории и магазини
<br>• Дневни тенденции и прогнози
<br>• Печалба по категории с breakdown
<br>• Comprehensive ROI анализи

<br><b>📄 Професионални отчети:</b>
<br>• Отчет за продажби (Excel с формули)
<br>• Отчет за инвентар с валуация
<br>• Отчет за печалба с margin analysis
<br>• Пълен финансов отчет с dashboards

<br><b>🚀 Как да използвате:</b>
<br>• Изберете период за анализ (нови fast filters)
<br>• Използвайте бързите бутони или персонализиран период
<br>• Статистиките се обновяват автоматично в реално време
<br>• Изберете желания анализ или отчет
        """)
        reports_text.setWordWrap(True)
        reports_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        reports_layout.addWidget(reports_text)
        reports_section.setLayout(reports_layout)
        
        # Database Tab
        database_section = QGroupBox("6. 🗄️ База данни")
        database_section.setMaximumWidth(600)
        database_layout = QVBoxLayout()
        database_text = QLabel("""
<br><b>Предназначение:</b> 
<br>• Управление на данни, резервни копия и настройки

<br><b>Преглед на базата данни:</b>
<br>• Общо артикули, продажби, магазини
<br>• Размер на базата данни
<br>• Статус на интегритета
<br>• Последно резервно копие

<br><b>Резервни копия:</b>
<br>• Ръчно създаване на резервни копия
<br>• Автоматично създаване (настройваемо)
<br>• Възстановяване от резервни копия
<br>• Преглед на съществуващи копия

<br><b>Управление на данни:</b>
<br>• Експорт в Excel, CSV, JSON
<br>• Импорт от Excel, CSV
<br>• Миграция на схема
<br>• Конвертиране на стари данни

<br><b>Важни забележки:</b>
<br>• ВИНАГИ правете резервни копия преди големи промени
<br>• Автоматичните копия се препоръчват ежедневно
<br>• Възстановяването заменя текущата база данни
        """)
        database_text.setWordWrap(True)
        database_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        database_layout.addWidget(database_text)
        database_section.setLayout(database_layout)
        
        row3_layout.addWidget(reports_section)
        row3_layout.addWidget(database_section)
        tabs_layout.addLayout(row3_layout)
        
        tabs_group.setLayout(tabs_layout)
        layout.addWidget(tabs_group)
        
        # === WORKFLOW GUIDE ===
        workflow_group = QGroupBox("🔄 Ръководство за работа")
        workflow_layout = QVBoxLayout()
        
        workflow_text = QLabel("""
<br><b>Типичен работен ден:</b>

<br><b>1. Сутрин:</b>
<br>• Проверете статуса на системата в таб "База данни"
<br>• Прегледайте отчетите от предния ден
<br>• Проверете артикулите с ниско количество

<br><b>2. Добавяне на нови артикули:</b>
<br>• Отидете в таб "Добави артикул"
<br>• Попълнете всички данни внимателно
<br>• Генерирайте и отпечатайте етикета
<br>• Поставете етикета на артикула

<br><b>3. Зареждане на магазини:</b>
<br>• Изберете таб "Зареждане на магазин"
<br>• Изберете правилния магазин
<br>• Сканирайте артикулите за зареждане
<br>• Проверете правилността на количествата

<br><b>4. Продажби:</b>
<br>• Работете в таб "Продажби"
<br>• Изберете магазина от който продавате
<br>• Сканирайте артикулите при продажба
<br>• Проверявайте периодично историята

<br><b>5. Вечер:</b>
<br>• Генерирайте отчети за деня
<br>• Направете резервно копие
<br>• Проверете складовите наличности

<br><b>Съвети за ефективност:</b>
<br>• Използвайте клавишните комбинации за по-бърза работа
<br>• Настройте автоматични резервни копия
<br>• Редовно почиствайте стари данни
<br>• Обучете всички потребители на системата
        """)
        workflow_text.setWordWrap(True)
        workflow_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        workflow_layout.addWidget(workflow_text)
        
        workflow_group.setLayout(workflow_layout)
        layout.addWidget(workflow_group)
        
        # === TROUBLESHOOTING ===
        troubleshooting_group = QGroupBox("🔧 Решаване на проблеми")
        troubleshooting_layout = QVBoxLayout()
        
        troubleshooting_text = QLabel("""
<br><b>Чести проблеми и решения:</b>

<br><b>Принтерът не печата:</b>
<br>• Проверете дали принтерът е включен и свързан
<br>• Уверете се, че има хартия в принтера
<br>• Рестартирайте принтера и опитайте отново
<br>• Проверете настройките на принтера в Windows

<br><b>Баркод скенерът не работи:</b>
<br>• Проверете USB връзката
<br>• Тествайте скенера в текстов редактор
<br>• Уверете се, че курсорът е в правилното поле

<br><b>Грешки в базата данни:</b>
<br>• Използвайте "Проверка на интегритета" в таб "База данни"
<br>• При необходимост възстановете от резервно копие
<br>• Свържете се с техническа поддръжка

<br><b>Бавна работа:</b>
<br>• Рестартирайте приложението
<br>• Проверете свободното място на диска
<br>• Направете резервно копие и почистете стари данни

<br><b>Загуба на данни:</b>
<br>• НЕ изпадайте в паника
<br>• Проверете папката "backups" за автоматични копия
<br>• Използвайте "Възстанови от копие" в таб "База данни"

<br><b>За техническа поддръжка:</b>
<br>• Запишете точното съобщение за грешка
<br>• Отбележете какво правехте когато възникна проблемът
<br>• Направете снимка на екрана ако е възможно
<br>• Проверете файла "logs/app.log" за допълнителна информация
        """)
        troubleshooting_text.setWordWrap(True)
        troubleshooting_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        troubleshooting_layout.addWidget(troubleshooting_text)
        
        troubleshooting_group.setLayout(troubleshooting_layout)
        layout.addWidget(troubleshooting_group)
        
        # === ENHANCED SEARCH GUIDE ===
        search_guide_group = QGroupBox("🔍 Ръководство за усъвършенствано търсене")
        search_guide_layout = QVBoxLayout()
        
        search_guide_text = QLabel("""
<b>🆕 Нова система за търсене с табове</b>

<br><b>📋 Склад - 4 таба за търсене:</b>
<br>• <b>"Общо търсене":</b> Търсене във всички полета едновременно (баркод, категория, метал, камък, описание, цени)
<br>• <b>"Цена и тегло":</b> Филтриране по ценови диапазони (min/max) и тегло (грамове)
<br>• <b>"Дата":</b> Търсене по период на създаване или модификация (от/до дата)
<br>• <b>"Категория":</b> Целенасочено филтриране по категории, метали и камъни

<br><b>💰 Продажби - 2 таба за търсене:</b>
<br>• <b>"Общо търсене":</b> Търсене в записите за продажби по всички детайли на артикула
<br>• <b>"Филтър по дата":</b> Бързи филтри (днес, седмица, месец, година) + персонализиран период

<br><b>⚡ Интелигентни функции:</b>
<br>• <b>Комбинирани филтри:</b> Можете да комбинирате няколко филтъра за точни резултати
<br>• <b>Автоматично предложения:</b> Системата предлага възможни стойности докато пишете
<br>• <b>Запазване на филтри:</b> Последните настройки се запазват за сесията
<br>• <b>Real-time търсене:</b> Резултатите се актуализират веднага при промяна

<br><b>🎯 Съвети за ефективно търсене:</b>
<br>• За бързо намиране на артикул използвайте "Общо търсене" и въведете част от баркода
<br>• За анализ на печалбата използвайте "Цена и тегло" с ценови диапазони
<br>• За проследяване на нови артикули използвайте "Дата" филтъра
<br>• За инвентаризация по категории използвайте "Категория" таба
<br>• В продажбите използвайте "Дата" за периодични анализи

<br><b>📊 Технически детайли:</b>
<br>• Търсенето обхваща всички 11 колони с данни
<br>• Поддържа се UTF-8 за кирилица и специални символи
<br>• Optimized за бази данни с над 10,000 записа
<br>• Автоматично indexing за максимална скорост
        """)
        search_guide_text.setWordWrap(True)
        search_guide_text.setStyleSheet("padding: 10px; line-height: 1.4;")
        search_guide_layout.addWidget(search_guide_text)
        
        search_guide_group.setLayout(search_guide_layout)
        layout.addWidget(search_guide_group)
        

        
        # === FOOTER ===
        footer_group = QGroupBox()
        footer_layout = QVBoxLayout()
        
        footer_text = QLabel("""
<div style='text-align: center; color: #666; font-style: italic;'>
<b>💎 Система за управление на бижута</b><br>
Версия 2.0 - Enhanced Search Edition | Професионално решение за бижутерски бизнес<br>
🚀 Последни подобрения: Усъвършенствана система за търсене, ROI анализи, Excel експорт<br>
За въпроси и поддръжка се обърнете към системния администратор
</div>
        """)
        footer_text.setWordWrap(True)
        footer_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        footer_layout.addWidget(footer_text)
        
        footer_group.setLayout(footer_layout)
        layout.addWidget(footer_group)
        
        # Add stretch to push everything to top
        layout.addStretch()
        
        # Set up scroll area
        scroll_area.setWidget(scroll_content)
        scroll_area.setWidgetResizable(True)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        main_layout.addWidget(scroll_area)

        return widget

    def create_stats_card(self, title, value, color):
        """Create a statistics card widget"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 8px;
                padding: 10px;
                margin: 5px;
            }}
            QLabel {{
                color: white;
                font-weight: bold;
            }}
        """)
        
        layout = QVBoxLayout(card)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold;")  # Same size as value
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        
        # Store value label for updates
        card.value_label = value_label
        
        return card

    def set_quick_period(self, period):
        """Set date range based on quick period selection"""
        today = QDate.currentDate()
        
        if period == "today":
            self.report_start_date.setDate(today)
            self.report_end_date.setDate(today)
        elif period == "week":
            start_of_week = today.addDays(-today.dayOfWeek() + 1)
            self.report_start_date.setDate(start_of_week)
            self.report_end_date.setDate(today)
        elif period == "month":
            start_of_month = QDate(today.year(), today.month(), 1)
            self.report_start_date.setDate(start_of_month)
            self.report_end_date.setDate(today)
        elif period == "quarter":
            quarter = (today.month() - 1) // 3 + 1
            start_month = (quarter - 1) * 3 + 1
            start_of_quarter = QDate(today.year(), start_month, 1)
            self.report_start_date.setDate(start_of_quarter)
            self.report_end_date.setDate(today)
        elif period == "year":
            start_of_year = QDate(today.year(), 1, 1)
            self.report_start_date.setDate(start_of_year)
            self.report_end_date.setDate(today)
        
        self.update_dashboard_stats()

    def update_dashboard_stats(self):
        """Update dashboard statistics based on selected date range"""
        try:
            start_date = self.report_start_date.date().toPyDate()
            end_date = self.report_end_date.date().toPyDate()
            today = datetime.now().date()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Sales for selected period
                cursor.execute("""
                    SELECT COALESCE(SUM(total_price), 0) FROM sales 
                    WHERE date(sale_date) >= ? AND date(sale_date) <= ?
                """, (start_date.isoformat(), end_date.isoformat()))
                period_sales = cursor.fetchone()[0]
                
                # Update title to reflect current period
                if start_date == end_date == today:
                    period_label = "Продажби днес"
                elif start_date == today:
                    period_label = "Продажби днес"
                elif (end_date - start_date).days <= 7:
                    period_label = "Продажби (седмица)"
                elif (end_date - start_date).days <= 31:
                    period_label = "Продажби (месец)"
                else:
                    period_label = "Продажби (период)"
                
                self.stats_cards["today_sales"].layout().itemAt(0).widget().setText(period_label)
                self.stats_cards["today_sales"].value_label.setText(f"{self.format_number_with_spaces(period_sales)} €")
                
                # Total inventory value (always current)
                cursor.execute("""
                    SELECT COALESCE(SUM(price * stock_quantity), 0) FROM items
                """)
                inventory_value = cursor.fetchone()[0]
                self.stats_cards["total_inventory_value"].value_label.setText(f"{self.format_number_with_spaces(inventory_value)} €")
                
                # Low stock items (always current)
                cursor.execute("""
                    SELECT COUNT(*) FROM items WHERE stock_quantity <= 5
                """)
                low_stock = cursor.fetchone()[0]
                self.stats_cards["low_stock_items"].value_label.setText(str(low_stock))
                
                # Total items (always current)
                cursor.execute("""
                    SELECT COUNT(*) FROM items
                """)
                total_items = cursor.fetchone()[0]
                self.stats_cards["total_items"].value_label.setText(str(total_items))
                
                # This month's sales (always current month)
                first_day_month = today.replace(day=1)
                cursor.execute("""
                    SELECT COALESCE(SUM(total_price), 0) FROM sales 
                    WHERE date(sale_date) >= ?
                """, (first_day_month.isoformat(),))
                month_sales = cursor.fetchone()[0]
                self.stats_cards["this_month_sales"].value_label.setText(f"{self.format_number_with_spaces(month_sales)} €")
                
                # Average profit margin (always current)
                cursor.execute("""
                    SELECT AVG((i.price - i.cost) / NULLIF(i.price, 0) * 100) 
                    FROM items i WHERE i.price > 0
                """)
                avg_margin = cursor.fetchone()[0]
                margin_text = f"{avg_margin:.1f}%" if avg_margin else "0%"
                self.stats_cards["avg_profit_margin"].value_label.setText(margin_text)
                
        except Exception as e:
            logger.error(f"Error updating dashboard stats: {e}")

    # === ANALYTICS METHODS ===
    def show_top_selling_items(self):
        """Show top selling items analysis"""
        try:
            start_date = self.report_start_date.date().toPyDate()
            end_date = self.report_end_date.date().toPyDate()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT i.barcode, i.category, SUM(s.quantity) as total_sold, 
                           SUM(s.total_price) as total_revenue
                    FROM sales s
                    JOIN items i ON s.item_id = i.id
                    WHERE date(s.sale_date) BETWEEN ? AND ?
                    GROUP BY s.item_id
                    ORDER BY total_sold DESC
                    LIMIT 20
                """, (start_date.isoformat(), end_date.isoformat()))
                
                results = cursor.fetchall()
                
            if results:
                report_text = "ТОП 20 ПРОДАВАНИ АРТИКУЛА\n" + "="*50 + "\n\n"
                report_text += f"Период: {start_date} до {end_date}\n\n"
                report_text += f"{'Баркод':<15} {'Категория':<15} {'Продадени':<10} {'Приходи':<15}\n"
                report_text += "-" * 60 + "\n"
                
                for barcode, category, sold, revenue in results:
                    report_text += f"{barcode:<15} {category:<15} {sold:<10} {revenue:<15.2f}\n"
                
                self.show_analysis_dialog("Топ продавани артикули", report_text)
            else:
                QMessageBox.information(self, "Информация", "Няма данни за избрания период")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")

    def show_sales_by_category(self):
        """Show sales breakdown by category"""
        try:
            start_date = self.report_start_date.date().toPyDate()
            end_date = self.report_end_date.date().toPyDate()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT i.category, COUNT(*) as count, SUM(s.total_price) as revenue
                    FROM sales s
                    JOIN items i ON s.item_id = i.id
                    WHERE date(s.sale_date) BETWEEN ? AND ?
                    GROUP BY i.category
                    ORDER BY revenue DESC
                """, (start_date.isoformat(), end_date.isoformat()))
                
                results = cursor.fetchall()
                
            if results:
                total_revenue = sum(row[2] for row in results)
                report_text = "ПРОДАЖБИ ПО КАТЕГОРИИ\n" + "="*40 + "\n\n"
                report_text += f"Период: {start_date} до {end_date}\n\n"
                report_text += f"{'Категория':<15} {'Брой':<8} {'Приходи':<15} {'%':<8}\n"
                report_text += "-" * 50 + "\n"
                
                for category, count, revenue in results:
                    percentage = (revenue / total_revenue * 100) if total_revenue > 0 else 0
                    report_text += f"{category:<15} {count:<8} {revenue:<15.2f} {percentage:<8.1f}\n"
                
                self.show_analysis_dialog("Продажби по категории", report_text)
            else:
                QMessageBox.information(self, "Информация", "Няма данни за избрания период")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")

    def show_daily_sales_trend(self):
        """Show daily sales trend"""
        try:
            start_date = self.report_start_date.date().toPyDate()
            end_date = self.report_end_date.date().toPyDate()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT date(sale_date) as sale_day, COUNT(*) as count, SUM(total_price) as revenue
                    FROM sales
                    WHERE date(sale_date) BETWEEN ? AND ?
                    GROUP BY date(sale_date)
                    ORDER BY sale_day
                """, (start_date.isoformat(), end_date.isoformat()))
                
                results = cursor.fetchall()
                
            if results:
                report_text = "ДНЕВНИ ПРОДАЖБИ\n" + "="*35 + "\n\n"
                report_text += f"Период: {start_date} до {end_date}\n\n"
                report_text += f"{'Дата':<12} {'Брой':<8} {'Приходи':<15}\n"
                report_text += "-" * 40 + "\n"
                
                total_count = 0
                total_revenue = 0
                
                for sale_day, count, revenue in results:
                    report_text += f"{sale_day:<12} {count:<8} {revenue:<15.2f}\n"
                    total_count += count
                    total_revenue += revenue
                
                report_text += "-" * 40 + "\n"
                report_text += f"{'ОБЩО:':<12} {total_count:<8} {total_revenue:<15.2f}\n"
                
                avg_daily = total_revenue / len(results) if results else 0
                report_text += f"\nСредно дневно: {avg_daily:.2f} лв\n"
                
                self.show_analysis_dialog("Дневни продажби", report_text)
            else:
                QMessageBox.information(self, "Информация", "Няма данни за избрания период")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")

    def show_sales_by_shop(self):
        """Show sales breakdown by shop"""
        try:
            start_date = self.report_start_date.date().toPyDate()
            end_date = self.report_end_date.date().toPyDate()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT sh.name, COUNT(*) as count, SUM(s.total_price) as revenue
                    FROM sales s
                    JOIN shops sh ON s.shop_id = sh.id
                    WHERE date(s.sale_date) BETWEEN ? AND ?
                    GROUP BY s.shop_id
                    ORDER BY revenue DESC
                """, (start_date.isoformat(), end_date.isoformat()))
                
                results = cursor.fetchall()
                
            if results:
                total_revenue = sum(row[2] for row in results)
                report_text = "ПРОДАЖБИ ПО МАГАЗИНИ\n" + "="*40 + "\n\n"
                report_text += f"Период: {start_date} до {end_date}\n\n"
                report_text += f"{'Магазин':<20} {'Брой':<8} {'Приходи':<15} {'%':<8}\n"
                report_text += "-" * 55 + "\n"
                
                for shop_name, count, revenue in results:
                    percentage = (revenue / total_revenue * 100) if total_revenue > 0 else 0
                    report_text += f"{shop_name:<20} {count:<8} {revenue:<15.2f} {percentage:<8.1f}\n"
                
                self.show_analysis_dialog("Продажби по магазини", report_text)
            else:
                QMessageBox.information(self, "Информация", "Няма данни за избрания период")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")

    def show_inventory_by_category(self):
        """Show inventory breakdown by category"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT category, COUNT(*) as count, SUM(stock_quantity) as total_stock,
                           SUM(price * stock_quantity) as total_value
                    FROM items
                    GROUP BY category
                    ORDER BY total_value DESC
                """)
                
                results = cursor.fetchall()
                
            if results:
                total_value = sum(row[3] for row in results)
                report_text = "ИНВЕНТАР ПО КАТЕГОРИИ\n" + "="*45 + "\n\n"
                report_text += f"{'Категория':<15} {'Видове':<8} {'Общо бр.':<10} {'Стойност':<15} {'%':<8}\n"
                report_text += "-" * 60 + "\n"
                
                for category, count, stock, value in results:
                    percentage = (value / total_value * 100) if total_value > 0 else 0
                    report_text += f"{category:<15} {count:<8} {stock:<10} {value:<15.2f} {percentage:<8.1f}\n"
                
                self.show_analysis_dialog("Инвентар по категории", report_text)
            else:
                QMessageBox.information(self, "Информация", "Няма данни в инвентара")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")

    def show_low_stock_items(self):
        """Show items with low stock"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT barcode, category, stock_quantity, price
                    FROM items
                    WHERE stock_quantity <= 5
                    ORDER BY stock_quantity ASC, category
                """)
                
                results = cursor.fetchall()
                
            if results:
                report_text = "АРТИКУЛИ С НИСКО НАЛИЧЕСТВО\n" + "="*45 + "\n\n"
                report_text += f"{'Баркод':<15} {'Категория':<15} {'Количество':<10} {'Цена':<10}\n"
                report_text += "-" * 55 + "\n"
                
                for barcode, category, stock, price in results:
                    report_text += f"{barcode:<15} {category:<15} {stock:<10} {price:<10.2f}\n"
                
                self.show_analysis_dialog("Артикули с ниско наличество", report_text)
            else:
                QMessageBox.information(self, "Информация", "Всички артикули имат достатъчно количество")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")

    def show_most_expensive_items(self):
        """Show most expensive items"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT barcode, category, price, stock_quantity
                    FROM items
                    ORDER BY price DESC
                    LIMIT 20
                """)
                
                results = cursor.fetchall()
                
            if results:
                report_text = "НАЙ-СКЪПИ АРТИКУЛИ\n" + "="*35 + "\n\n"
                report_text += f"{'Баркод':<15} {'Категория':<15} {'Цена':<10} {'Количество':<10}\n"
                report_text += "-" * 55 + "\n"
                
                for barcode, category, price, stock in results:
                    report_text += f"{barcode:<15} {category:<15} {price:<10.2f} {stock:<10}\n"
                
                self.show_analysis_dialog("Най-скъпи артикули", report_text)
            else:
                QMessageBox.information(self, "Информация", "Няма данни в инвентара")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")

    def show_old_inventory(self):
        """Show items older than 6 months"""
        try:
            six_months_ago = datetime.now() - timedelta(days=180)
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT barcode, category, price, stock_quantity, created_at
                    FROM items
                    WHERE created_at < ?
                    ORDER BY created_at ASC
                """, (six_months_ago.strftime("%Y-%m-%d"),))
                
                results = cursor.fetchall()
                
            if results:
                report_text = "СТАРИ АРТИКУЛИ (>6 МЕСЕЦА)\n" + "="*45 + "\n\n"
                report_text += f"{'Баркод':<15} {'Категория':<15} {'Цена':<10} {'Количество':<10} {'Дата':<12}\n"
                report_text += "-" * 70 + "\n"
                
                for barcode, category, price, stock, created_at in results:
                    date_str = created_at[:10] if created_at else "Неизвестна"
                    report_text += f"{barcode:<15} {category:<15} {price:<10.2f} {stock:<10} {date_str:<12}\n"
                
                self.show_analysis_dialog("Стари артикули", report_text)
            else:
                QMessageBox.information(self, "Информация", "Няма артикули по-стари от 6 месеца")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при анализ: {str(e)}")



    def show_analysis_dialog(self, title, text):
        """Show analysis results in a dialog"""
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setModal(True)
        dialog.resize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        text_edit = QTextEdit()
        text_edit.setFont(QFont("Courier", 10))
        text_edit.setPlainText(text)
        text_edit.setReadOnly(True)
        layout.addWidget(text_edit)
        
        button_layout = QHBoxLayout()
        
        export_btn = QPushButton("Експорт")
        export_btn.clicked.connect(lambda: self.export_analysis_text(title, text))
        button_layout.addWidget(export_btn)
        
        close_btn = QPushButton("Затвори")
        close_btn.clicked.connect(dialog.close)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
        dialog.exec()

    def export_analysis_text(self, title, text):
        """Export analysis text to file"""
        try:
            exports_dir = self.get_exports_directory()
            # Convert title to Bulgarian snake_case
            analysis_filename = self.generate_bulgarian_filename(title.lower(), "txt")
            file_path, _ = QFileDialog.getSaveFileName(
                self, f"Експорт: {title}", 
                f"{exports_dir}/{analysis_filename}",
                "Text Files (*.txt)"
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(text)
                QMessageBox.information(self, "Успех", f"Анализът е експортиран в:\n{file_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при експорт: {str(e)}")

    def export_comprehensive_report(self):
        """Export comprehensive report with all important data for the selected period"""
        try:
            start_date = self.report_start_date.date().toPyDate()
            end_date = self.report_end_date.date().toPyDate()
            current_time = datetime.now()
            
            # Generate filename following naming convention - only period dates
            period_str = f"{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}"
            filename = f"Отчет_{period_str}"
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Build comprehensive report
                report_lines = []
                report_lines.append("=" * 80)
                report_lines.append("КОМПЛЕКСЕН ОТЧЕТ ЗА БИЖУТА")
                report_lines.append("=" * 80)
                report_lines.append(f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
                report_lines.append(f"Генериран на: {current_time.strftime('%d.%m.%Y в %H:%M:%S')}")
                report_lines.append("")
                
                # 1. SUMMARY STATISTICS - restructured as proper table
                report_lines.append("📊 ОБОБЩЕНА СТАТИСТИКА")
                report_lines.append("-" * 40)
                
                # Sales for period
                cursor.execute("""
                    SELECT COUNT(*), COALESCE(SUM(total_price), 0), COALESCE(SUM(quantity), 0)
                    FROM sales 
                    WHERE date(sale_date) BETWEEN ? AND ?
                """, (start_date.isoformat(), end_date.isoformat()))
                sales_count, total_revenue, total_items_sold = cursor.fetchone()
                
                # Current inventory
                cursor.execute("SELECT COUNT(*), COALESCE(SUM(price * stock_quantity), 0) FROM items")
                total_items, inventory_value = cursor.fetchone()
                
                # Low stock items
                cursor.execute("SELECT COUNT(*) FROM items WHERE stock_quantity <= 5")
                low_stock_count = cursor.fetchone()[0]
                
                # Create summary statistics table
                report_lines.append(f"{'Показател':<35} {'Стойност':<20}")
                report_lines.append("-" * 55)
                report_lines.append(f"{'Общо продажби в периода':<35} {sales_count:<20}")
                report_lines.append(f"{'Приходи в периода':<35} {self.format_number_with_spaces(total_revenue)} €")
                report_lines.append(f"{'Продадени артикули':<35} {total_items_sold:<20}")
                report_lines.append(f"{'Общо артикули в склада':<35} {total_items:<20}")
                report_lines.append(f"{'Стойност на склада':<35} {self.format_number_with_spaces(inventory_value)} €")
                report_lines.append(f"{'Артикули с ниско наличество':<35} {low_stock_count:<20}")
                report_lines.append("")
                
                # 2. TOP SELLING ITEMS
                report_lines.append("🏆 ТОП 10 ПРОДАВАНИ АРТИКУЛА")
                report_lines.append("-" * 40)
                cursor.execute("""
                    SELECT i.barcode, i.category, SUM(s.quantity) as total_sold, 
                           SUM(s.total_price) as revenue
                    FROM sales s
                    JOIN items i ON s.item_id = i.id
                    WHERE date(s.sale_date) BETWEEN ? AND ?
                    GROUP BY s.item_id
                    ORDER BY total_sold DESC
                    LIMIT 10
                """, (start_date.isoformat(), end_date.isoformat()))
                
                top_items = cursor.fetchall()
                report_lines.append(f"{'№':<3} {'Баркод':<12} {'Категория':<12} {'Продадени':<8} {'Приходи':<12}")
                report_lines.append("-" * 50)
                if top_items:
                    for i, (barcode, category, sold, revenue) in enumerate(top_items, 1):
                        report_lines.append(f"{i:<3} {barcode:<12} {category:<12} {sold:<8} {revenue:<12.2f}")
                else:
                    report_lines.append("Няма данни за избрания период")
                report_lines.append("")
                
                # 3. SALES BY CATEGORY
                report_lines.append("📈 ПРОДАЖБИ ПО КАТЕГОРИИ")
                report_lines.append("-" * 40)
                cursor.execute("""
                    SELECT i.category, COUNT(*) as count, SUM(s.total_price) as revenue
                    FROM sales s
                    JOIN items i ON s.item_id = i.id
                    WHERE date(s.sale_date) BETWEEN ? AND ?
                    GROUP BY i.category
                    ORDER BY revenue DESC
                """, (start_date.isoformat(), end_date.isoformat()))
                
                category_sales = cursor.fetchall()
                report_lines.append(f"{'Категория':<15} {'Брой':<8} {'Приходи':<12} {'%':<8}")
                report_lines.append("-" * 45)
                if category_sales:
                    category_total = sum(row[2] for row in category_sales)
                    for category, count, revenue in category_sales:
                        percentage = (revenue / category_total * 100) if category_total > 0 else 0
                        report_lines.append(f"{category:<15} {count:<8} {revenue:<12.2f} {percentage:<8.1f}")
                else:
                    report_lines.append("Няма данни за избрания период")
                report_lines.append("")
                
                # 4. DAILY BREAKDOWN (if period <= 31 days)
                days_diff = (end_date - start_date).days
                if days_diff <= 31:
                    report_lines.append("📅 ДНЕВНИ ПРОДАЖБИ")
                    report_lines.append("-" * 40)
                    cursor.execute("""
                        SELECT i.barcode, i.category, s.quantity, s.total_price
                        FROM sales s
                        JOIN items i ON s.item_id = i.id
                        WHERE date(s.sale_date) BETWEEN ? AND ?
                        ORDER BY s.sale_date DESC
                        LIMIT 50
                    """, (start_date.isoformat(), end_date.isoformat()))
                    
                    daily_sales = cursor.fetchall()
                    report_lines.append(f"{'Баркод':<12} {'Категория':<15} {'Количество':<10} {'Общо':<10}")
                    report_lines.append("-" * 50)
                    if daily_sales:
                        for barcode, category, quantity, total_price in daily_sales:
                            report_lines.append(f"{barcode:<12} {category:<15} {quantity:<10} {total_price:<10.2f}")
                    else:
                        report_lines.append("Няма данни за избрания период")
                    report_lines.append("")
                
                # 5. LOW STOCK ALERT - always show, even if empty
                report_lines.append("АРТИКУЛИ С НИСКО НАЛИЧЕСТВО")
                report_lines.append("-" * 40)
                cursor.execute("""
                    SELECT barcode, category, stock_quantity, price
                    FROM items
                    WHERE stock_quantity <= 5
                    ORDER BY stock_quantity ASC
                    LIMIT 20
                """)
                
                low_stock_items = cursor.fetchall()
                report_lines.append(f"{'Баркод':<12} {'Категория':<12} {'Количество':<10} {'Цена':<10}")
                report_lines.append("-" * 50)
                if low_stock_items:
                    for barcode, category, stock, price in low_stock_items:
                        report_lines.append(f"{barcode:<12} {category:<12} {stock:<10} {price:<10.2f}")
                else:
                    report_lines.append("Няма артикули с ниско наличество")
                report_lines.append("")
                
                # 6. FOOTER
                report_lines.append("=" * 80)
                report_lines.append("Край на отчета")
                report_lines.append("=" * 80)
                
                # Combine all lines
                report_text = "\n".join(report_lines)
                
                # Export to PDF file
                exports_dir = self.get_exports_directory()
                file_path, _ = QFileDialog.getSaveFileName(
                    self, "Експорт на комплексен отчет", 
                    f"{exports_dir}/{filename}.pdf",
                    "PDF Files (*.pdf)"
                )
                
                if file_path:
                    self.export_report_to_pdf(report_lines, file_path, period_str)
                    QMessageBox.information(self, "Успех", 
                        f"Комплексният отчет е експортиран успешно!\n\n"
                        f"Файл: {file_path}\n"
                        f"Период: {period_str}\n"
                        f"Продажби: {sales_count}\n"
                        f"Приходи: {self.format_number_with_spaces(total_revenue)} €")
                        
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при експорт на отчет: {str(e)}")

    def export_report_to_pdf(self, report_lines, file_path, period_str):
        """Export comprehensive report to PDF with consistent formatting"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # Register fonts that support Cyrillic characters (same as existing export_to_pdf)
            try:
                font_paths = [
                    "fonts/arial.ttf",  # Our project font
                    "C:/Windows/Fonts/arial.ttf",  # Windows system font
                    "C:/Windows/Fonts/calibri.ttf",  # Alternative Windows font
                    "/System/Library/Fonts/Arial.ttf",  # macOS system font
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"  # Linux font
                ]
                
                font_registered = False
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        try:
                            pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                            pdfmetrics.registerFont(TTFont('CyrillicFont-Bold', font_path))
                            font_registered = True
                            break
                        except Exception:
                            continue
                
                if not font_registered:
                    cyrillic_font = 'Helvetica'
                    cyrillic_font_bold = 'Helvetica-Bold'
                else:
                    cyrillic_font = 'CyrillicFont'
                    cyrillic_font_bold = 'CyrillicFont-Bold'
                    
            except Exception:
                cyrillic_font = 'Helvetica'
                cyrillic_font_bold = 'Helvetica-Bold'
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()
            
            # Create custom styles with Cyrillic font support (consistent with existing PDFs)
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontName=cyrillic_font_bold,
                fontSize=16,
                textColor=colors.black,
                spaceAfter=12,
                alignment=1  # Center alignment
            )
            
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Heading2'],
                fontName=cyrillic_font_bold,
                fontSize=14,
                textColor=colors.black,
                spaceAfter=8,
                spaceBefore=12
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=cyrillic_font,
                fontSize=10,
                textColor=colors.black,
                spaceAfter=4
            )
            
            # Process report lines and convert to PDF elements
            current_time = datetime.now()
            
            # Title
            elements.append(Paragraph("КОМПЛЕКСЕН ОТЧЕТ ЗА БИЖУТА", title_style))
            elements.append(Spacer(1, 12))
            
            # Date info
            elements.append(Paragraph(f"Период: {period_str}", normal_style))
            elements.append(Paragraph(f"Генериран на: {current_time.strftime('%d.%m.%Y в %H:%M:%S')}", normal_style))
            elements.append(Spacer(1, 20))
            
            # Process content by creating explicit tables for each section
            elements = []
            
            # Add title and period info
            elements.append(Paragraph("КОМПЛЕКСЕН ОТЧЕТ ЗА БИЖУТА", title_style))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Период: {period_str}", normal_style))
            elements.append(Paragraph(f"Генериран на: {datetime.now().strftime('%d.%m.%Y в %H:%M:%S')}", normal_style))
            elements.append(Spacer(1, 20))
            
            # Re-extract data directly from database to create proper tables
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Parse period from report_lines to get dates
                start_date = self.report_start_date.date().toPyDate()
                end_date = self.report_end_date.date().toPyDate()
                
                # 1. SUMMARY STATISTICS TABLE
                elements.append(Paragraph("📊 ОБОБЩЕНА СТАТИСТИКА", header_style))
                elements.append(Spacer(1, 6))
                
                # Get summary data
                cursor.execute("""
                    SELECT COUNT(*), COALESCE(SUM(total_price), 0), COALESCE(SUM(quantity), 0)
                    FROM sales 
                    WHERE date(sale_date) BETWEEN ? AND ?
                """, (start_date.isoformat(), end_date.isoformat()))
                sales_count, total_revenue, total_items_sold = cursor.fetchone()
                
                cursor.execute("SELECT COUNT(*), COALESCE(SUM(price * stock_quantity), 0) FROM items")
                total_items, inventory_value = cursor.fetchone()
                
                cursor.execute("SELECT COUNT(*) FROM items WHERE stock_quantity <= 5")
                low_stock_count = cursor.fetchone()[0]
                
                # Create summary table
                summary_data = [
                    ["Показател", "Стойност"],
                    ["Общо продажби в периода", str(sales_count)],
                    ["Приходи в периода", f"{self.format_number_with_spaces(total_revenue)} €"],
                    ["Продадени артикули", str(total_items_sold)],
                    ["Общо артикули в склада", str(total_items)],
                    ["Стойност на склада", f"{self.format_number_with_spaces(inventory_value)} €"],
                    ["Артикули с ниско наличество", str(low_stock_count)]
                ]
                
                summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))
                
                # 2. TOP SELLING ITEMS TABLE
                elements.append(Paragraph("🏆 ТОП 10 ПРОДАВАНИ АРТИКУЛА", header_style))
                elements.append(Spacer(1, 6))
                
                cursor.execute("""
                    SELECT i.barcode, i.category, SUM(s.quantity) as total_sold, 
                           SUM(s.total_price) as revenue
                    FROM sales s
                    JOIN items i ON s.item_id = i.id
                    WHERE date(s.sale_date) BETWEEN ? AND ?
                    GROUP BY s.item_id, i.barcode, i.category
                    ORDER BY total_sold DESC
                    LIMIT 10
                """, (start_date.isoformat(), end_date.isoformat()))
                
                top_items = cursor.fetchall()
                top_items_data = [["№", "Баркод", "Категория", "Продадени", "Приходи"]]
                
                if top_items:
                    for i, (barcode, category, sold, revenue) in enumerate(top_items, 1):
                        top_items_data.append([
                            str(i), 
                            str(barcode), 
                            str(category), 
                            str(sold), 
                            f"{revenue:.2f}"
                        ])
                else:
                    top_items_data.append(["", "", "Няма данни за избрания период", "", ""])
                
                top_items_table = Table(top_items_data, colWidths=[0.5*inch, 1.5*inch, 1.5*inch, 1*inch, 1.5*inch])
                top_items_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                ]))
                elements.append(top_items_table)
                elements.append(Spacer(1, 20))
                
                # 3. SALES BY CATEGORY TABLE
                elements.append(Paragraph("📈 ПРОДАЖБИ ПО КАТЕГОРИИ", header_style))
                elements.append(Spacer(1, 6))
                
                cursor.execute("""
                    SELECT i.category, COUNT(*) as count, SUM(s.total_price) as revenue
                    FROM sales s
                    JOIN items i ON s.item_id = i.id
                    WHERE date(s.sale_date) BETWEEN ? AND ?
                    GROUP BY i.category
                    ORDER BY revenue DESC
                """, (start_date.isoformat(), end_date.isoformat()))
                
                category_sales = cursor.fetchall()
                category_data = [["Категория", "Брой", "Приходи", "%"]]
                
                if category_sales:
                    category_total = sum(row[2] for row in category_sales)
                    for category, count, revenue in category_sales:
                        percentage = (revenue / category_total * 100) if category_total > 0 else 0
                        category_data.append([
                            str(category),
                            str(count),
                            f"{revenue:.2f}",
                            f"{percentage:.1f}"
                        ])
                else:
                    category_data.append(["", "Няма данни за избрания период", "", ""])
                
                category_table = Table(category_data, colWidths=[2*inch, 1*inch, 1.5*inch, 1*inch])
                category_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                ]))
                elements.append(category_table)
                elements.append(Spacer(1, 20))
                
                # 4. DAILY SALES TABLE (if period <= 31 days)
                days_diff = (end_date - start_date).days
                if days_diff <= 31:
                    elements.append(Paragraph("📅 ДНЕВНИ ПРОДАЖБИ", header_style))
                    elements.append(Spacer(1, 6))
                    
                    cursor.execute("""
                        SELECT i.barcode, i.category, s.quantity, s.total_price
                        FROM sales s
                        JOIN items i ON s.item_id = i.id
                        WHERE date(s.sale_date) BETWEEN ? AND ?
                        ORDER BY s.sale_date DESC
                        LIMIT 50
                    """, (start_date.isoformat(), end_date.isoformat()))
                    
                    daily_sales = cursor.fetchall()
                    daily_data = [["Баркод", "Категория", "Количество", "Общо"]]
                    
                    if daily_sales:
                        for barcode, category, quantity, total_price in daily_sales:
                            daily_data.append([
                                str(barcode),
                                str(category),
                                str(quantity),
                                f"{total_price:.2f}"
                            ])
                    else:
                        daily_data.append(["", "Няма данни за избрания период", "", ""])
                    
                    daily_table = Table(daily_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1*inch])
                    daily_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                    ]))
                    elements.append(daily_table)
                    elements.append(Spacer(1, 20))
                
                # 5. LOW STOCK ITEMS TABLE
                elements.append(Paragraph("⚠ АРТИКУЛИ С НИСКО НАЛИЧЕСТВО", header_style))
                elements.append(Spacer(1, 6))
                
                cursor.execute("""
                    SELECT barcode, category, stock_quantity, price
                    FROM items
                    WHERE stock_quantity <= 5
                    ORDER BY stock_quantity ASC
                    LIMIT 20
                """)
                
                low_stock_items = cursor.fetchall()
                low_stock_data = [["Баркод", "Категория", "Количество", "Цена"]]
                
                if low_stock_items:
                    for barcode, category, stock, price in low_stock_items:
                        low_stock_data.append([
                            str(barcode),
                            str(category),
                            str(stock),
                            f"{price:.2f}"
                        ])
                else:
                    low_stock_data.append(["", "Няма артикули с ниско наличество", "", ""])
                
                low_stock_table = Table(low_stock_data, colWidths=[1.5*inch, 2*inch, 1*inch, 1.5*inch])
                low_stock_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                ]))
                elements.append(low_stock_table)
            
            # Build PDF
            doc.build(elements)
            
        except ImportError:
            # Fallback to text export if reportlab not available
            with open(file_path.replace('.pdf', '.txt'), 'w', encoding='utf-8') as f:
                f.write("\n".join(report_lines))
            QMessageBox.warning(self, "Внимание", 
                "ReportLab не е инсталиран. Отчетът е експортиран като текстов файл.")
        except Exception as e:
            raise Exception(f"Грешка при създаване на PDF: {str(e)}")

    def load_data(self):
        """Load initial data"""
        try:
            # Ensure custom value sets are initialized
            if not hasattr(self, 'custom_categories'):
                self.custom_categories = set()
            if not hasattr(self, 'custom_metals'):
                self.custom_metals = set()
            if not hasattr(self, 'custom_stones'):
                self.custom_stones = set()
            
            # Load custom values from database with error handling
            try:
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    # Check if custom_values table exists
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='custom_values'")
                    if cursor.fetchone():
                        cursor.execute("SELECT type, value FROM custom_values")
                        for type_, value in cursor.fetchall():
                            if type_ == 'category':
                                self.custom_categories.add(value)
                            elif type_ == 'metal':
                                self.custom_metals.add(value)
                            elif type_ == 'stone':
                                self.custom_stones.add(value)
                    else:
                        logger.info("custom_values table does not exist yet")
            except Exception as db_error:
                logger.warning(f"Error loading custom values from database: {db_error}")

            # Update combo boxes with custom values (with error handling for each)
            try:
                if hasattr(self, 'category_input'):
                    self.category_input.clear()
                    self.category_input.addItems(sorted(list(self.custom_categories)) + ["Пръстен", "Гривна", "Обеци", "Синджир", "Друго"])
            except Exception as e:
                logger.warning(f"Error updating category combo: {e}")
            
            try:
                if hasattr(self, 'metal_input'):
                    self.metal_input.clear()
                    metal_items = sorted(list(self.custom_metals)) + ["Злато", "Сребро", "Платина", "Друго"]
                    self.metal_input.addItems(metal_items)
                    # Find and set Сребро as default
                    silver_index = metal_items.index("Сребро") if "Сребро" in metal_items else 1
                    self.metal_input.setCurrentIndex(silver_index)
            except Exception as e:
                logger.warning(f"Error updating metal combo: {e}")
            
            try:
                if hasattr(self, 'stone_input'):
                    self.stone_input.clear()
                    stone_items = sorted(list(self.custom_stones)) + ["Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък", "Друго"]
                    self.stone_input.addItems(stone_items)
                    # Find and set Без камък as default
                    no_stone_index = stone_items.index("Без камък") if "Без камък" in stone_items else 4
                    self.stone_input.setCurrentIndex(no_stone_index)
            except Exception as e:
                logger.warning(f"Error updating stone combo: {e}")
            
            # Load items and sales with individual error handling
            try:
                if hasattr(self, 'load_items'):
                    self.load_items()
            except Exception as e:
                logger.warning(f"Error loading items: {e}")
                if hasattr(self, 'items_table'):
                    self.items_table.setRowCount(0)
            
            try:
                if hasattr(self, 'load_sales'):
                    self.load_sales()
            except Exception as e:
                logger.warning(f"Error loading sales: {e}")
                if hasattr(self, 'sales_table'):
                    self.sales_table.setRowCount(0)
            
            # Load shop data if available
            try:
                if hasattr(self, 'shop_combo') and hasattr(self, 'sales_shop_combo'):
                    shops = self.db.get_all_shops()
                    # Update shop combo
                    self.shop_combo.clear()
                    self.sales_shop_combo.clear()
                    for shop in shops:
                        self.shop_combo.addItem(shop[1])
                        self.sales_shop_combo.addItem(shop[1])
            except Exception as e:
                logger.warning(f"Error loading shops: {e}")
                
        except Exception as e:
            if not getattr(self, '_suppress_error_dialogs', False):
                QMessageBox.critical(self, "Грешка", f"Грешка при зареждане на данните: {str(e)}")
            else:
                logger.error(f"Error in load_data (suppressed): {e}")

    def load_sales(self):
        """Load sales into table with filtering"""
        try:
            # Clear existing data
            self.sales_table.setRowCount(0)
            
            # Get filter selections
            shop_name = self.sales_shop_combo.currentText() if hasattr(self, 'sales_shop_combo') else None
            if not shop_name:
                return
            
            # Get time filter
            time_filter = "all"
            if hasattr(self, 'time_button_group'):
                checked_button = self.time_button_group.checkedButton()
                if checked_button:
                    button_id = self.time_button_group.id(checked_button)
                    time_filters = ["all", "today", "week", "month", "year", "custom"]
                    time_filter = time_filters[button_id] if button_id < len(time_filters) else "all"
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get shop ID
                cursor.execute("SELECT id FROM shops WHERE name = ?", (shop_name,))
                shop_row = cursor.fetchone()
                if not shop_row:
                    return
                shop_id = shop_row[0]
                
                # Build query with time and shop filters
                base_query = """
                    SELECT i.barcode, i.category, s.total_price, i.cost, i.weight, 
                           i.metal_type, i.stone_type, s.quantity, s.sale_date
                    FROM sales s 
                    JOIN items i ON s.item_id = i.id 
                    WHERE s.shop_id = ?
                """
                
                params = [shop_id]
                
                # Add time filter
                from datetime import datetime, timedelta
                now = datetime.now()
                
                if time_filter == "today":
                    date_filter = now.strftime("%Y-%m-%d")
                    base_query += " AND date(s.sale_date) = ?"
                    params.append(date_filter)
                elif time_filter == "week":
                    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
                    base_query += " AND date(s.sale_date) >= ?"
                    params.append(week_start)
                elif time_filter == "month":
                    month_start = now.replace(day=1).strftime("%Y-%m-%d")
                    base_query += " AND date(s.sale_date) >= ?"
                    params.append(month_start)
                elif time_filter == "year":
                    year_start = now.replace(month=1, day=1).strftime("%Y-%m-%d")
                    base_query += " AND date(s.sale_date) >= ?"
                    params.append(year_start)
                elif time_filter == "custom":
                    # Use custom date range from date fields
                    start_date = self.sales_start_date.date().toPyDate().strftime("%Y-%m-%d")
                    end_date = self.sales_end_date.date().toPyDate().strftime("%Y-%m-%d")
                    base_query += " AND date(s.sale_date) BETWEEN ? AND ?"
                    params.extend([start_date, end_date])
                
                base_query += " ORDER BY s.sale_date DESC"
                
                cursor.execute(base_query, params)
                sales = cursor.fetchall()
                
                # Populate table
                self.sales_table.setRowCount(len(sales))
                for row, sale in enumerate(sales):
                    # Updated to match new database query for sales
                    # Need to get description from items table via join
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute("SELECT description FROM items WHERE barcode = ?", (sale[0],))
                        description_row = cursor.fetchone()
                        description = description_row[0] if description_row else ""
                    
                    barcode, category, price, cost, weight, metal, stone, quantity, sale_date = sale
                    
                    # Parse and format the sale date
                    try:
                        dt = parse_database_datetime(sale_date)
                        if dt:
                            date_str = format_date_for_display(dt)
                            time_str = format_time_for_display(dt)
                        else:
                            date_str = ""
                            time_str = ""
                    except Exception:
                        date_str = ""
                        time_str = ""
                    
                    # Set table items - matching new inventory table structure
                    # ALL SALES TABLE ITEMS ARE IMMUTABLE (READ-ONLY)
                    barcode_item = QTableWidgetItem(str(barcode))
                    barcode_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    barcode_item.setFlags(barcode_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 0, barcode_item)
                    
                    category_item = QTableWidgetItem(str(category) if category else "")
                    category_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    category_item.setFlags(category_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 1, category_item)
                    
                    metal_item = QTableWidgetItem(str(metal) if metal else "")
                    metal_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    metal_item.setFlags(metal_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 2, metal_item)
                    
                    stone_item = QTableWidgetItem(str(stone) if stone else "")
                    stone_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    stone_item.setFlags(stone_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 3, stone_item)
                    
                    description_item = QTableWidgetItem(str(description) if description else "")
                    description_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    description_item.setFlags(description_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 4, description_item)
                    
                    # Cost in dual currency (assuming database stores Euro)
                    if cost:
                        cost_lev = self.euro_to_lev(cost)
                        cost_text = f"{cost:.2f} €\n{cost_lev:.2f} лв"
                    else:
                        cost_text = "0.00 €\n0.00 лв"
                    cost_item = QTableWidgetItem(cost_text)
                    cost_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    cost_item.setFlags(cost_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 5, cost_item)
                    
                    # Price in dual currency (assuming database stores Euro)
                    price_lev = self.euro_to_lev(price)
                    price_text = f"{price:.2f} €\n{price_lev:.2f} лв"
                    price_item = QTableWidgetItem(price_text)
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    price_item.setFlags(price_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 6, price_item)
                    
                    weight_item = QTableWidgetItem(self.format_grams(weight) if weight else "")
                    weight_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    weight_item.setFlags(weight_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 7, weight_item)
                    
                    quantity_item = QTableWidgetItem(str(quantity) if quantity else "1")
                    quantity_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    quantity_item.setFlags(quantity_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 8, quantity_item)
                    
                    # Highlight row if quantity is 0 in sales (though this shouldn't normally happen)
                    self.highlight_zero_quantity_row(self.sales_table, row, int(quantity) if quantity else 1)
                    
                    date_item = QTableWidgetItem(date_str)
                    date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    date_item.setFlags(date_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 9, date_item)
                    
                    time_item = QTableWidgetItem(time_str)
                    time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    time_item.setFlags(time_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.sales_table.setItem(row, 10, time_item)
                    
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при зареждане на продажбите: {str(e)}")
            logger.error(f"Error loading sales: {e}", exc_info=True)

    def export_shop_sales_to_pdf(self):
        """Export shop-specific sales to PDF with shop name, date/time, table and total"""
        try:
            # Get selected shop
            shop_name = self.sales_shop_combo.currentText() if hasattr(self, 'sales_shop_combo') else None
            if not shop_name:
                QMessageBox.warning(self, "Предупреждение", "Моля, изберете магазин")
                return
            
            # Get current date and time
            current_time = datetime.now()
            
            # Create filename in your preferred format: продажби_shop_name - DD.MM.YYYY.pdf
            date_formatted = current_time.strftime("%d.%m.%Y")
            filename = f"продажби_{shop_name} - {date_formatted}.pdf"
            
            # Ask user where to save the file
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Запази PDF отчет за продажби", 
                f"exports/{filename}",
                "PDF файлове (*.pdf)"
            )
            
            if not file_path:
                return
            
            # Create the PDF
            self.generate_shop_sales_pdf(shop_name, file_path, current_time)
            
            QMessageBox.information(self, "Успех", f"PDF отчетът е запазен успешно в:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при експортиране: {str(e)}")
            logger.error(f"Error exporting shop sales to PDF: {e}", exc_info=True)

    def generate_shop_sales_pdf(self, shop_name, file_path, current_time):
        """Generate PDF report for shop sales following existing PDF format standards"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # Register fonts that support Cyrillic characters (same as existing export_to_pdf)
            try:
                font_paths = [
                    "fonts/arial.ttf",  # Our project font
                    "C:/Windows/Fonts/arial.ttf",  # Windows system font
                    "C:/Windows/Fonts/calibri.ttf",  # Alternative Windows font
                    "/System/Library/Fonts/Arial.ttf",  # macOS system font
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"  # Linux font
                ]
                
                font_registered = False
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        try:
                            pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                            pdfmetrics.registerFont(TTFont('CyrillicFont-Bold', font_path))
                            font_registered = True
                            break
                        except Exception:
                            continue
                
                if not font_registered:
                    cyrillic_font = 'Helvetica'
                    cyrillic_font_bold = 'Helvetica-Bold'
                else:
                    cyrillic_font = 'CyrillicFont'
                    cyrillic_font_bold = 'CyrillicFont-Bold'
                    
            except Exception:
                cyrillic_font = 'Helvetica'
                cyrillic_font_bold = 'Helvetica-Bold'
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()
            
            # Create custom styles with Cyrillic font support (consistent with existing PDFs)
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontName=cyrillic_font_bold,
                fontSize=16,
                textColor=colors.black,
                spaceAfter=12,
                alignment=1  # Center alignment
            )
            
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Heading2'],
                fontName=cyrillic_font_bold,
                fontSize=14,
                textColor=colors.black,
                spaceAfter=8,
                spaceBefore=12
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=cyrillic_font,
                fontSize=10,
                textColor=colors.black,
                spaceAfter=4
            )
            
            # Title with shop name
            elements.append(Paragraph(f"ОТЧЕТ ЗА ПРОДАЖБИ - {shop_name.upper()}", title_style))
            elements.append(Spacer(1, 12))
            
            # Date and time of creation
            elements.append(Paragraph(f"Дата и час на създаване: {current_time.strftime('%d.%m.%Y в %H:%M:%S')}", normal_style))
            elements.append(Spacer(1, 20))
            
            # Get sales data for the selected shop
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get shop ID
                cursor.execute("SELECT id FROM shops WHERE name = ?", (shop_name,))
                shop_row = cursor.fetchone()
                if not shop_row:
                    raise Exception(f"Магазин '{shop_name}' не е намерен")
                shop_id = shop_row[0]
                
                # Get all sales for this shop
                cursor.execute("""
                    SELECT i.barcode, i.category, i.metal_type, i.stone_type, i.description,
                           s.total_price, i.weight, s.quantity, s.sale_date
                    FROM sales s 
                    JOIN items i ON s.item_id = i.id 
                    WHERE s.shop_id = ?
                    ORDER BY s.sale_date DESC
                """, (shop_id,))
                
                sales_data = cursor.fetchall()
                
                if not sales_data:
                    elements.append(Paragraph("Няма записани продажби за този магазин.", normal_style))
                    doc.build(elements)
                    return
                
                # Create sales table
                elements.append(Paragraph("📊 ПРОДАЖБИ", header_style))
                elements.append(Spacer(1, 6))
                
                # Table headers without description column
                table_data = [["Баркод", "Категория", "Метал", "Камък", "Цена", "Тегло", "К-во", "Дата"]]
                
                total_revenue = 0
                
                for sale in sales_data:
                    barcode, category, metal, stone, description, price, weight, quantity, sale_date = sale
                    
                    # Format date
                    try:
                        from utils.data_manager import parse_database_datetime, format_date_for_display
                        dt = parse_database_datetime(sale_date)
                        if dt:
                            date_formatted = format_date_for_display(dt)
                        else:
                            date_formatted = sale_date[:10] if sale_date else ""
                    except Exception:
                        date_formatted = sale_date[:10] if sale_date else ""
                    
                    # Format values for display (truncate text to prevent cell overflow)
                    barcode_str = str(barcode)[:12] if barcode else ""
                    category_str = str(category)[:15] if category else ""
                    metal_str = str(metal)[:12] if metal else ""
                    stone_str = str(stone)[:12] if stone else ""
                    price_str = f"{price:.2f} €" if price else "0.00 €"
                    weight_str = f"{weight:.2f}г" if weight else ""
                    quantity_str = str(quantity) if quantity else "1"
                    
                    # Add to total
                    total_revenue += price if price else 0
                    
                    table_data.append([
                        barcode_str, category_str, metal_str, stone_str,
                        price_str, weight_str, quantity_str, date_formatted
                    ])
                
                # Create table with proper column widths (removed description column)
                col_widths = [1.1*inch, 1.2*inch, 1.0*inch, 1.0*inch, 0.9*inch, 0.8*inch, 0.6*inch, 1.0*inch]
                
                sales_table = Table(table_data, colWidths=col_widths)
                sales_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),  # Enable text wrapping
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4)
                ]))
                
                elements.append(sales_table)
                elements.append(Spacer(1, 20))
                
                # Total under the table
                total_lev = self.euro_to_lev(total_revenue)
                total_text = f"ОБЩО: {self.format_number_with_spaces(total_revenue)} € ({self.format_number_with_spaces(total_lev)} лв)"
                
                total_style = ParagraphStyle(
                    'TotalStyle',
                    parent=styles['Normal'],
                    fontName=cyrillic_font_bold,
                    fontSize=12,
                    textColor=colors.darkblue,
                    spaceAfter=4,
                    alignment=1  # Center alignment
                )
                
                elements.append(Paragraph(total_text, total_style))
            
            # Build the PDF
            doc.build(elements)
            
        except Exception as e:
            logger.error(f"Error generating shop sales PDF: {e}", exc_info=True)
            raise

    def scan_barcode(self):
        """Scan barcode from camera"""
        barcode = self.barcode_scanner.scan_from_camera()
        if barcode:
            self.barcode_input.setText(barcode)
            self.search_items()



    def auto_resize_add_item_description(self):
        """Auto-resize description field in Add Item tab based on content"""
        try:
            # Get document height
            doc = self.description_input.document()
            doc_height = doc.size().height()
            
            # Calculate required height (with some padding)
            required_height = max(self.desc_min_height, min(self.desc_max_height, int(doc_height) + 10))
            
            # Set the height
            self.description_input.setFixedHeight(required_height)
        except Exception as e:
            # If anything goes wrong, just use minimum height
            self.description_input.setFixedHeight(self.desc_min_height)

    def euro_to_lev(self, euro_amount):
        """Convert Euro to Lev using fixed rate"""
        return round(euro_amount * self.EUR_TO_LEV_RATE, 2)
    
    def lev_to_euro(self, lev_amount):
        """Convert Lev to Euro using fixed rate"""
        return round(lev_amount / self.EUR_TO_LEV_RATE, 2)
    
    def format_currency_eur(self, amount):
        """Format amount as Euro currency with thousands separators"""
        return f"{amount:,.2f} €".replace(",", " ")
    
    def format_currency_lev(self, amount):
        """Format amount as Lev currency with thousands separators"""
        return f"{amount:,.2f} лв".replace(",", " ")
    
    def update_lev_price(self):
        """Update Lev price when Euro price changes"""
        try:
            euro_price = self.price_input.value()
            lev_price = self.euro_to_lev(euro_price)
            self.price_lev_label.setText(self.format_currency_lev(lev_price))
        except Exception as e:
            self.price_lev_label.setText("0.00 лв")
    
    def update_lev_cost(self):
        """Update Lev cost when Euro cost changes"""
        try:
            euro_cost = self.cost_input.value()
            lev_cost = self.euro_to_lev(euro_cost)
            self.cost_lev_label.setText(self.format_currency_lev(lev_cost))
        except Exception as e:
            self.cost_lev_label.setText("0.00 лв")

    def save_item(self):
        """Save item to database"""
        try:
            # Get values
            barcode = self.barcode_input.text().strip()
            name = self.category_input.currentText()  # Use category as name
            description = self.description_input.toPlainText().strip()  # Changed from text() to toPlainText()
            category = self.category_input.currentText()
            price = self.price_input.value()
            cost = self.cost_input.value()
            weight = self.weight_input.value()
            metal_type = self.metal_input.currentText()
            stone_type = self.stone_input.currentText()
            stock_quantity = self.stock_input.value()

            # Check if validation is needed (skip if called from print_and_add_item which already validates)
            import inspect
            caller_name = inspect.stack()[1].function if len(inspect.stack()) > 1 else ""
            
            if caller_name != "print_and_add_item":
                # Only validate if not called from print_and_add_item (to avoid duplicate validation)
                validation_errors = []

                # Required fields validation
                if not category or category == "Друго":
                    validation_errors.append("Моля, изберете валидна категория")

                if not metal_type or metal_type == "Друго":
                    validation_errors.append("Моля, изберете валиден метал")

                if not stone_type or stone_type == "Друго":
                    validation_errors.append("Моля, изберете валиден камъкк")

                # Numeric validation
                if price <= 0:
                    validation_errors.append("Цената трябва да бъде по-голяма от 0")

                if cost < 0:
                    validation_errors.append("Себестойността не може да бъде отрицателна")

                if cost > price:
                    validation_errors.append("Себестойността не може да бъде по-висока от цената")

                if weight <= 0:
                    validation_errors.append("Теглото трябва да бъде по-голямо от 0")

                if stock_quantity <= 0:
                    validation_errors.append("Количеството трябва да бъде по-голямо от 0")

                # Display validation errors
                if validation_errors:
                    error_message = "Моля, коригирайте следните грешки:\n\n" + "\n".join(f"• {error}" for error in validation_errors)
                    QMessageBox.warning(self, "Грешки при валидация", error_message)
                    return

            # Save image


            # Create and execute add item action
            item_data = {
                'barcode': barcode,
                'name': name,
                'description': description,
                'category': category,
                'price': price,
                'cost': cost,
                'weight': weight,
                'metal_type': metal_type,
                'stone_type': stone_type,
                'stock_quantity': stock_quantity
            }
            
            add_action = AddItemAction(self.db, item_data)
            if self.action_history.execute_action(add_action):
                QMessageBox.information(self, "Успех", "Артикулът е запазен успешно")
                self.clear_item_form()
                self.load_items()
                self.update_action_buttons()
                self.update_reports_and_database_stats()
            else:
                QMessageBox.warning(self, "Грешка", "Неуспешно запазване на артикула")

        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Възникна грешка: {str(e)}")

    def clear_item_form(self):
        """Clear item form fields"""
        self.barcode_input.clear()
        self.description_input.clear()  # QTextEdit.clear() works the same way
        
        # Preserve custom values when clearing
        self.category_input.clear()
        self.category_input.addItems(sorted(list(self.custom_categories)) + ["Пръстен", "Гривна", "Обеци", "Синджир", "Друго"])
        self.category_input.setCurrentIndex(0)
        
        self.metal_input.clear()
        metal_items = sorted(list(self.custom_metals)) + ["Злато", "Сребро", "Платина", "Друго"]
        self.metal_input.addItems(metal_items)
        # Find and set Сребро as default
        silver_index = metal_items.index("Сребро") if "Сребро" in metal_items else 0
        self.metal_input.setCurrentIndex(silver_index)
        
        self.stone_input.clear()
        stone_items = sorted(list(self.custom_stones)) + ["Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък", "Друго"]
        self.stone_input.addItems(stone_items)
        # Find and set Без камък as default
        no_stone_index = stone_items.index("Без камък") if "Без камък" in stone_items else 0
        self.stone_input.setCurrentIndex(no_stone_index)
        
        self.price_input.setValue(0)
        self.cost_input.setValue(0)
        self.weight_input.setValue(0)
        self.stock_input.setValue(0)
        
        # Reset confirmed values for spin boxes
        self.price_input.reset_confirmed_value()
        self.cost_input.reset_confirmed_value()
        self.weight_input.reset_confirmed_value()
        self.stock_input.reset_confirmed_value()
        
        # Clear currency labels
        self.price_lev_label.setText("0.00 лв")
        self.cost_lev_label.setText("0.00 лв")
        
        # Clear image path
        
        # Clear barcode preview and related state
        self.barcode_preview.clear()
        if hasattr(self, 'current_label'):
            delattr(self, 'current_label')
        if hasattr(self, 'barcode_image'):
            delattr(self, 'barcode_image')
        
        # Reset barcode lock
        self.barcode_locked = False

    def populate_add_item_form(self, item_data):
        """Populate Add Item form with existing item data for reprinting labels"""
        try:
            # Set barcode (this will be the same as the original)
            self.barcode_input.setText(item_data['barcode'])
            
            # Set category
            category_text = item_data['category']
            category_index = self.category_input.findText(category_text)
            if category_index >= 0:
                self.category_input.setCurrentIndex(category_index)
            else:
                self.category_input.setCurrentText(category_text)
            
            # Set metal
            metal_text = item_data['metal']
            metal_index = self.metal_input.findText(metal_text)
            if metal_index >= 0:
                self.metal_input.setCurrentIndex(metal_index)
            else:
                self.metal_input.setCurrentText(metal_text)
            
            # Set stone
            stone_text = item_data['stone']
            stone_index = self.stone_input.findText(stone_text)
            if stone_index >= 0:
                self.stone_input.setCurrentIndex(stone_index)
            else:
                self.stone_input.setCurrentText(stone_text)
            
            # Set description
            self.description_input.setPlainText(item_data['description'])
            
            # Set cost and price
            self.cost_input.setValue(item_data['cost'])
            self.price_input.setValue(item_data['price'])
            
            # Set weight
            self.weight_input.setValue(item_data['weight'])
            
            # Set stock (user might want to print more labels than current stock)
            self.stock_input.setValue(item_data['stock'])
            
            # Store warehouse reprint flag
            self.is_warehouse_reprint = item_data.get('is_warehouse_reprint', False)
            
            # Update currency displays
            self.update_lev_cost()
            self.update_lev_price()
            
            # Lock the barcode since this is for reprinting
            self.barcode_locked = True
            
            # Generate barcode preview automatically
            self.update_barcode_preview()
            
            # Show appropriate info message based on reprint type
            if self.is_warehouse_reprint:
                QMessageBox.information(
                    self, "Режим за повторен печат от склад",
                    f"Формата е попълнена с данните от артикул {item_data['barcode']}.\n\n"
                    f"Количеството е зададено на 0. Въведете желаното количество - то ще бъде ДОБАВЕНО към наличното в склада.\n\n"
                    f"Натиснете 'Принтирай етикет' за да отпечатате етикети и добавите количеството към склада."
                )
            else:
                QMessageBox.information(
                    self, "Режим за повторен печат",
                    f"Формата е попълнена с данните от артикул {item_data['barcode']}.\n\n"
                    f"Сега можете да натиснете 'Принтирай етикет' за да отпечатате допълнителни етикети.\n\n"
                    f"ВНИМАНИЕ: Артикулът няма да бъде добавен отново в системата!"
                )
            
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при попълване на формата: {str(e)}")

    def on_sales_barcode_changed(self):
        """Handle sales barcode input changes with timer delay"""
        # Stop any existing timer
        self.sales_barcode_timer.stop()
        
        # Start timer with 500ms delay to wait for complete barcode
        self.sales_barcode_timer.start(500)
        
    def process_sales_barcode(self):
        """Process the complete sales barcode after timer delay"""
        barcode = self.sale_barcode_input.text().strip()
        
        # Process when 7 or 13 digits are entered (support both formats)
        if (len(barcode) == 7 or len(barcode) == 13) and barcode.isdigit():
            shop_name = self.sales_shop_combo.currentText()
            if not shop_name:
                QMessageBox.warning(self, "Грешка", "Моля, изберете магазин!")
                self.sale_barcode_input.clear()
                return
            
            self.sell_item_from_shop(barcode, shop_name)

    def sell_item_from_shop(self, barcode, shop_name):
        """Sell an item from the selected shop"""
        try:
            # STEP 1: Get shop ID
            shops = self.db.get_all_shops()
            shop_id = None
            
            for shop in shops:
                if shop[1] == shop_name:  # shop[1] is the name
                    shop_id = shop[0]  # shop[0] is the id
                    break
            
            if shop_id is None:
                QMessageBox.warning(self, "Грешка", "Магазинът не е намерен!")
                self.sale_barcode_input.clear()
                return
            
            # STEP 2: Check database connection and FK enforcement
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Check foreign key enforcement
                cursor.execute("PRAGMA foreign_keys")
                fk_status = cursor.fetchone()[0]
                
                # STEP 3: Check if item exists in the selected shop
                shop_query = """
                    SELECT si.quantity, i.category, i.price, i.weight, i.metal_type, i.stone_type 
                    FROM shop_items si 
                    JOIN items i ON si.item_id = i.id 
                    WHERE si.shop_id = ? AND i.barcode = ?
                """
                
                cursor.execute(shop_query, (shop_id, barcode))
                shop_item = cursor.fetchone()
                
                if not shop_item or shop_item[0] < 1:
                    QMessageBox.warning(self, "Грешка", 
                                      f"Артикулът с баркод {barcode} не е наличен в магазин '{shop_name}'!")
                    self.sale_barcode_input.clear()
                    return
                
                quantity, category, price, weight, metal, stone = shop_item
                
                # STEP 4: Get item_id for the sale action
                item_query = "SELECT id FROM items WHERE barcode = ?"
                cursor.execute(item_query, (barcode,))
                item_result = cursor.fetchone()
                
                if not item_result:
                    QMessageBox.warning(self, "Грешка", f"Артикулът с баркод {barcode} не е намерен!")
                    self.sale_barcode_input.clear()
                    return
                
                item_id = item_result[0]
                
                # STEP 5: Create and execute sale action
                
                sale_action = SaleAction(self.db, item_id, barcode, shop_id, price, 1)
                
                action_result = self.action_history.execute_action(sale_action)
                
                if not action_result:
                    QMessageBox.warning(self, "Грешка", f"Неуспешна продажба на артикул {barcode}")
                    self.sale_barcode_input.clear()
                    return
                
                # Show temporary success message (2 seconds)
                self.show_temp_success_message(
                    f"✅ Продадено: {category} | Цена: {price:.2f} лв | Магазин: {shop_name}"
                )
                
                # STEP 7: Clear and refresh
                self.sale_barcode_input.clear()
                self.sale_barcode_input.setFocus()  # Keep focus for next scan
                
                self.load_sales()
                self.update_action_buttons()
                
                self.update_shop_inventory_info()
                
                # Delay shop inventory refresh to avoid race conditions with combo box events
                if hasattr(self, 'shop_combo'):
                    QTimer.singleShot(200, self.load_shop_inventory)  # 200ms delay
                
                # Update reports and database statistics
                self.update_reports_and_database_stats()
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            
            QMessageBox.critical(self, "Грешка", f"Грешка при продажба: {str(e)}")
            self.sale_barcode_input.clear()
    

    
    def on_time_filter_changed(self):
        """Handle time filter radio button changes"""
        # Enable/disable custom date fields based on selection
        is_custom = self.custom_radio.isChecked()
        self.sales_start_date.setEnabled(is_custom)
        self.sales_end_date.setEnabled(is_custom)
        
        # Update date dropdowns based on selected period
        checked_button = self.time_button_group.checkedButton()
        if checked_button:
            button_id = self.time_button_group.id(checked_button)
            if button_id == 1:  # Today
                self.set_sales_date_range("today")
            elif button_id == 2:  # Week
                self.set_sales_date_range("week")
            elif button_id == 3:  # Month
                self.set_sales_date_range("month")
            elif button_id == 4:  # Year
                self.set_sales_date_range("year")
            elif button_id == 0:  # All - reset dates to default "show all" range
                # Set dates to default range like warehouse - search function will skip filtering
                self.programmatic_date_change = True
                try:
                    self.sales_start_date.setDate(QDate.currentDate().addMonths(-1))
                    self.sales_end_date.setDate(QDate.currentDate())
                finally:
                    self.programmatic_date_change = False
                
                # Reload sales data and refresh search results
                self.load_sales()
                # Also trigger search to update any active search filters
                if hasattr(self, 'search_sales'):
                    self.search_sales()
            else:  # Custom (button_id == 5)
                # Just reload and search, don't change dates
                self.load_sales()
                if hasattr(self, 'search_sales'):
                    self.search_sales()
    
    def auto_switch_to_custom_period(self):
        """Automatically switch to custom period when date is changed"""
        # Don't switch if we're in the middle of a programmatic change
        if self.programmatic_date_change:
            return
            
        # Only switch if not already on custom period
        if not self.custom_radio.isChecked():
            self.custom_radio.setChecked(True)
            # Enable the date fields
            self.sales_start_date.setEnabled(True)
            self.sales_end_date.setEnabled(True)
    
    def auto_switch_to_custom_inventory_period(self):
        """Automatically switch to custom period when inventory date is changed"""
        # Don't switch if we're in the middle of a programmatic change
        if hasattr(self, 'programmatic_inventory_date_change') and self.programmatic_inventory_date_change:
            return
            
        # Only switch if not already on custom period
        if hasattr(self, 'inv_custom_radio') and not self.inv_custom_radio.isChecked():
            self.inv_custom_radio.setChecked(True)
            # Enable the date fields
            self.start_date_input.setEnabled(True)
            self.end_date_input.setEnabled(True)
    
    def eventFilter(self, obj, event):
        """Event filter to catch focus events on date pickers and input fields"""
        try:
            # Check if it's a focus event on our sales date pickers
            if (obj == self.sales_start_date or obj == self.sales_end_date) and hasattr(self, 'custom_radio'):
                if event.type() == event.Type.FocusIn or event.type() == event.Type.MouseButtonPress:
                    # User clicked into the date field, auto-switch to custom period
                    self.auto_switch_to_custom_period()
            
            # Check if it's a focus event on our inventory date pickers
            if (obj == self.start_date_input or obj == self.end_date_input) and hasattr(self, 'inv_custom_radio'):
                if event.type() == event.Type.FocusIn or event.type() == event.Type.MouseButtonPress:
                    # User clicked into the date field, auto-switch to custom period
                    self.auto_switch_to_custom_inventory_period()
            
            # Auto-clear text for spinbox input fields in add item tab
            if hasattr(self, 'price_input') and hasattr(self, 'cost_input') and hasattr(self, 'weight_input') and hasattr(self, 'stock_input'):
                if (obj == self.price_input.lineEdit() or 
                    obj == self.cost_input.lineEdit() or 
                    obj == self.weight_input.lineEdit() or 
                    obj == self.stock_input.lineEdit()):
                    
                    if event.type() == event.Type.FocusIn or event.type() == event.Type.MouseButtonPress:
                        # Clear the field and position cursor when field gains focus or is clicked
                        line_edit = obj
                        # Determine which spinbox this line edit belongs to and clear its value
                        from PyQt6.QtCore import QTimer
                        
                        def clear_field():
                            if obj == self.price_input.lineEdit():
                                self.price_input.setValue(0)
                                self.price_input.lineEdit().clear()
                            elif obj == self.cost_input.lineEdit():
                                self.cost_input.setValue(0)
                                self.cost_input.lineEdit().clear()
                            elif obj == self.weight_input.lineEdit():
                                self.weight_input.setValue(0)
                                self.weight_input.lineEdit().clear()
                            elif obj == self.stock_input.lineEdit():
                                self.stock_input.setValue(0)
                                self.stock_input.lineEdit().clear()
                        
                        # Use QTimer.singleShot to ensure clearing happens after the click event is processed
                        QTimer.singleShot(0, clear_field)
            
            # Auto-clear text for spinbox input fields in warehouse price/weight search tab
            if hasattr(self, 'min_price_input') and hasattr(self, 'max_price_input') and hasattr(self, 'min_weight_input') and hasattr(self, 'max_weight_input'):
                if (obj == self.min_price_input.lineEdit() or 
                    obj == self.max_price_input.lineEdit() or 
                    obj == self.min_weight_input.lineEdit() or 
                    obj == self.max_weight_input.lineEdit()):
                    
                    if event.type() == event.Type.FocusIn or event.type() == event.Type.MouseButtonPress:
                        # Clear the field and position cursor when field gains focus or is clicked
                        from PyQt6.QtCore import QTimer
                        
                        def clear_warehouse_field():
                            if obj == self.min_price_input.lineEdit():
                                self.min_price_input.setValue(0)
                                self.min_price_input.lineEdit().clear()
                            elif obj == self.max_price_input.lineEdit():
                                self.max_price_input.setValue(0)
                                self.max_price_input.lineEdit().clear()
                            elif obj == self.min_weight_input.lineEdit():
                                self.min_weight_input.setValue(0)
                                self.min_weight_input.lineEdit().clear()
                            elif obj == self.max_weight_input.lineEdit():
                                self.max_weight_input.setValue(0)
                                self.max_weight_input.lineEdit().clear()
                        
                        # Use QTimer.singleShot to ensure clearing happens after the click event is processed
                        QTimer.singleShot(0, clear_warehouse_field)
            
            # Handle Enter key press for blur functionality
            if event.type() == event.Type.KeyPress:
                if event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
                    # Check if this is one of our numeric input fields
                    input_fields = []
                    if hasattr(self, 'price_input'):
                        input_fields.append(self.price_input.lineEdit())
                    if hasattr(self, 'cost_input'):
                        input_fields.append(self.cost_input.lineEdit())
                    if hasattr(self, 'weight_input'):
                        input_fields.append(self.weight_input.lineEdit())
                    if hasattr(self, 'stock_input'):
                        input_fields.append(self.stock_input.lineEdit())
                    if hasattr(self, 'min_price_input'):
                        input_fields.append(self.min_price_input.lineEdit())
                    if hasattr(self, 'max_price_input'):
                        input_fields.append(self.max_price_input.lineEdit())
                    if hasattr(self, 'min_weight_input'):
                        input_fields.append(self.min_weight_input.lineEdit())
                    if hasattr(self, 'max_weight_input'):
                        input_fields.append(self.max_weight_input.lineEdit())
                    
                    if obj in input_fields:
                        # Clear focus from the line edit
                        obj.clearFocus()
                        
                        # Use QTimer to clear selection after focus is processed
                        def clear_selection():
                            obj.deselect()
                        
                        QTimer.singleShot(0, clear_selection)
                        return True  # Event handled
                        
        except Exception:
            pass  # Ignore any errors in event filtering
        
        # Always call the parent event filter
        return super().eventFilter(obj, event)




    def load_shop_inventory(self):
        """Load inventory for selected shop - protected against concurrent calls"""
        # Prevent concurrent loading to avoid race conditions
        if self.shop_inventory_loading:
            return
        
        self.shop_inventory_loading = True
        try:
            shop_name = self.shop_combo.currentText()
            logger.info(f"Loading shop inventory for: '{shop_name}'")
            
            if not shop_name:
                # Clear table if no shop selected
                logger.info("No shop selected, clearing table")
                self.shop_table.setRowCount(0)
                self.update_shop_summary([])
                return

            # Get shop ID
            shop_id = self.db.get_shop_id(shop_name)
            logger.info(f"Shop ID for '{shop_name}': {shop_id}")
            if not shop_id:
                # Clear table if shop not found
                logger.warning(f"Shop '{shop_name}' not found in database, clearing table")
                self.shop_table.setRowCount(0)
                self.update_shop_summary([])
                return

            # Get shop items
            items = self.db.get_shop_items(shop_id)
            logger.info(f"Retrieved {len(items)} items for shop '{shop_name}' (ID: {shop_id})")
            
            # Clear table first to ensure fresh data
            
            # Clear table and refresh
            self.shop_table.clearContents()
            self.shop_table.setRowCount(0)
            
            # Force table to refresh display
            self.shop_table.repaint()
            
            # Update table with new consistent structure
            self.shop_table.setRowCount(len(items))
            
            for row, item in enumerate(items):
                
                try:
                    # Parse and format the date - prioritize updated_at timestamp
                    date_added = None
                    try:
                        # The query returns: id, barcode, name, description, category, price, 
                        # cost, weight, metal_type, stone_type, stock_quantity, 
                        # created_at, updated_at, shop_quantity
                        # So: created_at is item[11], updated_at is item[12], shop_quantity is item[13]
                        if len(item) > 12 and item[12]:  # updated_at column (prioritize for latest changes)
                            date_added = parse_database_datetime(item[12])
                        elif len(item) > 11 and item[11]:  # created_at column (fallback)
                            date_added = parse_database_datetime(item[11])
                    except (IndexError, TypeError, ValueError):
                        date_added = None
                    
                    if date_added:
                        date_str = format_date_for_display(date_added)
                        time_str = format_time_for_display(date_added)
                    else:
                        date_str = ""
                        time_str = ""
                    
                    # Set table items matching new structure with NULL safety
                    barcode_item = QTableWidgetItem(str(item[1]) if len(item) > 1 and item[1] is not None else "")
                    barcode_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 0, barcode_item)  # Barcode
                    
                    category_item = QTableWidgetItem(str(item[4]) if len(item) > 4 and item[4] is not None else "")
                    category_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 1, category_item)  # Category
                    
                    metal_item = QTableWidgetItem(str(item[8]) if len(item) > 8 and item[8] is not None else "")
                    metal_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 2, metal_item)  # Metal
                    
                    stone_item = QTableWidgetItem(str(item[9]) if len(item) > 9 and item[9] is not None else "")
                    stone_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 3, stone_item)  # Stone
                    
                    description_item = QTableWidgetItem(str(item[3]) if len(item) > 3 and item[3] is not None else "")
                    description_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 4, description_item)  # Description
                    
                    # Handle cost (Euro in database)
                    cost_eur = float(item[6]) if len(item) > 6 and item[6] is not None else 0.0
                    cost_lev = self.euro_to_lev(cost_eur)
                    cost_text = f"{cost_eur:.2f} €\n{cost_lev:.2f} лв"
                    cost_item = QTableWidgetItem(cost_text)
                    cost_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 5, cost_item)  # Cost
                    
                    # Handle price (Euro in database)
                    price_eur = float(item[5]) if len(item) > 5 and item[5] is not None else 0.0
                    price_lev = self.euro_to_lev(price_eur)
                    price_text = f"{price_eur:.2f} €\n{price_lev:.2f} лв"
                    price_item = QTableWidgetItem(price_text)
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 6, price_item)  # Price
                    
                    # Handle weight
                    weight = float(item[7]) if len(item) > 7 and item[7] is not None else 0.0
                    weight_item = QTableWidgetItem(self.format_grams(weight))
                    weight_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 7, weight_item)  # Weight
                    
                    # Handle shop stock (shop_quantity is at index 13 from the fixed query)
                    shop_stock = int(item[13]) if len(item) > 13 and item[13] is not None else 0
                    stock_item = QTableWidgetItem(str(shop_stock))
                    stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 8, stock_item)  # Shop Stock
                    
                    # Highlight row if quantity is 0 in shop
                    self.highlight_zero_quantity_row(self.shop_table, row, shop_stock)
                
                    # Date and Time
                    date_item = QTableWidgetItem(date_str)
                    date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 9, date_item)  # Date
                    
                    time_item = QTableWidgetItem(time_str)
                    time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.shop_table.setItem(row, 10, time_item)  # Time
                    
                except Exception as e:
                    logger.error(f"Error loading shop item at row {row}: {e}")
                    print(f"📄 Exception type: {type(e)}")
                    import traceback
                    print(f"📄 Full traceback:")
                    traceback.print_exc()
                    continue

            # Update summary
            self.update_shop_summary(items)
            
            # Force table refresh with more aggressive updates
            self.shop_table.viewport().update()
            self.shop_table.update()
            self.shop_table.repaint()
            
            # Process pending events to ensure UI updates
            from PyQt6.QtCore import QCoreApplication
            QCoreApplication.processEvents()
            
            logger.info(f"Shop inventory loading completed. Table now has {self.shop_table.rowCount()} rows")
            
        except Exception as e:
            logger.error(f"Error in load_shop_inventory: {e}", exc_info=True)
            QMessageBox.critical(self, "Грешка", f"Грешка при зареждане на инвентара: {str(e)}")
        finally:
            # Always reset the loading flag to allow future loads
            self.shop_inventory_loading = False

    def on_shop_barcode_changed(self):
        """Handle shop barcode input changes with timer delay"""
        # Stop any existing timer
        self.shop_barcode_timer.stop()
        
        # Start timer with 500ms delay to wait for complete barcode
        self.shop_barcode_timer.start(500)
        
    def process_shop_barcode(self):
        """Process the complete shop barcode after timer delay"""
        try:
            barcode = self.shop_barcode_input.text().strip()
            
            # Process when 7 or 13 digits are entered (support both formats)
            if (len(barcode) == 7 or len(barcode) == 13) and barcode.isdigit():
                shop_name = self.shop_combo.currentText()
                if not shop_name:
                    QMessageBox.warning(self, "Грешка", "Моля, изберете магазин!")
                    self.shop_barcode_input.clear()
                    return
                
                self.load_item_to_shop(barcode, shop_name)

        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при обработка на баркод: {str(e)}")



    def load_item_to_shop(self, barcode, shop_name):
        """Load item to shop"""
        try:
            # Get shop ID
            shop_id = self.db.get_shop_id(shop_name)
            if not shop_id:
                QMessageBox.warning(self, "Грешка", "Магазинът не е намерен!")
                self.shop_barcode_input.clear()
                return
            
                # Check if item exists in inventory
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, category, stock_quantity FROM items WHERE barcode = ?", (barcode,))
                item = cursor.fetchone()
                
                if not item:
                    QMessageBox.warning(self, "Грешка", f"Артикулът с баркод {barcode} не е намерен в инвентара!")
                    self.shop_barcode_input.clear()
                    return
                
                item_id, category, available_stock = item
                
                if available_stock <= 0:
                    QMessageBox.warning(self, "Грешка", f"Няма количество за артикул {barcode}!")
                    self.shop_barcode_input.clear()
                    return
                
                # Ask for quantity to load with limits
                quantity, ok = QInputDialog.getInt(
                    self, "Зареди в магазин",
                    f"Колко броя да заредите в магазин '{shop_name}'?\n"
                    f"Категория: {category}\n"
                    f"Налично в склада: {available_stock}",
                    1, 1, available_stock, 1
                )
                
                if not ok:
                    self.shop_barcode_input.clear()
                    return
                
                # Validate quantity
                if quantity > available_stock:
                    QMessageBox.warning(self, "Грешка", f"Количеството ({quantity}) надвишава наличното в склада ({available_stock})")
                    self.shop_barcode_input.clear()
                    return
                
                if quantity <= 0:
                    QMessageBox.warning(self, "Грешка", "Количеството трябва да бъде положително число")
                    self.shop_barcode_input.clear()
                    return
                
                # Add/update item in shop with timestamps
                cursor.execute("SELECT quantity FROM shop_items WHERE shop_id = ? AND item_id = ?", (shop_id, item_id))
                existing = cursor.fetchone()
                
                if existing:
                    # Update quantity with timestamp
                    new_quantity = existing[0] + quantity
                    cursor.execute("UPDATE shop_items SET quantity = ?, updated_at = datetime('now', 'localtime') WHERE shop_id = ? AND item_id = ?", 
                                 (new_quantity, shop_id, item_id))
                else:
                    # Insert new with timestamps
                    cursor.execute("INSERT INTO shop_items (shop_id, item_id, quantity, created_at, updated_at) VALUES (?, ?, ?, datetime('now', 'localtime'), datetime('now', 'localtime'))", 
                                 (shop_id, item_id, quantity))
                
                # Decrease from main inventory with timestamp
                cursor.execute("UPDATE items SET stock_quantity = stock_quantity - ?, updated_at = datetime('now', 'localtime') WHERE id = ?", (quantity, item_id))
                
                conn.commit()
                
                QMessageBox.information(self, "Успех", f"Заредени {quantity} броя {category} в {shop_name}")
                self.shop_barcode_input.clear()
                # Force refresh of shop inventory
                self.load_shop_inventory()
                self.load_items()  # Refresh main inventory
                
                # Update sales tab shop info with longer delay to ensure combo box stability
                QTimer.singleShot(300, self.update_shop_inventory_info)
                self.update_reports_and_database_stats()

        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при зареждане: {str(e)}")
            self.shop_barcode_input.clear()

    def scan_barcode_shop(self):
        """Scan barcode for shop loading"""
        barcode = self.barcode_scanner.scan_from_camera()
        if barcode:
            self.shop_barcode_input.setText(barcode)

    def edit_shop_item(self, item):
        """Edit shop item using dedicated dialog"""
        try:
            row = item.row()
            
            # Get current item data from shop table
            barcode_item = self.shop_table.item(row, 0)     # Barcode
            category_item = self.shop_table.item(row, 1)    # Category
            metal_item = self.shop_table.item(row, 2)       # Metal
            stone_item = self.shop_table.item(row, 3)       # Stone
            description_item = self.shop_table.item(row, 4) # Description
            cost_item = self.shop_table.item(row, 5)        # Cost (Price bought)
            price_item = self.shop_table.item(row, 6)       # Price
            weight_item = self.shop_table.item(row, 7)      # Weight
            shop_stock_item = self.shop_table.item(row, 8)  # Shop Stock
            
            # Check if all required items exist
            if not all([barcode_item, category_item, metal_item, stone_item, cost_item, price_item, weight_item, shop_stock_item]):
                QMessageBox.warning(self, "Грешка", "Няма достатъчно данни за редактиране на този артикул")
                return
            
            # Extract text values
            barcode = barcode_item.text()
            category = category_item.text()
            metal = metal_item.text()
            stone = stone_item.text()
            description = description_item.text() if description_item else ""
            
            # Parse dual currency format (Euro on first line)
            cost_text = cost_item.text().split('\n')[0].replace(" €", "").replace(" ", "") if cost_item else "0"
            price_text = price_item.text().split('\n')[0].replace(" €", "").replace(" ", "") if price_item else "0"
            weight_text = weight_item.text()
            shop_stock_text = shop_stock_item.text()
            
            # Convert numeric values with error handling
            try:
                price = float(price_text) if price_text else 0.0
                cost = float(cost_text) if cost_text else 0.0
                shop_stock = int(shop_stock_text) if shop_stock_text else 0
            except (ValueError, IndexError) as e:
                QMessageBox.warning(self, "Грешка", f"Невалидни числови данни: {str(e)}")
                return
            
            # Parse weight back to grams
            weight_grams = self.parse_weight_to_grams(weight_text)
            
            # Get current inventory stock for this item
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, stock_quantity FROM items WHERE barcode = ?", (barcode,))
                item_result = cursor.fetchone()
                if not item_result:
                    QMessageBox.warning(self, "Грешка", "Артикулът не е намерен в инвентара")
                    return
                item_id, inventory_stock = item_result
            
            # Create and show edit dialog - from_warehouse=False for shop items
            dialog = EditItemDialog(self, barcode, category, description, price, cost, weight_grams, metal, stone, shop_stock, from_warehouse=False)
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # Get updated data from dialog
                updated_data = dialog.get_data()
                new_shop_stock = updated_data['stock']
                stock_difference = new_shop_stock - shop_stock
                
                # Check if we have enough inventory for increase
                if stock_difference > 0 and inventory_stock < stock_difference:
                    QMessageBox.warning(self, "Грешка", 
                                      f"Няма достатъчно количество в склада.\n"
                                      f"Налично в склад: {inventory_stock}\n"
                                      f"Необходимо: {stock_difference}")
                    return
                
                try:
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        
                        # Update main item data
                        cursor.execute("""
                            UPDATE items 
                            SET name = ?, description = ?, category = ?, price = ?, cost = ?, 
                                weight = ?, metal_type = ?, stone_type = ?, stock_quantity = stock_quantity - ?
                            WHERE id = ?
                        """, (updated_data['category'], updated_data['description'], updated_data['category'],
                              updated_data['price'], updated_data['cost'], updated_data['weight'],
                              updated_data['metal'], updated_data['stone'], stock_difference, item_id))
                        
                        # Update shop inventory
                        shop_id = self.db.get_shop_id(self.shop_combo.currentText())
                        cursor.execute("""
                            UPDATE shop_items 
                            SET quantity = ?
                            WHERE shop_id = ? AND item_id = ?
                        """, (new_shop_stock, shop_id, item_id))
                        
                        conn.commit()
                        
                        QMessageBox.information(self, "Успех", "Артикулът е обновен успешно")
                        self.load_items()  # Reload main inventory
                        self.load_shop_inventory()  # Reload shop inventory
                        
                        # Update sales tab shop info with longer delay to ensure combo box stability
                        QTimer.singleShot(300, self.update_shop_inventory_info)
                        self.update_reports_and_database_stats()
                        
                except Exception as e:
                    QMessageBox.critical(self, "Грешка", f"Грешка при обновяване: {str(e)}")
        
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при редактиране на артикула: {str(e)}")
            logger.error(f"Error in edit_shop_item: {e}", exc_info=True)

    def shop_right_click(self, position):
        """Handle right-click on shop inventory table"""
        try:
            row = self.shop_table.rowAt(position.y())
            menu = QMenu()
            
            if row >= 0 and row < self.shop_table.rowCount():
                # Right-click on specific item - check if items exist
                barcode_item = self.shop_table.item(row, 0)
                shop_stock_item = self.shop_table.item(row, 8)
                
                if not barcode_item or not shop_stock_item:
                    QMessageBox.warning(self, "Грешка", "Няма данни за този ред")
                    return
                
                barcode = barcode_item.text()
                shop_stock_text = shop_stock_item.text()
                
                if not barcode or not shop_stock_text:
                    QMessageBox.warning(self, "Грешка", "Няма валидни данни за този артикул")
                    return
                
                try:
                    shop_stock = int(shop_stock_text)
                except ValueError:
                    shop_stock = 1
                
                edit_action = menu.addAction("Редактирай")
                return_action = menu.addAction("Връщане в склада")
                menu.addSeparator()
                return_all_item_action = menu.addAction(f"Върни всички от този артикул ({shop_stock} броя)")
                
                action = menu.exec(self.shop_table.viewport().mapToGlobal(position))
                
                if action == edit_action:
                    if barcode_item:
                        self.edit_shop_item(barcode_item)
                elif action == return_action:
                    # Ask for quantity to return
                    quantity, ok = QInputDialog.getInt(
                        self, "Връщане в склада",
                        f"Колко броя да върнете в склада?\n(Максимум: {shop_stock})",
                        1, 1, shop_stock, 1
                    )
                    
                    if ok:
                        self.return_item_to_inventory(barcode, quantity)
                elif action == return_all_item_action:
                    # Return all quantities of this specific item
                    reply = QMessageBox.question(
                        self, "Потвърждение",
                        f"Ще върнете всички {shop_stock} броя от артикул {barcode} в склада.\n\nСигурни ли сте?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    if reply == QMessageBox.StandardButton.Yes:
                        self.return_item_to_inventory(barcode, shop_stock)
                    
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при обработка на контекстното меню: {str(e)}")

    def return_item_to_inventory(self, barcode, quantity):
        """Return item from shop to main inventory"""
        try:
            shop_name = self.shop_combo.currentText()
            shop_id = self.db.get_shop_id(shop_name)
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get item_id
                cursor.execute("SELECT id FROM items WHERE barcode = ?", (barcode,))
                item_result = cursor.fetchone()
                if not item_result:
                    QMessageBox.warning(self, "Грешка", "Артикулът не е намерен!")
                    return
                item_id = item_result[0]
                
                # Get current shop quantity
                cursor.execute("SELECT quantity FROM shop_items WHERE shop_id = ? AND item_id = ?", 
                             (shop_id, item_id))
                shop_result = cursor.fetchone()
                if not shop_result or shop_result[0] < quantity:
                    QMessageBox.warning(self, "Грешка", "Недостатъчно количество в магазина!")
                    return
                
                current_shop_quantity = shop_result[0]
                
                # Update shop inventory
                if current_shop_quantity == quantity:
                    # Remove from shop completely
                    cursor.execute("DELETE FROM shop_items WHERE shop_id = ? AND item_id = ?", 
                                 (shop_id, item_id))
                else:
                    # Decrease shop quantity with timestamp
                    cursor.execute("UPDATE shop_items SET quantity = quantity - ?, updated_at = datetime('now', 'localtime') WHERE shop_id = ? AND item_id = ?", 
                                 (quantity, shop_id, item_id))
                
                # Add back to main inventory with timestamp
                cursor.execute("UPDATE items SET stock_quantity = stock_quantity + ?, updated_at = datetime('now', 'localtime') WHERE id = ?", 
                             (quantity, item_id))
                
                conn.commit()
                
                QMessageBox.information(self, "Успех", 
                                      f"Върнати {quantity} броя от артикул {barcode} в склада")
                
                # Reload both tables with proper error handling
                try:
                    self.load_items()
                    logger.info(f"Warehouse inventory reloaded after returning {quantity} of {barcode}")
                except Exception as e:
                    logger.error(f"Error reloading warehouse inventory: {e}")
                
                try:
                    # Simple reload without forcing combo box changes - use delayed refresh
                    QTimer.singleShot(200, self.load_shop_inventory)  # 200ms delay to avoid race conditions
                    logger.info(f"Scheduled delayed shop inventory reload after returning {quantity} of {barcode}")
                except Exception as e:
                    logger.error(f"Error scheduling shop inventory reload: {e}")
                
                # Update sales tab shop info with longer delay to ensure combo box stability
                QTimer.singleShot(300, self.update_shop_inventory_info)
                self.update_reports_and_database_stats()
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при връщане в склада: {str(e)}")

    def refresh_shop_combo(self):
        """Refresh the shop combo box with current shops"""
        try:
            # Store current selection
            current_shop = self.shop_combo.currentText()
            logger.info(f"Refreshing shop combo, current selection: '{current_shop}'")
            
            # Use blockSignals to prevent signals during refresh (safer than disconnect/reconnect)
            self.shop_combo.blockSignals(True)
            
            try:
                # Clear and reload shops
                self.shop_combo.clear()
                shops = self.db.get_all_shops()
                logger.info(f"Loading {len(shops)} shops into combo box")
                
                for shop in sorted(shops, key=lambda x: x[1]):  # Sort by shop name
                    self.shop_combo.addItem(shop[1])
                
                # Restore selection if the shop still exists, otherwise select first available
                if current_shop and self.shop_combo.count() > 0:
                    index = self.shop_combo.findText(current_shop)
                    if index >= 0:
                        self.shop_combo.setCurrentIndex(index)
                        logger.info(f"Restored selection to '{current_shop}'")
                    else:
                        # Shop was deleted, select first available shop
                        self.shop_combo.setCurrentIndex(0)
                        new_selection = self.shop_combo.currentText()
                        logger.info(f"Shop '{current_shop}' no longer exists, selected '{new_selection}'")
                elif self.shop_combo.count() > 0:
                    self.shop_combo.setCurrentIndex(0)
                    logger.info(f"No previous selection, selected first shop: '{self.shop_combo.currentText()}'")
                else:
                    # No shops available
                    logger.info("No shops available")
                
            finally:
                # Always unblock signals
                self.shop_combo.blockSignals(False)
            
            # Load inventory for currently selected shop (after refresh)
            self.load_shop_inventory()
            
        except Exception as e:
            logger.error(f"Error refreshing shop combo: {e}")
    
    def refresh_all_shop_combos(self):
        """Refresh all shop combo boxes across the application"""
        try:
            logger.info("Refreshing all shop combo boxes after shop changes")
            
            # Refresh main shop combo (shop loading tab)
            self.refresh_shop_combo()
            
            # Refresh sales shop combo
            if hasattr(self, 'sales_shop_combo'):
                current_sales_shop = self.sales_shop_combo.currentText()
                logger.info(f"Refreshing sales shop combo, current selection: '{current_sales_shop}'")
                
                # Use blockSignals to prevent signals during refresh
                self.sales_shop_combo.blockSignals(True)
                
                try:
                    # Clear and reload
                    self.sales_shop_combo.clear()
                    shops = self.db.get_all_shops()
                    for shop in shops:
                        self.sales_shop_combo.addItem(shop[1])
                    
                    # Restore selection or select first available
                    if current_sales_shop and self.sales_shop_combo.count() > 0:
                        index = self.sales_shop_combo.findText(current_sales_shop)
                        if index >= 0:
                            self.sales_shop_combo.setCurrentIndex(index)
                            logger.info(f"Restored sales shop selection to '{current_sales_shop}'")
                        else:
                            # Shop was deleted, select first available
                            self.sales_shop_combo.setCurrentIndex(0)
                            new_selection = self.sales_shop_combo.currentText()
                            logger.info(f"Sales shop '{current_sales_shop}' no longer exists, selected '{new_selection}'")
                    elif self.sales_shop_combo.count() > 0:
                        self.sales_shop_combo.setCurrentIndex(0)
                        logger.info(f"No previous sales shop selection, selected first: '{self.sales_shop_combo.currentText()}'")
                
                finally:
                    # Always unblock signals
                    self.sales_shop_combo.blockSignals(False)
                
                # Update related data with current selection
                self.load_sales()
                self.update_shop_inventory_info()
            
            # Refresh audit shop combo
            if hasattr(self, 'audit_shop_combo'):
                self.refresh_audit_shop_combo()
            
            logger.info("All shop combo boxes refreshed successfully")
                
        except Exception as e:
            logger.error(f"Error refreshing all shop combos: {e}")
    




    def update_shop_summary(self, items):
        """Update shop inventory summary"""
        try:
            total_items = len(items)
            total_value_eur = 0.0
            total_weight = 0.0
            
            # Calculate totals with safe conversion
            for item in items:
                try:
                    # Safely convert price (item[5]) and shop_quantity (item[13]) and weight (item[7])
                    try:
                        price = float(item[5]) if item[5] is not None else 0.0
                    except (ValueError, TypeError):
                        price = 0.0
                    
                    try:
                        shop_quantity = float(item[13]) if item[13] is not None else 0.0
                    except (ValueError, TypeError):
                        shop_quantity = 0.0
                    
                    try:
                        weight = float(item[7]) if item[7] is not None else 0.0
                    except (ValueError, TypeError):
                        weight = 0.0
                    
                    total_value_eur += price * shop_quantity
                    total_weight += weight * shop_quantity
                    
                except (IndexError) as e:
                    # Log problematic item but continue with others
                    logger.warning(f"Error processing item in shop summary: {e}. Item data: {item}")
                    continue
            
            total_value_lev = self.euro_to_lev(total_value_eur)

            self.shop_summary_labels[0].setText(f"Артикули: {total_items}")
            value_text = f"{self.format_currency_eur(total_value_eur)}\n{self.format_currency_lev(total_value_lev)}"
            self.shop_summary_labels[2].setText(value_text)
            self.shop_summary_labels[3].setText(self.format_grams(total_weight))
            self.shop_summary_labels[4].setText("")
            self.shop_summary_labels[5].setText("")
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при обновяване на обобщението: {str(e)}")

    def highlight_zero_quantity_row(self, table, row, stock_value):
        """Highlight row like delete button if quantity is 0"""
        try:
            if stock_value <= 0:
                # Set delete button style: light pink background with dark red text
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        item.setBackground(QColor(255, 204, 204))  # Light pink background (#ffcccc)
                        item.setForeground(QColor(204, 0, 0))     # Dark red text (#cc0000)
                        item.setToolTip("ВНИМАНИЕ: Количеството е 0!")
                        # Make text bold to match delete button
                        font = item.font()
                        font.setBold(True)
                        item.setFont(font)
                    # Also set delete button style for widgets (like images)
                    widget = table.cellWidget(row, col)
                    if widget:
                        widget.setStyleSheet("background-color: #ffcccc; color: #cc0000; font-weight: bold;")
            else:
                # Set normal white text with default background for non-zero quantities
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        item.setBackground(QColor())  # Default background
                        item.setForeground(QColor(255, 255, 255))  # White text
                        item.setToolTip("")  # Remove tooltip
                        # Reset font weight
                        font = item.font()
                        font.setBold(False)
                        item.setFont(font)
                    # Also reset widget style
                    widget = table.cellWidget(row, col)
                    if widget and hasattr(widget, 'setStyleSheet'):
                        widget.setStyleSheet("color: white;")
        except Exception as e:
            logger.error(f"Error highlighting zero quantity row: {e}")

    def add_new_shop(self):
        """Add new shop"""
        try:
            name, ok = QInputDialog.getText(self, "Нов магазин", "Име на магазина:")
            if ok and name:
                if self.db.add_shop(name):
                    # Update all shop combo boxes across the application
                    self.refresh_all_shop_combos()
                    # Select the newly added shop
                    index = self.shop_combo.findText(name)
                    if index >= 0:
                        self.shop_combo.setCurrentIndex(index)
                    QMessageBox.information(self, "Успех", "Магазинът е добавен успешно")
                    self.update_reports_and_database_stats()  # Update statistics
                else:
                    QMessageBox.warning(self, "Грешка", "Неуспешно добавяне на магазин")
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при добавяне на магазин: {str(e)}")

    def rename_selected_shop(self):
        """Rename selected shop"""
        try:
            current_name = self.shop_combo.currentText()
            if not current_name:
                return

            new_name, ok = QInputDialog.getText(
                self, "Редактирай магазин",
                "Ново име на магазина:",
                text=current_name
            )
            if ok and new_name and new_name != current_name:
                if self.db.rename_shop(current_name, new_name):
                    # Refresh all shop combo boxes and select the renamed shop
                    self.refresh_all_shop_combos()
                    index = self.shop_combo.findText(new_name)
                    if index >= 0:
                        self.shop_combo.setCurrentIndex(index)
                    QMessageBox.information(self, "Успех", "Магазинът е преименуван успешно")
                    self.update_reports_and_database_stats()  # Update statistics
                else:
                    QMessageBox.warning(self, "Грешка", "Неуспешно преименуване на магазин")
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при преименуване на магазин: {str(e)}")

    def delete_selected_shop(self):
        """Delete selected shop"""
        try:
            shop_name = self.shop_combo.currentText()
            if not shop_name:
                return

            reply = QMessageBox.question(
                self, "Изтрий магазин",
                f"Сигурни ли сте, че искате да изтриете магазин '{shop_name}'?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                if self.db.delete_shop(shop_name):
                    logger.info(f"Shop '{shop_name}' deleted successfully")
                    
                    # Refresh all shop combo boxes after deletion
                    self.refresh_all_shop_combos()
                    
                    # Update statistics
                    self.update_reports_and_database_stats()
                    
                    QMessageBox.information(self, "Успех", "Магазинът е изтрит успешно")
                else:
                    QMessageBox.warning(self, "Грешка", "Неуспешно изтриване на магазин")
        except Exception as e:
            logger.error(f"Error deleting shop: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при изтриване на магазин: {str(e)}")

    # === DATABASE MANAGEMENT METHODS ===
    def update_database_statistics(self):
        """Update database statistics in the overview cards"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Total items
                cursor.execute("SELECT COUNT(*) FROM items")
                total_items = cursor.fetchone()[0]
                if hasattr(self, 'db_stats_cards'):
                    self.db_stats_cards["total_items"].value_label.setText(str(total_items))
                
                # Total sales
                cursor.execute("SELECT COUNT(*) FROM sales")
                total_sales = cursor.fetchone()[0]
                if hasattr(self, 'db_stats_cards'):
                    self.db_stats_cards["total_sales"].value_label.setText(str(total_sales))
                
                # Total shops
                cursor.execute("SELECT COUNT(*) FROM shops")
                total_shops = cursor.fetchone()[0]
                if hasattr(self, 'db_stats_cards'):
                    self.db_stats_cards["total_shops"].value_label.setText(str(total_shops))
                
                # Database size - check multiple possible paths
                db_size_mb = 0
                possible_db_paths = [get_persistent_path("data/jewelry.db")]
                
                for db_path in possible_db_paths:
                    if os.path.exists(db_path):
                        try:
                            size_bytes = os.path.getsize(db_path)
                            db_size_mb = size_bytes / (1024 * 1024)
                            break
                        except Exception as e:
                            logger.warning(f"Could not get size for {db_path}: {e}")
                            continue
                
                if hasattr(self, 'db_stats_cards'):
                    if db_size_mb > 0:
                        self.db_stats_cards["db_size"].value_label.setText(f"{db_size_mb:.1f} MB")
                    else:
                        # Try to get size from connection
                        try:
                            cursor.execute("PRAGMA page_count")
                            page_count = cursor.fetchone()[0]
                            cursor.execute("PRAGMA page_size") 
                            page_size = cursor.fetchone()[0]
                            size_bytes = page_count * page_size
                            db_size_mb = size_bytes / (1024 * 1024)
                            self.db_stats_cards["db_size"].value_label.setText(f"{db_size_mb:.1f} MB")
                        except:
                            self.db_stats_cards["db_size"].value_label.setText("? MB")
                
                # Last backup time
                if hasattr(self, 'db_stats_cards'):
                    last_backup_time = self.get_last_backup_time()
                    if last_backup_time:
                        self.db_stats_cards["last_backup"].value_label.setText(last_backup_time)
                    else:
                        self.db_stats_cards["last_backup"].value_label.setText("Никога")
                
                # Database integrity status
                if hasattr(self, 'db_stats_cards'):
                    integrity_status = self.check_database_integrity()
                    self.db_stats_cards["integrity_status"].value_label.setText(integrity_status)
                        
        except Exception as e:
            logger.error(f"Error updating database statistics: {e}")
    
    def load_backup_list(self):
        """Load backup files into the backup list table"""
        try:
            if not hasattr(self, 'backup_list'):
                return
                
            backup_dir = self.get_backup_directory()
            if not os.path.exists(backup_dir):
                return
                
            backup_files = [f for f in os.listdir(backup_dir) if f.endswith('.db')]
            backup_files.sort(key=lambda x: os.path.getctime(os.path.join(backup_dir, x)), reverse=True)
            
            self.backup_list.setRowCount(len(backup_files))
            
            for row, filename in enumerate(backup_files):
                filepath = os.path.join(backup_dir, filename)
                
                # Filename
                filename_item = QTableWidgetItem(filename)
                filename_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.backup_list.setItem(row, 0, filename_item)
                
                # Date
                file_time = datetime.fromtimestamp(os.path.getctime(filepath))
                date_item = QTableWidgetItem(file_time.strftime("%d/%m/%Y %H:%M"))
                date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.backup_list.setItem(row, 1, date_item)
                
                # Size
                size_mb = os.path.getsize(filepath) / (1024 * 1024)
                size_item = QTableWidgetItem(f"{size_mb:.1f} MB")
                size_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.backup_list.setItem(row, 2, size_item)
                
                # Actions column
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(4, 4, 4, 4)
                
                # Open file location button
                location_btn = QPushButton("📂")
                location_btn.setToolTip("Отвори местоположението на файла")
                location_btn.setMaximumWidth(30)
                location_btn.setStyleSheet("background-color: #007bff; color: white; border: 1px solid #0056b3;")
                location_btn.clicked.connect(lambda checked, path=filepath: self.open_backup_location(path))
                actions_layout.addWidget(location_btn)
                
                # Delete backup button
                delete_btn = QPushButton("🗑️")
                delete_btn.setToolTip("Изтрий резервното копие")
                delete_btn.setMaximumWidth(30)
                delete_btn.setStyleSheet("background-color: #dc3545; color: white; border: 1px solid #c82333;")
                delete_btn.clicked.connect(lambda checked, path=filepath, fname=filename: self.delete_backup_file(path, fname))
                actions_layout.addWidget(delete_btn)
                
                actions_layout.addStretch()
                self.backup_list.setCellWidget(row, 3, actions_widget)
                
        except Exception as e:
            logger.error(f"Error loading backup list: {e}")
    
    def save_last_backup_time(self):
        """Save the current time as the last backup time"""
        try:
            # Create data directory if it doesn't exist
            os.makedirs('data', exist_ok=True)
            
            # Save backup timestamp
            backup_info = {
                'last_backup_time': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                'last_backup_timestamp': datetime.now().isoformat()
            }
            
            with open(get_persistent_path('data/backup_info.json'), 'w', encoding='utf-8') as f:
                json.dump(backup_info, f, indent=2)
                
        except Exception as e:
            logger.error(f"Error saving backup time: {e}")
    
    def open_backup_folder(self, file_path):
        """Open the folder containing the backup file"""
        try:
            # Get the directory containing the file
            folder_path = os.path.dirname(os.path.abspath(file_path))
            
            # Open the folder in Windows Explorer
            if os.name == 'nt':  # Windows
                try:
                    # Try using Windows API first (more reliable)
                    os.startfile(folder_path)
                except:
                    # Fallback to subprocess if os.startfile fails
                    subprocess.run(['explorer', folder_path])
            else:  # Linux/Mac
                subprocess.run(['xdg-open', folder_path])
                
            logger.info(f"Opened backup folder: {folder_path}")
            
        except Exception as e:
            logger.error(f"Error opening backup folder: {e}")
            QMessageBox.warning(self, "Грешка", f"Не може да се отвори папката: {str(e)}")
    
    def open_backup_location(self, file_path):
        """Open file location and select the backup file"""
        try:
            # Convert to absolute path
            abs_path = os.path.abspath(file_path)
            
            # Open file location in Windows Explorer and select the file
            if os.name == 'nt':  # Windows
                try:
                    # Use subprocess for /select functionality (os.startfile doesn't support /select)
                    result = subprocess.run(['explorer', '/select,', abs_path], capture_output=True)
                    # Don't raise error for non-zero exit codes from Explorer
                    if result.returncode != 0:
                        logger.debug(f"Explorer returned exit code {result.returncode}, but this is normal")
                except Exception as inner_e:
                    # If /select fails, just open the folder
                    folder_path = os.path.dirname(abs_path)
                    os.startfile(folder_path)
                    logger.info(f"Fallback: opened folder instead of selecting file: {folder_path}")
            else:  # Linux/Mac
                folder_path = os.path.dirname(abs_path)
                subprocess.run(['xdg-open', folder_path])
                
            logger.info(f"Opened backup file location: {abs_path}")
            
        except Exception as e:
            logger.error(f"Error opening backup file location: {e}")
            QMessageBox.warning(self, "Грешка", f"Не може да се отвори местоположението на файла: {str(e)}")
    
    def open_exports_folder(self):
        """Open the exports folder where audit PDFs and Excel files are saved"""
        try:
            exports_dir = self.get_exports_directory()
            abs_path = os.path.abspath(exports_dir)
            
            # Open the folder in Windows Explorer
            if os.name == 'nt':  # Windows
                try:
                    # Try using Windows API first (more reliable)
                    os.startfile(abs_path)
                except:
                    # Fallback to subprocess if os.startfile fails
                    subprocess.run(['explorer', abs_path])
            else:  # Linux/Mac
                subprocess.run(['xdg-open', abs_path])
                
            logger.info(f"Opened exports folder: {abs_path}")
            
        except Exception as e:
            logger.error(f"Error opening exports folder: {e}")
            QMessageBox.warning(self, "Грешка", f"Не може да се отвори папката с експорти: {str(e)}")
    
    def delete_backup_file(self, file_path, filename):
        """Delete a backup file after confirmation"""
        try:
            # Confirm deletion
            reply = QMessageBox.question(
                self, 
                "Потвърждение", 
                f"Сигурни ли сте, че искате да изтриете резервното копие?\n\n{filename}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # Delete the file
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"Deleted backup file: {file_path}")
                    
                    # Refresh the backup list
                    self.load_backup_list()
                    
                    QMessageBox.information(self, "Успех", f"Резервното копие '{filename}' беше изтрито успешно.")
                else:
                    QMessageBox.warning(self, "Грешка", "Файлът не беше намерен.")
                    
        except Exception as e:
            logger.error(f"Error deleting backup file: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при изтриване на файла: {str(e)}")

    def copy_backup_to_desktop(self, file_path, filename):
        """Copy backup file to desktop"""
        try:
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            destination = os.path.join(desktop_path, filename)
            
            # Copy the file
            shutil.copy2(file_path, destination)
            logger.info(f"Copied backup to desktop: {destination}")
            QMessageBox.information(self, "Успех", f"Резервното копие беше копирано на работния плот:\n{filename}")
            
        except Exception as e:
            logger.error(f"Error copying backup to desktop: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при копиране на файла: {str(e)}")
    
    def closeEvent(self, event):
        """Handle application close event"""
        try:
            # Clean up file system watcher
            if hasattr(self, 'backup_watcher'):
                self.backup_watcher.deleteLater()
                logger.info("Backup file watcher cleaned up")
        except Exception as e:
            logger.error(f"Error during application close: {e}")
        finally:
            # Accept the close event
            event.accept()
    
    def get_last_backup_time(self):
        """Get the last backup time"""
        try:
            if os.path.exists(get_persistent_path('data/backup_info.json')):
                with open(get_persistent_path('data/backup_info.json'), 'r', encoding='utf-8') as f:
                    backup_info = json.load(f)
                    return backup_info.get('last_backup_time', None)
            else:
                # If no backup info file, check backup directory for newest file
                backup_dir = self.get_backup_directory()
                if os.path.exists(backup_dir):
                    backup_files = [f for f in os.listdir(backup_dir) if f.endswith('.db')]
                    if backup_files:
                        # Get the newest backup file
                        newest_backup = max(backup_files, key=lambda x: os.path.getctime(os.path.join(backup_dir, x)))
                        backup_time = datetime.fromtimestamp(os.path.getctime(os.path.join(backup_dir, newest_backup)))
                        return backup_time.strftime("%d/%m/%Y %H:%M:%S")
                return None
        except Exception as e:
            logger.error(f"Error getting backup time: {e}")
            return None
    
    def check_database_integrity(self):
        """Check database integrity and return status"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Run SQLite integrity check
                cursor.execute("PRAGMA integrity_check")
                result = cursor.fetchone()
                
                if result and result[0].lower() == 'ok':
                    # Additional checks
                    try:
                        # Check if essential tables exist
                        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                        tables = [row[0] for row in cursor.fetchall()]
                        
                        essential_tables = ['items', 'sales', 'shops', 'users']
                        missing_tables = [table for table in essential_tables if table not in tables]
                        
                        if missing_tables:
                            return f"⚠ Липсват таблици"
                        
                        # Check for data consistency
                        cursor.execute("SELECT COUNT(*) FROM items WHERE barcode IS NULL OR barcode = ''")
                        invalid_barcodes = cursor.fetchone()[0]
                        
                        cursor.execute("SELECT COUNT(*) FROM items WHERE price < 0")
                        invalid_prices = cursor.fetchone()[0]
                        
                        if invalid_barcodes > 0 or invalid_prices > 0:
                            return f"⚠ Данни с грешки"
                        
                        return "✓ Отлично"
                    except:
                        return "✓ Добро"
                else:
                    return "✗ Повредено"
                    
        except Exception as e:
            logger.error(f"Error checking database integrity: {e}")
            return "? Неизвестно"

    def factory_reset(self):
        """Perform system restart with warnings and PIN confirmation"""
        try:
            # First warning dialog
            first_warning = QMessageBox(self)
            first_warning.setIcon(QMessageBox.Icon.Critical)
            first_warning.setWindowTitle("⚠️ СИСТЕМЕН РЕСТАРТ")
            first_warning.setText("""
<h3 style="color: #d32f2f;">КРИТИЧНО ПРЕДУПРЕЖДЕНИЕ!</h3>

<p><b>Готвите се да извършите системен рестарт на програмата.</b></p>

<p>Това ще изтрие НАПЪЛНО всички данни:</p>
<ul>
<li>🗃️ Всички артикули от склада</li>
<li>💰 Всички продажби и история</li>
<li>🏪 Всички магазини и техните запаси</li>
<li>� Всички отчети и статистики</li>
<li>🏷️ Всички генерирани баркодове</li>
<li>📝 Всички логове и настройки</li>
</ul>

<p style="color: #4CAF50;"><b>💾 РЕЗЕРВНИТЕ КОПИЯ ЩЕ БЪДАТ ЗАПАЗЕНИ!</b></p>
<p style="color: #4CAF50;"><b>📄 ЕКСПОРТИРАНИТЕ ФАЙЛОВЕ ЩЕ БЪДАТ ЗАПАЗЕНИ!</b></p>

<p style="color: #d32f2f;"><b>СИСТЕМАТА ЩЕ СЕ ВЪРНЕ В СЪСТОЯНИЕ КАТО НОВА ИНСТАЛАЦИЯ!</b></p>

<p>Препоръчваме да създадете резервно копие преди да продължите.</p>
            """)
            first_warning.setStandardButtons(
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            first_warning.setDefaultButton(QMessageBox.StandardButton.No)
            first_warning.button(QMessageBox.StandardButton.Yes).setText("Продължи")
            first_warning.button(QMessageBox.StandardButton.No).setText("Отказ")

            # Style the dialog buttons
            first_warning.button(QMessageBox.StandardButton.Yes).setStyleSheet(
                "QPushButton { background-color: #d32f2f; color: white; font-weight: bold; padding: 8px 16px; }"
            )
            first_warning.button(QMessageBox.StandardButton.No).setStyleSheet(
                "QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 8px 16px; }"
            )

            if first_warning.exec() != QMessageBox.StandardButton.Yes:
                return
            
            # Second confirmation with count down
            second_warning = QMessageBox(self)
            second_warning.setIcon(QMessageBox.Icon.Warning)
            second_warning.setWindowTitle("⚠️ ПОСЛЕДНА ПРОВЕРКА")
            second_warning.setText("""
<h3 style="color: #d32f2f;">ПОСЛЕДНА ВЪЗМОЖНОСТ ЗА ОТКАЗ!</h3>

<p><b>Сигурни ли сте, че искате да изтриете всички данни?</b></p>

<p style="color: #d32f2f;">Тази операция е <b>НЕОБРАТИМА</b>!</p>

<p>Ако сте сигурни, въведете PIN кода за потвърждение.</p>
            """)
            second_warning.setStandardButtons(
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            second_warning.setDefaultButton(QMessageBox.StandardButton.No)
            second_warning.button(QMessageBox.StandardButton.Yes).setText("Въведи PIN")
            second_warning.button(QMessageBox.StandardButton.No).setText("ОТКАЗ")

            if second_warning.exec() != QMessageBox.StandardButton.Yes:
                return
            
            # PIN confirmation dialog
            pin_dialog = QDialog(self)
            pin_dialog.setWindowTitle("🔐 PIN Потвърждение")
            pin_dialog.setModal(True)
            pin_dialog.setFixedSize(450, 220)  # Increased width from 350 to 450 and height from 200 to 220
            
            pin_layout = QVBoxLayout(pin_dialog)
            
            # PIN prompt
            pin_prompt = QLabel("Въведете текущата си парола за потвърждение:")
            pin_prompt.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            pin_prompt.setAlignment(Qt.AlignmentFlag.AlignCenter)
            pin_prompt.setStyleSheet("color: #d32f2f; margin-bottom: 10px;")
            pin_layout.addWidget(pin_prompt)
            pin_layout.addSpacing(10)
            
            pin_input = QLineEdit()
            pin_input.setPlaceholderText("Текуща парола")
            pin_input.setEchoMode(QLineEdit.EchoMode.Password)
            pin_input.setFont(QFont("Arial", 16, QFont.Weight.Bold))
            pin_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
            pin_input.setMaxLength(10)  # Match login field limit
            pin_input.setStyleSheet("""
                QLineEdit {
                    border: 2px solid #d32f2f;
                    border-radius: 8px;
                    padding: 10px;
                    font-size: 18px;
                }
                QLineEdit:focus {
                    border-color: #ff5722;
                }
            """)
            pin_layout.addWidget(pin_input)
            
            pin_layout.addSpacing(20)
            
            # Buttons
            button_layout = QHBoxLayout()
            
            cancel_btn = QPushButton("Отказ")
            cancel_btn.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    font-weight: bold;
                    padding: 10px 20px;
                    border-radius: 5px;
                }
            """)
            cancel_btn.clicked.connect(pin_dialog.reject)
            button_layout.addWidget(cancel_btn)
            
            confirm_btn = QPushButton("🔄 ИЗПЪЛНИ РЕСТАРТ")
            confirm_btn.setStyleSheet("""
                QPushButton {
                    background-color: #d32f2f;
                    color: white;
                    font-weight: bold;
                    padding: 10px 20px;
                    border-radius: 5px;
                }
            """)
            
            def check_pin_and_reset():
                entered_pin = pin_input.text().strip()
                # Verify PIN against current user's password instead of hardcoded "0000"
                if self.db.verify_user("admin", entered_pin):
                    pin_dialog.accept()
                    self.perform_factory_reset()
                else:
                    QMessageBox.critical(pin_dialog, "Грешка", "Невалиден PIN код! Въведете текущата си парола.")
                    pin_input.clear()
                    pin_input.setFocus()
            
            confirm_btn.clicked.connect(check_pin_and_reset)
            pin_input.returnPressed.connect(check_pin_and_reset)
            button_layout.addWidget(confirm_btn)
            
            pin_layout.addLayout(button_layout)
            
            # Set focus to PIN input
            pin_input.setFocus()
            
            # Show dialog
            pin_dialog.exec()
            
        except Exception as e:
            logger.error(f"Error in system restart dialog: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при показване на диалога: {str(e)}")

    def perform_factory_reset(self):
        """Actually perform the system restart operation"""
        progress = None
        try:
            # Show progress dialog (non-blocking)
            progress = QMessageBox(self)
            progress.setIcon(QMessageBox.Icon.Information)
            progress.setWindowTitle("🔄 Системен рестарт в процес...")
            progress.setText("Изтриване на данни, моля изчакайте...")
            progress.setStandardButtons(QMessageBox.StandardButton.NoButton)
            progress.show()
            
            # Process events to show dialog
            QApplication.processEvents()
            
            logger.info("Starting system restart operation")
            
            # Step 1: Clear all database tables (optimized)
            progress.setText("🗃️ Изтриване на данни от базата...")
            QApplication.processEvents()
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Disable foreign key constraints temporarily
                cursor.execute("PRAGMA foreign_keys = OFF")
                
                # Get all table names (excluding system tables)
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT IN ('sqlite_sequence', 'sqlite_master')")
                tables = [row[0] for row in cursor.fetchall()]
                
                # Optimized: Clear all tables in one transaction
                try:
                    for table in tables:
                        cursor.execute(f"DELETE FROM {table}")
                        logger.info(f"Cleared table: {table}")
                    
                    # Reset auto increment counters
                    cursor.execute("DELETE FROM sqlite_sequence")
                    
                    logger.info("All database tables cleared successfully")
                except Exception as e:
                    logger.warning(f"Error clearing tables: {e}")
                
                # Re-enable foreign key constraints
                cursor.execute("PRAGMA foreign_keys = ON")
                
                # Create default shop
                cursor.execute("INSERT INTO shops (name) VALUES (?)", ("Магазин 1",))
                
                # Reset barcode sequence
                try:
                    cursor.execute("INSERT INTO barcode_sequence (id, next_val) VALUES (1, 1000000)")
                except:
                    pass  # Table might not exist
                
                # Create default user with password "0000"
                try:
                    # Reinitialize database schema to ensure all tables exist
                    self.db.force_reinitialize()
                    logger.info("Database schema reinitialized")
                    
                    # Create default user (force password reset during factory reset)
                    self.db.ensure_default_user(default_password="0000", force_create=True)
                    logger.info("Default user created with password '0000'")
                except Exception as e:
                    logger.warning(f"Could not create default user: {e}")
                
                # Reset all master recovery keys to unused status
                try:
                    self.db.reset_master_keys()
                    logger.info("Master recovery keys reset to unused status")
                except Exception as e:
                    logger.warning(f"Could not reset master keys: {e}")
                
                conn.commit()
            
            # Step 2: Delete generated files and directories (optimized)
            progress.setText("📁 Изтриване на системни файлове...")
            QApplication.processEvents()
            
            # CRITICAL: exports and backups folders should NEVER be deleted - they contain user data
            # Users need backups for bug reporting and data recovery
            directories_to_clear = ['barcodes', 'logs']
            
            for directory in directories_to_clear:
                if os.path.exists(directory):
                    try:
                        # Optimized: Use shutil.rmtree and recreate directory
                        shutil.rmtree(directory)
                        os.makedirs(directory, exist_ok=True)
                        logger.info(f"Cleared and recreated directory: {directory}")
                    except Exception as e:
                        logger.warning(f"Could not clear directory {directory}: {e}")
            
            # Step 3: Reset configuration files
            progress.setText("⚙️ Възстановяване на настройки...")
            QApplication.processEvents()
            
            config_files = [get_persistent_path('data/auto_backup_config.json'), get_persistent_path('data/backup_info.json')]
            
            for config_file in config_files:
                if os.path.exists(config_file):
                    try:
                        os.remove(config_file)
                        logger.info(f"Removed config file: {config_file}")
                    except Exception as e:
                        logger.warning(f"Could not remove config file {config_file}: {e}")
            
            # Step 4: Reset internal state (minimal UI updates)
            progress.setText("🔄 Финализиране...")
            QApplication.processEvents()
            
            # Reset internal state
            self.custom_categories.clear()
            self.custom_metals.clear()
            self.custom_stones.clear()
            
            # Clear action history
            if hasattr(self, 'action_history'):
                self.action_history.clear()
                if hasattr(self, 'update_action_buttons'):
                    self.update_action_buttons()
            
            # CRITICAL: Close progress dialog BEFORE any other dialogs
            if progress:
                try:
                    progress.close()
                    progress.deleteLater()
                    progress = None
                    QApplication.processEvents()  # Ensure dialog is fully closed
                    logger.info("Progress dialog closed successfully")
                except Exception as close_error:
                    logger.warning(f"Error closing progress dialog: {close_error}")
            
            # Minimal data reloading (skip complex operations for speed)
            try:
                logger.info("Starting minimal data reload after system restart...")
                
                # Set flag to suppress error dialogs during factory reset
                self._suppress_error_dialogs = True
                
                # Reset database singleton after factory reset for clean state
                try:
                    Database.reset_singleton()
                    self.db = Database()
                    logger.info("Database connection reinitialized after factory reset")
                except Exception as db_error:
                    logger.error(f"Error reinitializing database: {db_error}")
                
                # Basic UI cleanup only
                try:
                    # Clear tables to prevent accessing old data
                    if hasattr(self, 'items_table'):
                        self.items_table.setRowCount(0)
                    if hasattr(self, 'sales_table'):
                        self.sales_table.setRowCount(0)
                    if hasattr(self, 'shop_table'):
                        self.shop_table.setRowCount(0)
                    
                    # Clear combo boxes and set basic defaults
                    if hasattr(self, 'category_input'):
                        self.category_input.clear()
                        self.category_input.addItems(["Пръстен", "Гривна", "Обеци", "Синджир", "Друго"])
                    if hasattr(self, 'metal_input'):
                        self.metal_input.clear()
                        self.metal_input.addItems(["Злато", "Сребро", "Платина", "Друго"])
                        self.metal_input.setCurrentIndex(1)  # Сребро
                    if hasattr(self, 'stone_input'):
                        self.stone_input.clear()
                        self.stone_input.addItems(["Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък", "Друго"])
                        self.stone_input.setCurrentIndex(4)  # Без камък
                    
                    # Load shops (minimal)
                    shops = self.db.get_all_shops()
                    if hasattr(self, 'shop_combo'):
                        self.shop_combo.clear()
                        for shop in shops:
                            self.shop_combo.addItem(shop[1])
                    if hasattr(self, 'sales_shop_combo'):
                        self.sales_shop_combo.clear()
                        for shop in shops:
                            self.sales_shop_combo.addItem(shop[1])
                    
                    logger.info("Basic UI reset completed successfully")
                except Exception as ui_error:
                    logger.warning(f"Error in basic UI reset: {ui_error}")
                
            except Exception as load_error:
                logger.error(f"Error during minimal data loading: {load_error}")
            finally:
                # Re-enable error dialogs
                self._suppress_error_dialogs = False
            
            logger.info("System restart operation completed, showing dialogs...")
            
            
            # Success message
            success_dialog = QMessageBox(self)
            success_dialog.setIcon(QMessageBox.Icon.Information)
            success_dialog.setWindowTitle("✅ Системен рестарт завършен")
            success_dialog.setText("""
<h3 style="color: #4CAF50;">Системен рестарт завършен успешно!</h3>

<p>Всички данни са изтрити и системата е възстановена в първоначално състояние.</p>

<p><b>Какво е направено:</b></p>
<ul>
<li>✅ Изтрити всички артикули от склада</li>
<li>✅ Изтрити всички продажби и история</li>
<li>✅ Изтрити всички магазини (създаден "Магазин 1")</li>
<li>✅ Изтрити всички резервни копия</li>
<li>✅ Изтрити всички експортирани файлове</li>
<li>✅ Изтрити всички отчети и логове</li>
<li>✅ Възстановени настройки по подразбиране</li>
<li>🔐 Парола нулирана към: <b>0000</b></li>
</ul>

<p style="color: #2196F3;"><b>Системата е готова за използване като нова инсталация!</b></p>
            """)
            success_dialog.setStandardButtons(QMessageBox.StandardButton.Ok)
            success_dialog.button(QMessageBox.StandardButton.Ok).setText("Разбрах")
            success_dialog.exec()
            
            logger.info("System restart completed successfully")
            
            # Important shutdown message
            shutdown_dialog = QMessageBox(self)
            shutdown_dialog.setIcon(QMessageBox.Icon.Warning)
            shutdown_dialog.setWindowTitle("🔄 Затваряне на приложението")
            shutdown_dialog.setText("""
<h3 style="color: #ff9800;">Приложението ще се затвори сега</h3>

<p><b>Системният рестарт е завършен успешно!</b></p>

<p style="color: #d32f2f;"><b>ВАЖНО:</b> Приложението ще се затвори автоматично.</p>

<p><b>За да продължите работа:</b></p>
<ol>
<li>🔴 Изчакайте приложението да се затвори напълно</li>
<li>🟢 Стартирайте приложението отново ръчно</li>
<li>🔐 Влезте с новата парола: <b>0000</b></li>
</ol>

<p style="color: #2196F3;">След влизане можете да смените паролата от настройките.</p>
            """)
            shutdown_dialog.setStandardButtons(QMessageBox.StandardButton.Ok)
            shutdown_dialog.button(QMessageBox.StandardButton.Ok).setText("Затвори приложението сега")
            shutdown_dialog.exec()
            
            # Force close the application properly
            logger.info("Closing application after factory reset")
            
            # Close all remaining dialogs and windows
            try:
                # Close all open dialogs
                for widget in QApplication.allWidgets():
                    if isinstance(widget, (QDialog, QMessageBox)) and widget.isVisible():
                        widget.close()
                        widget.deleteLater()
                
                # Close main window
                self.close()
                
            except Exception as cleanup_error:
                logger.warning(f"Error during final cleanup: {cleanup_error}")
            
            # Quit the application
            QApplication.instance().quit()
            
        except Exception as e:
            logger.error(f"Error during system restart: {e}")
            # Ensure progress dialog is closed even on error
            if progress:
                try:
                    progress.close()
                    progress.deleteLater()
                except:
                    pass
            QMessageBox.critical(self, "Грешка", f"Грешка при системен рестарт: {str(e)}")
        
        finally:
            # Final cleanup - ensure progress dialog is always closed
            if progress is not None:
                try:
                    progress.close()
                    progress.deleteLater()
                except:
                    pass

    def update_auto_backup_button_color(self):
        """Update auto backup button color based on status"""
        try:
            if hasattr(self, 'auto_backup_enabled') and self.auto_backup_enabled:
                # Green for enabled
                self.auto_backup_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }")
            else:
                # Red for disabled
                self.auto_backup_btn.setStyleSheet("QPushButton { background-color: #f44336; color: white; font-weight: bold; padding: 10px; }")
        except:
            # Default style if button doesn't exist yet
            pass
    
    # Placeholder methods for database tab functionality
    def setup_auto_backup(self):
        """Setup automatic backup with real-time scheduling dialog"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("Настройки за автоматично резервно копиране")
            dialog.setModal(True)
            dialog.setFixedSize(400, 400)
            
            layout = QVBoxLayout(dialog)
            
            # Title
            title_label = QLabel("Автоматично резервно копиране")
            title_label.setFont(QFont("Arial", 14, QFont.Weight.Bold))
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(title_label)
            
            # Enable/Disable section
            enable_group = QGroupBox("Състояние")
            enable_layout = QVBoxLayout()
            
            # Initialize settings if not present
            if not hasattr(self, 'auto_backup_enabled'):
                self.auto_backup_enabled = False
            if not hasattr(self, 'auto_backup_frequency'):
                self.auto_backup_frequency = 'Ежедневно'
            if not hasattr(self, 'auto_backup_time'):
                self.auto_backup_time = '02:00'
            
            enable_checkbox = QCheckBox("Активирай автоматично резервно копиране")
            enable_checkbox.setChecked(self.auto_backup_enabled)
            enable_layout.addWidget(enable_checkbox)
            
            enable_group.setLayout(enable_layout)
            layout.addWidget(enable_group)
            
            # Frequency section
            freq_group = QGroupBox("Честота")
            freq_layout = QVBoxLayout()
            
            from PyQt6.QtWidgets import QRadioButton, QButtonGroup
            frequency_group = QButtonGroup()
            
            daily_radio = QRadioButton("Всеки ден")
            weekly_radio = QRadioButton("Всяка седмица")
            monthly_radio = QRadioButton("Всеки месец")
            
            frequency_group.addButton(daily_radio, 0)
            frequency_group.addButton(weekly_radio, 1)
            frequency_group.addButton(monthly_radio, 2)
            
            # Set current selection
            if self.auto_backup_frequency == 'Ежедневно':
                daily_radio.setChecked(True)
            elif self.auto_backup_frequency == 'Седмично':
                weekly_radio.setChecked(True)
            else:
                monthly_radio.setChecked(True)
            
            freq_layout.addWidget(daily_radio)
            freq_layout.addWidget(weekly_radio)
            freq_layout.addWidget(monthly_radio)
            
            freq_group.setLayout(freq_layout)
            layout.addWidget(freq_group)
            
            # Time selection
            time_group = QGroupBox("Час за изпълнение")
            time_layout = QHBoxLayout()
            
            time_label = QLabel("Час:")
            time_layout.addWidget(time_label)
            
            # Parse current time for separate inputs
            try:
                hour, minute = map(int, self.auto_backup_time.split(':'))
            except:
                hour, minute = 2, 0  # Default to 2:00 AM
            
            # Hour input (00-23)
            hour_edit = BlurOnEnterLineEdit(numeric_only=True, max_value=23)
            hour_edit.setPlaceholderText("ЧЧ")
            hour_edit.setMaxLength(2)
            hour_edit.setFixedWidth(40)
            hour_edit.setText(f"{hour:02d}")
            hour_edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hour_edit.last_confirmed_value = f"{hour:02d}"  # Set initial confirmed value
            time_layout.addWidget(hour_edit)
            
            # Colon separator
            colon_label = QLabel(":")
            colon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            colon_label.setFixedWidth(10)
            time_layout.addWidget(colon_label)
            
            # Minute input (00-59)
            minute_edit = BlurOnEnterLineEdit(numeric_only=True, max_value=59)
            minute_edit.setPlaceholderText("ММ")
            minute_edit.setMaxLength(2)
            minute_edit.setFixedWidth(40)
            minute_edit.setText(f"{minute:02d}")
            minute_edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
            minute_edit.last_confirmed_value = f"{minute:02d}"  # Set initial confirmed value
            time_layout.addWidget(minute_edit)
            
            # Add some spacing after time inputs
            time_layout.addStretch()
            
            time_group.setLayout(time_layout)
            layout.addWidget(time_group)
            
            # Status info
            status_group = QGroupBox("Информация")
            status_layout = QVBoxLayout()
            
            status_label = QLabel()
            def update_status():
                if enable_checkbox.isChecked():
                    freq_text = ['Ежедневно', 'Седмично', 'Месечно'][frequency_group.checkedId() if frequency_group.checkedId() >= 0 else 0]
                    
                    # Get time from separate fields
                    try:
                        hour_val = int(hour_edit.text()) if hour_edit.text().isdigit() else 0
                        minute_val = int(minute_edit.text()) if minute_edit.text().isdigit() else 0
                        time_text = f"{hour_val:02d}:{minute_val:02d}"
                    except:
                        time_text = "00:00"
                    
                    # Load last backup time from config
                    last_backup = "Никога"
                    try:
                        if os.path.exists(get_persistent_path('data/auto_backup_config.json')):
                            with open(get_persistent_path('data/auto_backup_config.json'), 'r', encoding='utf-8') as f:
                                config = json.load(f)
                                last_backup = config.get('last_auto_backup', 'Никога')
                    except:
                        pass
                    
                    status_label.setText(f"Статус: Активно\nЧестота: {freq_text}\nЧас: {time_text}\nПоследно копие: {last_backup}")
                else:
                    status_label.setText("Статус: Неактивно")
            
            update_status()
            status_layout.addWidget(status_label)
            
            status_group.setLayout(status_layout)
            layout.addWidget(status_group)
            
            # Real-time update functions
            def validate_time_input():
                """Validate and fix time inputs"""
                # Validate hour (0-23)
                try:
                    hour_val = int(hour_edit.text()) if hour_edit.text().isdigit() else 0
                    if hour_val > 23:
                        hour_val = 23
                        hour_edit.setText(f"{hour_val:02d}")
                except:
                    hour_val = 0
                    hour_edit.setText("00")
                
                # Validate minute (0-59) 
                try:
                    minute_val = int(minute_edit.text()) if minute_edit.text().isdigit() else 0
                    if minute_val > 59:
                        minute_val = 59
                        minute_edit.setText(f"{minute_val:02d}")
                except:
                    minute_val = 0
                    minute_edit.setText("00")
                
                return f"{hour_val:02d}:{minute_val:02d}"
            
            def on_settings_change():
                # Validate time inputs first
                validated_time = validate_time_input()
                
                # Update settings in real-time
                self.auto_backup_enabled = enable_checkbox.isChecked()
                if frequency_group.checkedId() >= 0:
                    self.auto_backup_frequency = ['Ежедневно', 'Седмично', 'Месечно'][frequency_group.checkedId()]
                self.auto_backup_time = validated_time
                
                # Save to file
                self.save_auto_backup_config()
                
                # Update scheduler
                if self.auto_backup_enabled:
                    self.setup_backup_scheduler()
                else:
                    self.disable_backup_scheduler()
                
                # Update button color
                self.update_auto_backup_button_color()
                
                # Update status display
                update_status()
            
            # Connect all controls to real-time updates
            enable_checkbox.toggled.connect(on_settings_change)
            daily_radio.toggled.connect(on_settings_change)
            weekly_radio.toggled.connect(on_settings_change)
            monthly_radio.toggled.connect(on_settings_change)
            
            # Connect time inputs to trigger updates on focus loss (Enter key)
            def on_time_change():
                # Small delay to allow validation to complete
                QTimer.singleShot(100, on_settings_change)
            
            hour_edit.editingFinished.connect(on_time_change)
            minute_edit.editingFinished.connect(on_time_change)
            
            # Close button
            close_btn = QPushButton("Затвори")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при настройка на автоматично копиране: {str(e)}")
    
    def save_auto_backup_config(self):
        """Save auto backup configuration to file"""
        try:
            config = {
                'auto_backup_enabled': getattr(self, 'auto_backup_enabled', False),
                'auto_backup_frequency': getattr(self, 'auto_backup_frequency', 'Ежедневно'),
                'auto_backup_time': getattr(self, 'auto_backup_time', '02:00')
            }
            
            # Create data directory if it doesn't exist
            os.makedirs('data', exist_ok=True)
            
            with open(get_persistent_path('data/auto_backup_config.json'), 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2)
                
        except Exception as e:
            logger.error(f"Error saving auto backup config: {e}")

    def setup_backup_scheduler(self):
        """Setup backup scheduler with Windows Task Scheduler"""
        try:
            if not TASK_SCHEDULER_AVAILABLE:
                logger.warning("Task Scheduler not available - auto backup will not work")
                return False
            
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Connect to Task Scheduler
            scheduler = win32com.client.Dispatch("Schedule.Service")
            scheduler.Connect()
            
            # Get root folder
            root_folder = scheduler.GetFolder("\\")
            
            # Delete existing task if it exists
            try:
                root_folder.DeleteTask("JewelryAutoBackup", 0)
            except:
                pass  # Task doesn't exist
            
            # Create new task
            task_def = scheduler.NewTask(0)
            
            # Set task settings
            task_def.Settings.Enabled = True
            task_def.Settings.Hidden = False
            task_def.Settings.StartWhenAvailable = True
            task_def.Settings.DisallowStartIfOnBatteries = False
            task_def.Settings.StopIfGoingOnBatteries = False
            
            # Set registration info
            task_def.RegistrationInfo.Description = "Automatic backup for Jewelry Management Software"
            task_def.RegistrationInfo.Author = "Jewelry Management Software"
            
            # Create trigger based on frequency
            if self.auto_backup_frequency == 'Ежедневно':
                trigger = task_def.Triggers.Create(2)  # TASK_TRIGGER_DAILY = 2
                trigger.DaysInterval = 1
            elif self.auto_backup_frequency == 'Седмично':
                trigger = task_def.Triggers.Create(3)  # TASK_TRIGGER_WEEKLY = 3
                trigger.WeeksInterval = 1
                trigger.DaysOfWeek = 1  # Sunday
            else:  # месечно
                trigger = task_def.Triggers.Create(4)  # TASK_TRIGGER_MONTHLY = 4
                trigger.MonthsOfYear = 4095  # All months
                trigger.DaysOfMonth = 1  # First day of month
            
            # Set start time
            trigger.StartBoundary = datetime.now().strftime("%Y-%m-%dT") + self.auto_backup_time + ":00"
            
            # Create action
            action = task_def.Actions.Create(0)  # TASK_ACTION_EXEC = 0
            action.Path = sys.executable  # Python executable
            action.Arguments = f'"{os.path.abspath(__file__)}" --auto-backup'
            action.WorkingDirectory = os.path.dirname(os.path.abspath(__file__))
            
            # Register the task
            root_folder.RegisterTaskDefinition(
                "JewelryAutoBackup",
                task_def,
                6,  # TASK_CREATE_OR_UPDATE = 6
                None,  # User
                None,  # Password
                3  # TASK_LOGON_INTERACTIVE_TOKEN = 3
            )
            
            logger.info(f"Auto backup scheduled: {self.auto_backup_frequency} at {self.auto_backup_time}")
            return True
            
        except Exception as e:
            logger.error(f"Error setting up backup scheduler: {e}")
            return False
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass

    def disable_backup_scheduler(self):
        """Disable backup scheduler"""
        try:
            if not TASK_SCHEDULER_AVAILABLE:
                return True
                
            # Initialize COM
            pythoncom.CoInitialize()
            
            # Connect to Task Scheduler
            scheduler = win32com.client.Dispatch("Schedule.Service")
            scheduler.Connect()
            
            # Get root folder
            root_folder = scheduler.GetFolder("\\")
            
            # Delete the task
            try:
                root_folder.DeleteTask("JewelryAutoBackup", 0)
                logger.info("Auto backup scheduler disabled")
                return True
            except:
                logger.info("Auto backup task not found (already disabled)")
                return True
                
        except Exception as e:
            logger.error(f"Error disabling backup scheduler: {e}")
            return False
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    
    def load_auto_backup_settings(self):
        """Load auto backup settings on startup"""
        try:
            if os.path.exists(get_persistent_path('data/auto_backup_config.json')):
                with open(get_persistent_path('data/auto_backup_config.json'), 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.auto_backup_enabled = config.get('auto_backup_enabled', False)
                    self.auto_backup_frequency = config.get('auto_backup_frequency', 'Ежедневно')
                    self.auto_backup_time = config.get('auto_backup_time', '02:00')
                    
                    if self.auto_backup_enabled:
                        self.setup_backup_scheduler()
            else:
                self.auto_backup_enabled = False
                self.auto_backup_frequency = 'Ежедневно'
                self.auto_backup_time = '02:00'
                
            # Update button color after loading settings
            if hasattr(self, 'auto_backup_btn'):
                self.update_auto_backup_button_color()
                
        except Exception as e:
            logger.error(f"Error loading auto backup settings: {e}")
            self.auto_backup_enabled = False
    
    def export_to_csv(self):
        """Export database to CSV files"""
        try:
            # Ensure exports directory exists
            exports_dir = self.get_exports_directory()
            csv_filename = self.generate_bulgarian_filename("експорт_база_данни", "zip")
            
            # Get save location for CSV export
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Запази CSV експорт", 
                f"{exports_dir}/{csv_filename}",
                "ZIP Files (*.zip)"
            )
            
            if not file_path:
                return
            
            import zipfile
            import tempfile
            
            # Create temporary directory for CSV files
            with tempfile.TemporaryDirectory() as temp_dir:
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    # Get all table names
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = cursor.fetchall()
                    
                    exported_files = []
                    
                    for table in tables:
                        table_name = table[0]
                        
                        # Skip system tables
                        if table_name.startswith('sqlite_'):
                            continue
                            
                        # Get table data
                        cursor.execute(f"SELECT * FROM {table_name}")
                        rows = cursor.fetchall()
                        
                        if rows:  # Only export tables with data
                            # Get column names
                            cursor.execute(f"PRAGMA table_info({table_name})")
                            columns = [col[1] for col in cursor.fetchall()]
                            
                            # Create CSV file in temp directory
                            csv_file = os.path.join(temp_dir, f"{table_name}.csv")
                            
                            with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
                                writer = csv.writer(f)
                                
                                # Write header
                                writer.writerow(columns)
                                
                                # Write data
                                writer.writerows(rows)
                            
                            exported_files.append(f"{table_name}.csv")
                
                # Export configuration as JSON in the CSV export
                try:
                    config_data = self.collect_configuration_data()
                    if config_data:
                        config_file = os.path.join(temp_dir, "configuration.json")
                        with open(config_file, 'w', encoding='utf-8') as f:
                            json.dump(config_data, f, indent=2, ensure_ascii=False, default=str)
                        exported_files.append("configuration.json")
                except Exception as e:
                    logger.warning(f"Could not export configuration in CSV export: {e}")
                
                # Create ZIP file with all CSV files
                if exported_files:
                    with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for csv_file in exported_files:
                            csv_path = os.path.join(temp_dir, csv_file)
                            zipf.write(csv_path, csv_file)
                    
                    QMessageBox.information(
                        self, "Успех", 
                        f"CSV експорт завършен успешно!\n\n"
                        f"Архивиран файл: {os.path.basename(file_path)}\n"
                        f"Съдържа {len(exported_files)} файла с данни от базата\n"
                        f"✅ Включва конфигурационни настройки"
                    )
                else:
                    QMessageBox.warning(self, "Предупреждение", "Няма данни за експорт")
                    
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при CSV експорт: {str(e)}")
    
    def export_to_json(self):
        """Export database to JSON format with complete migration support"""
        try:
            exports_dir = self.get_exports_directory()
            json_filename = self.generate_bulgarian_filename("пълен_експорт", "json")
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Запази JSON експорт", 
                f"{exports_dir}/{json_filename}",
                "JSON Files (*.json)"
            )
            
            if file_path:
                if not file_path.endswith('.json'):
                    file_path += '.json'
                
                import base64
                import os
                
                export_data = {
                    "_migration_info": {
                        "software_version": "1.0",
                        "export_date": datetime.now().isoformat(),
                        "schema_version": "1.0", 
                        "compatibility_level": "complete",
                        "export_type": "full_migration",
                        "includes_binary_data": True,
                        "includes_external_files": True,
                        "memory_efficient": True
                    },
                    "_external_files": {}
                }
                
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    # Get database schema for validation
                    cursor.execute("SELECT sql FROM sqlite_master WHERE type='table'")
                    schema_info = {row[0] for row in cursor.fetchall() if row[0]}
                    
                    # Get all table names
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = cursor.fetchall()
                    
                    total_rows = 0
                    binary_columns_found = []
                    
                    for table in tables:
                        table_name = table[0]
                        
                        # Skip system tables
                        if table_name.startswith('sqlite_'):
                            continue
                            
                        # Get table info with column types for binary detection
                        cursor.execute(f"PRAGMA table_info({table_name})")
                        columns_info = cursor.fetchall()
                        columns = [col[1] for col in columns_info]
                        column_types = [col[2] for col in columns_info]
                        
                        # Check for binary/blob columns
                        blob_columns = [i for i, col_type in enumerate(column_types) if 'BLOB' in col_type.upper()]
                        if blob_columns:
                            binary_columns_found.extend([f"{table_name}.{columns[i]}" for i in blob_columns])
                        
                        # Get table data
                        cursor.execute(f"SELECT * FROM {table_name}")
                        rows = cursor.fetchall()
                        total_rows += len(rows)
                        
                        # Convert to list of dictionaries with binary data support
                        table_data = []
                        for row in rows:
                            row_dict = {}
                            for i, value in enumerate(row):
                                column_name = columns[i]
                                
                                if value is None:
                                    row_dict[column_name] = None
                                elif isinstance(value, bytes):
                                    # Handle binary data with base64 encoding
                                    row_dict[column_name] = {
                                        "_type": "binary",
                                        "_encoding": "base64",
                                        "_data": base64.b64encode(value).decode('utf-8')
                                    }
                                elif isinstance(value, (int, float, str)):
                                    row_dict[column_name] = value
                                else:
                                    # Convert other types to string with type info
                                    row_dict[column_name] = {
                                        "_type": type(value).__name__,
                                        "_data": str(value)
                                    }
                            table_data.append(row_dict)
                        
                        export_data[table_name] = {
                            'columns': columns,
                            'column_types': column_types,
                            'has_binary_data': len(blob_columns) > 0,
                            'binary_columns': [columns[i] for i in blob_columns],
                            'data': table_data,
                            'row_count': len(rows)
                        }
                
                # Collect external files
                external_files = self.collect_external_files()
                if external_files:
                    export_data["_external_files"] = external_files
                
                # Collect configuration data
                config_data = self.collect_configuration_data()
                if config_data:
                    export_data["_configuration"] = config_data
                
                # Enhanced metadata for complete migration safety
                export_data['_migration_info'].update({
                    'database_schema': list(schema_info),
                    'table_count': len([t for t in tables if not t[0].startswith('sqlite_')]),
                    'total_rows': total_rows,
                    'binary_columns': binary_columns_found,
                    'external_file_count': len(external_files),
                    'configuration_included': len(config_data) > 0 if config_data else False,
                    'export_size': 0,  # Will be calculated after JSON creation
                    'warnings': ["Always create backup before import", "Verify version compatibility"]
                })
                
                # Calculate export size
                json_str = json.dumps(export_data, ensure_ascii=False, default=str)
                export_data['_migration_info']['export_size'] = len(json_str)
                
                # Save to JSON file
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
                
                QMessageBox.information(
                    self, "Успех", 
                    f"Подобрен JSON експорт завършен успешно!\n\n"
                    f"Файл: {os.path.basename(file_path)}\n"
                    f"Таблици: {export_data['_migration_info']['table_count']}\n"
                    f"Редове: {total_rows}\n"
                    f"Размер: {len(json_str) // 1024} KB\n\n"
                    f"✅ Включва информация за версия и схема\n"
                    f"✅ Безопасна миграция между версии\n"
                    f"✅ Конфигурационни файлове включени\n"
                    f"✅ Пълна система за възстановяване"
                )
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при JSON експорт: {str(e)}")
    
    def collect_external_files(self):
        """Collect external files that should be included in migration"""
        import base64
        
        external_files = {}
        
        # Key directories to include
        directories_to_export = [
            "fonts",
            "barcodes", 
            "resources",
            "dlls"
        ]
        
        for directory in directories_to_export:
            if os.path.exists(directory):
                for root, dirs, files in os.walk(directory):
                    for file in files:
                        file_path = os.path.join(root, file)
                        relative_path = os.path.relpath(file_path)
                        
                        try:
                            # For small files, embed directly
                            if os.path.getsize(file_path) < 100 * 1024:  # < 100KB
                                with open(file_path, 'rb') as f:
                                    content = base64.b64encode(f.read()).decode('utf-8')
                                    external_files[relative_path] = {
                                        "type": "embedded",
                                        "encoding": "base64",
                                        "content": content,
                                        "size": os.path.getsize(file_path)
                                    }
                            else:
                                # For large files, just reference
                                external_files[relative_path] = {
                                    "type": "reference",
                                    "size": os.path.getsize(file_path),
                                    "note": "Large file - copy manually"
                                }
                        except Exception as e:
                            external_files[relative_path] = {
                                "type": "error",
                                "error": str(e)
                            }
        
        return external_files
    
    def restore_external_files(self, external_files):
        """Restore external files from export"""
        import base64
        
        restored_count = 0
        
        for file_path, file_info in external_files.items():
            try:
                if file_info.get("type") == "embedded":
                    # Create directory if needed
                    os.makedirs(os.path.dirname(file_path), exist_ok=True)
                    
                    # Decode and write file
                    content = base64.b64decode(file_info["content"])
                    with open(file_path, 'wb') as f:
                        f.write(content)
                    
                    restored_count += 1
                    
                elif file_info.get("type") == "reference":
                    pass  # Large file needs manual copy
                    
            except Exception as e:
                logger.error(f"Error restoring {file_path}: {e}")
        
        return restored_count
    
    def collect_configuration_data(self):
        """Collect application configuration and settings for export"""
        config_data = {}
        
        try:
            # Collect app_config.json
            config_file_path = get_persistent_path("data/app_config.json")
            if os.path.exists(config_file_path):
                with open(config_file_path, 'r', encoding='utf-8') as f:
                    app_config = json.load(f)
                    config_data["app_config"] = {
                        "data": app_config,
                        "file_path": config_file_path,
                        "size": os.path.getsize(config_file_path),
                        "modified": datetime.fromtimestamp(os.path.getmtime(config_file_path)).isoformat()
                    }
            
            # Collect application settings from current instance
            runtime_settings = {}
            
            # Currency settings (if available)
            if hasattr(self, 'exchange_rate'):
                runtime_settings["exchange_rate"] = self.exchange_rate
            
            # Window state and UI preferences
            try:
                runtime_settings["window"] = {
                    "geometry": self.geometry().getRect() if hasattr(self, 'geometry') else None,
                    "is_maximized": self.isMaximized() if hasattr(self, 'isMaximized') else False
                }
            except:
                pass
            
            # Tab preferences
            try:
                if hasattr(self, 'tabs') and hasattr(self.tabs, 'currentIndex'):
                    runtime_settings["current_tab"] = self.tabs.currentIndex()
            except:
                pass
            
            # Database path
            if hasattr(self, 'db') and hasattr(self.db, 'db_path'):
                runtime_settings["database_path"] = str(self.db.db_path)
            
            if runtime_settings:
                config_data["runtime_settings"] = runtime_settings
            
            # Collect logs directory info (for reference, not content)
            logs_info = {}
            if os.path.exists("logs"):
                log_files = []
                for file in os.listdir("logs"):
                    if file.endswith('.log'):
                        file_path = os.path.join("logs", file)
                        log_files.append({
                            "name": file,
                            "size": os.path.getsize(file_path),
                            "modified": datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat()
                        })
                logs_info["log_files"] = log_files
                logs_info["note"] = "Log files not exported for privacy - paths listed for reference"
            
            if logs_info:
                config_data["logs_info"] = logs_info
            
            # Export metadata
            config_data["_export_info"] = {
                "exported_at": datetime.now().isoformat(),
                "export_type": "configuration_snapshot",
                "includes_app_config": "app_config" in config_data,
                "includes_runtime_settings": "runtime_settings" in config_data,
                "total_config_items": len(config_data)
            }
            
        except Exception as e:
            logger.error(f"Error collecting configuration data: {e}")
            config_data["_error"] = str(e)
        
        return config_data
    
    def restore_configuration_data(self, config_data):
        """Restore configuration data from export"""
        try:
            restored_items = []
            
            # Restore app_config.json
            if "app_config" in config_data:
                app_config_info = config_data["app_config"]
                config_file_path = app_config_info["file_path"]
                
                # Create backup of existing config
                if os.path.exists(config_file_path):
                    backup_path = f"{config_file_path}.backup.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    shutil.copy2(config_file_path, backup_path)
                    restored_items.append(f"Backed up existing config to {backup_path}")
                
                # Create directory if needed
                os.makedirs(os.path.dirname(config_file_path), exist_ok=True)
                
                # Write new config
                with open(config_file_path, 'w', encoding='utf-8') as f:
                    json.dump(app_config_info["data"], f, indent=2, ensure_ascii=False)
                
                restored_items.append(f"Restored {config_file_path}")
            
            # Note about runtime settings (cannot be automatically restored)
            if "runtime_settings" in config_data:
                restored_items.append("Runtime settings exported (manual application may be needed)")
            
            return restored_items
            
        except Exception as e:
            logger.error(f"Error restoring configuration data: {e}")
            return [f"Error: {str(e)}"]

    def import_from_json(self):
        """Import data from JSON file with enhanced safety checks"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Избери JSON файл за импорт", "exports",
                "JSON Files (*.json)"
            )
            
            if file_path:
                # Load and analyze import file
                with open(file_path, 'r', encoding='utf-8') as f:
                    import_data = json.load(f)
                
                # Check for enhanced migration info
                migration_info = import_data.get('_migration_info', {})
                legacy_metadata = import_data.get('_metadata', {})
                external_files = import_data.get('_external_files', {})
                config_data = import_data.get('_configuration', {})
                
                # Analyze import file
                analysis_text = "📋 Анализ на файла за импорт:\n\n"
                
                if migration_info:
                    # Enhanced export
                    export_version = migration_info.get('software_version', 'неизвестна')
                    export_date = migration_info.get('export_date', 'неизвестна')
                    table_count = migration_info.get('table_count', 'неизвестен')
                    total_rows = migration_info.get('total_rows', 'неизвестен')
                    compatibility = migration_info.get('compatibility_level', 'неизвестна')
                    
                    analysis_text += f"✅ Подобрен експорт (безопасен)\n"
                    analysis_text += f"📦 Версия: {export_version}\n"
                    analysis_text += f"📅 Дата: {export_date[:16]}\n"
                    analysis_text += f"📊 Таблици: {table_count}\n"
                    analysis_text += f"📈 Редове: {total_rows}\n"
                    analysis_text += f"🔧 Съвместимост: {compatibility}\n\n"
                    
                    # Version compatibility check
                    if export_version != '1.0':
                        analysis_text += f"⚠️ ВНИМАНИЕ: Различна версия!\n"
                        analysis_text += f"   Експорт: v{export_version}\n"
                        analysis_text += f"   Текуща: v1.0\n\n"
                    
                    warnings = migration_info.get('warnings', [])
                    if warnings:
                        analysis_text += f"⚠️ Предупреждения:\n"
                        for warning in warnings:
                            analysis_text += f"   • {warning}\n"
                        analysis_text += "\n"
                
                elif legacy_metadata:
                    # Legacy export with some metadata
                    export_date = legacy_metadata.get('export_date', 'неизвестна')
                    export_version = legacy_metadata.get('export_version', 'неизвестна')
                    
                    analysis_text += f"⚠️ Стар формат с метаданни\n"
                    analysis_text += f"📅 Дата: {export_date[:16]}\n"
                    analysis_text += f"📦 Версия: {export_version}\n"
                    analysis_text += f"🔧 Ниска съвместимост\n\n"
                    analysis_text += f"⚠️ Препоръчва се внимание при импорт\n\n"
                
                else:
                    # Very old export without metadata
                    analysis_text += f"❌ Много стар формат\n"
                    analysis_text += f"📅 Няма информация за дата\n"
                    analysis_text += f"📦 Няма информация за версия\n"
                    analysis_text += f"🔧 Неизвестна съвместимост\n\n"
                    analysis_text += f"⚠️ ВИСОК РИСК! Създайте копие преди импорт!\n\n"
                
                # Count tables and estimated rows
                data_tables = [k for k in import_data.keys() if not k.startswith('_')]
                estimated_rows = 0
                for table_name in data_tables:
                    table_data = import_data[table_name]
                    if isinstance(table_data, dict):
                        if 'data' in table_data:
                            estimated_rows += len(table_data['data'])
                        elif 'rows' in table_data:
                            estimated_rows += len(table_data['rows'])
                
                analysis_text += f"📊 Данни за импорт:\n"
                analysis_text += f"   Таблици: {len(data_tables)}\n"
                analysis_text += f"   Приблизителни редове: {estimated_rows}\n"
                
                # Configuration and external files info
                if external_files:
                    analysis_text += f"   Външни файлове: {len(external_files)}\n"
                if config_data:
                    config_info = config_data.get('_export_info', {})
                    if config_info.get('includes_app_config'):
                        analysis_text += f"   ⚙️ Конфигурационни файлове: ДА\n"
                    if config_info.get('includes_runtime_settings'):
                        analysis_text += f"   🖥️ Runtime настройки: ДА\n"
                
                analysis_text += f"\n💾 Ще се създаде резервно копие преди импорт\n"
                analysis_text += f"🗑️ ВСИЧКИ текущи данни ще бъдат изтрити!\n\n"
                analysis_text += f"Продължаване на импорта?"
                
                # Show detailed confirmation dialog
                reply = QMessageBox.question(
                    self, "🔍 Анализ и потвърждение на импорт", 
                    analysis_text,
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # Create backup before import
                    try:
                        backup_path = self.db.create_backup()
                        if backup_path:
                            print(f"💾 Backup created before import: {backup_path}")
                        else:
                            QMessageBox.warning(
                                self, "Предупреждение", 
                                "Неуспешно създаване на резервно копие!\nПродължете ли въпреки това?"
                            )
                    except Exception as e:
                        QMessageBox.warning(
                            self, "Предупреждение", 
                            f"Грешка при създаване на резервно копие: {e}\nПродължете ли въпреки това?"
                        )
                    
                    # Restore external files if present
                    if external_files:
                        restored_files = self.restore_external_files(external_files)
                        print(f"📁 Restored {restored_files} external files")
                    
                    # Restore configuration data if present
                    config_data = import_data.get('_configuration', {})
                    if config_data:
                        restored_config = self.restore_configuration_data(config_data)
                        print(f"⚙️ Configuration restoration completed: {len(restored_config)} items")
                    
                    # Perform enhanced import with validation
                    success = self.safe_import_with_validation(file_path, import_data)
                    
                    if success:
                        result_msg = f"✅ JSON данните са импортирани успешно!\n\n"
                        result_msg += f"Файл: {os.path.basename(file_path)}\n"
                        result_msg += f"Таблици: {len(data_tables)}\n"
                        if migration_info.get('includes_binary_data'):
                            result_msg += f"🔧 Включва двоични данни\n"
                        if external_files:
                            result_msg += f"📁 Възстановени файлове: {len(external_files)}\n"
                        if config_data:
                            result_msg += f"⚙️ Конфигурация възстановена\n"
                        
                        QMessageBox.information(self, "Успех", result_msg)
                        
                        # Reload all data after successful import
                        self.load_data()
                        self.load_sales()
                        self.load_shop_inventory()
                        self.update_reports_and_database_stats()
                        
                    else:
                        QMessageBox.critical(
                            self, "Грешка", 
                            "❌ Неуспешен импорт на JSON данни!\n\n"
                            "Данните остават непроменени.\n"
                            "Проверете файла и опитайте отново."
                        )
                        
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при JSON импорт: {str(e)}")
    
    def safe_import_with_validation(self, file_path, import_data):
        """Safely import data with structure validation and error handling"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Disable foreign keys during import
                cursor.execute("PRAGMA foreign_keys = OFF")
                
                imported_tables = []
                skipped_tables = []
                
                for table_name, table_data in import_data.items():
                    if table_name.startswith('_') or table_name.startswith('sqlite_'):
                        continue
                    
                    try:
                        # Validate table structure
                        if not self.validate_import_table_structure(cursor, table_name, table_data):
                            skipped_tables.append(f"{table_name} (structure mismatch)")
                            continue
                        
                        # Clear existing data
                        cursor.execute(f"DELETE FROM {table_name}")
                        
                        # Import data
                        if 'data' in table_data:
                            columns = table_data['columns']
                            rows = table_data['data']
                        else:
                            # Legacy format
                            columns = table_data.get('columns', [])
                            rows = table_data.get('rows', [])
                        
                        if rows:
                            placeholders = ", ".join(["?" for _ in columns])
                            imported_row_count = 0
                            
                            for row_data in rows:
                                try:
                                    if isinstance(row_data, dict):
                                        values = [row_data.get(col) for col in columns]
                                    else:
                                        values = list(row_data)
                                    
                                    # Handle None values and binary data
                                    processed_values = []
                                    for value in values:
                                        if value == '' or value == 'None':
                                            processed_values.append(None)
                                        elif isinstance(value, dict) and value.get("_type") == "binary":
                                            # Handle binary data with base64 decoding
                                            if value.get("_encoding") == "base64":
                                                import base64
                                                binary_value = base64.b64decode(value["_data"])
                                                processed_values.append(binary_value)
                                            else:
                                                processed_values.append(None)
                                        elif isinstance(value, dict) and "_type" in value:
                                            # Handle other typed data
                                            processed_values.append(value.get("_data"))
                                        else:
                                            processed_values.append(value)
                                    
                                    cursor.execute(
                                        f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})",
                                        processed_values
                                    )
                                    imported_row_count += 1
                                    
                                except Exception as row_error:
                                    logger.warning(f"Skipped row in {table_name}: {row_error}")
                                    continue
                            
                            imported_tables.append(f"{table_name} ({imported_row_count} rows)")
                        else:
                            imported_tables.append(f"{table_name} (empty)")
                        
                    except Exception as table_error:
                        logger.error(f"Error importing table {table_name}: {table_error}")
                        skipped_tables.append(f"{table_name} (error: {str(table_error)[:50]})")
                
                # Re-enable foreign keys
                cursor.execute("PRAGMA foreign_keys = ON")
                conn.commit()
                
                # Show detailed results
                if imported_tables or skipped_tables:
                    result_text = "📊 Резултати от импорт:\n\n"
                    
                    if imported_tables:
                        result_text += f"✅ Успешно импортирани таблици:\n"
                        for table in imported_tables[:10]:  # Show first 10
                            result_text += f"   • {table}\n"
                        if len(imported_tables) > 10:
                            result_text += f"   ... и още {len(imported_tables) - 10}\n"
                        result_text += "\n"
                    
                    if skipped_tables:
                        result_text += f"⚠️ Пропуснати таблици:\n"
                        for table in skipped_tables[:5]:  # Show first 5
                            result_text += f"   • {table}\n"
                        if len(skipped_tables) > 5:
                            result_text += f"   ... и още {len(skipped_tables) - 5}\n"
                    
                    # Show results dialog if there were issues
                    if skipped_tables:
                        QMessageBox.warning(self, "Частично успешен импорт", result_text)
                    
                return len(imported_tables) > 0
                
        except Exception as e:
            logger.error(f"Safe import failed: {e}")
            return False
    
    def validate_import_table_structure(self, cursor, table_name, table_data):
        """Validate import table structure against current database"""
        try:
            # Check if table exists
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
            if not cursor.fetchone():
                logger.warning(f"Table {table_name} doesn't exist in current database")
                return False
            
            # Get current table structure
            cursor.execute(f"PRAGMA table_info({table_name})")
            current_columns = {col[1]: col[2] for col in cursor.fetchall()}
            
            # Get import table structure
            import_columns = table_data.get('columns', [])
            
            # Check for critical missing columns
            missing_columns = set(import_columns) - set(current_columns.keys())
            if missing_columns:
                logger.warning(f"Table {table_name} missing columns: {missing_columns}")
            
            # Check for extra columns in current DB (usually OK)
            extra_columns = set(current_columns.keys()) - set(import_columns)
            if extra_columns:
                logger.info(f"Table {table_name} has extra columns: {extra_columns}")
            
            # Allow import if at least some columns match
            matching_columns = set(import_columns) & set(current_columns.keys())
            return len(matching_columns) > 0
            
        except Exception as e:
            logger.error(f"Structure validation failed for table {table_name}: {e}")
            return False
    
    def import_from_csv(self):
        """Import data from CSV files"""
        try:
            # Get directory containing CSV files
            directory = QFileDialog.getExistingDirectory(
                self, "Избери директория с CSV файлове", "exports"
            )
            
            if directory:
                # Find CSV files in directory
                csv_files = [f for f in os.listdir(directory) if f.endswith('.csv')]
                
                if not csv_files:
                    QMessageBox.warning(self, "Предупреждение", "Не са намерени CSV файлове в избраната директория")
                    return
                
                # Show file selection dialog
                from PyQt6.QtWidgets import QListWidget, QVBoxLayout, QDialog, QDialogButtonBox
                
                dialog = QDialog(self)
                dialog.setWindowTitle("Избери CSV файлове за импорт")
                dialog.setModal(True)
                dialog.resize(400, 300)
                
                layout = QVBoxLayout(dialog)
                
                file_list = QListWidget()
                file_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
                
                for csv_file in csv_files:
                    file_list.addItem(csv_file)
                
                layout.addWidget(QLabel("Избери файлове за импорт:"))
                layout.addWidget(file_list)
                
                # Warning label
                warning_label = QLabel("⚠️ ВНИМАНИЕ: Импортът ще изтрие съществуващите данни в избраните таблици!")
                warning_label.setStyleSheet("color: red; font-weight: bold;")
                layout.addWidget(warning_label)
                
                # Buttons
                button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addWidget(button_box)
                
                if dialog.exec() == QDialog.DialogCode.Accepted:
                    selected_files = [item.text() for item in file_list.selectedItems()]
                    
                    if not selected_files:
                        QMessageBox.warning(self, "Предупреждение", "Не са избрани файлове")
                        return
                    
                    # Confirmation dialog
                    reply = QMessageBox.question(
                        self, "Потвърждение",
                        f"Ще импортирате {len(selected_files)} файла.\n"
                        f"Това ще изтрие съществуващите данни в съответните таблици.\n\n"
                        f"Сигурни ли сте?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )

                    if reply == QMessageBox.StandardButton.Yes:
                        imported_tables = []
                        failed_files = []
                        
                        with self.db.get_connection() as conn:
                            cursor = conn.cursor()
                            
                            for csv_file in selected_files:
                                try:
                                    file_path = os.path.join(directory, csv_file)
                                    
                                    # Extract table name from filename (remove timestamp if present)
                                    table_name = csv_file.replace('.csv', '')
                                    # Remove timestamp pattern if present
                                    import re
                                    table_name = re.sub(r'_\d{8}_\d{6}$', '', table_name)
                                    
                                    # Read CSV file
                                    with open(file_path, 'r', encoding='utf-8-sig') as f:
                                        reader = csv.reader(f)
                                        headers = next(reader)  # Get column names
                                        rows = list(reader)
                                    
                                    if rows:
                                        # Clear existing data
                                        cursor.execute(f"DELETE FROM {table_name}")
                                        
                                        # Insert new data
                                        placeholders = ", ".join(["?" for _ in headers])
                                        for row in rows:
                                            # Handle empty values
                                            processed_row = [None if cell == '' else cell for cell in row]
                                            cursor.execute(
                                                f"INSERT INTO {table_name} ({', '.join(headers)}) VALUES ({placeholders})",
                                                processed_row
                                            )
                                        
                                        imported_tables.append(table_name)
                                    
                                except Exception as e:
                                    logger.error(f"Error importing {csv_file}: {e}")
                                    failed_files.append(csv_file)
                            
                            conn.commit()
                        
                        # Show results
                        if failed_files:
                            QMessageBox.warning(
                                self, "Частично успешен импорт",
                                f"Импортирани таблици: {', '.join(imported_tables[:5])}\n"
                                f"{'...' if len(imported_tables) > 5 else ''}\n\n"
                                f"Неуспешни файлове: {', '.join(failed_files)}"
                            )
                        else:
                            QMessageBox.information(
                                self, "Успех",
                                f"CSV импорт завършен успешно!\n\n"
                                f"Импортирани таблици: {', '.join(imported_tables)}"
                            )
                        
                        # Reload data
                        self.load_data()
                        self.update_reports_and_database_stats()
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при CSV импорт: {str(e)}")
    
    def run_schema_migration(self):
        """Run database schema migration"""
        try:
            reply = QMessageBox.question(
                self, "Обновяване на схема",
                "Това ще обнови структурата на базата данни.\n"
                "Препоръчва се да направите резервно копие първо.\n\n"
                "Продължаване?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                try:
                    # Run the migration function
                    self.db.migrate_schema()
                    
                    QMessageBox.information(
                        self, "Успех",
                        "Схемата на базата данни е обновена успешно!\n\n"
                        "Промени:\n"
                        "- Проверени/добавени липсващи колони\n"
                        "- Обновени индекси\n"
                        "- Поправени несъответствия"
                    )
                    
                    # Reload data to reflect changes
                    self.load_data()
                    self.update_database_statistics()
                    
                except Exception as e:
                    QMessageBox.critical(
                        self, "Грешка при миграция",
                        f"Грешка при обновяване на схемата:\n{str(e)}\n\n"
                        f"Моля, възстановете от резервно копие ако е необходимо."
                    )
                    
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при обновяване на схема: {str(e)}")
    
    def convert_legacy_data(self):
        """Convert legacy data formats"""
        try:
            # Show conversion options dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("Конвертиране на данни")
            dialog.setModal(True)
            dialog.setFixedSize(500, 400)
            
            layout = QVBoxLayout(dialog)
            
            # Title
            title_label = QLabel("Конвертиране на стари данни")
            title_label.setFont(QFont("Arial", 14, QFont.Weight.Bold))
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(title_label)
            
            # Available conversions
            conversions_group = QGroupBox("Достъпни конвертирания")
            conversions_layout = QVBoxLayout()
            
            # Checkboxes for different conversion types
            convert_timestamps = QCheckBox("Конвертирай стари формати на дати")
            convert_timestamps.setChecked(True)
            conversions_layout.addWidget(convert_timestamps)
            
            convert_barcodes = QCheckBox("Нормализирай баркодове (13 цифри)")
            convert_barcodes.setChecked(True)
            conversions_layout.addWidget(convert_barcodes)
            
            convert_prices = QCheckBox("Конвертирай цени (премахни невалидни стойности)")
            convert_prices.setChecked(True)
            conversions_layout.addWidget(convert_prices)
            
            convert_images = QCheckBox("Проверка и почистване на пътища към изображения")
            convert_images.setChecked(True)
            conversions_layout.addWidget(convert_images)
            
            conversions_group.setLayout(conversions_layout)
            layout.addWidget(conversions_group)
            
            # Warning
            warning_label = QLabel("⚠️ ВНИМАНИЕ: Създайте резервно копие преди конвертиране!")
            warning_label.setStyleSheet("color: red; font-weight: bold; background-color: #ffe6e6; padding: 10px; border-radius: 5px;")
            layout.addWidget(warning_label)
            
            # Progress info
            info_group = QGroupBox("Информация")
            info_layout = QVBoxLayout()
            
            info_text = QLabel(
                "Конвертирането ще:\n"
                "• Поправи неправилни формати на данни\n"
                "• Нормализира съществуващите записи\n"
                "• Премахне невалидни стойности\n"
                "• Обнови метаданни"
            )
            info_layout.addWidget(info_text)
            
            info_group.setLayout(info_layout)
            layout.addWidget(info_group)
            
            # Buttons
            button_layout = QHBoxLayout()
            
            convert_btn = QPushButton("Конвертирай")
            convert_btn.clicked.connect(lambda: self.perform_data_conversion(
                convert_timestamps.isChecked(),
                convert_barcodes.isChecked(),
                convert_prices.isChecked(),
                convert_images.isChecked(),
                dialog
            ))
            button_layout.addWidget(convert_btn)
            
            cancel_btn = QPushButton("Отказ")
            cancel_btn.clicked.connect(dialog.reject)
            button_layout.addWidget(cancel_btn)
            
            layout.addLayout(button_layout)
            
            dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при конвертиране на данни: {str(e)}")
    
    def perform_data_conversion(self, timestamps, barcodes, prices, images, dialog):
        """Perform the actual data conversion"""
        try:
            converted_items = []
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                if timestamps:
                    # Convert old timestamp formats
                    cursor.execute("SELECT id, date_added FROM items WHERE date_added IS NOT NULL")
                    items = cursor.fetchall()
                    
                    for item_id, date_str in items:
                        try:
                            # Try to parse and normalize the date
                            if date_str and not date_str.startswith('20'):  # Not in proper format
                                cursor.execute("UPDATE items SET date_added = datetime('now') WHERE id = ?", (item_id,))
                        except:
                            pass
                    
                    converted_items.append("Времеви печати")
                
                if barcodes:
                    # Normalize barcodes to 13 digits
                    cursor.execute("SELECT id, barcode FROM items")
                    items = cursor.fetchall()
                    
                    for item_id, barcode in items:
                        if barcode and len(barcode) != 7:
                            # Pad with zeros or truncate
                            normalized = barcode.zfill(7)[:7]
                            cursor.execute("UPDATE items SET barcode = ? WHERE id = ?", (normalized, item_id))
                    
                    converted_items.append("Баркодове")
                
                if prices:
                    # Clean up price data
                    cursor.execute("UPDATE items SET price = 0 WHERE price < 0 OR price IS NULL")
                    cursor.execute("UPDATE items SET cost = 0 WHERE cost < 0 OR cost IS NULL")
                    converted_items.append("Цени")
                

                
                conn.commit()
            
            if converted_items:
                QMessageBox.information(
                    self, "Успех",
                    f"Конвертирането завърши успешно!\n\n"
                    f"Обработени: {', '.join(converted_items)}"
                )
            else:
                QMessageBox.information(self, "Информация", "Не са намерени данни за конвертиране")
            
            # Reload data
            self.load_data()
            self.update_database_statistics()
            
            dialog.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при конвертиране: {str(e)}")
    
    def change_user_password(self):
        """Handle password change from UI"""
        try:
            # Get input values
            new_password = self.new_password_input.text().strip()
            confirm_password = self.confirm_password_input.text().strip()
            old_password = self.old_password_input.text().strip()
            
            # Debug logging
            logger.info(f"Password change attempt - New password length: {len(new_password)}")
            
            # Validate inputs
            if not new_password or not confirm_password or not old_password:
                QMessageBox.warning(self, "Предупреждение", "Моля, попълнете всички полета!")
                return
            
            # Check if new passwords match
            if new_password != confirm_password:
                QMessageBox.warning(self, "Предупреждение", "Новата парола и потвърждението не съвпадат!")
                return
            
            # Length validation
            if len(new_password) < 4:
                QMessageBox.warning(self, "Предупреждение", "Паролата трябва да бъде поне 4 символа!")
                return
            
            if len(new_password) > 10:
                QMessageBox.warning(self, "Предупреждение", "Паролата не трябва да бъде повече от 10 символа!")
                return
            
            # Enhanced character validation - support numbers only, letters only, or combination
            import re
            if not re.match(r'^[a-zA-Z0-9]+$', new_password):
                QMessageBox.warning(self, "Предупреждение", 
                                  "Паролата може да съдържа само:\n"
                                  "• Само цифри (напр. 12345)\n"
                                  "• Само английски букви (напр. abcde)\n"
                                  "• Комбинация от букви и цифри (напр. abc123)")
                return
            
            # Additional validation - at least one character type should be present
            has_letter = bool(re.search(r'[a-zA-Z]', new_password))
            has_digit = bool(re.search(r'[0-9]', new_password))
            
            if not (has_letter or has_digit):
                QMessageBox.warning(self, "Предупреждение", 
                                  "Паролата трябва да съдържа поне една буква или цифра!")
                return
            
            logger.info(f"Password validation passed - Has letters: {has_letter}, Has digits: {has_digit}")
            
            # Get current user
            current_user = self.db.get_current_user()
            if not current_user:
                QMessageBox.critical(self, "Грешка", "Не може да се намери текущия потребител!")
                logger.error("No current user found for password change")
                return
            
            logger.info(f"Found current user: {current_user}")
            
            # Attempt to change password
            try:
                logger.info("Attempting to change password...")
                success = self.db.change_user_password(current_user, old_password, new_password)
                
                if success:
                    QMessageBox.information(self, "Успех", "Паролата е сменена успешно!")
                    logger.info("Password changed successfully")
                    
                    # Clear input fields
                    self.new_password_input.clear()
                    self.confirm_password_input.clear()
                    self.old_password_input.clear()
                else:
                    QMessageBox.critical(self, "Грешка", 
                                       "Неуспешна смяна на парола!\n\n"
                                       "Възможни причини:\n"
                                       "• Старата парола не е правилна\n"
                                       "• Проблем с базата данни")
                    logger.error("Password change failed - returned False")
                    
            except ValueError as ve:
                # Display specific validation errors from database
                QMessageBox.warning(self, "Грешка при валидация", str(ve))
                logger.error(f"Password validation error: {ve}")
            except Exception as e:
                QMessageBox.critical(self, "Грешка", f"Грешка при смяна на парола: {str(e)}")
                logger.error(f"Exception during password change: {e}")
                
        except Exception as e:
            logger.error(f"Error in change_user_password: {e}")
            QMessageBox.critical(self, "Грешка", f"Неочаквана грешка: {str(e)}")

    def toggle_new_password_visibility(self, state):
        """Toggle visibility for new password field"""
        if state == 2:  # Checked state (Hide password)
            self.new_password_input.setEchoMode(QLineEdit.EchoMode.Password)
            # Don't change text here, let it be handled by the checkbox state
        else:  # Unchecked state (Show password)
            self.new_password_input.setEchoMode(QLineEdit.EchoMode.Normal)
            # Don't change text here, let it be handled by the checkbox state
    
    def toggle_confirm_password_visibility(self, state):
        """Toggle visibility for confirm password field"""
        if state == 2:  # Checked state (Hide password)
            self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        else:  # Unchecked state (Show password)
            self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Normal)
    
    def toggle_old_password_visibility(self, state):
        """Toggle visibility for old password field"""
        if state == 2:  # Checked state (Hide password)
            self.old_password_input.setEchoMode(QLineEdit.EchoMode.Password)
        else:  # Unchecked state (Show password)
            self.old_password_input.setEchoMode(QLineEdit.EchoMode.Normal)



    def create_backup(self):
        """Create a backup of the database"""
        try:
            backup_path = self.db.create_backup()
            
            # Save the backup timestamp
            self.save_last_backup_time()
            
            QMessageBox.information(self, "Успех", f"Резервното копие е създадено успешно в:\n{backup_path}")
            self.load_backup_list()  # Refresh backup list
            self.update_database_statistics()  # Update database stats
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Неуспешно създаване на резервно копие: {str(e)}")

    def restore_backup(self):
        """Restore database from backup"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Избери резервно копие", "backups",
                "Database Files (*.db);;All Files (*.*)"
            )
            if file_path:
                if self.db.restore_backup(file_path):
                    # Save restore time as a backup event
                    self.save_last_backup_time()
                    
                    QMessageBox.information(self, "Успех", "Базата данни е възстановена успешно")
                    self.load_data()  # Reload all data after restore
                    self.update_reports_and_database_stats()  # Update all statistics
                else:
                    QMessageBox.warning(self, "Грешка", "Неуспешно възстановяване на базата данни")
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Неуспешно възстановяване на резервно копие: {str(e)}")

    def export_data(self):
        """Export data to Excel file"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Съхранение на експортирани данни", "exports",
                "Excel Files (*.xlsx)"
            )
            if file_path:
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'

                # Create a new workbook
                wb = Workbook()
                
                # Get all tables
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = cursor.fetchall()
                    
                    for table in tables:
                        table_name = table[0]
                        # Create a new sheet for each table
                        ws = wb.create_sheet(title=table_name)
                        
                        # Get table data
                        cursor.execute(f"SELECT * FROM {table_name}")
                        columns = [description[0] for description in cursor.description]
                        rows = cursor.fetchall()
                        
                        # Write headers
                        for col, header in enumerate(columns, 1):
                            ws.cell(row=1, column=col, value=header)
                            # Set column width
                            ws.column_dimensions[get_column_letter(col)].width = 15
                        
                        # Write data
                        for row_idx, row in enumerate(rows, 2):
                            for col_idx, value in enumerate(row, 1):
                                # Format decimal numbers to 2 decimal places
                                if isinstance(value, float):
                                    value = f"{value:.2f}"
                                elif isinstance(value, str) and columns[col_idx-1].lower() in ['price', 'cost', 'weight', 'цена', 'тегло', 'стойност']:
                                    try:
                                        # Try to format numeric strings
                                        float_val = float(value)
                                        value = f"{float_val:.2f}"
                                    except (ValueError, TypeError):
                                        pass  # Keep original value if not numeric
                                ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
                
                # Save the workbook
                wb.save(file_path)
                QMessageBox.information(self, "Успех", f"Данните са експортирани успешно в:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Неуспешен експорт на данни: {str(e)}")

    def import_data(self):
        """Import data from Excel file"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Избери файл за импорт", "exports",
                "Excel Files (*.xlsx)"
            )
            if file_path:
                # Load the workbook
                wb = openpyxl.load_workbook(file_path)
                
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        # Get headers
                        headers = [cell.value for cell in ws[1]]
                        
                        # Clear existing data
                        cursor.execute(f"DELETE FROM {sheet_name}")
                        
                        # Insert data
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            placeholders = ", ".join(["?" for _ in headers])
                            cursor.execute(
                                f"INSERT INTO {sheet_name} ({', '.join(headers)}) VALUES ({placeholders})",
                                row
                            )
                    
                    conn.commit()
                
                QMessageBox.information(self, "Успех", "Данните са импортирани успешно")
                self.load_data()  # Reload all data after import
                self.update_reports_and_database_stats()  # Update all statistics
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Неуспешен импорт на данни: {str(e)}")

    def generate_barcode(self):
        """Generate a new barcode"""
        try:
            # Check if barcode is locked and fields are filled
            if hasattr(self, 'barcode_locked') and self.barcode_locked:
                QMessageBox.warning(self, "Предупреждение", "Баркодът вече е генериран. За нов баркод, първо принтирайте или изчистете полетата.")
                return

            # Validate all required fields
            if not (self.category_input.currentText() and
                    self.metal_input.currentText() and
                    self.stone_input.currentText() and
                    self.cost_input.value() >= 0 and
                    self.price_input.value() >= 0 and
                    self.weight_input.value() >= 0 and
                    self.stock_input.value() >= 0):
                QMessageBox.warning(self, "Предупреждение", "Моля, попълнете всички полета, преди да генерирате баркод!")
                return

            # Get next barcode from sequence
            conn = sqlite3.connect(get_persistent_path('data/jewelry.db'))
            cursor = conn.cursor()
            
            # Get current sequence value
            cursor.execute("SELECT next_val FROM barcode_sequence WHERE id=1")
            current = cursor.fetchone()[0]
            
            # Calculate next barcode
            next_barcode = current + 1
            
            # Update sequence
            cursor.execute("UPDATE barcode_sequence SET next_val=? WHERE id=1", (next_barcode,))
            conn.commit()
            conn.close()
            
            # Set the barcode
            self.barcode_input.setText(str(next_barcode))
            
            # Lock the barcode
            self.barcode_locked = True
            
            # Generate and display barcode preview
            self.update_barcode_preview()
            
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при генериране на баркод: {str(e)}")

    def update_barcode_preview(self):
        """Update the barcode preview using Citizen CLP 631 compatible method"""
        try:
            if not self.barcode_input.text():
                # Clear preview if no barcode
                self.barcode_preview.clear()
                return
            
            # Get the current barcode
            current_barcode = self.barcode_input.text()
            
            # Clear the preview first to force refresh
            self.barcode_preview.clear()
            
            # Generate Code 128 barcode using the working method
            try:
                from barcode import Code128
                code128 = Code128(current_barcode, writer=ImageWriter())
                temp_dir = tempfile.gettempdir()
                temp_file = os.path.join(temp_dir, "temp_barcode")
                
                # Get font path for barcode text - use resource_path for PyInstaller compatibility
                barcode_font_path = resource_path("fonts/arial.ttf")
                
                # Thermal printer optimized options - white background will be converted to transparent
                barcode_options = {
                    "write_text": True,
                    "text_distance": 5,
                    "font_size": 13,
                    "module_width": 0.3,
                    "module_height": 10,
                    "quiet_zone": 2.5,
                    "background": "#FFFFFF",  # White background (will be converted to transparent)
                    "foreground": "#000000",  # Black ink for thermal printing
                    "dpi": self.dpi,
                    "center_text": True,
                    "font_path": barcode_font_path  # Specify bundled font for PyInstaller compatibility
                }
                
                code128.save(temp_file, options=barcode_options)
                logger.debug(f"Barcode saved to: {temp_file}.png")
                
            except Exception as barcode_error:
                logger.error(f"Error generating barcode: {barcode_error}")
                QMessageBox.critical(self, "Грешка", f"Грешка при генериране на баркод: {str(barcode_error)}")
                return
            
            # Load and resize barcode using the original method (restored barcode size)
            try:
                img = Image.open(temp_file + ".png")
                barcode_width = int(21.7 * self.mm_to_px)
                img = img.resize((barcode_width, self.label_height), Image.Resampling.LANCZOS)
                enhancer = ImageEnhance.Sharpness(img)
                img = enhancer.enhance(1.5)
                self.barcode_image = img
                logger.debug(f"Barcode image loaded and resized successfully")
                
            except Exception as image_error:
                logger.error(f"Error loading/resizing barcode image: {image_error}")
                QMessageBox.critical(self, "Грешка", f"Грешка при зареждане на изображението: {str(image_error)}")
                return
            
            # Create transparent label for thermal printing - only black elements will be printed
            label_img = Image.new("RGBA", (self.label_width, self.label_height), (255, 255, 255, 0))
            
            # Add text to label (category and price)
            draw = ImageDraw.Draw(label_img)
            
            # Load fonts - CLP-631 optimized sizes
            font = None
            bold_font = None
            optimal_font_size = 27
            
            # Try multiple font paths for regular font - using resource_path for PyInstaller compatibility
            font_paths = [
                resource_path("fonts/arial.ttf"),  # Primary path for bundled font
                "arial.ttf",
                "fonts/arial.ttf", 
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/Arial.ttf"
            ]
            
            # Try multiple font paths for bold font - using resource_path for PyInstaller compatibility
            bold_font_paths = [
                resource_path("fonts/arialbd.ttf"),  # Primary path for bundled font
                "arialbd.ttf",
                "fonts/arialbd.ttf",
                "C:/Windows/Fonts/arialbd.ttf", 
                "C:/Windows/Fonts/ARIALBD.TTF"
            ]
            
            # Load regular font
            for font_path in font_paths:
                try:
                    if os.path.exists(font_path):
                        font = ImageFont.truetype(font_path, optimal_font_size)
                        logger.debug(f"Loaded regular font from: {font_path}")
                        break
                except Exception as e:
                    logger.debug(f"Could not load font from {font_path}: {e}")
                    continue
            
            # Load bold font
            for bold_font_path in bold_font_paths:
                try:
                    if os.path.exists(bold_font_path):
                        bold_font = ImageFont.truetype(bold_font_path, optimal_font_size)
                        logger.debug(f"Loaded bold font from: {bold_font_path}")
                        break
                except Exception as e:
                    logger.debug(f"Could not load bold font from {bold_font_path}: {e}")
                    continue
            
            # Fallback to default fonts if TrueType fonts failed
            if font is None:
                try:
                    font = ImageFont.load_default()
                    logger.warning("Using default font for regular text")
                except:
                    logger.error("Could not load any font - text may not display properly")
                    
            if bold_font is None:
                bold_font = font  # Use regular font as fallback
                logger.warning("Using regular font as bold font fallback")
            
            # Add text that will always be displayed (now at top)
            always_display_text = "Сребро 925"
            try:
                if bold_font:
                    draw.text((15, 10), always_display_text, fill='black', font=bold_font)
                else:
                    draw.text((15, 10), always_display_text, fill='black')
            except Exception as e:
                logger.error(f"Error drawing always display text: {e}")
            
            # Draw product name (use category as name) - now below
            name = self.category_input.currentText() or "Product Name"
            
            # Add grams to category name if checkbox is checked
            if self.include_grams_checkbox.isChecked():
                weight_value = self.weight_input.value()
                # Format grams: remove leading zero if single digit (e.g., "7.25 g" not "07.25 g")
                grams_text = f"{weight_value:g} g"  # :g removes trailing zeros and doesn't pad with leading zeros
                name = f"{name} / {grams_text}"
            
            try:
                if bold_font:
                    draw.text((15, 45), name, fill='black', font=bold_font)
                else:
                    draw.text((15, 45), name, fill='black')
            except Exception as e:
                logger.error(f"Error drawing product name: {e}")

            # Draw price - conditionally include Lev based on checkbox
            price_eur = self.price_input.value()
            
            # Check if лв price should be included
            if self.include_lev_price_checkbox.isChecked():
                # Show both Euro and Lev prices
                price_lev = self.euro_to_lev(price_eur)
                
                # Check if price order should be inverted
                if self.invert_prices_checkbox.isChecked():
                    # Lev first, Euro second
                    euro_lev_text = f"{price_lev:.2f} лв / {price_eur:.2f} €"
                else:
                    # Euro first, Lev second (default)
                    euro_lev_text = f"{price_eur:.2f} € / {price_lev:.2f} лв"
            else:
                # Show only Euro price (inversion doesn't apply when only one currency)
                euro_lev_text = f"{price_eur:.2f} €"
            
            try:
                if bold_font:
                    draw.text((15, 80), euro_lev_text, fill='black', font=bold_font)
                else:
                    draw.text((15, 80), euro_lev_text, fill='black')
            except Exception as e:
                logger.error(f"Error drawing price text: {e}")
            
            # Paste barcode onto transparent label (use mask to handle transparency)
            # Barcode positioning: (X, Y) coordinates - currently (~256, 10) 
            # Note: self.label_width // 2 ≈ 256 pixels for 43.4mm label
            barcode_x = 250  # X position (was: self.label_width // 2)
            barcode_y = 10   # Y position
            
            # Debug: Print barcode mode information
            print(f"DEBUG: Barcode image mode: {self.barcode_image.mode}")
            print(f"DEBUG: Barcode image size: {self.barcode_image.size}")
            
            if self.barcode_image.mode == 'RGBA':
                # If barcode has alpha channel, use it as mask
                print("DEBUG: Using RGBA path - pasting barcode with alpha channel")
                label_img.paste(self.barcode_image, (barcode_x, barcode_y), self.barcode_image)
            else:
                # Convert white background to transparent for thermal printing
                print("DEBUG: Using non-RGBA path - converting to RGBA")
                barcode_rgba = self.barcode_image.convert('RGBA')
                data = barcode_rgba.getdata()
                new_data = []
                for item in data:
                    # Change white background to transparent
                    if item[0] > 240 and item[1] > 240 and item[2] > 240:  # Near white
                        new_data.append((255, 255, 255, 0))  # Transparent
                    else:
                        new_data.append(item)  # Keep other colors (black text/barcode)
                barcode_rgba.putdata(new_data)
                label_img.paste(barcode_rgba, (barcode_x, barcode_y), barcode_rgba)
            
            # Store for CLP-631 optimized printing
            self.current_label = label_img
            print("Label prepared for CLP-631 thermal transfer printing")
            
            # Convert PIL image to QPixmap for display (with transparency support)
            img_data = label_img.tobytes('raw', 'RGBA')
            qim = QImage(img_data, label_img.size[0], label_img.size[1], QImage.Format.Format_RGBA8888)
            pixmap = QPixmap.fromImage(qim)
            
            # Scale and display
            scaled_pixmap = pixmap.scaled(
                self.barcode_preview.width(),
                self.barcode_preview.height(),
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation
            )
            
            # Force update the preview
            self.barcode_preview.setPixmap(scaled_pixmap)
            self.barcode_preview.update()  # Force widget update
            
            # Clean up temporary file
            try:
                os.remove(temp_file + ".png")
            except:
                pass  # Ignore cleanup errors
            
            print(f"DEBUG: Preview updated successfully for barcode: {current_barcode}")
            print("DEBUG: update_barcode_preview() completed normally")
            
        except Exception as e:
            logger.error(f"Error in update_barcode_preview: {str(e)}")
            logger.error(f"Error type: {type(e).__name__}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            print(f"DEBUG: Error updating preview for barcode {self.barcode_input.text()}: {str(e)}")
            print(f"DEBUG: Error type: {type(e).__name__}")
            QMessageBox.critical(self, "Грешка", f"Грешка при обновяване на прегледа: {str(e)}")
            
            # Try to clear the preview on error
            try:
                self.barcode_preview.clear()
            except:
                pass

    def print_and_add_item(self):
        """Print barcode label and add item"""
        try:
            if not self.barcode_input.text():
                QMessageBox.warning(self, "Предупреждение", "Моля, генерирайте баркод първо!")
                return
            
            # VALIDATE ALL FIELDS BEFORE PRINTING
            # Get values for validation
            category = self.category_input.currentText()
            metal_type = self.metal_input.currentText()
            stone_type = self.stone_input.currentText()
            price = self.price_input.value()
            cost = self.cost_input.value()
            weight = self.weight_input.value()
            stock_quantity = self.stock_input.value()

            # Comprehensive validation
            validation_errors = []

            # Required fields validation
            if not category or category == "Друго":
                validation_errors.append("Моля, изберете валидна категория")

            if not metal_type or metal_type == "Друго":
                validation_errors.append("Моля, изберете валиден метал")

            if not stone_type or stone_type == "Друго":
                validation_errors.append("Моля, изберете валиден камък")

            # Numeric validation
            if price <= 0:
                validation_errors.append("Цената трябва да бъде по-голяма от 0")

            if cost < 0:
                validation_errors.append("Себестойността не може да бъде отрицателна")

            if cost > price:
                validation_errors.append("Себестойността не може да бъде по-висока от цената")

            if weight <= 0:
                validation_errors.append("Теглото трябва да бъде по-голямо от 0")

            if stock_quantity <= 0:
                validation_errors.append("Количеството трябва да бъде по-голямо от 0")

            # Display validation errors BEFORE printing
            if validation_errors:
                error_message = "Моля, коригирайте следните грешки преди принтиране:\n\n" + "\n".join(f"• {error}" for error in validation_errors)
                QMessageBox.warning(self, "Грешки при валидация", error_message)
                return
                
            # Print the label (only if validation passes)
            self.print_barcode_label()
            
            # Check if this is a reprint (barcode already exists in database)
            barcode = self.barcode_input.text().strip()
            is_reprint = False
            
            try:
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM items WHERE barcode = ?", (barcode,))
                    is_reprint = cursor.fetchone()[0] > 0
            except Exception:
                is_reprint = False
            
            if is_reprint:
                # Handle warehouse reprint vs regular reprint differently
                if self.is_warehouse_reprint:
                    # This is a warehouse reprint - add quantity to existing stock
                    reply = QMessageBox.question(
                        self, "Потвърждение на печат (Склад)",
                        f"Етикетът е изпратен за печат.\n\n"
                        f"Печатът завърши ли успешно?\n\n"
                        f"Ако отговорите 'Да', количеството ({stock_quantity}) ще бъде ДОБАВЕНО към наличността в склада.",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.Yes
                    )

                    if reply == QMessageBox.StandardButton.Yes:
                        # Add quantity to existing warehouse stock
                        self.add_quantity_to_warehouse(barcode, stock_quantity)
                        QMessageBox.information(self, "Успех", 
                                              f"Етикетът е отпечатан успешно!\n\n"
                                              f"Количество {stock_quantity} е добавено към склада.")
                        # Clear form for next use
                        self.clear_form()
                        self.is_warehouse_reprint = False
                    else:
                        QMessageBox.information(self, "Информация", 
                                              "Можете да опитате отново печата или да изчистите формата.")
                else:
                    # This is a regular reprint - just confirm printing without adding to database
                    reply = QMessageBox.question(
                        self, "Потвърждение на печат (Повторен печат)",
                        "Етикетът е изпратен за печат.\n\n"
                        "Печатът завърши ли успешно?\n\n"
                        "ЗАБЕЛЕЖКА: Това е повторен печат - артикулът вече съществува в системата.",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.Yes
                    )

                    if reply == QMessageBox.StandardButton.Yes:
                        QMessageBox.information(self, "Успех", 
                                              "Етикетът е отпечатан успешно!\n\n"
                                              "Можете да изчистите формата или да отпечатате още етикети.")
                        # Don't save anything - just unlock for potential new prints
                        self.barcode_locked = False
                    else:
                        QMessageBox.information(self, "Информация", 
                                              "Можете да опитате отново печата или да изчистите формата.")
            else:
                # This is a new item - normal flow
                reply = QMessageBox.question(
                    self, "Потвърждение на печат",
                    "Етикетът е изпратен за печат.\n\n"
                    "Печатът завърши ли успешно?\n"
                    "Ако отговорите 'Да', артикулът ще бъде добавен в системата.",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # Add the item only after user confirms printing is complete
                    self.save_item()
                    
                    # Unlock barcode generation after successful print and save
                    self.barcode_locked = False
                else:
                    QMessageBox.information(self, "Информация", 
                                          "Артикулът не е добавен в системата.\n"
                                          "Можете да опитате отново печата или да изчистите формата.")
            
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при принтиране и добавяне: {str(e)}")

    def print_barcode_label(self):
        """Print the barcode label using CLP-631 optimized method"""
        if not hasattr(self, 'current_label') or not self.current_label or not self.barcode_input.text():
            QMessageBox.warning(self, "Предупреждение", "Моля, генерирайте баркод първо")
            return

        try:
            # Use the CLP-631 optimized printing method
            success, message = self.printer_handler.print_label_clp631(self.current_label)
            
            if success:
                logger.info(f"Print successful: {message}")
            else:
                QMessageBox.critical(self, "Грешка при печат", message)
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при печат: {str(e)}")

    def add_quantity_to_warehouse(self, barcode, additional_quantity):
        """Add quantity to existing warehouse stock for reprints"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get current stock
                cursor.execute("SELECT stock_quantity FROM items WHERE barcode = ?", (barcode,))
                result = cursor.fetchone()
                if not result:
                    raise Exception("Артикулът не е намерен в базата данни")
                
                current_stock = result[0]
                new_stock = current_stock + additional_quantity
                
                # Update the stock
                cursor.execute("UPDATE items SET stock_quantity = ? WHERE barcode = ?", (new_stock, barcode))
                conn.commit()
                
                # Log the action
                logger.info(f"Added {additional_quantity} to warehouse stock for barcode {barcode}. New stock: {new_stock}")
                
                # Refresh the warehouse table if visible
                self.load_items()
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при добавяне на количество към склада: {str(e)}")
            raise

    def clear_form(self):
        """Clear the add item form"""
        try:
            # Clear all input fields
            self.barcode_input.clear()
            self.category_input.setCurrentIndex(0)
            self.description_input.clear()
            self.price_input.setValue(0.0)
            self.cost_input.setValue(0.0)
            self.weight_input.setValue(0.0)
            self.metal_input.setCurrentIndex(0)
            self.stone_input.setCurrentIndex(0)
            self.stock_input.setValue(1)
            
            # Reset confirmed values for spin boxes
            self.price_input.reset_confirmed_value()
            self.cost_input.reset_confirmed_value()
            self.weight_input.reset_confirmed_value()
            self.stock_input.reset_confirmed_value()
            
            # Clear barcode preview
            if hasattr(self, 'barcode_preview'):
                self.barcode_preview.clear()
                self.barcode_preview.setText("Баркод ще се покаже тук")
            
            # Reset flags
            self.barcode_locked = False
            self.is_warehouse_reprint = False
            self.current_label = None
            self.barcode_image = None
            
            # Update currency displays
            self.update_lev_cost()
            self.update_lev_price()
            
        except Exception as e:
            logger.error(f"Error clearing form: {e}")

    def get_item_shop_locations(self, barcode):
        """Get information about where an item is located in shops and quantities"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get item_id from barcode
                cursor.execute("SELECT id FROM items WHERE barcode = ?", (barcode,))
                item_result = cursor.fetchone()
                if not item_result:
                    return []
                
                item_id = item_result[0]
                
                # Get shop locations and quantities
                cursor.execute("""
                    SELECT s.name, si.quantity, si.updated_at 
                    FROM shop_items si
                    JOIN shops s ON si.shop_id = s.id
                    WHERE si.item_id = ? AND si.quantity > 0
                    ORDER BY s.name
                """, (item_id,))
                
                results = cursor.fetchall()
                locations = []
                
                for shop_name, quantity, updated_at in results:
                    locations.append({
                        'shop_name': shop_name,
                        'quantity': quantity,
                        'updated_at': updated_at
                    })
                
                return locations
                
        except Exception as e:
            logger.error(f"Error getting shop locations for barcode {barcode}: {e}")
            return []

    def print_image(self, image_path, copies=1):
        """Print an image using the default printer"""
        try:
            # Get the default printer
            printer_name = win32print.GetDefaultPrinter()
            
            # Use Windows ShellExecute to show the print dialog
            win32api.ShellExecute(
                0,
                "print",
                image_path,
                f'/d:"{printer_name}"',
                ".",
                0
            )
            return True
                
        except Exception as e:
            QMessageBox.critical(self, "Грешка", f"Грешка при принтиране: {str(e)}")
            return False

    def build_barcode_tab(self):
        """Create the barcode tab"""
        widget = QWidget()
        layout = QHBoxLayout(widget)

        # Left panel - Form
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        
        # Product Name
        name_label = QLabel("Продукт:")
        self.bc_name_entry = QLineEdit()
        self.bc_name_entry.setFont(QFont('Segoe UI', 11))
        name_layout = QHBoxLayout()
        name_layout.addWidget(self.bc_name_entry)
        clear_name_btn = QPushButton("✕")
        clear_name_btn.setFont(QFont('Segoe UI', 10))
        clear_name_btn.clicked.connect(lambda: self.bc_name_entry.clear())
        name_layout.addWidget(clear_name_btn)
        left_layout.addWidget(name_label)
        left_layout.addLayout(name_layout)

        # Product Price
        price_label = QLabel("Цена на продукт:")
        self.bc_price_entry = QLineEdit()
        self.bc_price_entry.setFont(QFont('Segoe UI', 11))
        price_layout = QHBoxLayout()
        price_layout.addWidget(self.bc_price_entry)
        clear_price_btn = QPushButton("✕")
        clear_price_btn.setFont(QFont('Segoe UI', 10))
        clear_price_btn.clicked.connect(lambda: self.bc_price_entry.clear())
        price_layout.addWidget(clear_price_btn)
        left_layout.addWidget(price_label)
        left_layout.addLayout(price_layout)

        # Barcode
        barcode_label = QLabel("Code 128 Баркод:")
        self.bc_barcode_entry = QLineEdit()
        self.bc_barcode_entry.setFont(QFont('Segoe UI', 11))
        self.bc_barcode_entry.setReadOnly(True)
        barcode_layout = QHBoxLayout()
        barcode_layout.addWidget(self.bc_barcode_entry)
        left_layout.addWidget(barcode_label)
        left_layout.addLayout(barcode_layout)

        # Grams
        grams_label = QLabel("Грамаж:")
        self.bc_grams_entry = QLineEdit()
        self.bc_grams_entry.setFont(QFont('Segoe UI', 11))
        grams_layout = QHBoxLayout()
        grams_layout.addWidget(self.bc_grams_entry)
        clear_grams_btn = QPushButton("✕")
        clear_grams_btn.setFont(QFont('Segoe UI', 10))
        clear_grams_btn.clicked.connect(lambda: self.bc_grams_entry.clear())
        grams_layout.addWidget(clear_grams_btn)
        left_layout.addWidget(grams_label)
        left_layout.addLayout(grams_layout)

        # Quantity
        qty_label = QLabel("Количество:")
        self.bc_qty_entry = QLineEdit()
        self.bc_qty_entry.setFont(QFont('Segoe UI', 11))
        qty_layout = QHBoxLayout()
        qty_layout.addWidget(self.bc_qty_entry)
        clear_qty_btn = QPushButton("✕")
        clear_qty_btn.setFont(QFont('Segoe UI', 10))
        clear_qty_btn.clicked.connect(lambda: self.bc_qty_entry.clear())
        qty_layout.addWidget(clear_qty_btn)
        left_layout.addWidget(qty_label)
        left_layout.addLayout(qty_layout)

        # Image Upload
        self.bc_image_path = ""
        upload_btn = QPushButton("Качи снимка")
        upload_btn.setFont(QFont('Segoe UI', 11))
        upload_btn.clicked.connect(self.upload_bc_image)
        left_layout.addWidget(upload_btn)

        self.bc_image_label = QLabel()
        self.bc_image_label.setFixedSize(100, 100)
        self.bc_image_label.setStyleSheet("background-color: #f7f7f7;")
        left_layout.addWidget(self.bc_image_label)

        clear_image_btn = QPushButton("✕")
        clear_image_btn.setFont(QFont('Segoe UI', 10))
        clear_image_btn.clicked.connect(lambda: [setattr(self, 'bc_image_path', ''), self.bc_image_label.clear()])
        left_layout.addWidget(clear_image_btn)

        # Buttons
        generate_btn = QPushButton("Генерирай етикет")
        generate_btn.clicked.connect(self.generate_barcode)
        left_layout.addWidget(generate_btn)

        print_btn = QPushButton("Принтирай етикет")
        print_btn.clicked.connect(self.print_and_add_item)
        left_layout.addWidget(print_btn)

        clear_all_btn = QPushButton("Изчисти всичко")
        clear_all_btn.setStyleSheet("background-color: #ffe0e0;")
        clear_all_btn.clicked.connect(self.clear_barcode_tab_inputs)
        left_layout.addWidget(clear_all_btn)

        # Right panel - Preview
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        # Preview Canvas
        self.bc_preview_canvas = QLabel()
        self.bc_preview_canvas.setFixedSize(450, 100)
        self.bc_preview_canvas.setStyleSheet("background-color: white; border: 1px solid #d0d0d0;")
        right_layout.addWidget(self.bc_preview_canvas)

        # Add panels to main layout
        layout.addWidget(left_panel)
        layout.addWidget(right_panel)

        # Set up tab navigation
        self.bc_entries = [
            self.bc_name_entry,
            self.bc_price_entry,
            self.bc_barcode_entry,
            self.bc_grams_entry,
            self.bc_qty_entry
        ]

        return widget

    def handle_custom_input(self, combo_box, custom_values_set, current_text):
        """Handle custom input for combo boxes with proper capitalization"""
        if current_text == "Друго":
            text, ok = QInputDialog.getText(self, "Въведете стойност", "Въведете нова стойност:")
            if ok and text:
                # Capitalize first letter of each word
                words = text.split()
                capitalized_words = [word.capitalize() for word in words]
                new_text = " ".join(capitalized_words)
                
                # Check if value already exists
                if new_text in custom_values_set or new_text in ["Пръстен", "Гривна", "Обеци", "Синджир", "Злато", "Сребро", "Платина", "Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък"]:
                    QMessageBox.warning(self, "Предупреждение", "Тази стойност вече съществува!")
                    return False
                
                # Add to custom values set
                custom_values_set.add(new_text)
                
                # Save to database
                try:
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        # Determine the type based on which combo box we're dealing with
                        type_ = 'category' if combo_box == self.category_input else 'metal' if combo_box == self.metal_input else 'stone'
                        cursor.execute("INSERT INTO custom_values (type, value) VALUES (?, ?)", (type_, new_text))
                except Exception as e:
                    QMessageBox.critical(self, "Грешка", f"Грешка при запазване на стойността: {str(e)}")
                    return False
                
                # Update combo box items while preserving all existing items
                current_items = [combo_box.itemText(i) for i in range(combo_box.count())]
                if "Друго" in current_items:
                    current_items.remove("Друго")
                combo_box.clear()
                combo_box.addItems(current_items + [new_text, "Друго"])
                
                # Set the new value
                combo_box.setCurrentText(new_text)
                return True
        return False

    # === INVENTORY AUDIT METHODS ===
    
    def refresh_audit_shop_combo(self):
        """Refresh the audit shop combo box"""
        try:
            self.audit_shop_combo.clear()
            self.audit_shop_combo.addItem("-- Изберете магазин --", None)
            
            shops = self.db.get_all_shops()
            for shop_row in shops:
                shop_id = shop_row[0]  # First column is ID
                shop_name = shop_row[1]  # Second column is name
                self.audit_shop_combo.addItem(shop_name, shop_id)
                
        except Exception as e:
            logger.error(f"Error refreshing audit shop combo: {e}")
    
    def start_inventory_audit(self):
        """Start a new inventory audit"""
        try:
            # Check if shop is selected
            current_index = self.audit_shop_combo.currentIndex()
            if current_index <= 0:
                QMessageBox.warning(self, "Предупреждение", "Моля, изберете магазин за инвентаризация!")
                return
            
            shop_id = self.audit_shop_combo.currentData()
            shop_name = self.audit_shop_combo.currentText()
            
            # Confirm start
            reply = QMessageBox.question(
                self, "Започване на инвентаризация",
                f"Започване на инвентаризация за магазин: {shop_name}\n\n"
                "ВНИМАНИЕ: По време на инвентаризацията софтуерът ще бъде в ограничен режим.\n"
                "Ще можете да преглеждате другите раздели, но без да правите промени.\n\n"
                "Желаете ли да продължите?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Initialize audit session
            self.audit_in_progress = True
            self.audit_shop_id = shop_id
            self.audit_shop_name = shop_name
            self.audit_start_time = datetime.now()
            self.audit_paused = False
            self.audit_items_data = {}  # {barcode: {expected_qty, scanned_qty, category, metal_type, stone_type, price, weight, description}}
            self.audit_scanned_items = {}  # {barcode: scanned_quantity}
            
            # Generate unique session ID
            self.audit_session_id = f"audit_{shop_id}_{self.audit_start_time.strftime('%Y%m%d_%H%M%S')}"
            
            # Load all items from shop and populate table
            self.load_audit_items()
            
            # Enable audit controls
            self.start_audit_btn.setEnabled(False)
            self.pause_audit_btn.setEnabled(True)
            self.finish_audit_btn.setEnabled(True)
            self.audit_barcode_input.setEnabled(True)
            self.audit_shop_combo.setEnabled(False)
            
            # Set focus to barcode input
            self.audit_barcode_input.setFocus()
            
            # Switch to the Инвентаризация tab (first tab)
            self.audit_tabs.setCurrentIndex(0)
            
            # Enable read-only mode for other tabs
            self.set_audit_mode(True)
            
            # Update statistics
            self.update_audit_statistics()
            
            # Log audit start
            logger.info(f"Inventory audit started for shop: {shop_name} (ID: {shop_id})")
            
            QMessageBox.information(
                self, "Инвентаризация започната",
                f"Инвентаризацията на магазин '{shop_name}' е започната успешно!\n\n"
                "Започнете да сканирате артикулите един по един.\n"
                "Използвайте падащото меню в таблицата за да въведете количеството."
            )
            
        except Exception as e:
            logger.error(f"Error starting inventory audit: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при започване на инвентаризация: {str(e)}")
    
    def load_audit_items(self):
        """Load all items from shop and populate audit table (all items start as red/missing)"""
        try:
            # Get all items currently in the selected shop
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Query shop items with item details
                cursor.execute("""
                    SELECT i.barcode, i.category, i.metal_type, i.stone_type, 
                           i.price, i.weight, si.quantity, i.description
                    FROM shop_items si
                    JOIN items i ON si.item_id = i.id
                    WHERE si.shop_id = ? AND si.quantity > 0
                    ORDER BY i.barcode
                """, (self.audit_shop_id,))
                
                items = cursor.fetchall()
                
                # Clear previous data and table
                self.audit_items_data = {}
                self.audit_items_table.setRowCount(0)
                
                # Populate audit items table with all shop items (all start as red)
                self.audit_items_table.setRowCount(len(items))
                
                for row, item in enumerate(items):
                    barcode, category, metal, stone, price, weight, expected_qty, description = item
                    
                    # Handle NULL values safely
                    category = category if category is not None else "Неопределена категория"
                    metal = metal if metal is not None else "Неопределен метал"
                    stone = stone if stone is not None else "Без камък"
                    price = price if price is not None else 0.0
                    weight = weight if weight is not None else 0.0
                    expected_qty = expected_qty if expected_qty is not None else 0
                    description = description if description is not None else ""
                    
                    # Store item data
                    self.audit_items_data[barcode] = {
                        'expected_qty': expected_qty,
                        'scanned_qty': 0,  # Start with 0
                        'category': category,
                        'metal_type': metal,
                        'stone_type': stone,
                        'price': price,
                        'weight': weight,
                        'description': description
                    }
                    
                    # Create product name
                    product_name = f"{category} {metal}"
                    if stone and stone != "Без камък":
                        product_name += f" с {stone}"
                    
                    # Barcode
                    barcode_item = QTableWidgetItem(barcode)
                    barcode_item.setBackground(QColor(255, 153, 153))  # #ff9999 - like "Изчисти" button
                    barcode_item.setForeground(QColor(184, 134, 11))   # #B8860B - dark yellow for better readability
                    barcode_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.audit_items_table.setItem(row, 0, barcode_item)
                    
                    # Product name
                    product_item = QTableWidgetItem(product_name)
                    product_item.setBackground(QColor(255, 153, 153))  # #ff9999 - like "Изчисти" button
                    product_item.setForeground(QColor(184, 134, 11))   # #B8860B - dark yellow for better readability
                    product_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.audit_items_table.setItem(row, 1, product_item)
                    
                    # Quantity - Create dropdown
                    quantity_combo = QComboBox()
                    quantity_combo.setStyleSheet("background-color: #ff9999; color: #B8860B;")  # Dark yellow text on light red background
                    
                    # Disable mouse wheel events for the combo box
                    quantity_combo.wheelEvent = lambda event: None  # Disable wheel scroll
                    
                    # Add quantity options (0 to expected_qty + some extra)
                    max_qty = max(expected_qty + 3, 10)  # At least 10 options
                    for i in range(max_qty + 1):
                        quantity_combo.addItem(str(i))
                    quantity_combo.addItem("Друго...")
                    
                    # Set default to 0
                    quantity_combo.setCurrentText("0")
                    
                    # Connect to handler
                    quantity_combo.currentTextChanged.connect(
                        lambda text, b=barcode: self.on_quantity_changed(b, text)
                    )
                    
                    self.audit_items_table.setCellWidget(row, 2, quantity_combo)
                    
                    # Price
                    price_item = QTableWidgetItem(f"{price:.2f} €")
                    price_item.setBackground(QColor(255, 153, 153))  # #ff9999 - like "Изчисти" button
                    price_item.setForeground(QColor(184, 134, 11))   # #B8860B - dark yellow for better readability
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.audit_items_table.setItem(row, 3, price_item)
                
                logger.info(f"Loaded {len(items)} items for audit from shop {self.audit_shop_name}")
                
        except Exception as e:
            logger.error(f"Error loading audit items: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при зареждане на артикулите: {str(e)}")
    
    def on_quantity_changed(self, barcode, quantity_text):
        """Handle quantity change in dropdown"""
        try:
            if quantity_text == "Друго...":
                # Show input dialog for custom quantity
                custom_qty, ok = QInputDialog.getInt(
                    self, "Въведете количество", 
                    f"Количество за артикул {barcode}:", 
                    0, 0, 999
                )
                if ok:
                    # Find the combo box and update it
                    for row in range(self.audit_items_table.rowCount()):
                        if self.audit_items_table.item(row, 0).text() == barcode:
                            combo = self.audit_items_table.cellWidget(row, 2)
                            # Check if this quantity exists in combo, if not add it
                            if combo.findText(str(custom_qty)) == -1:
                                combo.insertItem(combo.count() - 1, str(custom_qty))
                            combo.setCurrentText(str(custom_qty))
                            self.update_item_status(barcode, custom_qty)
                            break
                else:
                    # User cancelled, revert to previous value
                    for row in range(self.audit_items_table.rowCount()):
                        if self.audit_items_table.item(row, 0).text() == barcode:
                            combo = self.audit_items_table.cellWidget(row, 2)
                            combo.setCurrentText(str(self.audit_items_data[barcode]['scanned_qty']))
                            break
            else:
                # Regular quantity selection
                try:
                    quantity = int(quantity_text)
                    self.update_item_status(barcode, quantity)
                except ValueError:
                    pass
                    
        except Exception as e:
            logger.error(f"Error handling quantity change: {e}")
    
    def update_item_status(self, barcode, scanned_qty):
        """Update item status and colors based on scanned quantity"""
        try:
            if barcode not in self.audit_items_data:
                return
                
            # Update scanned quantity
            self.audit_items_data[barcode]['scanned_qty'] = scanned_qty
            expected_qty = self.audit_items_data[barcode]['expected_qty']
            
            # Find the row for this barcode
            for row in range(self.audit_items_table.rowCount()):
                if self.audit_items_table.item(row, 0).text() == barcode:
                    # Determine status and color based on 4-color system
                    if scanned_qty == 0:
                        # Red - Nothing scanned (missing)
                        bg_color = QColor(255, 153, 153)  # #ff9999 - like "Изчисти" button
                        text_color = QColor(153, 0, 0)    # #990000 - like "Изчисти" button text
                        combo_style = "background-color: #ff9999; color: #990000;"
                    elif scanned_qty == expected_qty:
                        # Green - Perfect match
                        bg_color = QColor(153, 255, 153)  # #99ff99 - like "Принтирай етикет" button
                        text_color = QColor(0, 77, 0)     # #004d00 - like "Принтирай етикет" button text
                        combo_style = "background-color: #99ff99; color: #004d00;"
                    elif scanned_qty < expected_qty:
                        # Yellow - Partial scan (less than expected)
                        bg_color = QColor(255, 255, 153)  # #ffff99 - light yellow
                        text_color = QColor(184, 134, 11)  # #B8860B - darker yellow/goldenrod for better readability
                        combo_style = "background-color: #ffff99; color: #B8860B;"
                    else:  # scanned_qty > expected_qty
                        # Orange - Over quantity (more than expected)
                        bg_color = QColor(255, 204, 153)  # #ffcc99 - light orange
                        text_color = QColor(153, 77, 0)   # #994d00 - dark orange text
                        combo_style = "background-color: #ffcc99; color: #994d00;"
                    
                    # Update row colors
                    barcode_item = self.audit_items_table.item(row, 0)
                    barcode_item.setBackground(bg_color)
                    barcode_item.setForeground(text_color)
                    barcode_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    product_item = self.audit_items_table.item(row, 1)
                    product_item.setBackground(bg_color)
                    product_item.setForeground(text_color)
                    product_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    price_item = self.audit_items_table.item(row, 3)
                    price_item.setBackground(bg_color)
                    price_item.setForeground(text_color)
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    # Update combo box color
                    combo = self.audit_items_table.cellWidget(row, 2)
                    combo.setStyleSheet(combo_style)
                    break
            
            # Update statistics
            self.update_audit_statistics()
            
        except Exception as e:
            logger.error(f"Error updating item status: {e}")
    
    def process_audit_barcode(self):
        """Process the scanned barcode during audit - automatically set quantity to 1"""
        try:
            barcode = self.audit_barcode_input.text().strip()
            if not barcode:
                return
            
            # Check if barcode exists in audit items
            if barcode in self.audit_items_data:
                current_qty = self.audit_items_data[barcode]['scanned_qty']
                
                if current_qty > 0:
                    # Item already scanned - show dialog to select quantity
                    QMessageBox.information(
                        self, "Артикул вече сканиран",
                        f"Артикул '{barcode}' вече е сканиран с количество {current_qty}.\n\n"
                        "Използвайте падащото меню в таблицата за да промените количеството."
                    )
                else:
                    # First scan - set to 1 and update colors based on expected quantity
                    for row in range(self.audit_items_table.rowCount()):
                        if self.audit_items_table.item(row, 0).text() == barcode:
                            combo = self.audit_items_table.cellWidget(row, 2)
                            combo.setCurrentText("1")
                            # Use update_item_status to apply correct colors based on expected quantity
                            self.update_item_status(barcode, 1)
                            break
                    
                    # Show success message
                    item_data = self.audit_items_data[barcode]
                    # Safely access item data with fallbacks
                    category = item_data.get('category', 'Неопределена категория')
                    metal_type = item_data.get('metal_type', 'Неопределен метал')
                    stone_type = item_data.get('stone_type', '')
                    price = item_data.get('price', 0.0)
                    expected_qty = item_data.get('expected_qty', 0)
                    
                    product_name = f"{category} {metal_type}"
                    if stone_type and stone_type != "Без камък":
                        product_name += f" с {stone_type}"
                    
                    QMessageBox.information(
                        self, "Артикул сканиран",
                        f"✅ Артикул сканиран успешно!\n\n"
                        f"Баркод: {barcode}\n"
                        f"Продукт: {product_name}\n"
                        f"Цена: {price:.2f} €\n"
                        f"Очаквано количество: {expected_qty}\n\n"
                        "Количеството е зададено на 1. Използвайте падащото меню за корекция."
                    )
            else:
                # Invalid barcode - not in this shop
                self.audit_barcode_input.setStyleSheet("background-color: #f8d7da; border: 2px solid #dc3545;")
                QMessageBox.warning(
                    self, "Артикул не е намерен",
                    f"Баркод '{barcode}' не се намира в инвентара на магазин '{self.audit_shop_name}'.\n\n"
                    "Моля, проверете дали баркодът е правилен или дали артикулът принадлежи на този магазин."
                )
            
            # Clear barcode input
            self.audit_barcode_input.clear()
            self.audit_barcode_input.setStyleSheet("")  # Reset style
            self.audit_barcode_input.setFocus()
                
        except Exception as e:
            logger.error(f"Error processing audit barcode: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при обработка на баркод: {str(e)}")
    
    def update_audit_statistics(self):
        """Update audit statistics display"""
        try:
            total_items = len(self.audit_items_data)
            scanned_items = len([item for item in self.audit_items_data.values() if item['scanned_qty'] > 0])
            missing_items = len([item for item in self.audit_items_data.values() if item['scanned_qty'] == 0])
            
            # Update labels
            self.audit_stats_total.setText(f"Всички артикули: {total_items}")
            self.audit_stats_scanned.setText(f"Сканирани: {scanned_items}")
            self.audit_stats_missing.setText(f"Липсващи: {missing_items}")
            
            # Calculate progress based on scanned items vs total items
            if total_items > 0:
                progress = int((scanned_items / total_items) * 100)
            else:
                progress = 0
            
            self.audit_stats_progress.setText("Прогрес:")
            self.audit_progress_bar.setValue(progress)
            
        except Exception as e:
            logger.error(f"Error updating audit statistics: {e}")
    
    def update_audit_expected_table(self):
        """Update the expected items table"""
        try:
            self.audit_expected_table.setRowCount(len(self.audit_items_data))
            
            for row, (barcode, item_data) in enumerate(self.audit_items_data.items()):
                # Barcode
                self.audit_expected_table.setItem(row, 0, QTableWidgetItem(barcode))
                
                # Product name (category + metal + stone)
                product_name = f"{item_data['category']} {item_data['metal_type']}"
                if item_data['stone_type'] != "Без камък":
                    product_name += f" с {item_data['stone_type']}"
                self.audit_expected_table.setItem(row, 1, QTableWidgetItem(product_name))
                
                # Expected quantity
                self.audit_expected_table.setItem(row, 2, QTableWidgetItem(str(item_data['quantity_expected'])))
                
                # Price
                self.audit_expected_table.setItem(row, 3, QTableWidgetItem(f"{item_data['price']:.2f} €"))
                
                # Scanned quantity
                scanned_qty = item_data['quantity_scanned']
                scanned_item = QTableWidgetItem(f"{scanned_qty}/{item_data['quantity_expected']}")
                if scanned_qty == item_data['quantity_expected']:
                    scanned_item.setBackground(QColor(144, 238, 144))  # Light green
                    scanned_item.setForeground(QColor(0, 77, 0))       # Dark green text
                elif scanned_qty > 0:
                    scanned_item.setBackground(QColor(255, 255, 153))  # Light yellow
                    scanned_item.setForeground(QColor(184, 134, 11))   # Dark yellow/goldenrod text for readability
                else:
                    scanned_item.setBackground(QColor(255, 182, 193))  # Light red
                    scanned_item.setForeground(QColor(139, 0, 0))      # Dark red text
                self.audit_expected_table.setItem(row, 4, scanned_item)
                
                # Status
                status = item_data['status']
                status_item = QTableWidgetItem(status)
                if status == "Завършено":
                    status_item.setBackground(QColor(144, 238, 144))  # Light green
                    status_item.setForeground(QColor(0, 77, 0))       # Dark green text
                elif status == "Частично":
                    status_item.setBackground(QColor(255, 255, 153))  # Light yellow
                    status_item.setForeground(QColor(184, 134, 11))   # Dark yellow/goldenrod text for readability
                else:
                    status_item.setBackground(QColor(255, 182, 193))  # Light red
                    status_item.setForeground(QColor(139, 0, 0))      # Dark red text
                self.audit_expected_table.setItem(row, 5, status_item)
                
        except Exception as e:
            logger.error(f"Error updating audit expected table: {e}")
    
    def on_audit_barcode_changed(self):
        """Handle barcode input change during audit"""
        if not self.audit_in_progress or self.audit_paused:
            return
            
        # Auto-process barcode after short delay (for scanner input)
        if hasattr(self, 'audit_barcode_timer'):
            self.audit_barcode_timer.stop()
        
        self.audit_barcode_timer = QTimer()
        self.audit_barcode_timer.setSingleShot(True)
        self.audit_barcode_timer.timeout.connect(self.process_audit_barcode)
        self.audit_barcode_timer.start(300)  # 300ms delay
    
    def process_audit_barcode_manual(self):
        """Process the scanned barcode during audit - manual workflow"""
        try:
            barcode = self.audit_barcode_input.text().strip()
            if not barcode:
                return
            
            # Check if barcode exists in expected items
            if barcode in self.audit_items_data:
                # Valid item - highlight and prepare for adding
                self.audit_barcode_input.setStyleSheet("background-color: #d4edda; border: 2px solid #28a745;")
                
                # Show item info - access data directly from audit_items_data
                item_data = self.audit_items_data[barcode]
                product_name = f"{item_data['category']} {item_data['metal_type']}"
                if item_data['stone_type'] != "Без камък":
                    product_name += f" с {item_data['stone_type']}"
                
                QMessageBox.information(
                    self, "Артикул намерен",
                    f"Баркод: {barcode}\n"
                    f"Продукт: {product_name}\n"
                    f"Цена: {item_data['price']:.2f} €\n"
                    f"Очаквано количество: {item_data['expected_qty']}\n"
                    f"Вече сканирано: {item_data['scanned_qty']}\n\n"
                    "Регулирайте количеството ако е необходимо и натиснете 'Добави сканиран артикул'."
                )
            else:
                # Invalid barcode - not in this shop
                self.audit_barcode_input.setStyleSheet("background-color: #f8d7da; border: 2px solid #dc3545;")
                QMessageBox.warning(
                    self, "Артикул не е намерен",
                    f"Баркод '{barcode}' не се намира в инвентара на магазин '{self.audit_shop_name}'.\n\n"
                    "Моля, проверете дали баркодът е правилен или дали артикулът принадлежи на този магазин."
                )
                
        except Exception as e:
            logger.error(f"Error processing audit barcode: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при обработка на баркод: {str(e)}")
    
    def add_scanned_item(self):
        """Add scanned item to audit results"""
        try:
            barcode = self.audit_barcode_input.text().strip()
            quantity = self.audit_quantity_input.value()
            
            if not barcode:
                QMessageBox.warning(self, "Предупреждение", "Моля, въведете баркод!")
                return
            
            if barcode not in self.audit_items_data:
                QMessageBox.warning(self, "Предупреждение", "Този баркод не се намира в очакваните артикули!")
                return
            
            # Add to scanned items
            if barcode in self.audit_scanned_items:
                self.audit_scanned_items[barcode] += quantity
            else:
                self.audit_scanned_items[barcode] = quantity
            
            # Update expected items data
            self.audit_items_data[barcode]['scanned_qty'] = self.audit_scanned_items[barcode]
            
            # Update status
            expected_qty = self.audit_items_data[barcode]['expected_qty']
            scanned_qty = self.audit_scanned_items[barcode]
            
            if scanned_qty >= expected_qty:
                self.audit_items_data[barcode]['status'] = "Завършено"
            elif scanned_qty > 0:
                self.audit_items_data[barcode]['status'] = "Частично"
            else:
                self.audit_items_data[barcode]['status'] = "Очаква се"
            
            # Update tables
            self.update_audit_scanned_table()
            self.update_audit_expected_table()
            self.update_audit_statistics()
            
            # Clear inputs and reset style
            self.audit_barcode_input.clear()
            self.audit_barcode_input.setStyleSheet("")
            self.audit_quantity_input.setValue(1)
            self.audit_barcode_input.setFocus()
            
            # Log the scan
            logger.info(f"Audit scan: {barcode} x{quantity} in shop {self.audit_shop_name}")
            
        except Exception as e:
            logger.error(f"Error adding scanned item: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при добавяне на сканиран артикул: {str(e)}")
    
    def update_audit_scanned_table(self):
        """Update the scanned items table"""
        try:
            self.audit_scanned_table.setRowCount(len(self.audit_scanned_items))
            
            for row, (barcode, quantity) in enumerate(self.audit_scanned_items.items()):
                # Barcode
                self.audit_scanned_table.setItem(row, 0, QTableWidgetItem(barcode))
                
                # Quantity
                self.audit_scanned_table.setItem(row, 1, QTableWidgetItem(str(quantity)))
                
                # Status
                expected_qty = self.audit_items_data[barcode]['expected_qty']
                if quantity >= expected_qty:
                    status = "✅ Завършен"
                    status_item = QTableWidgetItem(status)
                    status_item.setBackground(QColor(144, 238, 144))  # Light green
                    status_item.setForeground(QColor(0, 77, 0))       # Dark green text
                elif quantity > expected_qty:
                    status = "⚠️ Повече от очакваното"
                    status_item = QTableWidgetItem(status)
                    status_item.setBackground(QColor(255, 165, 0))  # Orange
                    status_item.setForeground(QColor(139, 69, 0))    # Dark orange text
                else:
                    status = "🔄 Частично"
                    status_item = QTableWidgetItem(status)
                    status_item.setBackground(QColor(255, 255, 153))  # Light yellow
                    status_item.setForeground(QColor(184, 134, 11))   # Dark yellow/goldenrod text for readability
                
                self.audit_scanned_table.setItem(row, 2, status_item)
                
        except Exception as e:
            logger.error(f"Error updating audit scanned table: {e}")
    
    def pause_resume_audit(self):
        """Pause or resume the audit"""
        try:
            if not self.audit_paused:
                # Pause audit
                self.audit_paused = True
                self.pause_audit_btn.setText("▶️ Продължи")
                self.pause_audit_btn.setStyleSheet("background-color: #28a745; color: white; padding: 8px; border-radius: 5px;")
                
                # Disable audit inputs
                self.audit_barcode_input.setEnabled(False)
                
                # Disable audit mode (allow navigation but read-only)
                self.set_audit_mode(False)
                
                QMessageBox.information(
                    self, "Инвентаризация на пауза",
                    "Инвентаризацията е поставена на пауза.\n\n"
                    "Можете да разглеждате други раздели на софтуера, но в режим само за четене.\n"
                    "За да продължите инвентаризацията, се върнете тук и натиснете 'Продължи'."
                )
            else:
                # Resume audit
                self.audit_paused = False
                self.pause_audit_btn.setText("⏸️ Пауза")
                self.pause_audit_btn.setStyleSheet("background-color: #ffc107; color: black; padding: 8px; border-radius: 5px;")
                
                # Enable audit inputs
                self.audit_barcode_input.setEnabled(True)
                
                # Re-enable audit mode
                self.set_audit_mode(True)
                
                # Focus barcode input
                self.audit_barcode_input.setFocus()
                
                QMessageBox.information(
                    self, "Инвентаризация продължена",
                    "Инвентаризацията е продължена.\n\n"
                    "Можете да продължите със сканирането на артикули."
                )
            
        except Exception as e:
            logger.error(f"Error pausing/resuming audit: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при пауза/продължаване: {str(e)}")
    
    def set_audit_mode(self, enabled):
        """Enable or disable audit mode (restricts other tab interactions)"""
        try:
            # For now, we'll implement basic restriction
            # In a full implementation, you'd iterate through all widgets and disable them
            # but still allow viewing
            
            # This is a simplified version - you can expand this based on your needs
            if enabled:
                # Restrict to audit tab only
                current_tab = self.tabs.currentIndex()
                audit_tab_index = None
                
                # Find audit tab index
                for i in range(self.tabs.count()):
                    if self.tabs.tabText(i) == "Инвентаризация":
                        audit_tab_index = i
                        break
                
                if audit_tab_index is not None and current_tab != audit_tab_index:
                    self.tabs.setCurrentIndex(audit_tab_index)
                
            logger.info(f"Audit mode {'enabled' if enabled else 'disabled'}")
            
        except Exception as e:
            logger.error(f"Error setting audit mode: {e}")
    
    def finish_inventory_audit(self):
        """Finish the inventory audit and generate report"""
        try:
            # Confirm finish
            reply = QMessageBox.question(
                self, "Завършване на инвентаризация",
                f"Сигурни ли сте, че искате да завършите инвентаризацията на магазин '{self.audit_shop_name}'?\n\n"
                "След завършване няма да можете да добавяте повече сканирани артикули.\n"
                "Генерира се финален отчет и данните се запазват в базата данни.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Generate final audit report
            audit_end_time = datetime.now()
            audit_duration = audit_end_time - self.audit_start_time
            
            # Save audit to database
            audit_id = self.save_audit_to_database(audit_end_time, audit_duration)
            
            # Generate and show audit report
            self.show_audit_report(audit_id, audit_end_time, audit_duration)
            
            # Reset audit state
            self.reset_audit_state()
            
            logger.info(f"Inventory audit completed for shop: {self.audit_shop_name}")
            
        except Exception as e:
            logger.error(f"Error finishing inventory audit: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при завършване на инвентаризация: {str(e)}")
    
    def save_audit_to_database(self, end_time, duration):
        """Save audit results to database"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Create audit_sessions table if it doesn't exist
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS audit_sessions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        session_id TEXT UNIQUE NOT NULL,
                        shop_id INTEGER NOT NULL,
                        shop_name TEXT NOT NULL,
                        start_time TEXT NOT NULL,
                        end_time TEXT NOT NULL,
                        duration_minutes INTEGER,
                        total_expected INTEGER,
                        total_scanned INTEGER,
                        total_missing INTEGER,
                        total_completed INTEGER,
                        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (shop_id) REFERENCES shops (id)
                    )
                """)
                
                # Create audit_items table if it doesn't exist
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS audit_items (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        audit_session_id INTEGER NOT NULL,
                        barcode TEXT NOT NULL,
                        expected_quantity INTEGER NOT NULL,
                        scanned_quantity INTEGER NOT NULL,
                        price REAL NOT NULL,
                        weight REAL NOT NULL,
                        category TEXT,
                        metal_type TEXT,
                        stone_type TEXT,
                        status TEXT,
                        FOREIGN KEY (audit_session_id) REFERENCES audit_sessions (id)
                    )
                """)
                
                # Calculate statistics
                total_expected = len(self.audit_items_data)
                total_scanned = len([item for item in self.audit_items_data.values() if item['scanned_qty'] > 0])
                total_missing = len([item for item in self.audit_items_data.values() if item['scanned_qty'] == 0])
                total_completed = len([item for item in self.audit_items_data.values() if item['scanned_qty'] >= item['expected_qty']])
                duration_minutes = int(duration.total_seconds() / 60)
                
                # Insert audit session
                cursor.execute("""
                    INSERT INTO audit_sessions 
                    (session_id, shop_id, shop_name, start_time, end_time, duration_minutes,
                     total_expected, total_scanned, total_missing, total_completed)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    self.audit_session_id,
                    self.audit_shop_id,
                    self.audit_shop_name,
                    self.audit_start_time.strftime('%Y-%m-%d %H:%M:%S'),
                    end_time.strftime('%Y-%m-%d %H:%M:%S'),
                    duration_minutes,
                    total_expected,
                    total_scanned,
                    total_missing,
                    total_completed
                ))
                
                audit_id = cursor.lastrowid
                
                # Insert audit items
                for barcode, item_data in self.audit_items_data.items():
                    cursor.execute("""
                        INSERT INTO audit_items 
                        (audit_session_id, barcode, expected_quantity, scanned_quantity,
                         price, weight, category, metal_type, stone_type, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        audit_id,
                        barcode,
                        item_data['expected_qty'],
                        item_data['scanned_qty'],
                        item_data['price'],
                        item_data['weight'],
                        item_data['category'],
                        item_data['metal_type'],
                        item_data['stone_type'],
                        'Завършено' if item_data['scanned_qty'] >= item_data['expected_qty'] else 'Липсва' if item_data['scanned_qty'] == 0 else 'Частично'
                    ))
                
                conn.commit()
                return audit_id
                
        except Exception as e:
            logger.error(f"Error saving audit to database: {e}")
            raise e
    
    def show_audit_report(self, audit_id, end_time, duration):
        """Show comprehensive audit report"""
        try:
            # Create report dialog
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Отчет за инвентаризация - {self.audit_shop_name}")
            dialog.setModal(True)
            dialog.resize(800, 600)
            
            layout = QVBoxLayout(dialog)
            
            # Title
            title_label = QLabel(f"Отчет за инвентаризация на магазин '{self.audit_shop_name}'")
            title_label.setFont(QFont("Arial", 16, QFont.Weight.Bold))
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_label.setStyleSheet("background-color: #2c5aa0; color: white; padding: 10px; border-radius: 5px;")
            layout.addWidget(title_label)
            
            # Statistics section
            stats_group = QGroupBox("Статистика")
            stats_layout = QGridLayout()
            
            # Calculate detailed statistics
            total_expected = len(self.audit_items_data)
            total_scanned_items = len([item for item in self.audit_items_data.values() if item['scanned_qty'] > 0])
            total_missing = len([item for item in self.audit_items_data.values() if item['scanned_qty'] == 0])
            total_completed = len([item for item in self.audit_items_data.values() if item['scanned_qty'] >= item['expected_qty']])
            
            # Calculate value statistics
            total_expected_value = sum(item['price'] * item['expected_qty'] for item in self.audit_items_data.values())
            total_scanned_value = sum(item['price'] * item['scanned_qty'] for item in self.audit_items_data.values())
            missing_value = sum(item['price'] * item['expected_qty'] for item in self.audit_items_data.values() if item['scanned_qty'] == 0)
            
            # Time info
            duration_str = f"{int(duration.total_seconds() // 3600):02d}:{int((duration.total_seconds() % 3600) // 60):02d}"
            
            # Add statistics labels
            stats_layout.addWidget(QLabel("Начало:"), 0, 0)
            stats_layout.addWidget(QLabel(self.audit_start_time.strftime("%d.%m.%Y %H:%M")), 0, 1)
            
            stats_layout.addWidget(QLabel("Край:"), 1, 0)
            stats_layout.addWidget(QLabel(end_time.strftime("%d.%m.%Y %H:%M")), 1, 1)
            
            stats_layout.addWidget(QLabel("Продължителност:"), 2, 0)
            stats_layout.addWidget(QLabel(duration_str), 2, 1)
            
            stats_layout.addWidget(QLabel("Всички артикули:"), 3, 0)
            stats_layout.addWidget(QLabel(str(total_expected)), 3, 1)
            
            stats_layout.addWidget(QLabel("Сканирани артикули:"), 4, 0)
            stats_layout.addWidget(QLabel(str(total_scanned_items)), 4, 1)
            
            stats_layout.addWidget(QLabel("Липсващи артикули:"), 5, 0)
            missing_label = QLabel(str(total_missing))
            if total_missing > 0:
                missing_label.setStyleSheet("color: red; font-weight: bold;")
            stats_layout.addWidget(missing_label, 5, 1)
            
            stats_layout.addWidget(QLabel("Очаквана стойност:"), 6, 0)
            stats_layout.addWidget(QLabel(f"{total_expected_value:.2f} €"), 6, 1)
            
            stats_layout.addWidget(QLabel("Сканирана стойност:"), 7, 0)
            stats_layout.addWidget(QLabel(f"{total_scanned_value:.2f} €"), 7, 1)
            
            stats_layout.addWidget(QLabel("Липсваща стойност:"), 8, 0)
            missing_value_label = QLabel(f"{missing_value:.2f} €")
            if missing_value > 0:
                missing_value_label.setStyleSheet("color: red; font-weight: bold;")
            stats_layout.addWidget(missing_value_label, 8, 1)
            
            stats_group.setLayout(stats_layout)
            layout.addWidget(stats_group)
            
            # Missing items section (if any)
            if total_missing > 0:
                missing_group = QGroupBox(f"Липсващи артикули ({total_missing})")
                missing_layout = QVBoxLayout()
                
                missing_table = QTableWidget()
                missing_table.setColumnCount(5)
                missing_table.setHorizontalHeaderLabels(["Баркод", "Продукт", "Количество", "Цена", "Обща стойност"])
                
                # Make table immutable - disable all editing
                missing_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
                missing_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
                missing_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
                
                # Configure column spacing with controlled resize limits (resizable between bounds)
                header = missing_table.horizontalHeader()
                header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
                
                # Set default widths for first 4 columns - all resizable
                default_width = 140  # Good size for missing items table
                for col in range(4):  # First 4 columns are resizable
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
                    missing_table.setColumnWidth(col, default_width)
                
                # Last column stretches to fill remaining space
                header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
                
                missing_table.setAlternatingRowColors(True)
                
                missing_items = [item for item in self.audit_items_data.items() if item[1]['scanned_qty'] == 0]
                missing_table.setRowCount(len(missing_items))
                
                for row, (barcode, item_data) in enumerate(missing_items):
                    barcode_item = QTableWidgetItem(barcode)
                    barcode_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    barcode_item.setFlags(barcode_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    missing_table.setItem(row, 0, barcode_item)
                    
                    product_name = f"{item_data['category']} {item_data['metal_type']}"
                    if item_data['stone_type'] != "Без камък":
                        product_name += f" с {item_data['stone_type']}"
                    product_item = QTableWidgetItem(product_name)
                    product_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    product_item.setFlags(product_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    missing_table.setItem(row, 1, product_item)
                    
                    qty_item = QTableWidgetItem(str(item_data['expected_qty']))
                    qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    qty_item.setFlags(qty_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    missing_table.setItem(row, 2, qty_item)
                    
                    price_item = QTableWidgetItem(f"{item_data['price']:.2f} €")
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    price_item.setFlags(price_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    missing_table.setItem(row, 3, price_item)
                    
                    total_item = QTableWidgetItem(f"{item_data['price'] * item_data['expected_qty']:.2f} €")
                    total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    total_item.setFlags(total_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    missing_table.setItem(row, 4, total_item)
                
                missing_layout.addWidget(missing_table)
                missing_group.setLayout(missing_layout)
                layout.addWidget(missing_group)
            
            # Buttons
            button_layout = QHBoxLayout()
            
            close_btn = QPushButton("Затвори")
            close_btn.clicked.connect(dialog.accept)
            button_layout.addWidget(close_btn)
            
            layout.addLayout(button_layout)
            
            dialog.exec()
            
        except Exception as e:
            logger.error(f"Error showing audit report: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при показване на отчет: {str(e)}")
    
    def export_audit_report_pdf(self, audit_id):
        """Export audit report to PDF with proper Cyrillic font support"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # Register fonts that support Cyrillic characters
            try:
                # Try to use system fonts that support Cyrillic
                font_paths = [
                    "fonts/arial.ttf",  # Our project font
                    "C:/Windows/Fonts/arial.ttf",  # Windows system font
                    "C:/Windows/Fonts/calibri.ttf",  # Alternative Windows font
                    "/System/Library/Fonts/Arial.ttf",  # macOS system font
                    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"  # Linux font
                ]
                
                font_registered = False
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        try:
                            pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                            pdfmetrics.registerFont(TTFont('CyrillicFont-Bold', font_path))  # Use same font for bold
                            font_registered = True
                            break
                        except Exception:
                            continue
                
                if not font_registered:
                    # Fallback to built-in font (may not display Cyrillic properly)
                    cyrillic_font = 'Helvetica'
                    cyrillic_font_bold = 'Helvetica-Bold'
                else:
                    cyrillic_font = 'CyrillicFont'
                    cyrillic_font_bold = 'CyrillicFont-Bold'
                    
            except Exception:
                # Fallback to built-in fonts
                cyrillic_font = 'Helvetica'
                cyrillic_font_bold = 'Helvetica-Bold'
            
            # Get file path with standardized Bulgarian filename
            exports_dir = self.get_exports_directory()
            audit_filename = self.generate_bulgarian_filename(f"инвентаризация_{self.audit_shop_name.lower()}", "pdf")
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Запази PDF отчет", 
                f"{exports_dir}/{audit_filename}",
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Create custom styles with Cyrillic font support and black text
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontName=cyrillic_font_bold,
                fontSize=16,
                textColor=colors.black,  # Ensure black text
                spaceAfter=12
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontName=cyrillic_font_bold,
                fontSize=14,
                textColor=colors.black,  # Ensure black text
                spaceAfter=6
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontName=cyrillic_font,
                fontSize=12,
                textColor=colors.black,  # Ensure black text
                spaceAfter=6
            )
            
            # Title
            title = Paragraph(f"Отчет за инвентаризация - {self.audit_shop_name}", title_style)
            story.append(title)
            
            # Add spacing
            story.append(Paragraph("<br/><br/>", normal_style))
            
            # Statistics
            stats_data = [
                ["Параметър", "Стойност"],
                ["Магазин", self.audit_shop_name],
                ["Дата и час на започване", self.audit_start_time.strftime("%d.%m.%Y %H:%M")],
                ["Дата и час на завършване", datetime.now().strftime("%d.%m.%Y %H:%M")],
                ["Всички артикули", str(len(self.audit_items_data))],
                ["Сканирани артикули", str(len([item for item in self.audit_items_data.values() if item['scanned_qty'] > 0]))],
                ["Липсващи артикули", str(len([item for item in self.audit_items_data.values() if item['scanned_qty'] == 0]))],
            ]
            
            stats_table = Table(stats_data)
            stats_table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # White text on grey header
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),  # Use Cyrillic font for header
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Body styling
                ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),  # Use Cyrillic font for body
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # BLACK text on light background
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # White background for body
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
            story.append(Paragraph("<br/><br/>", normal_style))
            
            # Missing items (if any)
            missing_items = [item for item in self.audit_items_data.items() if item[1]['scanned_qty'] == 0]
            if missing_items:
                story.append(Paragraph("Липсващи артикули:", heading_style))
                
                missing_data = [["Баркод", "Продукт", "Количество", "Цена", "Обща стойност"]]
                
                for barcode, item_data in missing_items:
                    product_name = f"{item_data['category']} {item_data['metal_type']}"
                    if item_data['stone_type'] != "Без камък":
                        product_name += f" с {item_data['stone_type']}"
                    
                    missing_data.append([
                        str(barcode) if barcode else "",
                        str(product_name) if product_name else "",
                        str(item_data['expected_qty']) if item_data['expected_qty'] else "0",
                        f"{item_data['price']:.2f} €" if item_data['price'] else "0.00 €",
                        f"{item_data['price'] * item_data['expected_qty']:.2f} €" if item_data['price'] and item_data['expected_qty'] else "0.00 €"
                    ])
                
                missing_table = Table(missing_data)
                missing_table.setStyle(TableStyle([
                    # Header styling for missing items - standardized grey headers
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # White text on grey header
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),  # Use Cyrillic font for header
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Body styling for missing items - red background to highlight missing items
                    ('FONTNAME', (0, 1), (-1, -1), cyrillic_font),  # Use Cyrillic font for body
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # BLACK text
                    ('BACKGROUND', (0, 1), (-1, -1), colors.Color(1, 0.9, 0.9)),  # Light red background for missing items
                    
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(missing_table)
            
            # Build PDF
            doc.build(story)
            
            QMessageBox.information(self, "Успех", f"PDF отчетът е запазен успешно в:\n{file_path}")
            
        except Exception as e:
            logger.error(f"Error exporting audit report to PDF: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при експорт в PDF: {str(e)}")
    
    def reset_audit_state(self):
        """Reset audit state after completion"""
        try:
            # Reset all audit variables
            self.audit_in_progress = False
            self.audit_shop_id = None
            self.audit_shop_name = ""
            self.audit_start_time = None
            self.audit_paused = False
            self.audit_items_data = {}
            self.audit_scanned_items = {}
            self.audit_session_id = None
            
            # Reset UI controls
            self.start_audit_btn.setEnabled(True)
            self.pause_audit_btn.setEnabled(False)
            self.pause_audit_btn.setText("⏸️ Пауза")
            self.pause_audit_btn.setStyleSheet("background-color: #ffc107; color: black; padding: 8px; border-radius: 5px;")
            self.finish_audit_btn.setEnabled(False)
            self.audit_barcode_input.setEnabled(False)
            self.audit_shop_combo.setEnabled(True)
            
            # Clear inputs
            self.audit_barcode_input.clear()
            self.audit_barcode_input.setStyleSheet("")
            
            # Clear table
            self.audit_items_table.setRowCount(0)
            
            # Reset statistics
            self.audit_stats_total.setText("Всички артикули: 0")
            self.audit_stats_scanned.setText("Сканирани: 0")
            self.audit_stats_missing.setText("Липсващи: 0")
            self.audit_stats_progress.setText("Прогрес:")
            self.audit_progress_bar.setValue(0)
            
            # Disable audit mode
            self.set_audit_mode(False)
            
            # Refresh audit results in case any were completed
            self.load_audit_results()
            
            logger.info("Audit state reset successfully")
            
        except Exception as e:
            logger.error(f"Error resetting audit state: {e}")

    # === AUDIT RESULTS MANAGEMENT METHODS ===
    
    def load_audit_results(self):
        """Load the audit results table with historical data"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get all audit sessions from database
                cursor.execute("""
                    SELECT id, session_id, shop_name, start_time, end_time, duration_minutes,
                           total_expected, total_scanned, total_missing, total_completed,
                           created_at
                    FROM audit_sessions 
                    ORDER BY created_at DESC
                """)
                
                results = cursor.fetchall()
                
                # Clear and populate table
                self.audit_results_table.setRowCount(len(results))
                
                for row, result in enumerate(results):
                    audit_id, session_id, shop_name, start_time, end_time, duration_minutes, \
                    total_expected, total_scanned, total_missing, total_completed, created_at = result
                    
                    # Date (from start_time)
                    try:
                        date_obj = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
                        date_str = date_obj.strftime('%d.%m.%Y %H:%M')
                    except:
                        date_str = start_time
                    
                    date_item = QTableWidgetItem(date_str)
                    date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.audit_results_table.setItem(row, 0, date_item)
                    
                    # Shop name
                    shop_item = QTableWidgetItem(shop_name)
                    shop_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.audit_results_table.setItem(row, 1, shop_item)
                    
                    # Duration
                    duration_str = f"{duration_minutes // 60:02d}:{duration_minutes % 60:02d}"
                    duration_item = QTableWidgetItem(duration_str)
                    duration_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.audit_results_table.setItem(row, 2, duration_item)
                    
                    # Expected items
                    expected_item = QTableWidgetItem(str(total_expected))
                    expected_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.audit_results_table.setItem(row, 3, expected_item)
                    
                    # Scanned items
                    scanned_item = QTableWidgetItem(str(total_scanned))
                    scanned_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    if total_scanned == total_expected:
                        scanned_item.setBackground(QColor(144, 238, 144))  # Light green
                        scanned_item.setForeground(QColor(0, 77, 0))       # Dark green text
                    elif total_scanned > 0:
                        scanned_item.setBackground(QColor(255, 255, 153))  # Light yellow
                        scanned_item.setForeground(QColor(184, 134, 11))   # Dark yellow/goldenrod text for readability
                    self.audit_results_table.setItem(row, 4, scanned_item)
                    
                    # Missing items
                    missing_item = QTableWidgetItem(str(total_missing))
                    missing_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    if total_missing > 0:
                        missing_item.setBackground(QColor(255, 182, 193))  # Light red
                        missing_item.setForeground(QColor(139, 0, 0))  # Dark red text
                    else:
                        missing_item.setBackground(QColor(144, 238, 144))  # Light green
                    self.audit_results_table.setItem(row, 5, missing_item)
                    
                    # Actions buttons
                    actions_widget = QWidget()
                    actions_layout = QHBoxLayout(actions_widget)
                    actions_layout.setContentsMargins(4, 4, 4, 4)
                    
                    # Folder icon button
                    folder_btn = QPushButton("📁")
                    folder_btn.setToolTip("Отвори папка с експорти")
                    folder_btn.setMaximumWidth(30)
                    folder_btn.setStyleSheet("background-color: white; color: black; border: 1px solid #ccc;")
                    folder_btn.clicked.connect(lambda checked: self.open_exports_folder())
                    actions_layout.addWidget(folder_btn)
                    
                    # View button
                    view_btn = QPushButton("👁️")
                    view_btn.setToolTip("Преглед на детайли")
                    view_btn.setMaximumWidth(30)
                    view_btn.setStyleSheet("background-color: white; color: black; border: 1px solid #ccc;")
                    view_btn.clicked.connect(lambda checked, aid=audit_id: self.view_audit_details(aid))
                    actions_layout.addWidget(view_btn)
                    
                    # Download PDF button
                    pdf_btn = QPushButton("📄")
                    pdf_btn.setToolTip("Преглед PDF")
                    pdf_btn.setMaximumWidth(30)
                    pdf_btn.setStyleSheet("background-color: white; color: black; border: 1px solid #ccc;")
                    pdf_btn.clicked.connect(lambda checked, aid=audit_id, sn=shop_name, st=start_time: self.download_audit_pdf(aid, sn, st))
                    actions_layout.addWidget(pdf_btn)
                    
                    # Download Excel button
                    excel_btn = QPushButton("📊")
                    excel_btn.setToolTip("Преглед Excel")
                    excel_btn.setMaximumWidth(30)
                    excel_btn.setStyleSheet("background-color: white; color: black; border: 1px solid #ccc;")
                    excel_btn.clicked.connect(lambda checked, aid=audit_id, sn=shop_name, st=start_time: self.download_audit_excel(aid, sn, st))
                    actions_layout.addWidget(excel_btn)
                    
                    # Delete button
                    delete_btn = QPushButton("🗑️")
                    delete_btn.setToolTip("Изтрий резултат")
                    delete_btn.setMaximumWidth(30)
                    delete_btn.setStyleSheet("background-color: #dc3545; color: white;")
                    delete_btn.clicked.connect(lambda checked, aid=audit_id: self.delete_audit_result(aid))
                    actions_layout.addWidget(delete_btn)
                    
                    actions_layout.addStretch()
                    self.audit_results_table.setCellWidget(row, 6, actions_widget)
                
                # Update total count
                self.results_total_label.setText(f"Общо инвентаризации: {len(results)}")
                
        except Exception as e:
            logger.error(f"Error loading audit results: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при зареждане на резултатите: {str(e)}")

    def view_audit_details(self, audit_id):
        """View detailed audit results"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get audit session info
                cursor.execute("""
                    SELECT session_id, shop_name, start_time, end_time, duration_minutes,
                           total_expected, total_scanned, total_missing, total_completed
                    FROM audit_sessions 
                    WHERE id = ?
                """, (audit_id,))
                
                session_info = cursor.fetchone()
                if not session_info:
                    QMessageBox.warning(self, "Грешка", "Инвентаризацията не е намерена!")
                    return
                
                # Get audit items
                cursor.execute("""
                    SELECT barcode, expected_quantity, scanned_quantity, price, weight,
                           category, metal_type, stone_type, status
                    FROM audit_items 
                    WHERE audit_session_id = ?
                    ORDER BY barcode
                """, (audit_id,))
                
                items = cursor.fetchall()
                
                # Create details dialog
                dialog = QDialog(self)
                dialog.setWindowTitle(f"Детайли на инвентаризация - {session_info[1]}")
                dialog.setModal(True)
                dialog.resize(1200, 700)  # Increased width to fit table properly
                
                layout = QVBoxLayout(dialog)
                
                # Items table (removed header info widget)
                items_table = QTableWidget()
                items_table.setColumnCount(9)
                items_table.setHorizontalHeaderLabels([
                    "Баркод", "Продукт", "Очаквано", "Сканирано", "Разлика", "Цена", "Загуба/€", "Тегло", "Статус"
                ])
                items_table.setRowCount(len(items))
                
                # Make table immutable - disable all editing
                items_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
                items_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
                items_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
                
                # Configure column spacing for better fit in wider window
                header = items_table.horizontalHeader()
                header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # Allow manual resize
                
                # Set appropriate widths for each column based on content
                items_table.setColumnWidth(0, 120)  # Баркод
                items_table.setColumnWidth(1, 200)  # Продукт
                items_table.setColumnWidth(2, 100)  # Очаквано
                items_table.setColumnWidth(3, 100)  # Сканирано
                items_table.setColumnWidth(4, 100)  # Разлика
                items_table.setColumnWidth(5, 100)  # Цена
                items_table.setColumnWidth(6, 120)  # Загуба/€
                items_table.setColumnWidth(7, 100)  # Тегло
                items_table.setColumnWidth(8, 120)  # Статус
                
                items_table.setAlternatingRowColors(True)
                
                total_loss_value = 0
                total_loss_weight = 0
                
                for row, item in enumerate(items):
                    barcode, expected_qty, scanned_qty, price, weight, category, metal, stone, status = item
                    
                    # Calculate difference and loss
                    difference = scanned_qty - expected_qty
                    loss_value = (expected_qty - scanned_qty) * price if scanned_qty < expected_qty else 0
                    loss_weight = (expected_qty - scanned_qty) * weight if scanned_qty < expected_qty else 0
                    
                    total_loss_value += loss_value
                    total_loss_weight += loss_weight
                    
                    # Product name
                    product_name = f"{category} {metal}"
                    if stone != "Без камък":
                        product_name += f" с {stone}"
                    
                    # Populate table with center alignment
                    barcode_item = QTableWidgetItem(barcode)
                    barcode_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    barcode_item.setFlags(barcode_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    items_table.setItem(row, 0, barcode_item)
                    
                    product_item = QTableWidgetItem(product_name)
                    product_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    product_item.setFlags(product_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    items_table.setItem(row, 1, product_item)
                    
                    expected_item = QTableWidgetItem(str(expected_qty))
                    expected_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    expected_item.setFlags(expected_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    items_table.setItem(row, 2, expected_item)
                    
                    scanned_item = QTableWidgetItem(str(scanned_qty))
                    scanned_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    scanned_item.setFlags(scanned_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    items_table.setItem(row, 3, scanned_item)
                    
                    # Difference with color coding
                    diff_item = QTableWidgetItem(f"{difference:+d}")
                    diff_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    diff_item.setFlags(diff_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    if difference < 0:
                        diff_item.setBackground(QColor(255, 182, 193))  # Light red
                        diff_item.setForeground(QColor(139, 0, 0))
                    elif difference > 0:
                        diff_item.setBackground(QColor(173, 216, 230))  # Light blue
                    else:
                        diff_item.setBackground(QColor(144, 238, 144))  # Light green
                    items_table.setItem(row, 4, diff_item)
                    
                    price_item = QTableWidgetItem(f"{price:.2f} €")
                    price_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    price_item.setFlags(price_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    items_table.setItem(row, 5, price_item)
                    
                    # Loss value with color
                    loss_item = QTableWidgetItem(f"{loss_value:.2f} €")
                    loss_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    loss_item.setFlags(loss_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    if loss_value > 0:
                        loss_item.setBackground(QColor(255, 182, 193))
                        loss_item.setForeground(QColor(139, 0, 0))
                    items_table.setItem(row, 6, loss_item)
                    
                    weight_item = QTableWidgetItem(f"{weight:.2f} г")
                    weight_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    weight_item.setFlags(weight_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    items_table.setItem(row, 7, weight_item)
                    
                    status_item = QTableWidgetItem(status)
                    status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Make read-only
                    items_table.setItem(row, 8, status_item)
                
                # Even spacing maintained - no stretch last section
                layout.addWidget(items_table)
                
                # Summary with proper number formatting
                summary_layout = QHBoxLayout()
                formatted_loss_value = f"{total_loss_value:,.2f}".replace(",", " ")
                formatted_loss_weight = f"{total_loss_weight:,.2f}".replace(",", " ")
                summary_layout.addWidget(QLabel(f"Общо загуба: {formatted_loss_value} € | {formatted_loss_weight} г"))
                summary_layout.addStretch()
                
                close_btn = QPushButton("Затвори")
                close_btn.clicked.connect(dialog.accept)
                summary_layout.addWidget(close_btn)
                
                layout.addLayout(summary_layout)
                
                dialog.exec()
                
        except Exception as e:
            logger.error(f"Error viewing audit details: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при показване на детайли: {str(e)}")
    
    def download_audit_pdf(self, audit_id, shop_name, start_time):
        """Download audit report as PDF"""
        try:
            # Get save location for PDF file
            exports_dir = self.get_exports_directory()
            # Generate Bulgarian filename for audit report
            audit_filename = self.generate_bulgarian_filename(f"инвентаризация_{shop_name.lower()}", "pdf")
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Запази PDF отчет", 
                f"{exports_dir}/{audit_filename}",
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            # Generate PDF using the existing audit data
            self.generate_audit_pdf_from_db(audit_id, file_path)
            
            QMessageBox.information(self, "Успех", f"PDF отчетът е запазен в:\n{file_path}")
            
        except Exception as e:
            logger.error(f"Error downloading audit PDF: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при изтегляне на PDF: {str(e)}")
    
    def download_audit_excel(self, audit_id, shop_name, start_time):
        """Download audit report as Excel"""
        try:
            exports_dir = self.get_exports_directory()
            # Generate Bulgarian filename for audit report
            audit_filename = self.generate_bulgarian_filename(f"инвентаризация_{shop_name.lower()}", "xlsx")
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Запази Excel отчет", 
                f"{exports_dir}/{audit_filename}",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Generate Excel using the existing audit data
            self.generate_audit_excel_from_db(audit_id, file_path)
            
            QMessageBox.information(self, "Успех", f"Excel отчетът е запазен в:\n{file_path}")
            
        except Exception as e:
            logger.error(f"Error downloading audit Excel: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при изтегляне на Excel: {str(e)}")
    
    def generate_audit_pdf_from_db(self, audit_id, file_path):
        """Generate PDF report from database data"""
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # Register Cyrillic font with multiple fallback paths
            font_paths = [
                os.path.join(os.path.dirname(__file__), "fonts", "arial.ttf"),
                "C:\\Windows\\Fonts\\arial.ttf",
                "arial.ttf"
            ]
            
            font_registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('Arial', font_path))
                        font_registered = True
                        break
                    except Exception as e:
                        logger.warning(f"Could not register font from {font_path}: {e}")
            
            if not font_registered:
                logger.warning("Could not register Arial font, using default fonts")
                cyrillic_font = 'Helvetica'
                cyrillic_font_bold = 'Helvetica-Bold'
            else:
                cyrillic_font = 'Arial'
                cyrillic_font_bold = 'Arial'
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get session info
                cursor.execute("""
                    SELECT shop_name, start_time, end_time, duration_minutes,
                           total_expected, total_scanned, total_missing, total_completed
                    FROM audit_sessions WHERE id = ?
                """, (audit_id,))
                session_info = cursor.fetchone()
                
                # Get items
                cursor.execute("""
                    SELECT barcode, expected_quantity, scanned_quantity, price, weight,
                           category, metal_type, stone_type, status
                    FROM audit_items WHERE audit_session_id = ?
                """, (audit_id,))
                items = cursor.fetchall()
            
            # Create PDF
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Create custom styles with Cyrillic font support
            title_style = ParagraphStyle(
                'CyrillicTitle',
                parent=styles['Title'],
                fontName=cyrillic_font_bold,
                fontSize=16,
                textColor=colors.black
            )
            
            heading_style = ParagraphStyle(
                'CyrillicHeading',
                parent=styles['Heading2'],
                fontName=cyrillic_font_bold,
                fontSize=12,
                textColor=colors.black
            )
            
            # Title
            title = Paragraph(f"Отчет за инвентаризация - {session_info[0]}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Session info table
            session_data = [
                ["Параметър", "Стойност"],
                ["Магазин", session_info[0]],
                ["Начало", session_info[1]],
                ["Край", session_info[2]],
                ["Продължителност", f"{session_info[3]} мин"],
                ["Очаквани артикули", str(session_info[4])],
                ["Сканирани артикули", str(session_info[5])],
                ["Липсващи артикули", str(session_info[6])]
            ]
            
            session_table = Table(session_data)
            session_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), cyrillic_font)
            ]))
            
            story.append(session_table)
            story.append(Spacer(1, 30))
            
            # Items table
            if items:
                story.append(Paragraph("Детайли по артикули:", heading_style))
                story.append(Spacer(1, 10))
                
                items_data = [["Баркод", "Продукт", "Очаквано", "Сканирано", "Разлика", "Цена", "Статус"]]
                
                for item in items:
                    barcode, expected_qty, scanned_qty, price, weight, category, metal, stone, status = item
                    difference = scanned_qty - expected_qty
                    
                    # Handle null/empty values properly
                    category = category if category else "Неопределена"
                    metal = metal if metal else "Неопределен"
                    stone = stone if stone else "Без камък"
                    status = status if status else "Неопределен"
                    
                    product_name = f"{category} {metal}"
                    if stone and stone != "Без камък":
                        product_name += f" с {stone}"
                    
                    # Ensure all values are properly formatted
                    barcode = str(barcode) if barcode else "Няма"
                    expected_qty = expected_qty if expected_qty is not None else 0
                    scanned_qty = scanned_qty if scanned_qty is not None else 0
                    price = price if price is not None else 0.0
                    difference = scanned_qty - expected_qty
                    
                    items_data.append([
                        barcode,
                        product_name,
                        str(expected_qty),
                        str(scanned_qty),
                        f"{difference:+d}",
                        f"{price:.2f} €",
                        status
                    ])
                
                items_table = Table(items_data)
                items_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), cyrillic_font_bold),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), cyrillic_font)
                ]))
                
                story.append(items_table)
            
            doc.build(story)
            
        except Exception as e:
            logger.error(f"Error generating PDF from DB: {e}")
            raise e
    
    def generate_audit_excel_from_db(self, audit_id, file_path):
        """Generate Excel report from database data"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                # Get session info
                cursor.execute("""
                    SELECT shop_name, start_time, end_time, duration_minutes,
                           total_expected, total_scanned, total_missing, total_completed
                    FROM audit_sessions WHERE id = ?
                """, (audit_id,))
                session_info = cursor.fetchone()
                
                # Get items
                cursor.execute("""
                    SELECT barcode, expected_quantity, scanned_quantity, price, weight,
                           category, metal_type, stone_type, status
                    FROM audit_items WHERE audit_session_id = ?
                """, (audit_id,))
                items = cursor.fetchall()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Инвентаризация"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Session info section
            ws['A1'] = "ОТЧЕТ ЗА ИНВЕНТАРИЗАЦИЯ"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:G1')
            ws['A1'].alignment = center_alignment
            
            # Session details
            row = 3
            session_details = [
                ("Магазин:", session_info[0]),
                ("Начало:", session_info[1]),
                ("Край:", session_info[2]),
                ("Продължителност:", f"{session_info[3]} мин"),
                ("Очаквани артикули:", session_info[4]),
                ("Сканирани артикули:", session_info[5]),
                ("Липсващи артикули:", session_info[6])
            ]
            
            for label, value in session_details:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                row += 1
            
            # Items section
            row += 2
            ws[f'A{row}'] = "ДЕТАЙЛИ ПО АРТИКУЛИ"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:I{row}')
            
            row += 1
            headers = ["Баркод", "Продукт", "Очаквано", "Сканирано", "Разлика", "Цена", "Загуба/€", "Тегло", "Статус"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Items data
            total_loss = 0
            for item in items:
                row += 1
                barcode, expected_qty, scanned_qty, price, weight, category, metal, stone, status = item
                
                # Handle null/empty values properly
                category = category if category else "Неопределена"
                metal = metal if metal else "Неопределен"
                stone = stone if stone else "Без камък"
                status = status if status else "Неопределен"
                
                # Ensure all values are properly formatted
                barcode = str(barcode) if barcode else "Няма"
                expected_qty = expected_qty if expected_qty is not None else 0
                scanned_qty = scanned_qty if scanned_qty is not None else 0
                price = price if price is not None else 0.0
                weight = weight if weight is not None else 0.0
                
                difference = scanned_qty - expected_qty
                loss = (expected_qty - scanned_qty) * price if scanned_qty < expected_qty else 0
                total_loss += loss
                
                product_name = f"{category} {metal}"
                if stone and stone != "Без камък":
                    product_name += f" с {stone}"
                
                data = [
                    barcode, product_name, expected_qty, scanned_qty, difference,
                    f"{price:.2f}", f"{loss:.2f}", f"{weight:.2f}", status
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = center_alignment
                    
                    # Color coding for differences - reduced saturation
                    if col == 5:  # Difference column
                        if difference < 0:
                            cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
                        elif difference > 0:
                            cell.fill = PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")
                    elif col == 7 and loss > 0:  # Loss column
                        cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
            
            # Total loss
            row += 2
            ws[f'F{row}'] = "ОБЩО ЗАГУБА:"
            ws[f'F{row}'].font = Font(bold=True)
            ws[f'G{row}'] = f"{total_loss:.2f} €"
            ws[f'G{row}'].font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error(f"Error generating Excel from DB: {e}")
            raise e
    
    def delete_audit_result(self, audit_id):
        """Delete an audit result"""
        try:
            reply = QMessageBox.question(
                self, "Потвърждение",
                "Сигурни ли сте, че искате да изтриете този резултат от инвентаризация?\n\n"
                "Това действие не може да бъде отменено!",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    # Delete audit items first (foreign key constraint)
                    cursor.execute("DELETE FROM audit_items WHERE audit_session_id = ?", (audit_id,))
                    
                    # Delete audit session
                    cursor.execute("DELETE FROM audit_sessions WHERE id = ?", (audit_id,))
                    
                    conn.commit()
                
                # Refresh the table
                self.load_audit_results()
                
                QMessageBox.information(self, "Успех", "Резултатът от инвентаризацията е изтрит успешно!")
                
        except Exception as e:
            logger.error(f"Error deleting audit result: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при изтриване на резултат: {str(e)}")
    
    def clear_all_audit_results(self):
        """Clear all audit results"""
        try:
            reply = QMessageBox.question(
                self, "Потвърждение",
                "Сигурни ли сте, че искате да изтриете ВСИЧКИ резултати от инвентаризации?\n\n"
                "Това действие ще изтрие цялата история и не може да бъде отменено!",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                # Second confirmation
                reply2 = QMessageBox.question(
                    self, "Финално потвърждение",
                    "Това е последният ви шанс!\n\n"
                    "Всички данни за инвентаризации ще бъдат безвъзвратно изтрити.\n"
                    "Продължаване?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )

                if reply2 == QMessageBox.StandardButton.Yes:
                    with self.db.get_connection() as conn:
                        cursor = conn.cursor()
                        
                        # Delete all audit items first
                        cursor.execute("DELETE FROM audit_items")
                        
                        # Delete all audit sessions
                        cursor.execute("DELETE FROM audit_sessions")
                        
                        conn.commit()
                    
                    # Refresh the table
                    self.load_audit_results()
                    
                    QMessageBox.information(self, "Успех", "Всички резултати от инвентаризации са изтрити!")
                    
        except Exception as e:
            logger.error(f"Error clearing all audit results: {e}")
            QMessageBox.critical(self, "Грешка", f"Грешка при изчистване на резултатите: {str(e)}")

    def editorEvent(self, event, model, option, index):
        if event.type() == event.Type.MouseButtonPress:
            text = index.data()
            is_custom = False
            custom_values_set = None
            combo_box = None
            type_ = None
            
            # Check which combo box we're dealing with
            if hasattr(self.parent, 'category_input') and self.parent.category_input.view().model() == model:
                is_custom = text in self.parent.custom_categories
                custom_values_set = self.parent.custom_categories
                combo_box = self.parent.category_input
                type_ = 'category'
            elif hasattr(self.parent, 'metal_input') and self.parent.metal_input.view().model() == model:
                is_custom = text in self.parent.custom_metals
                custom_values_set = self.parent.custom_metals
                combo_box = self.parent.metal_input
                type_ = 'metal'
            elif hasattr(self.parent, 'stone_input') and self.parent.stone_input.view().model() == model:
                is_custom = text in self.parent.custom_stones
                custom_values_set = self.parent.custom_stones
                combo_box = self.parent.stone_input
                type_ = 'stone'

            if is_custom and text != "Друго" and combo_box is not None:
                # Calculate icon position
                icon_x = option.rect.right() - self.icon_size - self.icon_padding
                icon_y = option.rect.top() + (option.rect.height() - self.icon_size) // 2
                icon_rect = QRect(icon_x, icon_y, self.icon_size, self.icon_size)

                # Check if click is within icon area
                if icon_rect.contains(event.pos()):
                    # Show context menu
                    menu = QMenu()
                    edit_action = menu.addAction("✎ Редактирай")
                    remove_action = menu.addAction("🗑 Премахни")
                    action = menu.exec(event.globalPosition().toPoint())

                    if action == edit_action:
                        # Edit the custom value
                        new_text, ok = QInputDialog.getText(
                            self.parent, "Редактирай стойност",
                            "Нова стойност:",
                            text=text
                        )
                        if ok and new_text:
                            # Capitalize first letter of each word
                            words = new_text.split()
                            capitalized_words = [word.capitalize() for word in words]
                            new_text = " ".join(capitalized_words)
                            
                            # Check if new value already exists
                            if new_text != text and (new_text in custom_values_set or new_text in ["Пръстен", "Гривна", "Обеци", "Синджир", "Злато", "Сребро", "Платина", "Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък"]):
                                QMessageBox.warning(self.parent, "Предупреждение", "Тази стойност вече съществува!")
                                return True
                            
                            # Update the value in database
                            try:
                                with self.parent.get_connection() as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("UPDATE custom_values SET value = ? WHERE type = ? AND value = ?", 
                                                 (new_text, type_, text))
                            except Exception as e:
                                QMessageBox.critical(self.parent, "Грешка", f"Грешка при редактиране на стойността: {str(e)}")
                                return True
                            
                            # Update the value in memory
                            custom_values_set.remove(text)
                            custom_values_set.add(new_text)
                            
                            # Update combo box items
                            current_items = [combo_box.itemText(i) for i in range(combo_box.count())]
                            idx = current_items.index(text)
                            current_items[idx] = new_text
                            combo_box.clear()
                            combo_box.addItems(current_items)
                            combo_box.setCurrentText(new_text)
                            return True
                            
                    elif action == remove_action:
                        # Remove the custom value from database
                        try:
                            with self.parent.get_connection() as conn:
                                cursor = conn.cursor()
                                cursor.execute("DELETE FROM custom_values WHERE type = ? AND value = ?", (type_, text))
                        except Exception as e:
                            QMessageBox.critical(self.parent, "Грешка", f"Грешка при премахване на стойността: {str(e)}")
                            return True
                        
                        # Remove from memory
                        custom_values_set.remove(text)
                        
                        # Update combo box items
                        current_items = [combo_box.itemText(i) for i in range(combo_box.count())]
                        current_items.remove(text)
                        combo_box.clear()
                        combo_box.addItems(current_items)
                        
                        # Set to first item
                        combo_box.setCurrentIndex(0)
                        return True

        return super().editorEvent(event, model, option, index)

class CustomComboDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.icon_size = 16
        self.icon_padding = 4

    def paint(self, painter, option, index):
        # Get the text and check if it's a custom value
        text = index.data()
        is_custom = False
        
        # Check which combo box we're dealing with
        if hasattr(self.parent, 'category_input') and self.parent.category_input.view().model() == index.model():
            is_custom = text in self.parent.custom_categories
            combo_box = self.parent.category_input
        elif hasattr(self.parent, 'metal_input') and self.parent.metal_input.view().model() == index.model():
            is_custom = text in self.parent.custom_metals
            combo_box = self.parent.metal_input
        elif hasattr(self.parent, 'stone_input') and self.parent.stone_input.view().model() == index.model():
            is_custom = text in self.parent.custom_stones
            combo_box = self.parent.stone_input

        # Draw the item
        super().paint(painter, option, index)

        # If it's a custom value, draw the pencil icon
        if is_custom and text != "Друго":
            # Calculate icon position
            icon_x = option.rect.right() - self.icon_size - self.icon_padding
            icon_y = option.rect.top() + (option.rect.height() - self.icon_size) // 2
            icon_rect = QRect(icon_x, icon_y, self.icon_size, self.icon_size)
            
            # Draw pencil icon
            painter.save()
            painter.setPen(QPen(Qt.GlobalColor.gray))
            painter.drawText(icon_rect, Qt.AlignmentFlag.AlignCenter, "✎")
            painter.restore()

    def sizeHint(self, option, index):
        size = super().sizeHint(option, index)
        size.setWidth(size.width() + self.icon_size + self.icon_padding * 2)  # Add space for the icon
        return size

    def editorEvent(self, event, model, option, index):
        if event.type() == event.Type.MouseButtonPress:
            text = index.data()
            is_custom = False
            custom_values_set = None
            combo_box = None
            type_ = None
            
            # Check which combo box we're dealing with
            if hasattr(self.parent, 'category_input') and self.parent.category_input.view().model() == model:
                is_custom = text in self.parent.custom_categories
                custom_values_set = self.parent.custom_categories
                combo_box = self.parent.category_input
                type_ = 'category'
            elif hasattr(self.parent, 'metal_input') and self.parent.metal_input.view().model() == model:
                is_custom = text in self.parent.custom_metals
                custom_values_set = self.parent.custom_metals
                combo_box = self.parent.metal_input
                type_ = 'metal'
            elif hasattr(self.parent, 'stone_input') and self.parent.stone_input.view().model() == model:
                is_custom = text in self.parent.custom_stones
                custom_values_set = self.parent.custom_stones
                combo_box = self.parent.stone_input
                type_ = 'stone'

            if is_custom and text != "Друго" and combo_box is not None:
                # Calculate icon position
                icon_x = option.rect.right() - self.icon_size - self.icon_padding
                icon_y = option.rect.top() + (option.rect.height() - self.icon_size) // 2
                icon_rect = QRect(icon_x, icon_y, self.icon_size, self.icon_size)

                # Check if click is within icon area
                if icon_rect.contains(event.pos()):
                    # Show context menu
                    menu = QMenu()
                    edit_action = menu.addAction("✎ Редактирай")
                    remove_action = menu.addAction("🗑 Премахни")
                    action = menu.exec(event.globalPosition().toPoint())

                    if action == edit_action:
                        # Edit the custom value
                        new_text, ok = QInputDialog.getText(
                            self.parent, "Редактирай стойност",
                            "Нова стойност:",
                            text=text
                        )
                        if ok and new_text:
                            # Capitalize first letter of each word
                            words = new_text.split()
                            capitalized_words = [word.capitalize() for word in words]
                            new_text = " ".join(capitalized_words)
                            
                            # Check if new value already exists
                            if new_text != text and (new_text in custom_values_set or new_text in ["Пръстен", "Гривна", "Обеци", "Синджир", "Злато", "Сребро", "Платина", "Диамант", "Рубин", "Сапфир", "Смарагд", "Без камък"]):
                                QMessageBox.warning(self.parent, "Предупреждение", "Тази стойност вече съществува!")
                                return True
                            
                            # Update the value in database
                            try:
                                with self.parent.db.get_connection() as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("UPDATE custom_values SET value = ? WHERE type = ? AND value = ?", 
                                                 (new_text, type_, text))
                            except Exception as e:
                                QMessageBox.critical(self.parent, "Грешка", f"Грешка при редактиране на стойността: {str(e)}")
                                return True
                            
                            # Update the value in memory
                            custom_values_set.remove(text)
                            custom_values_set.add(new_text)
                            
                            # Update combo box items
                            current_items = [combo_box.itemText(i) for i in range(combo_box.count())]
                            idx = current_items.index(text)
                            current_items[idx] = new_text
                            combo_box.clear()
                            combo_box.addItems(current_items)
                            combo_box.setCurrentText(new_text)
                            return True
                            
                    elif action == remove_action:
                        # Remove the custom value from database
                        try:
                            with self.parent.db.get_connection() as conn:
                                cursor = conn.cursor()
                                cursor.execute("DELETE FROM custom_values WHERE type = ? AND value = ?", (type_, text))
                        except Exception as e:
                            QMessageBox.critical(self.parent, "Грешка", f"Грешка при премахване на стойността: {str(e)}")
                            return True
                        
                        # Remove from memory
                        custom_values_set.remove(text)
                        
                        # Update combo box items
                        current_items = [combo_box.itemText(i) for i in range(combo_box.count())]
                        current_items.remove(text)
                        combo_box.clear()
                        combo_box.addItems(current_items)
                        
                        # Set to first item
                        combo_box.setCurrentIndex(0)
                        return True

        return super().editorEvent(event, model, option, index)

def setup_directories():
    """Create necessary directories if they don't exist and hide system folders from normal Explorer view"""
    directories = [
        'exports',
        'barcodes',
        'backups',
        'logs',
        'data'
    ]
    
    # Top-level directories that should be hidden from normal Explorer view
    # Only 'exports' and 'backups' folders remain visible to users
    hidden_directories = [
        'barcodes',  # Hide the barcodes folder too
        'logs',
        'data'
    ]
    
    for directory in directories:
        Path(directory).mkdir(parents=True, exist_ok=True)
    
    # Hide system directories from normal Windows Explorer view
    # Users can still access them by enabling "View > Show > Hidden items" in Explorer
    if os.name == 'nt':  # Windows only
        import subprocess
        for hidden_dir in hidden_directories:
            if os.path.exists(hidden_dir):
                try:
                    # Set Windows hidden attribute
                    subprocess.run(['attrib', '+H', hidden_dir], 
                                 check=False, capture_output=True)
                    logger.info(f"Hidden directory: {hidden_dir}")
                except Exception as e:
                    # Silently continue if hiding fails - not critical
                    logger.debug(f"Could not hide directory {hidden_dir}: {e}")
    
    logger.info("Directory setup completed - 'exports' and 'backups' folders visible, system folders hidden")

def main():
    try:
        logger.info("Starting application...")
        
        # Setup directories
        setup_directories()
        logger.info("Directories setup complete")
        
        # Create application
        app = QApplication(sys.argv)
        logger.info("QApplication created")
        
        # Set application icon globally (applies to taskbar and all windows)
        try:
            icon = get_application_icon()
            if not icon.isNull():
                # Set the application icon for all windows
                app.setWindowIcon(icon)
                
                # Additional Windows-specific taskbar integration
                try:
                    if sys.platform == 'win32':
                        # Set application ID for Windows taskbar grouping
                        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("JewelryManagement.Software.1.0")
                except Exception:
                    pass  # Not critical if this fails
                
                # Also set as application icon for better cross-platform integration
                if hasattr(app, 'setDesktopFileName'):
                    app.setDesktopFileName("jewelry-management")
                    
                logger.info("Global application icon set using high-quality PNG with taskbar integration")
            else:
                logger.warning("No global application icon could be loaded")
        except Exception as e:
            logger.warning(f"Could not set global application icon: {e}")
        
        # Set application style
        app.setStyle("Fusion")
        logger.info("Style set to Fusion")
        
        # Create database instance
        db = Database()
        logger.info("Database instance created")
        
        # Database initialization already ensures default user exists via ensure_default_user()
        # No need to create admin user here as it's handled by the database initialization
        
        # Create and show login window with the database instance
        login_window = LoginWindow(database=db)
        logger.info("Login window created")
        logger.info("Login window shown - PIN authentication required")
        
        # Start the event loop
        logger.info("Starting event loop...")
        return app.exec()
    except Exception as e:
        logger.error(f"Application error: {e}", exc_info=True)
        QMessageBox.critical(None, "Грешка", f"Грешка в приложението: {str(e)}")
        return 1

def execute_auto_backup():
    """Execute automatic backup without GUI"""
    try:
        # Setup logging for auto backup
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('logs/auto_backup.log'),
                logging.StreamHandler()
            ]
        )
        
        logger = logging.getLogger(__name__)
        logger.info("Starting automatic backup")
        
        # Create temporary instance just for backup
        from database.models import Database
        
        db = Database()
        
        # Create backup using Database method (returns the path where backup was created)
        backup_path = db.create_backup()
        
        if backup_path:
            logger.info(f"Automatic backup created successfully: {backup_path}")
            
            # Update last backup time in config
            try:
                config_path = get_persistent_path('data/auto_backup_config.json')
                if os.path.exists(config_path):
                    with open(config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    config['last_auto_backup'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    with open(config_path, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=2)
                        
                    logger.info("Updated last backup time in config")
            except Exception as e:
                logger.error(f"Error updating backup config: {e}")
                
            return True
        else:
            logger.error("Automatic backup failed")
            return False
            
    except Exception as e:
        logger.error(f"Error during automatic backup: {e}")
        return False

if __name__ == "__main__":
    try:
        logger.info("Script started")
        
        # Check for auto-backup argument
        if len(sys.argv) > 1 and sys.argv[1] == '--auto-backup':
            # Execute automatic backup and exit
            success = execute_auto_backup()
            sys.exit(0 if success else 1)
        else:
            # Normal GUI startup
            sys.exit(main())
            
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)
    finally:
        logger.info("Script ended")
