import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self, database_or_output_dir="reports"):
        """Initialize ReportGenerator with either a Database object or output directory path"""
        # Handle both Database object and string path for backward compatibility
        if hasattr(database_or_output_dir, 'get_all_items'):  # It's a Database object
            self.database = database_or_output_dir
            self.output_dir = Path("reports")
        else:  # It's a string path
            self.database = None
            self.output_dir = Path(database_or_output_dir)
        
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_inventory_report(self):
        """Generate inventory report - returns list of items"""
        if self.database:
            return self.database.get_all_items()
        else:
            # For backward compatibility, return empty list
            logger.warning("No database provided to ReportGenerator")
            return []

    def generate_low_stock_report(self, threshold=5):
        """Generate low stock report - returns list of items below threshold"""
        if self.database:
            all_items = self.database.get_all_items()
            # Assuming stock_quantity is at index 10 (based on get_all_items structure)
            low_stock_items = [item for item in all_items if item[10] < threshold]
            return low_stock_items
        else:
            logger.warning("No database provided to ReportGenerator")
            return []

    def generate_value_report(self):
        """Generate value report - returns dictionary with total value information"""
        if self.database:
            all_items = self.database.get_all_items()
            total_value = 0
            total_cost = 0
            total_items = len(all_items)
            
            for item in all_items:
                # Assuming price is at index 5, cost at index 6, stock_quantity at index 10
                price = item[5] or 0
                cost = item[6] or 0
                quantity = item[10] or 0
                
                total_value += price * quantity
                total_cost += cost * quantity
            
            return {
                'total_value': total_value,
                'total_cost': total_cost,
                'total_items': total_items,
                'profit_margin': total_value - total_cost if total_value > 0 else 0
            }
        else:
            logger.warning("No database provided to ReportGenerator")
            return {'total_value': 0, 'total_cost': 0, 'total_items': 0, 'profit_margin': 0}

    def generate_sales_report(self, sales_data, start_date=None, end_date=None):
        """Generate sales report in Excel format"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Продажби"

            # Add title
            ws['A1'] = "Отчет за продажби"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:E1')

            # Add date range
            date_range = ""
            if start_date and end_date:
                date_range = f"Период: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            ws['A2'] = date_range
            ws.merge_cells('A2:E2')

            # Add headers
            headers = ["Дата", "Баркод", "Артикул", "Количество", "Общо"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Add data
            row = 5
            total_sales = 0
            for sale in sales_data:
                ws.cell(row=row, column=1).value = sale[4].strftime("%d/%m/%Y")  # Date
                ws.cell(row=row, column=2).value = sale[5]  # Barcode
                ws.cell(row=row, column=3).value = sale[6]  # Name
                ws.cell(row=row, column=4).value = sale[2]  # Quantity
                ws.cell(row=row, column=5).value = sale[3]  # Total price
                total_sales += sale[3]
                row += 1

            # Add total
            ws.cell(row=row, column=1).value = "Общо:"
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=5).value = total_sales
            ws.cell(row=row, column=5).font = Font(bold=True)

            # Adjust column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = self.output_dir / f"sales_report_{timestamp}.xlsx"
            wb.save(filename)
            return str(filename)
        except Exception as e:
            logger.error(f"Error generating sales report: {e}")
            return None

    def generate_inventory_report(self, inventory_data=None):
        """Generate inventory report - can use provided data or fetch from database"""
        if inventory_data is None:
            if self.database:
                inventory_data = self.database.get_all_items()
            else:
                logger.warning("No inventory data provided and no database available")
                return []
        
        # If this was called to generate an Excel report, do that
        if hasattr(self, '_generate_excel_report'):
            return self._generate_inventory_excel_report(inventory_data)
        
        # Otherwise, return the data for testing
        return inventory_data

    def _generate_inventory_excel_report(self, inventory_data):
        """Generate inventory report in Excel format"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Наличности"

            # Add title
            ws['A1'] = "Отчет за наличности"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:H1')

            # Add headers
            headers = ["Баркод", "Име", "Категория", "Цена", "Себестойност", "Тегло", "Метал", "Наличност"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Add data
            row = 4
            total_value = 0
            for item in inventory_data:
                ws.cell(row=row, column=1).value = item[1]  # Barcode
                ws.cell(row=row, column=2).value = item[2]  # Name
                ws.cell(row=row, column=3).value = item[4]  # Category
                ws.cell(row=row, column=4).value = item[5]  # Price
                ws.cell(row=row, column=5).value = item[6]  # Cost
                ws.cell(row=row, column=6).value = item[7]  # Weight
                ws.cell(row=row, column=7).value = item[8]  # Metal type
                ws.cell(row=row, column=8).value = item[11]  # Stock
                total_value += item[5] * item[11]  # Price * Stock
                row += 1

            # Add total
            ws.cell(row=row, column=1).value = "Обща стойност:"
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=4).value = total_value
            ws.cell(row=row, column=4).font = Font(bold=True)

            # Adjust column widths
            for col in range(1, 9):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = self.output_dir / f"inventory_report_{timestamp}.xlsx"
            wb.save(filename)
            return str(filename)
        except Exception as e:
            logger.error(f"Error generating inventory report: {e}")
            return None

    def generate_profit_report(self, sales_data, inventory_data):
        """Generate profit report in Excel format"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Печалба"

            # Add title
            ws['A1'] = "Отчет за печалба"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:F1')

            # Add headers
            headers = ["Дата", "Артикул", "Количество", "Приход", "Разход", "Печалба"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Add data
            row = 4
            total_profit = 0
            for sale in sales_data:
                # Find item in inventory
                item = next((i for i in inventory_data if i[0] == sale[1]), None)
                if item:
                    revenue = sale[3]  # Total price
                    cost = item[6] * sale[2]  # Cost * Quantity
                    profit = revenue - cost
                    total_profit += profit

                    ws.cell(row=row, column=1).value = sale[4].strftime("%d/%m/%Y")  # Date
                    ws.cell(row=row, column=2).value = sale[6]  # Name
                    ws.cell(row=row, column=3).value = sale[2]  # Quantity
                    ws.cell(row=row, column=4).value = revenue
                    ws.cell(row=row, column=5).value = cost
                    ws.cell(row=row, column=6).value = profit
                    row += 1

            # Add total
            ws.cell(row=row, column=1).value = "Обща печалба:"
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=6).value = total_profit
            ws.cell(row=row, column=6).font = Font(bold=True)

            # Adjust column widths
            for col in range(1, 7):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = self.output_dir / f"profit_report_{timestamp}.xlsx"
            wb.save(filename)
            return str(filename)
        except Exception as e:
            logger.error(f"Error generating profit report: {e}")
            return None 